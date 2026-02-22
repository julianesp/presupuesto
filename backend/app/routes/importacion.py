import io
import csv as csv_module
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.services import importacion as svc
from app.auth.dependencies import get_current_user, require_escritura, require_admin
from app.models.tenant import User

router = APIRouter(prefix="/api/importacion", tags=["Importacion"])


def _decode(raw: bytes) -> str:
    """Intenta UTF-8, cae a Latin-1 si falla (archivos Windows)."""
    try:
        return raw.decode("utf-8-sig")  # utf-8-sig elimina BOM si existe
    except UnicodeDecodeError:
        return raw.decode("latin-1")


@router.get("/plantillas/excel")
async def plantilla_excel(user: User = Depends(get_current_user)):
    """Descarga plantilla Excel con hojas GASTOS e INGRESOS lista para diligenciar."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    AZUL = "1E3A5F"
    AZUL_CLARO = "D6E4F0"
    GRIS = "F5F5F5"
    VERDE = "1A5276"
    VERDE_CLARO = "D5F5E3"

    def _estilo_header(ws, fila_headers: list[tuple[str, int]], row=1, color_fondo=AZUL):
        """Aplica estilo a la fila de encabezados y ajusta anchos."""
        thin = Side(style="thin", color="CCCCCC")
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, (titulo, ancho) in enumerate(fila_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=titulo)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=color_fondo)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho
        ws.row_dimensions[row].height = 30

    def _fila_ejemplo(ws, valores: list, row: int, fondo=GRIS):
        thin = Side(style="thin", color="CCCCCC")
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = borde
            cell.fill = PatternFill("solid", fgColor=fondo)
            if isinstance(val, (int, float)) and val > 0:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:  # código
                cell.alignment = Alignment(horizontal="center")

    wb = openpyxl.Workbook()

    # ── Hoja GASTOS ─────────────────────────────────────────────────────────
    ws_g = wb.active
    ws_g.title = "GASTOS"
    ws_g.freeze_panes = "A3"

    # Título
    ws_g.merge_cells("A1:I1")
    titulo_g = ws_g["A1"]
    titulo_g.value = "PLAN PRESUPUESTAL DE GASTOS — Plantilla de carga"
    titulo_g.font = Font(bold=True, color="FFFFFF", size=12)
    titulo_g.fill = PatternFill("solid", fgColor=AZUL)
    titulo_g.alignment = Alignment(horizontal="center", vertical="center")
    ws_g.row_dimensions[1].height = 22

    # Encabezados fila 2 (coincide con columnas que lee el importador: B=código, C=cuenta, I=aprop)
    headers_g = [
        ("N°", 5),
        ("CÓDIGO *", 14),
        ("NOMBRE DE LA CUENTA *", 45),
        ("", 5), ("", 5), ("", 5), ("", 5), ("", 5),
        ("APROPIACIÓN INICIAL *", 20),
    ]
    _estilo_header(ws_g, headers_g, row=2, color_fondo=AZUL)

    # Nota en columnas D-H
    ws_g.merge_cells("D2:H2")
    nota_g = ws_g["D2"]
    nota_g.value = "← columnas sin uso (dejar vacías o con cualquier valor)"
    nota_g.font = Font(italic=True, color="888888", size=8)
    nota_g.alignment = Alignment(horizontal="center", vertical="center")

    # Datos de ejemplo
    ejemplos_g = [
        [1, "2", "GASTOS DE FUNCIONAMIENTO", "", "", "", "", "", 0],
        [2, "2.1", "SERVICIOS PERSONALES ASOCIADOS A LA NÓMINA", "", "", "", "", "", 0],
        [3, "2.1.1", "SUELDOS Y SALARIOS", "", "", "", "", "", 150_000_000],
        [4, "2.1.2", "HORAS EXTRAS Y FESTIVOS", "", "", "", "", "", 8_000_000],
        [5, "2.1.3", "PRIMA DE SERVICIOS", "", "", "", "", "", 12_500_000],
        [6, "2.2", "ADQUISICIÓN DE BIENES Y SERVICIOS", "", "", "", "", "", 0],
        [7, "2.2.1", "MATERIALES Y SUMINISTROS", "", "", "", "", "", 25_000_000],
        [8, "2.2.2", "MANTENIMIENTO Y REPARACIONES", "", "", "", "", "", 15_000_000],
    ]
    for i, fila in enumerate(ejemplos_g, 3):
        _fila_ejemplo(ws_g, fila, row=i, fondo="FFFFFF" if i % 2 == 1 else GRIS)

    # Instrucciones al final
    fila_inst_g = 3 + len(ejemplos_g) + 1
    ws_g.merge_cells(f"A{fila_inst_g}:I{fila_inst_g}")
    inst_g = ws_g.cell(row=fila_inst_g, column=1,
        value="✏ INSTRUCCIONES: Columnas marcadas con * son obligatorias. "
              "Para rubros padre (agrupadores) coloca apropiación 0. "
              "No borres la fila de encabezados (fila 2). "
              "Guarda como .xlsx antes de cargar.")
    inst_g.font = Font(italic=True, color="555555", size=9)
    inst_g.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    inst_g.alignment = Alignment(wrap_text=True, vertical="center")
    ws_g.row_dimensions[fila_inst_g].height = 40

    # ── Hoja INGRESOS ────────────────────────────────────────────────────────
    ws_i = wb.create_sheet("INGRESOS")
    ws_i.freeze_panes = "A3"

    ws_i.merge_cells("A1:G1")
    titulo_i = ws_i["A1"]
    titulo_i.value = "PLAN PRESUPUESTAL DE INGRESOS — Plantilla de carga"
    titulo_i.font = Font(bold=True, color="FFFFFF", size=12)
    titulo_i.fill = PatternFill("solid", fgColor=VERDE)
    titulo_i.alignment = Alignment(horizontal="center", vertical="center")
    ws_i.row_dimensions[1].height = 22

    headers_i = [
        ("N°", 5),
        ("CÓDIGO *", 14),
        ("NOMBRE DE LA CUENTA *", 45),
        ("", 5), ("", 5), ("", 5),
        ("PRESUPUESTO INICIAL *", 22),
    ]
    _estilo_header(ws_i, headers_i, row=2, color_fondo=VERDE)

    ws_i.merge_cells("D2:F2")
    nota_i = ws_i["D2"]
    nota_i.value = "← columnas sin uso"
    nota_i.font = Font(italic=True, color="888888", size=8)
    nota_i.alignment = Alignment(horizontal="center", vertical="center")

    ejemplos_i = [
        [1, "1", "INGRESOS CORRIENTES", "", "", "", 0],
        [2, "1.1", "INGRESOS TRIBUTARIOS", "", "", "", 0],
        [3, "1.1.1", "IMPUESTO PREDIAL UNIFICADO", "", "", "", 45_000_000],
        [4, "1.1.2", "INDUSTRIA Y COMERCIO", "", "", "", 30_000_000],
        [5, "1.2", "TRANSFERENCIAS CORRIENTES", "", "", "", 0],
        [6, "1.2.1", "RECURSOS SGP FUNCIONAMIENTO", "", "", "", 180_000_000],
        [7, "1.2.2", "RECURSOS SGP CALIDAD", "", "", "", 25_000_000],
        [8, "1.3", "RECURSOS PROPIOS", "", "", "", 12_000_000],
    ]
    for i, fila in enumerate(ejemplos_i, 3):
        _fila_ejemplo(ws_i, fila, row=i, fondo="FFFFFF" if i % 2 == 1 else GRIS)

    fila_inst_i = 3 + len(ejemplos_i) + 1
    ws_i.merge_cells(f"A{fila_inst_i}:G{fila_inst_i}")
    inst_i = ws_i.cell(row=fila_inst_i, column=1,
        value="✏ INSTRUCCIONES: Columnas marcadas con * son obligatorias. "
              "Para rubros agrupadores coloca presupuesto 0. "
              "No borres la fila de encabezados (fila 2). "
              "Guarda como .xlsx antes de cargar.")
    inst_i.font = Font(italic=True, color="555555", size=9)
    inst_i.fill = PatternFill("solid", fgColor=VERDE_CLARO)
    inst_i.alignment = Alignment(wrap_text=True, vertical="center")
    ws_i.row_dimensions[fila_inst_i].height = 40

    # ── Hoja TERCEROS ────────────────────────────────────────────────────────
    ws_t = wb.create_sheet("TERCEROS (CSV)")
    ws_t.merge_cells("A1:J1")
    titulo_t = ws_t["A1"]
    titulo_t.value = "REFERENCIA — Formato CSV para Terceros (no se importa esta hoja)"
    titulo_t.font = Font(bold=True, color="FFFFFF", size=11)
    titulo_t.fill = PatternFill("solid", fgColor="7D3C98")
    titulo_t.alignment = Alignment(horizontal="center", vertical="center")
    ws_t.row_dimensions[1].height = 22

    headers_t = [
        ("NIT *", 14), ("DV", 5), ("NOMBRE *", 35), ("DIRECCIÓN", 25),
        ("TELÉFONO", 14), ("EMAIL", 25), ("TIPO", 12),
        ("BANCO", 20), ("TIPO CUENTA", 14), ("No. CUENTA", 20),
    ]
    _estilo_header(ws_t, headers_t, row=2, color_fondo="7D3C98")

    ejemplos_t = [
        ["900123456", "7", "EMPRESA EJEMPLO SAS", "Calle 1 # 2-3", "3001234567", "info@ejemplo.co", "Juridica", "BANCOLOMBIA", "CORRIENTE", "12345678901"],
        ["12345678", "9", "JUAN PEREZ GARCIA", "Carrera 5 # 10-20", "3209876543", "", "Natural", "DAVIVIENDA", "AHORROS", "98765432100"],
        ["70000001", "5", "MARIA LOPEZ RUIZ", "", "", "", "Natural", "", "", ""],
    ]
    for i, fila in enumerate(ejemplos_t, 3):
        _fila_ejemplo(ws_t, fila, row=i, fondo="FFFFFF" if i % 2 == 1 else GRIS)

    fila_inst_t = 3 + len(ejemplos_t) + 1
    ws_t.merge_cells(f"A{fila_inst_t}:J{fila_inst_t}")
    inst_t = ws_t.cell(row=fila_inst_t, column=1,
        value="✏ Para cargar terceros usa un archivo CSV separado con ; (punto y coma). "
              "Esta hoja es solo de referencia del formato. "
              "Columnas opcionales pueden dejarse vacías.")
    inst_t.font = Font(italic=True, color="555555", size=9)
    inst_t.fill = PatternFill("solid", fgColor="E8DAEF")
    inst_t.alignment = Alignment(wrap_text=True, vertical="center")
    ws_t.row_dimensions[fila_inst_t].height = 40

    # ── Serializar y retornar ────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_presupuestal.xlsx"},
    )


@router.post("/sincronizar-padres")
async def sincronizar_padres(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_escritura),
):
    """Recalcula los totales de los rubros agrupadores sumando sus hojas hijas."""
    from app.services import rubros_gastos, rubros_ingresos
    await rubros_gastos._recalcular_hojas(db, user.tenant_id)
    await rubros_ingresos._recalcular_hojas(db, user.tenant_id)
    await rubros_gastos.sincronizar_padres(db, user.tenant_id)
    await rubros_ingresos.sincronizar_padres(db, user.tenant_id)
    return {"ok": True, "mensaje": "Rubros agrupadores sincronizados correctamente"}


@router.post("/catalogo-excel")
async def importar_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_escritura),
):
    try:
        content = await file.read()
        result = await svc.importar_catalogo_excel(db, user.tenant_id, content)
        return result
    except Exception as e:
        raise HTTPException(400, f"Error al importar: {e}")


@router.post("/csv/rubros-gastos")
async def importar_rubros_gastos(
    file: UploadFile = File(...),
    separador: str = Query(default=";", description="Separador de columnas (ej: ; o ,)"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_escritura),
):
    try:
        content = _decode(await file.read())
        return await svc.importar_rubros_gastos_csv(db, user.tenant_id, content, separador)
    except Exception as e:
        raise HTTPException(400, f"Error al importar: {e}")


@router.post("/csv/rubros-ingresos")
async def importar_rubros_ingresos(
    file: UploadFile = File(...),
    separador: str = Query(default=";", description="Separador de columnas (ej: ; o ,)"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_escritura),
):
    try:
        content = _decode(await file.read())
        return await svc.importar_rubros_ingresos_csv(db, user.tenant_id, content, separador)
    except Exception as e:
        raise HTTPException(400, f"Error al importar: {e}")


@router.post("/csv/terceros")
async def importar_terceros(
    file: UploadFile = File(...),
    separador: str = Query(default=";", description="Separador de columnas (ej: ; o ,)"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_escritura),
):
    try:
        content = _decode(await file.read())
        return await svc.importar_terceros_csv(db, user.tenant_id, content, separador)
    except Exception as e:
        raise HTTPException(400, f"Error al importar: {e}")
