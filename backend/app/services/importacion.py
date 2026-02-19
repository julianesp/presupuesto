import io
import csv
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.rubros import RubroGasto, RubroIngreso
from app.models.terceros import Tercero
from app.models.conceptos import Concepto
from app.services import rubros_gastos, rubros_ingresos


async def importar_catalogo_excel(db: AsyncSession, file_content: bytes) -> dict:
    """Import budget catalog from Excel file."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)

    def buscar_hoja(nombre):
        for sheet_name in wb.sheetnames:
            if nombre.lower() in sheet_name.lower():
                return wb[sheet_name]
        return None

    skip_words = {"total", "codigo", "cuenta", "presupuesto", "desequilibrio"}
    count_gastos = count_ingresos = 0
    total_gastos = total_ingresos = 0

    # Import gastos
    ws_gastos = buscar_hoja("GASTOS")
    if ws_gastos:
        for row in ws_gastos.iter_rows(min_row=2):
            codigo = str(row[1].value or "").strip()
            if not codigo or any(w in codigo.lower() for w in skip_words):
                continue
            cuenta = str(row[2].value or "").strip()
            if not cuenta or any(w in cuenta.lower() for w in skip_words):
                continue
            aprop_def = float(row[8].value or 0) if len(row) > 8 and row[8].value else 0

            all_rubros = await rubros_gastos.get_rubros(db)
            codigos = {r.codigo for r in all_rubros}
            has_children = any(c.startswith(codigo + ".") for c in codigos)

            rubro = RubroGasto(
                codigo=codigo, cuenta=cuenta, es_hoja=0 if has_children else 1,
                apropiacion_inicial=aprop_def, apropiacion_definitiva=aprop_def,
            )
            await db.merge(rubro)
            count_gastos += 1
            if not has_children:
                total_gastos += aprop_def

    # Import ingresos
    ws_ingresos = buscar_hoja("INGRESOS")
    if ws_ingresos:
        for row in ws_ingresos.iter_rows(min_row=2):
            codigo = str(row[1].value or "").strip()
            if not codigo or any(w in codigo.lower() for w in skip_words):
                continue
            cuenta = str(row[2].value or "").strip()
            if not cuenta or any(w in cuenta.lower() for w in skip_words):
                continue
            ppto_def = float(row[6].value or 0) if len(row) > 6 and row[6].value else 0

            all_rubros = await rubros_ingresos.get_rubros(db)
            codigos = {r.codigo for r in all_rubros}
            has_children = any(c.startswith(codigo + ".") for c in codigos)

            rubro = RubroIngreso(
                codigo=codigo, cuenta=cuenta, es_hoja=0 if has_children else 1,
                presupuesto_inicial=ppto_def, presupuesto_definitivo=ppto_def,
            )
            await db.merge(rubro)
            count_ingresos += 1
            if not has_children:
                total_ingresos += ppto_def

    await db.commit()

    # Recalculate leaf flags
    await rubros_gastos._recalcular_hojas(db)
    await rubros_ingresos._recalcular_hojas(db)

    return {
        "rubros_gastos": count_gastos,
        "rubros_ingresos": count_ingresos,
        "total_gastos": total_gastos,
        "total_ingresos": total_ingresos,
        "diferencia": total_ingresos - total_gastos,
    }


async def importar_rubros_gastos_csv(db: AsyncSession, content: str, separador: str = ";") -> dict:
    reader = csv.reader(io.StringIO(content), delimiter=separador)
    cantidad = 0
    errores = []
    for i, row in enumerate(reader, 1):
        try:
            if len(row) < 3:
                continue
            codigo, cuenta, aprop_ini = row[0].strip(), row[1].strip(), float(row[2].strip() or "0")
            if not codigo or not cuenta:
                continue
            rubro = RubroGasto(
                codigo=codigo, cuenta=cuenta, es_hoja=1,
                apropiacion_inicial=aprop_ini, apropiacion_definitiva=aprop_ini,
            )
            await db.merge(rubro)
            cantidad += 1
        except Exception as e:
            errores.append(f"Fila {i}: {e}")
    await db.commit()
    await rubros_gastos._recalcular_hojas(db)
    return {"cantidad": cantidad, "errores": errores}


async def importar_rubros_ingresos_csv(db: AsyncSession, content: str, separador: str = ";") -> dict:
    reader = csv.reader(io.StringIO(content), delimiter=separador)
    cantidad = 0
    errores = []
    for i, row in enumerate(reader, 1):
        try:
            if len(row) < 3:
                continue
            codigo, cuenta, ppto_ini = row[0].strip(), row[1].strip(), float(row[2].strip() or "0")
            if not codigo or not cuenta:
                continue
            rubro = RubroIngreso(
                codigo=codigo, cuenta=cuenta, es_hoja=1,
                presupuesto_inicial=ppto_ini, presupuesto_definitivo=ppto_ini,
            )
            await db.merge(rubro)
            cantidad += 1
        except Exception as e:
            errores.append(f"Fila {i}: {e}")
    await db.commit()
    await rubros_ingresos._recalcular_hojas(db)
    return {"cantidad": cantidad, "errores": errores}


async def importar_terceros_csv(db: AsyncSession, content: str, separador: str = ";") -> dict:
    reader = csv.reader(io.StringIO(content), delimiter=separador)
    cantidad = 0
    errores = []
    for i, row in enumerate(reader, 1):
        try:
            if len(row) < 3:
                continue
            nit = row[0].strip()
            dv = row[1].strip() if len(row) > 1 else ""
            nombre = row[2].strip() if len(row) > 2 else ""
            if not nit or not nombre:
                continue
            tercero = Tercero(
                nit=nit, dv=dv, nombre=nombre.upper(),
                direccion=row[3].strip() if len(row) > 3 else "",
                telefono=row[4].strip() if len(row) > 4 else "",
                email=row[5].strip() if len(row) > 5 else "",
                tipo=row[6].strip() if len(row) > 6 else "Natural",
                banco=row[7].strip() if len(row) > 7 else "",
                tipo_cuenta=row[8].strip() if len(row) > 8 else "",
                no_cuenta=row[9].strip() if len(row) > 9 else "",
            )
            await db.merge(tercero)
            cantidad += 1
        except Exception as e:
            errores.append(f"Fila {i}: {e}")
    await db.commit()
    return {"cantidad": cantidad, "errores": errores}
