from sqlalchemy import select, func, and_, Integer
from sqlalchemy.ext.asyncio import AsyncSession
import csv
import io
import zipfile
from app.models.rubros import RubroGasto, RubroIngreso
from app.models.cdp import CDP
from app.models.rp import RP
from app.models.obligacion import Obligacion
from app.models.pago import Pago
from app.models.recaudo import Recaudo
from app.models.reconocimiento import Reconocimiento
from app.models.cuentas_bancarias import CuentaBancaria
from app.models.modificaciones import ModificacionPresupuestal, DetalleModificacion
from app.models.pac import PAC
from app.services import config as config_svc


def _nivel(codigo: str) -> int:
    return codigo.count(".") + 1


async def _sum_by_rubro_and_period(db, tenant_id, model, campo_valor, codigo_rubro,
                                    campo_estado, estado_excluir, mes_antes=None, mes_desde=None, mes_hasta=None):
    """Helper to sum values by rubro optionally filtering by month period."""
    mes_col = func.cast(func.substr(model.fecha, 6, 2), Integer)
    conditions = [model.tenant_id == tenant_id, model.codigo_rubro == codigo_rubro]
    if campo_estado:
        conditions.append(getattr(model, campo_estado) != estado_excluir)
    if mes_antes is not None:
        conditions.append(mes_col < mes_antes)
    if mes_desde is not None:
        conditions.append(mes_col >= mes_desde)
    if mes_hasta is not None:
        conditions.append(mes_col <= mes_hasta)

    result = await db.execute(
        select(func.coalesce(func.sum(getattr(model, campo_valor)), 0)).where(and_(*conditions))
    )
    return result.scalar()


async def resumen_rubro(db: AsyncSession, tenant_id: str, codigo_rubro: str,
                        mes_inicio: int = 1, mes_fin: int = 12) -> dict:
    rubro = await db.execute(select(RubroGasto).where(RubroGasto.tenant_id == tenant_id, RubroGasto.codigo == codigo_rubro))
    rubro = rubro.scalar_one_or_none()
    if not rubro:
        raise ValueError(f"Rubro {codigo_rubro} no encontrado")

    # If parent, aggregate children
    if rubro.es_hoja == 0:
        children = await db.execute(
            select(RubroGasto).where(
                and_(RubroGasto.tenant_id == tenant_id, RubroGasto.codigo.like(f"{codigo_rubro}.%"), RubroGasto.es_hoja == 1)
            )
        )
        codigos = [c.codigo for c in children.scalars().all()]
    else:
        codigos = [codigo_rubro]

    totals = {
        "disp_anteriores": 0, "disp_periodo": 0,
        "comp_anteriores": 0, "comp_periodo": 0,
        "obl_anteriores": 0, "obl_periodo": 0,
        "pago_anteriores": 0, "pago_periodo": 0,
    }

    for cod in codigos:
        totals["disp_anteriores"] += await _sum_by_rubro_and_period(
            db, tenant_id, CDP, "valor", cod, "estado", "ANULADO", mes_antes=mes_inicio)
        totals["disp_periodo"] += await _sum_by_rubro_and_period(
            db, tenant_id, CDP, "valor", cod, "estado", "ANULADO", mes_desde=mes_inicio, mes_hasta=mes_fin)
        totals["comp_anteriores"] += await _sum_by_rubro_and_period(
            db, tenant_id, RP, "valor", cod, "estado", "ANULADO", mes_antes=mes_inicio)
        totals["comp_periodo"] += await _sum_by_rubro_and_period(
            db, tenant_id, RP, "valor", cod, "estado", "ANULADO", mes_desde=mes_inicio, mes_hasta=mes_fin)
        totals["obl_anteriores"] += await _sum_by_rubro_and_period(
            db, tenant_id, Obligacion, "valor", cod, "estado", "ANULADA", mes_antes=mes_inicio)
        totals["obl_periodo"] += await _sum_by_rubro_and_period(
            db, tenant_id, Obligacion, "valor", cod, "estado", "ANULADA", mes_desde=mes_inicio, mes_hasta=mes_fin)
        totals["pago_anteriores"] += await _sum_by_rubro_and_period(
            db, tenant_id, Pago, "valor", cod, "estado", "ANULADO", mes_antes=mes_inicio)
        totals["pago_periodo"] += await _sum_by_rubro_and_period(
            db, tenant_id, Pago, "valor", cod, "estado", "ANULADO", mes_desde=mes_inicio, mes_hasta=mes_fin)

    total_disp = totals["disp_anteriores"] + totals["disp_periodo"]
    total_comp = totals["comp_anteriores"] + totals["comp_periodo"]
    total_obl = totals["obl_anteriores"] + totals["obl_periodo"]
    total_pago = totals["pago_anteriores"] + totals["pago_periodo"]

    return {
        "rubro": {"codigo": rubro.codigo, "cuenta": rubro.cuenta},
        "apropiacion_inicial": rubro.apropiacion_inicial,
        "adiciones": rubro.adiciones,
        "reducciones": rubro.reducciones,
        "creditos": rubro.creditos,
        "contracreditos": rubro.contracreditos,
        "apropiacion_definitiva": rubro.apropiacion_definitiva,
        "disp_anteriores": totals["disp_anteriores"],
        "disp_periodo": totals["disp_periodo"],
        "total_disp": total_disp,
        "saldo_disponible": rubro.apropiacion_definitiva - total_disp,
        "disp_sin_compromiso": total_disp - total_comp,
        "comp_anteriores": totals["comp_anteriores"],
        "comp_periodo": totals["comp_periodo"],
        "total_comp": total_comp,
        "comp_sin_obligacion": total_comp - total_obl,
        "obl_anteriores": totals["obl_anteriores"],
        "obl_periodo": totals["obl_periodo"],
        "total_obl": total_obl,
        "obl_x_pagar": total_obl - total_pago,
        "pago_anteriores": totals["pago_anteriores"],
        "pago_periodo": totals["pago_periodo"],
        "total_pago": total_pago,
        "aprop_x_afectar": rubro.apropiacion_definitiva - total_disp,
    }


async def informe_ejecucion_gastos(db: AsyncSession, tenant_id: str, mes_consulta: int | None = None) -> list[dict]:
    if mes_consulta is None:
        mes_consulta = int(await config_svc.get_config(db, tenant_id, "mes_actual") or "1")

    rubros_result = await db.execute(select(RubroGasto).where(RubroGasto.tenant_id == tenant_id).order_by(RubroGasto.codigo))
    rubros = rubros_result.scalars().all()
    filas = []

    for r in rubros:
        if r.es_hoja == 1:
            codigos = [r.codigo]
        else:
            children = await db.execute(
                select(RubroGasto).where(
                    and_(RubroGasto.tenant_id == tenant_id, RubroGasto.codigo.like(f"{r.codigo}.%"), RubroGasto.es_hoja == 1)
                )
            )
            codigos = [c.codigo for c in children.scalars().all()]

        comp_ant = comp_mes = pago_ant = pago_mes = 0
        for cod in codigos:
            comp_ant += await _sum_by_rubro_and_period(db, tenant_id, RP, "valor", cod, "estado", "ANULADO", mes_antes=mes_consulta)
            comp_mes += await _sum_by_rubro_and_period(db, tenant_id, RP, "valor", cod, "estado", "ANULADO", mes_desde=mes_consulta, mes_hasta=mes_consulta)
            pago_ant += await _sum_by_rubro_and_period(db, tenant_id, Pago, "valor", cod, "estado", "ANULADO", mes_antes=mes_consulta)
            pago_mes += await _sum_by_rubro_and_period(db, tenant_id, Pago, "valor", cod, "estado", "ANULADO", mes_desde=mes_consulta, mes_hasta=mes_consulta)

        comp_acum = comp_ant + comp_mes
        pago_acum = pago_ant + pago_mes

        filas.append({
            "codigo": r.codigo,
            "cuenta": r.cuenta,
            "es_hoja": r.es_hoja,
            "nivel": _nivel(r.codigo),
            "ppto_inicial": r.apropiacion_inicial,
            "adiciones": r.adiciones,
            "reducciones": r.reducciones,
            "creditos": r.creditos,
            "contracreditos": r.contracreditos,
            "ppto_definitivo": r.apropiacion_definitiva,
            "comp_anterior": comp_ant,
            "comp_mes": comp_mes,
            "comp_acumulado": comp_acum,
            "pago_anterior": pago_ant,
            "pago_mes": pago_mes,
            "pago_acumulado": pago_acum,
            "saldo_apropiacion": r.apropiacion_definitiva - comp_acum,
            "saldo_comp_pagar": comp_acum - pago_acum,
        })

    return filas


async def informe_ejecucion_ingresos(db: AsyncSession, tenant_id: str, mes_consulta: int | None = None) -> list[dict]:
    if mes_consulta is None:
        mes_consulta = int(await config_svc.get_config(db, tenant_id, "mes_actual") or "1")

    rubros_result = await db.execute(select(RubroIngreso).where(RubroIngreso.tenant_id == tenant_id).order_by(RubroIngreso.codigo))
    rubros = rubros_result.scalars().all()
    filas = []

    for r in rubros:
        if r.es_hoja == 1:
            codigos = [r.codigo]
        else:
            children = await db.execute(
                select(RubroIngreso).where(
                    and_(RubroIngreso.tenant_id == tenant_id, RubroIngreso.codigo.like(f"{r.codigo}.%"), RubroIngreso.es_hoja == 1)
                )
            )
            codigos = [c.codigo for c in children.scalars().all()]

        rec_ant = rec_mes = 0
        recon_ant = recon_mes = 0
        for cod in codigos:
            rec_ant   += await _sum_by_rubro_and_period(db, tenant_id, Recaudo,        "valor", cod, "estado", "ANULADO",  mes_antes=mes_consulta)
            rec_mes   += await _sum_by_rubro_and_period(db, tenant_id, Recaudo,        "valor", cod, "estado", "ANULADO",  mes_desde=mes_consulta, mes_hasta=mes_consulta)
            recon_ant += await _sum_by_rubro_and_period(db, tenant_id, Reconocimiento, "valor", cod, "estado", "ANULADO",  mes_antes=mes_consulta)
            recon_mes += await _sum_by_rubro_and_period(db, tenant_id, Reconocimiento, "valor", cod, "estado", "ANULADO",  mes_desde=mes_consulta, mes_hasta=mes_consulta)

        rec_acum   = rec_ant   + rec_mes
        recon_acum = recon_ant + recon_mes

        filas.append({
            "codigo": r.codigo,
            "cuenta": r.cuenta,
            "es_hoja": r.es_hoja,
            "nivel": _nivel(r.codigo),
            "ppto_inicial": r.presupuesto_inicial,
            "adiciones": r.adiciones,
            "reducciones": r.reducciones,
            "ppto_definitivo": r.presupuesto_definitivo,
            "recon_anterior": recon_ant,
            "recon_mes": recon_mes,
            "recon_acumulado": recon_acum,
            "recaudo_anterior": rec_ant,
            "recaudo_mes": rec_mes,
            "recaudo_acumulado": rec_acum,
            "saldo_por_recaudar": r.presupuesto_definitivo - rec_acum,
        })

    return filas


async def generar_tarjeta(db: AsyncSession, tenant_id: str, codigo_rubro: str) -> dict:
    rubro = await db.execute(select(RubroGasto).where(RubroGasto.tenant_id == tenant_id, RubroGasto.codigo == codigo_rubro))
    rubro = rubro.scalar_one_or_none()
    if not rubro:
        raise ValueError(f"Rubro {codigo_rubro} no encontrado")

    movimientos = []

    cdps = await db.execute(
        select(CDP).where(and_(CDP.tenant_id == tenant_id, CDP.codigo_rubro == codigo_rubro, CDP.estado != "ANULADO"))
        .order_by(CDP.fecha)
    )
    for c in cdps.scalars().all():
        movimientos.append({
            "fecha": c.fecha, "tipo": "CDP", "numero": c.numero,
            "nit": "", "tercero": "", "concepto": c.objeto,
            "v_cdp": c.valor, "v_rp": 0, "v_obl": 0, "v_pago": 0,
        })

    rps = await db.execute(
        select(RP).where(and_(RP.tenant_id == tenant_id, RP.codigo_rubro == codigo_rubro, RP.estado != "ANULADO"))
        .order_by(RP.fecha)
    )
    for r in rps.scalars().all():
        tercero = r.tercero
        movimientos.append({
            "fecha": r.fecha, "tipo": "RP", "numero": r.numero,
            "nit": r.nit_tercero, "tercero": tercero.nombre if tercero else "",
            "concepto": r.objeto,
            "v_cdp": 0, "v_rp": r.valor, "v_obl": 0, "v_pago": 0,
        })

    obls = await db.execute(
        select(Obligacion).where(
            and_(Obligacion.tenant_id == tenant_id, Obligacion.codigo_rubro == codigo_rubro, Obligacion.estado != "ANULADA")
        ).order_by(Obligacion.fecha)
    )
    for o in obls.scalars().all():
        tercero = o.tercero
        movimientos.append({
            "fecha": o.fecha, "tipo": "OBL", "numero": o.numero,
            "nit": o.nit_tercero, "tercero": tercero.nombre if tercero else "",
            "concepto": o.factura,
            "v_cdp": 0, "v_rp": 0, "v_obl": o.valor, "v_pago": 0,
        })

    pagos = await db.execute(
        select(Pago).where(
            and_(Pago.tenant_id == tenant_id, Pago.codigo_rubro == codigo_rubro, Pago.estado != "ANULADO")
        ).order_by(Pago.fecha)
    )
    for p in pagos.scalars().all():
        tercero = p.tercero
        movimientos.append({
            "fecha": p.fecha, "tipo": "PAGO", "numero": p.numero,
            "nit": p.nit_tercero, "tercero": tercero.nombre if tercero else "",
            "concepto": p.concepto,
            "v_cdp": 0, "v_rp": 0, "v_obl": 0, "v_pago": p.valor,
        })

    movimientos.sort(key=lambda m: (m["fecha"], ["CDP", "RP", "OBL", "PAGO"].index(m["tipo"])))

    return {
        "rubro": {
            "codigo": rubro.codigo, "cuenta": rubro.cuenta,
            "apropiacion_definitiva": rubro.apropiacion_definitiva,
        },
        "movimientos": movimientos,
    }


async def informe_cadena_presupuestal(db: AsyncSession, tenant_id: str) -> list[dict]:
    cdps = await db.execute(
        select(CDP).where(CDP.tenant_id == tenant_id, CDP.estado != "ANULADO").order_by(CDP.numero)
    )
    resultado = []

    for cdp in cdps.scalars().all():
        rps_data = []
        rps = await db.execute(
            select(RP).where(and_(RP.tenant_id == tenant_id, RP.cdp_numero == cdp.numero, RP.estado != "ANULADO"))
            .order_by(RP.numero)
        )
        for rp in rps.scalars().all():
            obls_data = []
            obls = await db.execute(
                select(Obligacion).where(
                    and_(Obligacion.tenant_id == tenant_id, Obligacion.rp_numero == rp.numero, Obligacion.estado != "ANULADA")
                ).order_by(Obligacion.numero)
            )
            for obl in obls.scalars().all():
                pagos_data = []
                pagos = await db.execute(
                    select(Pago).where(
                        and_(Pago.tenant_id == tenant_id, Pago.obligacion_numero == obl.numero, Pago.estado != "ANULADO")
                    ).order_by(Pago.numero)
                )
                for pag in pagos.scalars().all():
                    pagos_data.append({
                        "numero": pag.numero, "fecha": pag.fecha,
                        "valor": pag.valor, "concepto": pag.concepto,
                    })
                obls_data.append({
                    "obligacion": {
                        "numero": obl.numero, "fecha": obl.fecha,
                        "valor": obl.valor, "factura": obl.factura, "estado": obl.estado,
                    },
                    "pagos": pagos_data,
                })
            rps_data.append({
                "rp": {
                    "numero": rp.numero, "fecha": rp.fecha,
                    "nit_tercero": rp.nit_tercero, "valor": rp.valor,
                    "objeto": rp.objeto, "estado": rp.estado,
                },
                "obligaciones": obls_data,
            })
        resultado.append({
            "cdp": {
                "numero": cdp.numero, "fecha": cdp.fecha,
                "codigo_rubro": cdp.codigo_rubro, "valor": cdp.valor,
                "objeto": cdp.objeto, "estado": cdp.estado,
            },
            "rps": rps_data,
        })

    return resultado


async def get_resumen(db: AsyncSession, tenant_id: str) -> dict:
    # Total apropiacion
    result = await db.execute(
        select(func.coalesce(func.sum(RubroGasto.apropiacion_definitiva), 0))
        .where(RubroGasto.tenant_id == tenant_id, RubroGasto.es_hoja == 1)
    )
    total_aprop = result.scalar()

    # Total CDPs
    result = await db.execute(
        select(func.coalesce(func.sum(CDP.valor), 0)).where(CDP.tenant_id == tenant_id, CDP.estado != "ANULADO")
    )
    total_cdp = result.scalar()

    # Total comprometido (RP)
    result = await db.execute(
        select(func.coalesce(func.sum(RP.valor), 0)).where(RP.tenant_id == tenant_id, RP.estado != "ANULADO")
    )
    total_comp = result.scalar()

    # Total obligado
    result = await db.execute(
        select(func.coalesce(func.sum(Obligacion.valor), 0)).where(Obligacion.tenant_id == tenant_id, Obligacion.estado != "ANULADA")
    )
    total_obl = result.scalar()

    # Total pagado
    result = await db.execute(
        select(func.coalesce(func.sum(Pago.valor), 0)).where(Pago.tenant_id == tenant_id, Pago.estado != "ANULADO")
    )
    total_pag = result.scalar()

    # Total ingresos
    result = await db.execute(
        select(func.coalesce(func.sum(RubroIngreso.presupuesto_definitivo), 0))
        .where(RubroIngreso.tenant_id == tenant_id, RubroIngreso.es_hoja == 1)
    )
    total_ing = result.scalar()

    # Total recaudado
    result = await db.execute(
        select(func.coalesce(func.sum(Recaudo.valor), 0)).where(Recaudo.tenant_id == tenant_id, Recaudo.estado != "ANULADO")
    )
    total_rec = result.scalar()

    def _pct(parte, total):
        return round(float(parte) / float(total) * 100, 1) if total and total > 0 else 0.0

    return {
        "apropiacion": total_aprop,
        "cdp": total_cdp,
        "comprometido": total_comp,
        "obligado": total_obl,
        "pagado": total_pag,
        "saldo_disponible": total_aprop - total_cdp,
        "saldo_por_pagar": total_comp - total_pag,
        "ppto_ingresos": total_ing,
        "recaudado": total_rec,
        "saldo_por_recaudar": total_ing - total_rec,
        "equilibrio": total_ing - total_aprop,
        "pct_cdp": _pct(total_cdp, total_aprop),
        "pct_comprometido": _pct(total_comp, total_aprop),
        "pct_obligado": _pct(total_obl, total_aprop),
        "pct_pagado": _pct(total_pag, total_aprop),
        "pct_recaudado": _pct(total_rec, total_ing),
    }


async def informe_sia_gastos(db: AsyncSession, tenant_id: str, mes_consulta: int | None = None) -> list[dict]:
    """Ejecución presupuestal de gastos en formato SIA Contraloría:
    Incluye Compromisos (RP), Obligaciones y Pagos con desagregación anterior/período/acumulado."""
    if mes_consulta is None:
        mes_consulta = int(await config_svc.get_config(db, tenant_id, "mes_actual") or "1")

    rubros_result = await db.execute(select(RubroGasto).where(RubroGasto.tenant_id == tenant_id).order_by(RubroGasto.codigo))
    rubros = rubros_result.scalars().all()
    filas = []

    for r in rubros:
        if r.es_hoja == 1:
            codigos = [r.codigo]
        else:
            children = await db.execute(
                select(RubroGasto).where(
                    and_(RubroGasto.tenant_id == tenant_id, RubroGasto.codigo.like(f"{r.codigo}.%"), RubroGasto.es_hoja == 1)
                )
            )
            codigos = [c.codigo for c in children.scalars().all()]

        comp_ant = comp_mes = obl_ant = obl_mes = pago_ant = pago_mes = 0
        for cod in codigos:
            comp_ant += await _sum_by_rubro_and_period(db, tenant_id, RP, "valor", cod, "estado", "ANULADO", mes_antes=mes_consulta)
            comp_mes += await _sum_by_rubro_and_period(db, tenant_id, RP, "valor", cod, "estado", "ANULADO", mes_desde=mes_consulta, mes_hasta=mes_consulta)
            obl_ant  += await _sum_by_rubro_and_period(db, tenant_id, Obligacion, "valor", cod, "estado", "ANULADA", mes_antes=mes_consulta)
            obl_mes  += await _sum_by_rubro_and_period(db, tenant_id, Obligacion, "valor", cod, "estado", "ANULADA", mes_desde=mes_consulta, mes_hasta=mes_consulta)
            pago_ant += await _sum_by_rubro_and_period(db, tenant_id, Pago, "valor", cod, "estado", "ANULADO", mes_antes=mes_consulta)
            pago_mes += await _sum_by_rubro_and_period(db, tenant_id, Pago, "valor", cod, "estado", "ANULADO", mes_desde=mes_consulta, mes_hasta=mes_consulta)

        comp_acum = comp_ant + comp_mes
        obl_acum  = obl_ant  + obl_mes
        pago_acum = pago_ant + pago_mes

        filas.append({
            "codigo": r.codigo, "cuenta": r.cuenta,
            "es_hoja": r.es_hoja, "nivel": _nivel(r.codigo),
            "ppto_inicial": r.apropiacion_inicial,
            "adiciones": r.adiciones,
            "reducciones": r.reducciones,
            "creditos": r.creditos,
            "contracreditos": r.contracreditos,
            "ppto_definitivo": r.apropiacion_definitiva,
            "comp_anterior": comp_ant, "comp_mes": comp_mes, "comp_acumulado": comp_acum,
            "obl_anterior": obl_ant,  "obl_mes": obl_mes,  "obl_acumulado": obl_acum,
            "pago_anterior": pago_ant, "pago_mes": pago_mes, "pago_acumulado": pago_acum,
            "saldo_x_comprometer": r.apropiacion_definitiva - comp_acum,
            "saldo_x_obligar": comp_acum - obl_acum,
            "saldo_x_pagar": obl_acum - pago_acum,
        })

    return filas


async def generar_sia_excel(
    db: AsyncSession,
    tenant_id: str,
    mes_consulta: int | None = None,
    nombre_institucion: str = "INSTITUCIÓN EDUCATIVA",
    vigencia: int = 2026,
) -> bytes:
    """Genera archivo Excel SIA Contraloría con hojas GASTOS e INGRESOS."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if mes_consulta is None:
        mes_consulta = int(await config_svc.get_config(db, tenant_id, "mes_actual") or "1")

    gastos  = await informe_sia_gastos(db, tenant_id, mes_consulta)
    ingresos = await informe_ejecucion_ingresos(db, tenant_id, mes_consulta)

    MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    periodo = MESES[mes_consulta] if 1 <= mes_consulta <= 12 else str(mes_consulta)

    AZUL   = "1E3A5F"
    AZUL2  = "2E4A6F"
    GRIS_H = "E8EDF2"
    GRIS_F = "F5F7FA"
    VERDE  = "1A5276"
    VERDE2 = "2A6286"
    GRIS_HI = "E8F4F0"

    thin = Side(style="thin", color="CCCCCC")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fmt(val: float) -> str:
        return f"{val:,.0f}" if val else "-"

    def _cell(ws, row, col, value, bold=False, fondo=None, align="right", size=8, color=None, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=color or "000000", italic=italic)
        c.border = borde
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if fondo:
            c.fill = PatternFill("solid", fgColor=fondo)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            c.number_format = '#,##0'
        return c

    wb = openpyxl.Workbook()

    # ─── Hoja GASTOS ──────────────────────────────────────────────────────────
    ws_g = wb.active
    ws_g.title = "GASTOS"
    ws_g.freeze_panes = "C4"

    # Fila 1: título
    ws_g.merge_cells("A1:V1")
    _cell(ws_g, 1, 1,
          f"INFORME SIA CONTRALORÍA — EJECUCIÓN PRESUPUESTAL DE GASTOS — {nombre_institucion} — Vigencia {vigencia}",
          bold=True, fondo=AZUL, align="center", size=10, color="FFFFFF")
    ws_g.row_dimensions[1].height = 22

    # Fila 2: subtítulo período
    ws_g.merge_cells("A2:V2")
    _cell(ws_g, 2, 1, f"Período: {periodo} de {vigencia}  |  Valores en pesos colombianos",
          bold=False, fondo=AZUL2, align="center", size=9, color="FFFFFF", italic=True)
    ws_g.row_dimensions[2].height = 16

    # Fila 3: encabezados
    HDRS_G = [
        ("CÓDIGO", 12, "left"), ("DENOMINACIÓN", 46, "left"),
        ("APROP.\nINICIAL", 13, "right"), ("ADICIONES", 13, "right"),
        ("REDUC-\nCIONES", 13, "right"), ("CRÉDITOS", 13, "right"),
        ("CONTRA-\nCRÉDITOS", 13, "right"), ("APROP.\nDEFINITIVA", 14, "right"),
        ("COMP.\nANTERIOR", 13, "right"), ("COMP.\nPERÍODO", 13, "right"),
        ("COMP.\nACUM.", 14, "right"),
        ("OBL.\nANTERIOR", 13, "right"), ("OBL.\nPERÍODO", 13, "right"),
        ("OBL.\nACUM.", 14, "right"),
        ("PAGOS\nANTERIOR", 13, "right"), ("PAGOS\nPERÍODO", 13, "right"),
        ("PAGOS\nACUM.", 14, "right"),
        ("SALDO X\nCOMPROMETER", 15, "right"),
        ("SALDO X\nOBLIGAR", 14, "right"),
        ("SALDO X\nPAGAR", 14, "right"),
    ]
    for col, (titulo, ancho, al) in enumerate(HDRS_G, 1):
        _cell(ws_g, 3, col, titulo, bold=True, fondo=GRIS_H, align="center", size=8)
        ws_g.column_dimensions[get_column_letter(col)].width = ancho
    ws_g.row_dimensions[3].height = 36

    # Filas de datos
    for i, r in enumerate(gastos):
        row = 4 + i
        es_padre = r["es_hoja"] == 0
        fondo_fila = GRIS_F if es_padre else None
        bold_fila = es_padre

        indent = "  " * (r["nivel"] - 1)
        _cell(ws_g, row, 1, r["codigo"], bold=bold_fila, fondo=fondo_fila, align="left", size=8)
        _cell(ws_g, row, 2, indent + r["cuenta"], bold=bold_fila, fondo=fondo_fila, align="left", size=8)
        for col, key in enumerate([
            "ppto_inicial", "adiciones", "reducciones", "creditos", "contracreditos",
            "ppto_definitivo",
            "comp_anterior", "comp_mes", "comp_acumulado",
            "obl_anterior", "obl_mes", "obl_acumulado",
            "pago_anterior", "pago_mes", "pago_acumulado",
            "saldo_x_comprometer", "saldo_x_obligar", "saldo_x_pagar",
        ], start=3):
            _cell(ws_g, row, col, r[key] or None, bold=bold_fila, fondo=fondo_fila, size=8)
        ws_g.row_dimensions[row].height = 14

    # ─── Hoja INGRESOS ────────────────────────────────────────────────────────
    ws_i = wb.create_sheet("INGRESOS")
    ws_i.freeze_panes = "C4"

    ws_i.merge_cells("A1:L1")
    _cell(ws_i, 1, 1,
          f"INFORME SIA CONTRALORÍA — EJECUCIÓN PRESUPUESTAL DE INGRESOS — {nombre_institucion} — Vigencia {vigencia}",
          bold=True, fondo=VERDE, align="center", size=10, color="FFFFFF")
    ws_i.row_dimensions[1].height = 22

    ws_i.merge_cells("A2:L2")
    _cell(ws_i, 2, 1, f"Período: {periodo} de {vigencia}  |  Valores en pesos colombianos",
          bold=False, fondo=VERDE2, align="center", size=9, color="FFFFFF", italic=True)
    ws_i.row_dimensions[2].height = 16

    HDRS_I = [
        ("CÓDIGO", 12, "left"), ("DENOMINACIÓN", 46, "left"),
        ("PPTO.\nINICIAL", 14, "right"), ("ADICIONES", 13, "right"),
        ("REDUC-\nCIONES", 13, "right"), ("PPTO.\nDEFINITIVO", 14, "right"),
        ("RECAUDOS\nANTERIOR", 14, "right"), ("RECAUDOS\nPERÍODO", 14, "right"),
        ("RECAUDOS\nACUM.", 14, "right"), ("SALDO X\nRECAUDAR", 14, "right"),
    ]
    for col, (titulo, ancho, al) in enumerate(HDRS_I, 1):
        _cell(ws_i, 3, col, titulo, bold=True, fondo=GRIS_HI, align="center", size=8)
        ws_i.column_dimensions[get_column_letter(col)].width = ancho
    ws_i.row_dimensions[3].height = 36

    for i, r in enumerate(ingresos):
        row = 4 + i
        es_padre = r["es_hoja"] == 0
        fondo_fila = GRIS_F if es_padre else None
        bold_fila = es_padre
        indent = "  " * (r["nivel"] - 1)

        _cell(ws_i, row, 1, r["codigo"], bold=bold_fila, fondo=fondo_fila, align="left", size=8)
        _cell(ws_i, row, 2, indent + r["cuenta"], bold=bold_fila, fondo=fondo_fila, align="left", size=8)
        for col, key in enumerate([
            "ppto_inicial", "adiciones", "reducciones", "ppto_definitivo",
            "recaudo_anterior", "recaudo_mes", "recaudo_acumulado", "saldo_por_recaudar",
        ], start=3):
            _cell(ws_i, row, col, r[key] or None, bold=bold_fila, fondo=fondo_fila, size=8)
        ws_i.row_dimensions[row].height = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ─── Helpers para exportación CSV SIA Contraloría ─────────────────────────────

def _sia_fecha_corte(mes: int, anio: str) -> tuple[str, str]:
    """Rango enero..mes como strings ISO: (fecha_ini, fecha_fin_exclusiva)."""
    fecha_ini = f"{anio}-01-01"
    if mes >= 12:
        fecha_fin = f"{int(anio) + 1}-01-01"
    else:
        fecha_fin = f"{anio}-{mes + 1:02d}-01"
    return fecha_ini, fecha_fin


def _csv_bytes(encabezados: list[str], filas: list[list]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(encabezados)
    w.writerows(filas)
    return buf.getvalue().encode("utf-8-sig")


_MESES_SIA = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ─── F03: Movimiento de Bancos ────────────────────────────────────────────────

async def generar_sia_csv_f03(db: AsyncSession, tenant_id: str, mes: int, anio: str, institucion: str) -> bytes:
    fecha_ini, fecha_fin = _sia_fecha_corte(mes, anio)

    cuentas_res = await db.execute(
        select(CuentaBancaria).where(
            CuentaBancaria.tenant_id == tenant_id,
            CuentaBancaria.estado == "ACTIVA",
        ).order_by(CuentaBancaria.banco, CuentaBancaria.numero_cuenta)
    )
    cuentas = cuentas_res.scalars().all()

    filas = []
    for cb in cuentas:
        ing_res = await db.execute(
            select(func.coalesce(func.sum(Recaudo.valor), 0)).where(
                Recaudo.tenant_id == tenant_id,
                Recaudo.estado != "ANULADO",
                Recaudo.cuenta_bancaria_id == cb.id,
                Recaudo.fecha >= fecha_ini,
                Recaudo.fecha < fecha_fin,
            )
        )
        ingresos = ing_res.scalar() or 0

        eg_res = await db.execute(
            select(func.coalesce(func.sum(Pago.valor), 0)).where(
                Pago.tenant_id == tenant_id,
                Pago.estado != "ANULADO",
                Pago.cuenta_bancaria_id == cb.id,
                Pago.fecha >= fecha_ini,
                Pago.fecha < fecha_fin,
            )
        )
        egresos = eg_res.scalar() or 0

        filas.append([
            cb.banco, cb.numero_cuenta, cb.denominacion or institucion,
            "FONDO DE SERVICIOS EDUCATIVOS",
            0, ingresos, egresos, 0, 0,
            ingresos - egresos, ingresos - egresos,
        ])

    if not filas:
        filas.append(["", "", institucion, "FONDO DE SERVICIOS EDUCATIVOS", 0, 0, 0, 0, 0, 0, 0])

    return _csv_bytes([
        "(C) Banco", "(C) No. De Cuenta", "(C) Denominación",
        "(C) Fuente De Financiación", "(D) Saldo Inicial A 1 De Enero",
        "(D) Ingresos", "(D) Egresos", "(D) Notas Débito", "(D) Notas Crédito",
        "(D) Saldo A 31 De Diciembre Según Libros",
        "(D) Saldo A 31 De Diciembre Según Extractos Bancarios",
    ], filas)


# ─── F7B: Formato de Pagos ───────────────────────────────────────────────────

async def generar_sia_csv_f7b(db: AsyncSession, tenant_id: str, mes: int, anio: str) -> bytes:
    fecha_ini, fecha_fin = _sia_fecha_corte(mes, anio)

    pagos_res = await db.execute(
        select(Pago).where(
            Pago.tenant_id == tenant_id,
            Pago.estado != "ANULADO",
            Pago.fecha >= fecha_ini,
            Pago.fecha < fecha_fin,
        ).order_by(Pago.fecha, Pago.numero)
    )
    pagos = pagos_res.scalars().all()

    filas = []
    for p in pagos:
        obl = p.obligacion
        rp_numero = obl.rp_numero if obl else ""
        t = p.tercero
        filas.append([
            p.fecha, p.codigo_rubro, rp_numero, p.medio_pago,
            "FONDO DE SERVICIOS EDUCATIVOS", p.no_comprobante,
            t.nombre if t else "", p.nit_tercero, p.concepto,
            p.valor, 0, 0, 0, p.valor,
            t.banco if t else "", t.no_cuenta if t else "", p.no_comprobante,
        ])

    return _csv_bytes([
        "(F) Fecha De Pago", "(G) Código Presupuestal",
        "(N) Número Registro Presupuestal", "(C) Clase De Pago",
        "(C) Fuente De Financiación", "(C) No. De Comprobante",
        "(C) Beneficiario", "(N) Cédula O Nit", "(C) Detalle De Pago",
        "(D) Valor Comprobante De Pago", "(D) Descuentos Seg. Social",
        "(D) Descuentos Retenciones", "(D) Otros Descuentos", "(D) Neto Pagado",
        "(C) Banco", "(C) No. De Cuenta", "(C) No. De Cheque O Nd",
    ], filas)


# ─── F08A: Modificaciones presupuestales ──────────────────────────────────────

async def generar_sia_csv_f08a(db: AsyncSession, tenant_id: str, mes: int, anio: str, tipo_rubro: str) -> bytes:
    fecha_ini, fecha_fin = _sia_fecha_corte(mes, anio)

    res = await db.execute(
        select(DetalleModificacion, ModificacionPresupuestal)
        .join(ModificacionPresupuestal, DetalleModificacion.id_modificacion == ModificacionPresupuestal.id)
        .where(
            ModificacionPresupuestal.tenant_id == tenant_id,
            DetalleModificacion.tenant_id == tenant_id,
            ModificacionPresupuestal.estado == "ACTIVO",
            ModificacionPresupuestal.fecha >= fecha_ini,
            ModificacionPresupuestal.fecha < fecha_fin,
            DetalleModificacion.tipo_rubro == tipo_rubro,
        ).order_by(ModificacionPresupuestal.fecha, DetalleModificacion.codigo_rubro)
    )

    filas = []
    for det, mod in res.all():
        campo = det.campo_afectado.lower()
        adicion   = det.valor if "adicion"   in campo else 0
        reduccion = det.valor if "reduccion" in campo else 0
        filas.append([
            det.codigo_rubro,
            mod.numero_acto or f"ACTO-{mod.id}",
            mod.fecha, adicion, reduccion,
        ])

    return _csv_bytes([
        "(I) Código Rubro Presupuestal", "(C) Acto Administrativo",
        "(F) Fecha", "(D) Adición", "(D) Reducción",
    ], filas)


# ─── F09: PAC ────────────────────────────────────────────────────────────────

async def generar_sia_csv_f09(db: AsyncSession, tenant_id: str, mes: int, anio: str) -> bytes:
    fecha_ini, fecha_fin = _sia_fecha_corte(mes, anio)

    rubros_res = await db.execute(
        select(RubroGasto).where(
            RubroGasto.tenant_id == tenant_id,
            RubroGasto.es_hoja == 1,
        ).order_by(RubroGasto.codigo)
    )
    rubros = rubros_res.scalars().all()

    filas = []
    for r in rubros:
        pac_res = await db.execute(
            select(func.coalesce(func.sum(PAC.valor_programado), 0)).where(
                PAC.tenant_id == tenant_id,
                PAC.codigo_rubro == r.codigo,
                PAC.mes >= 1,
                PAC.mes <= mes,
            )
        )
        pac_acum = pac_res.scalar() or 0

        pago_res = await db.execute(
            select(func.coalesce(func.sum(Pago.valor), 0)).where(
                Pago.tenant_id == tenant_id,
                Pago.codigo_rubro == r.codigo,
                Pago.estado != "ANULADO",
                Pago.fecha >= fecha_ini,
                Pago.fecha < fecha_fin,
            )
        )
        pago_acum = pago_res.scalar() or 0

        filas.append([
            r.codigo, r.cuenta, pac_acum, 0,
            r.adiciones, r.reducciones, 0, pac_acum, pago_acum,
        ])

    return _csv_bytes([
        "(G) Código Rubro Presupuestal", "(C) Nombre Rubro Presupuestal",
        "(D) Pac Período Rendido", "(D) Anticipos", "(D) Adiciones",
        "(D) Reducciones", "(D) Aplazamientos", "(D) Pac Situado", "(D) Pago",
    ], filas)


# ─── F13A: Contratación ──────────────────────────────────────────────────────

async def generar_sia_csv_f13a(db: AsyncSession, tenant_id: str, mes: int, anio: str) -> bytes:
    fecha_ini, fecha_fin = _sia_fecha_corte(mes, anio)

    rps_res = await db.execute(
        select(RP).where(
            RP.tenant_id == tenant_id,
            RP.estado != "ANULADO",
            RP.fecha >= fecha_ini,
            RP.fecha < fecha_fin,
        ).order_by(RP.numero)
    )
    rps = rps_res.scalars().all()

    filas = []
    for rp in rps:
        pagos_res = await db.execute(
            select(func.coalesce(func.sum(Pago.valor), 0))
            .join(Obligacion, and_(
                Pago.obligacion_numero == Obligacion.numero,
                Obligacion.rp_numero == rp.numero,
            ))
            .where(
                Pago.tenant_id == tenant_id,
                Pago.estado != "ANULADO",
                Pago.fecha >= fecha_ini,
                Pago.fecha < fecha_fin,
            )
        )
        pagos_total = pagos_res.scalar() or 0

        cdp = rp.cdp
        t = rp.tercero
        filas.append([
            f"RP-{rp.numero}", "FONDO DE SERVICIOS EDUCATIVOS",
            rp.objeto, "CONTRATO DE COMPRAVENTA", rp.valor,
            t.nombre if t else "", rp.nit_tercero,
            cdp.numero if cdp else "", cdp.fecha if cdp else "",
            cdp.valor if cdp else 0,
            rp.fecha, "CONTRATACION DIRECTA",
            rp.fecha, rp.numero, rp.codigo_rubro, rp.valor,
            "", rp.fecha, "30 DIAS", "", "", 0,
            pagos_total, "", "",
        ])

    return _csv_bytes([
        "(C) Número Del Contrato", "(C) Fuente De Recurso", "(C) Objeto",
        "(C) Clase", "(D) Valor Del Contrato", "(C) Nombre Del Contratista",
        "(C) Nit O Cédula Del Contratista", "(N) No Disponibilidad Presupuestal",
        "(F) Fecha Disponibilidad", "(D) Valor Disponibilidad",
        "(F) Fecha Firma", "(C) Forma De Contratación",
        "(F) Fecha Registro Presupuestal", "(N) No De Registro Presupuestal",
        "(C) Rubro Registro Presupuestal", "(D) Valor Registro Presupuestal",
        "(F) Fecha Aprobación Garantía Única", "(F) Fecha Iniciación",
        "(C) Plazo Contrato", "(F) Fecha Adición", "(C) Plazo Adición",
        "(D) Valor Adición", "(D) Valor Pagos Efectuados",
        "(F) Fecha De Terminación", "(F) Fecha De Acta De Liquidación",
    ], filas)


# ─── ZIP con todos los formatos ───────────────────────────────────────────────

async def generar_sia_zip(db: AsyncSession, tenant_id: str, mes_consulta: int | None = None) -> tuple[bytes, str, str]:
    """Genera ZIP con los 6 formatos SIA. Retorna (bytes, anio, nombre_mes)."""
    if mes_consulta is None:
        mes_consulta = int(await config_svc.get_config(db, tenant_id, "mes_actual") or "1")
    anio = await config_svc.get_config(db, tenant_id, "vigencia") or "2026"
    institucion = await config_svc.get_config(db, tenant_id, "nombre_institucion") or "INSTITUCIÓN"
    nombre_mes = _MESES_SIA.get(mes_consulta, str(mes_consulta))
    m = f"{mes_consulta:02d}"

    archivos = {
        f"F03_MovBancos_{anio}_Ene_a_{m}.csv":
            await generar_sia_csv_f03(db, tenant_id, mes_consulta, anio, institucion),
        f"F7B_Pagos_{anio}_Ene_a_{m}.csv":
            await generar_sia_csv_f7b(db, tenant_id, mes_consulta, anio),
        f"F08A_Modif_Gastos_{anio}_Ene_a_{m}.csv":
            await generar_sia_csv_f08a(db, tenant_id, mes_consulta, anio, "GASTO"),
        f"F08A_Modif_Ingresos_{anio}_Ene_a_{m}.csv":
            await generar_sia_csv_f08a(db, tenant_id, mes_consulta, anio, "INGRESO"),
        f"F09_PAC_{anio}_Ene_a_{m}.csv":
            await generar_sia_csv_f09(db, tenant_id, mes_consulta, anio),
        f"F13A_Contratacion_{anio}_Ene_a_{m}.csv":
            await generar_sia_csv_f13a(db, tenant_id, mes_consulta, anio),
    }

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nombre, contenido in archivos.items():
            zf.writestr(nombre, contenido)
    buf.seek(0)
    return buf.read(), anio, nombre_mes


async def cuentas_por_pagar(db: AsyncSession, tenant_id: str) -> list[dict]:
    """Obligaciones activas con saldo pendiente de pago."""
    obls_res = await db.execute(
        select(Obligacion).where(
            Obligacion.tenant_id == tenant_id,
            Obligacion.estado != "ANULADA",
        ).order_by(Obligacion.nit_tercero, Obligacion.fecha, Obligacion.numero)
    )
    obls = obls_res.scalars().all()

    resultado = []
    for obl in obls:
        pago_res = await db.execute(
            select(func.coalesce(func.sum(Pago.valor), 0)).where(
                Pago.tenant_id == tenant_id,
                Pago.obligacion_numero == obl.numero,
                Pago.estado != "ANULADO",
            )
        )
        total_pagado = float(pago_res.scalar() or 0)
        saldo = obl.valor - total_pagado
        if saldo > 0.01:
            t = obl.tercero
            resultado.append({
                "obl_numero": obl.numero,
                "fecha": obl.fecha,
                "codigo_rubro": obl.codigo_rubro,
                "nit": obl.nit_tercero,
                "tercero": t.nombre if t else "",
                "factura": obl.factura,
                "valor_obl": obl.valor,
                "total_pagado": total_pagado,
                "saldo_por_pagar": saldo,
            })
    return resultado


async def pac_vs_ejecutado(db: AsyncSession, tenant_id: str, mes_corte: int | None = None) -> dict:
    """PAC programado vs pagos reales por rubro, acumulado hasta mes_corte."""
    if mes_corte is None:
        mes_corte = int(await config_svc.get_config(db, tenant_id, "mes_actual") or "1")

    rubros_res = await db.execute(
        select(RubroGasto).where(
            RubroGasto.tenant_id == tenant_id,
            RubroGasto.es_hoja == 1,
        ).order_by(RubroGasto.codigo)
    )
    rubros = rubros_res.scalars().all()

    resultado = []
    for r in rubros:
        pac_total = 0.0
        pago_total = 0.0
        meses_data = []
        for mes in range(1, mes_corte + 1):
            pac_res = await db.execute(
                select(func.coalesce(func.sum(PAC.valor_programado), 0)).where(
                    PAC.tenant_id == tenant_id,
                    PAC.codigo_rubro == r.codigo,
                    PAC.mes == mes,
                )
            )
            pac_mes = float(pac_res.scalar() or 0)

            mes_col = func.cast(func.substr(Pago.fecha, 6, 2), Integer)
            pago_res = await db.execute(
                select(func.coalesce(func.sum(Pago.valor), 0)).where(
                    Pago.tenant_id == tenant_id,
                    Pago.codigo_rubro == r.codigo,
                    Pago.estado != "ANULADO",
                    mes_col == mes,
                )
            )
            pago_mes = float(pago_res.scalar() or 0)

            pac_total += pac_mes
            pago_total += pago_mes
            meses_data.append({"mes": mes, "pac": pac_mes, "pagado": pago_mes})

        pct = round(pago_total / pac_total * 100, 1) if pac_total > 0 else 0.0
        resultado.append({
            "codigo": r.codigo,
            "cuenta": r.cuenta,
            "pac_total": pac_total,
            "pagado_total": pago_total,
            "saldo_pac": pac_total - pago_total,
            "pct_ejecucion": pct,
            "meses": meses_data,
        })

    return {"mes_corte": mes_corte, "rubros": resultado}


async def informe_tercero(
    db: AsyncSession, tenant_id: str, nit: str,
    mes_inicio: int = 1, mes_fin: int = 12
) -> dict:
    """Todos los documentos (RP, Obligaciones, Pagos) de un tercero en un período."""
    from app.models.terceros import Tercero

    tercero_res = await db.execute(
        select(Tercero).where(Tercero.tenant_id == tenant_id, Tercero.nit == nit)
    )
    tercero = tercero_res.scalar_one_or_none()
    if not tercero:
        raise ValueError(f"Tercero {nit} no encontrado")

    from app.models.rp import RP as RPModel

    mes_col_rp = func.cast(func.substr(RPModel.fecha, 6, 2), Integer)
    rps_res = await db.execute(
        select(RPModel).where(
            RPModel.tenant_id == tenant_id,
            RPModel.nit_tercero == nit,
            RPModel.estado != "ANULADO",
            mes_col_rp >= mes_inicio,
            mes_col_rp <= mes_fin,
        ).order_by(RPModel.fecha, RPModel.numero)
    )
    rps = [
        {
            "numero": r.numero, "fecha": r.fecha,
            "codigo_rubro": r.codigo_rubro, "valor": r.valor,
            "objeto": r.objeto, "estado": r.estado,
        }
        for r in rps_res.scalars().all()
    ]

    mes_col_obl = func.cast(func.substr(Obligacion.fecha, 6, 2), Integer)
    obls_res = await db.execute(
        select(Obligacion).where(
            Obligacion.tenant_id == tenant_id,
            Obligacion.nit_tercero == nit,
            Obligacion.estado != "ANULADA",
            mes_col_obl >= mes_inicio,
            mes_col_obl <= mes_fin,
        ).order_by(Obligacion.fecha, Obligacion.numero)
    )
    obls = [
        {
            "numero": o.numero, "fecha": o.fecha,
            "codigo_rubro": o.codigo_rubro, "rp_numero": o.rp_numero,
            "valor": o.valor, "factura": o.factura, "estado": o.estado,
        }
        for o in obls_res.scalars().all()
    ]

    mes_col_pago = func.cast(func.substr(Pago.fecha, 6, 2), Integer)
    pagos_res = await db.execute(
        select(Pago).where(
            Pago.tenant_id == tenant_id,
            Pago.nit_tercero == nit,
            Pago.estado != "ANULADO",
            mes_col_pago >= mes_inicio,
            mes_col_pago <= mes_fin,
        ).order_by(Pago.fecha, Pago.numero)
    )
    pagos = [
        {
            "numero": p.numero, "fecha": p.fecha,
            "codigo_rubro": p.codigo_rubro, "obligacion_numero": p.obligacion_numero,
            "valor": p.valor, "concepto": p.concepto,
            "medio_pago": p.medio_pago, "no_comprobante": p.no_comprobante,
        }
        for p in pagos_res.scalars().all()
    ]

    return {
        "tercero": {"nit": tercero.nit, "nombre": tercero.nombre},
        "mes_inicio": mes_inicio,
        "mes_fin": mes_fin,
        "rps": rps,
        "obligaciones": obls,
        "pagos": pagos,
        "total_rp": sum(r["valor"] for r in rps),
        "total_obl": sum(o["valor"] for o in obls),
        "total_pagos": sum(p["valor"] for p in pagos),
    }


async def verificar_equilibrio(db: AsyncSession, tenant_id: str) -> dict:
    result = await db.execute(
        select(func.coalesce(func.sum(RubroGasto.apropiacion_definitiva), 0))
        .where(RubroGasto.tenant_id == tenant_id, RubroGasto.es_hoja == 1)
    )
    total_gastos = result.scalar()

    result = await db.execute(
        select(func.coalesce(func.sum(RubroIngreso.presupuesto_definitivo), 0))
        .where(RubroIngreso.tenant_id == tenant_id, RubroIngreso.es_hoja == 1)
    )
    total_ingresos = result.scalar()

    return {
        "total_gastos": total_gastos,
        "total_ingresos": total_ingresos,
        "diferencia": total_ingresos - total_gastos,
    }
