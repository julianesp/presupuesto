"""
database.py - Modulo de base de datos SQLite
Sistema Presupuestal IE 2026
"""
import sqlite3
import os
import shutil
from datetime import datetime, date

DB_PATH = os.path.join(os.path.dirname(__file__), "presupuesto_ie_2026.db")


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_connection()
    c = conn.cursor()

    c.executescript("""
    CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY,
        valor TEXT
    );

    CREATE TABLE IF NOT EXISTS rubros_gastos (
        codigo TEXT PRIMARY KEY,
        cuenta TEXT NOT NULL,
        es_hoja INTEGER DEFAULT 0,
        apropiacion_inicial REAL DEFAULT 0,
        adiciones REAL DEFAULT 0,
        reducciones REAL DEFAULT 0,
        creditos REAL DEFAULT 0,
        contracreditos REAL DEFAULT 0,
        apropiacion_definitiva REAL DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS rubros_ingresos (
        codigo TEXT PRIMARY KEY,
        cuenta TEXT NOT NULL,
        es_hoja INTEGER DEFAULT 0,
        presupuesto_inicial REAL DEFAULT 0,
        adiciones REAL DEFAULT 0,
        reducciones REAL DEFAULT 0,
        presupuesto_definitivo REAL DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS terceros (
        nit TEXT PRIMARY KEY,
        dv TEXT DEFAULT '',
        nombre TEXT NOT NULL,
        direccion TEXT DEFAULT '',
        telefono TEXT DEFAULT '',
        email TEXT DEFAULT '',
        tipo TEXT DEFAULT 'Natural',
        banco TEXT DEFAULT '',
        tipo_cuenta TEXT DEFAULT '',
        no_cuenta TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS cdp (
        numero INTEGER PRIMARY KEY,
        fecha TEXT NOT NULL,
        codigo_rubro TEXT NOT NULL,
        objeto TEXT NOT NULL,
        valor REAL NOT NULL,
        estado TEXT DEFAULT 'ACTIVO',
        FOREIGN KEY (codigo_rubro) REFERENCES rubros_gastos(codigo)
    );

    CREATE TABLE IF NOT EXISTS rp (
        numero INTEGER PRIMARY KEY,
        fecha TEXT NOT NULL,
        cdp_numero INTEGER NOT NULL,
        codigo_rubro TEXT NOT NULL,
        nit_tercero TEXT NOT NULL,
        valor REAL NOT NULL,
        objeto TEXT NOT NULL,
        estado TEXT DEFAULT 'ACTIVO',
        FOREIGN KEY (cdp_numero) REFERENCES cdp(numero),
        FOREIGN KEY (nit_tercero) REFERENCES terceros(nit)
    );

    CREATE TABLE IF NOT EXISTS obligacion (
        numero INTEGER PRIMARY KEY,
        fecha TEXT NOT NULL,
        rp_numero INTEGER NOT NULL,
        codigo_rubro TEXT NOT NULL,
        nit_tercero TEXT NOT NULL,
        valor REAL NOT NULL,
        factura TEXT DEFAULT '',
        estado TEXT DEFAULT 'ACTIVO',
        FOREIGN KEY (rp_numero) REFERENCES rp(numero)
    );

    CREATE TABLE IF NOT EXISTS pago (
        numero INTEGER PRIMARY KEY,
        fecha TEXT NOT NULL,
        obligacion_numero INTEGER NOT NULL,
        codigo_rubro TEXT NOT NULL,
        nit_tercero TEXT NOT NULL,
        valor REAL NOT NULL,
        concepto TEXT DEFAULT '',
        medio_pago TEXT DEFAULT 'Transferencia',
        no_comprobante TEXT DEFAULT '',
        estado TEXT DEFAULT 'PAGADO',
        FOREIGN KEY (obligacion_numero) REFERENCES obligacion(numero)
    );

    CREATE TABLE IF NOT EXISTS conceptos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_rubro TEXT NOT NULL,
        concepto TEXT NOT NULL,
        veces_usado INTEGER DEFAULT 1,
        ultimo_uso TEXT,
        UNIQUE(codigo_rubro, concepto)
    );

    CREATE TABLE IF NOT EXISTS consolidacion_mensual (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mes INTEGER NOT NULL,
        anio INTEGER NOT NULL,
        codigo_rubro TEXT NOT NULL,
        compromisos_mes REAL DEFAULT 0,
        pagos_mes REAL DEFAULT 0,
        fecha_consolidacion TEXT,
        UNIQUE(mes, anio, codigo_rubro)
    );

    CREATE TABLE IF NOT EXISTS recaudo (
        numero INTEGER PRIMARY KEY,
        fecha TEXT NOT NULL,
        codigo_rubro TEXT NOT NULL,
        valor REAL NOT NULL,
        concepto TEXT DEFAULT '',
        no_comprobante TEXT DEFAULT '',
        estado TEXT DEFAULT 'ACTIVO',
        FOREIGN KEY (codigo_rubro) REFERENCES rubros_ingresos(codigo)
    );

    CREATE TABLE IF NOT EXISTS consolidacion_mensual_ingresos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mes INTEGER NOT NULL,
        anio INTEGER NOT NULL,
        codigo_rubro TEXT NOT NULL,
        recaudo_mes REAL DEFAULT 0,
        fecha_consolidacion TEXT,
        UNIQUE(mes, anio, codigo_rubro)
    );

    CREATE TABLE IF NOT EXISTS modificaciones_presupuestales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        tipo TEXT NOT NULL,
        numero_acto TEXT DEFAULT '',
        descripcion TEXT DEFAULT '',
        valor REAL NOT NULL,
        estado TEXT DEFAULT 'ACTIVO'
    );

    CREATE TABLE IF NOT EXISTS detalle_modificacion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_modificacion INTEGER NOT NULL,
        codigo_rubro TEXT NOT NULL,
        tipo_rubro TEXT NOT NULL,
        campo_afectado TEXT NOT NULL,
        valor REAL NOT NULL,
        FOREIGN KEY (id_modificacion) REFERENCES modificaciones_presupuestales(id)
    );

    CREATE TABLE IF NOT EXISTS pac (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_rubro TEXT NOT NULL,
        mes INTEGER NOT NULL,
        valor_programado REAL DEFAULT 0,
        UNIQUE(codigo_rubro, mes),
        FOREIGN KEY (codigo_rubro) REFERENCES rubros_gastos(codigo)
    );

    CREATE TABLE IF NOT EXISTS catalogo_sifse_fuentes (
        codigo INTEGER PRIMARY KEY,
        descripcion TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS catalogo_sifse_items (
        codigo INTEGER PRIMARY KEY,
        descripcion TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS mapeo_sifse_ingresos (
        codigo_rubro TEXT PRIMARY KEY,
        sifse_fuente INTEGER NOT NULL,
        FOREIGN KEY (codigo_rubro) REFERENCES rubros_ingresos(codigo)
    );

    CREATE TABLE IF NOT EXISTS mapeo_sifse_gastos (
        codigo_rubro TEXT PRIMARY KEY,
        sifse_item INTEGER NOT NULL,
        FOREIGN KEY (codigo_rubro) REFERENCES rubros_gastos(codigo)
    );

    CREATE TABLE IF NOT EXISTS cuentas_bancarias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        banco TEXT NOT NULL,
        tipo_cuenta TEXT DEFAULT 'Ahorros',
        numero_cuenta TEXT NOT NULL,
        denominacion TEXT DEFAULT '',
        estado TEXT DEFAULT 'ACTIVA'
    );
    """)

    # Migracion: agregar columna fuente_sifse a tablas de ejecucion
    for tabla in ("cdp", "rp", "obligacion", "pago"):
        try:
            c.execute(f"ALTER TABLE {tabla} ADD COLUMN fuente_sifse INTEGER DEFAULT 0")
        except Exception:
            pass  # Ya existe la columna

    # Migracion: agregar columna item_sifse a tablas de ejecucion
    for tabla in ("cdp", "rp", "obligacion", "pago"):
        try:
            c.execute(f"ALTER TABLE {tabla} ADD COLUMN item_sifse INTEGER DEFAULT 0")
        except Exception:
            pass  # Ya existe la columna

    # Migracion: agregar columna cuenta_bancaria_id a pago
    try:
        c.execute("ALTER TABLE pago ADD COLUMN cuenta_bancaria_id INTEGER DEFAULT 0")
    except Exception:
        pass  # Ya existe la columna

    # Migracion: agregar columna cuenta_bancaria_id a recaudo
    try:
        c.execute("ALTER TABLE recaudo ADD COLUMN cuenta_bancaria_id INTEGER DEFAULT 0")
    except Exception:
        pass  # Ya existe la columna

    # Inicializar config si esta vacia
    c.execute("SELECT COUNT(*) FROM config")
    if c.fetchone()[0] == 0:
        defaults = [
            ("vigencia", "2026"),
            ("institucion", "(Nombre de la Institucion Educativa)"),
            ("nit_institucion", "(NIT)"),
            ("rector", "(Nombre del Rector)"),
            ("tesorero", "(Nombre del Tesorero)"),
            ("mes_actual", "1"),
            ("consecutivo_cdp", "0"),
            ("consecutivo_rp", "0"),
            ("consecutivo_obligacion", "0"),
            ("consecutivo_pago", "0"),
            ("consecutivo_recaudo", "0"),
            ("consecutivo_modificacion", "0"),
        ]
        c.executemany("INSERT OR IGNORE INTO config VALUES (?, ?)", defaults)
    else:
        # Asegurar que claves nuevas existan en BD existentes
        c.execute("INSERT OR IGNORE INTO config VALUES ('consecutivo_recaudo', '0')")
        c.execute("INSERT OR IGNORE INTO config VALUES ('consecutivo_modificacion', '0')")

    # Agregar codigo_dane a config si no existe
    c.execute("INSERT OR IGNORE INTO config VALUES ('codigo_dane', '186755000015')")

    # Poblar catalogos SIFSE si estan vacios
    _poblar_catalogos_sifse(c)

    conn.commit()
    conn.close()


def _poblar_catalogos_sifse(cursor):
    """Pobla los catalogos SIFSE con los datos oficiales si estan vacios."""
    # Verificar si ya hay datos
    if cursor.execute("SELECT COUNT(*) FROM catalogo_sifse_fuentes").fetchone()[0] > 0:
        return

    fuentes = [
        (1, "Recursos Propios - Venta Bienes y Servicios"),
        (2, "SGP Calidad por Gratuidad (MEN)"),
        (3, "Otras Transferencias"),
        (6, "SGP Matricula - Calidad"),
        (32, "Rec. Balance SGP Gratuidad"),
        (33, "Rec. Balance Venta Bienes y Servicios"),
        (35, "Rendimientos Financieros SGP Gratuidad"),
        (36, "Rendimientos Financieros Distintos SGP"),
        (38, "Reintegros SGP Gratuidad"),
        (39, "Reintegros Recursos Propios"),
        (42, "Rec. Balance SGP Matricula"),
        (43, "Rendimientos Financieros Trans. Calidad"),
        (44, "Reintegros Trans. Calidad"),
        (52, "Donaciones"),
        (57, "Rec. Balance Otras Transferencias"),
        (61, "Otros Reintegros"),
    ]
    cursor.executemany(
        "INSERT OR IGNORE INTO catalogo_sifse_fuentes VALUES (?, ?)", fuentes
    )

    items = [
        (7, "Adquisicion de bienes - Funcionamiento basico"),
        (8, "Arrendamientos"),
        (9, "Servicios publicos (Acueducto/Aseo)"),
        (10, "Energia"),
        (11, "Telefono"),
        (12, "Internet / Hosting"),
        (13, "Otros servicios publicos"),
        (14, "Seguros"),
        (15, "Servicios profesionales y tecnicos"),
        (16, "Impresos y publicaciones"),
        (17, "Personal - Docentes"),
        (19, "Construccion / Ampliacion infraestructura"),
        (20, "Mantenimiento infraestructura"),
        (21, "Dotacion institucional"),
        (22, "Material pedagogico"),
        (23, "Transporte escolar"),
        (24, "Semovientes / Proyectos pedagogicos"),
        (25, "Alimentacion escolar"),
        (26, "Actividades pedagogicas"),
        (86, "Servicios financieros"),
    ]
    cursor.executemany(
        "INSERT OR IGNORE INTO catalogo_sifse_items VALUES (?, ?)", items
    )

    # Mapeo por defecto: ingresos CCPET -> fuente SIFSE
    mapeo_ingresos = [
        ("1.1.02.05", 1),    # Venta bienes/servicios -> Recursos Propios
        ("1.1.02.06.001.01.03.02", 2),  # Calidad por Gratuidad MEN
        ("1.1.02.06.006.01", 3),  # Aportes Nacion Alimentacion
        ("1.1.02.06.006.06.01.01", 6),  # Calidad SGP Dptos
        ("1.1.02.06.006.06.01.02", 3),  # Otros aportes Dptos
        ("1.1.02.06.006.06.02.01", 6),  # Calidad SGP Municipio
        ("1.1.02.06.006.06.02.02", 3),  # Otros Aportes Muni
        ("1.1.02.06.006.06.03", 3),  # Otras Entidades Publicas
        ("1.1.1.01.01", 6),  # Transferencias SGP
        ("1.2.05.02.01", 35),  # Rend.Fin. SGP Gratuidad
        ("1.2.05.02.02", 36),  # Rend.Fin. Propios
        ("1.2.05.02.03", 43),  # Rend.Fin. SGP Matricula
        ("1.2.05.02.04", 36),  # Rend.Fin. Otros
        ("1.2.08", 52),  # Donaciones
        ("1.2.10.02.01", 32),  # Rec.Bal. SGP Gratuidad
        ("1.2.10.02.02", 33),  # Rec.Bal. Propios
        ("1.2.10.02.03", 42),  # Rec.Bal. SGP Matricula
        ("1.2.10.02.04", 57),  # Rec.Bal. Otros
        ("1.2.13.01.01", 38),  # Reintegros SGP Gratuidad
        ("1.2.13.01.02", 39),  # Reintegro Propios
        ("1.2.13.01.03", 44),  # Reintegros SGP Matricula
        ("1.2.13.01.04", 61),  # Reintegros Otros
    ]

    # Mapeo por defecto: gastos CCPET -> item SIFSE
    mapeo_gastos = [
        ("2.1.1.01.01", 17),  # Sueldos -> Personal-Docentes
        ("2.1.2.01.01.003", 21),  # Maquinaria/Equipos -> Dotacion
        ("2.1.2.01.01.004", 21),  # Muebles/Mobiliario -> Dotacion
        ("2.1.2.01.01.005.01", 24),  # Animales/Arboles -> Semovientes
        ("2.1.2.01.01.005.02", 21),  # Software -> Dotacion
        ("2.1.2.02.01.000", 24),  # Insumos agro -> Semovientes
        ("2.1.2.02.01.003.01", 22),  # Papeleria -> Material pedagogico
        ("2.1.2.02.01.003.02", 7),   # Aseo/Botiquin -> Funcionamiento basico
        ("2.1.2.02.01.003.03", 13),  # Combustibles -> Otros serv publicos
        ("2.1.2.02.01.003.04", 20),  # Mat electricos/ferreteria -> Mant infraestructura
        ("2.1.2.02.01.003.05", 16),  # Cartuchos/Toners -> Impresos
        ("2.1.2.02.01.003.06", 22),  # Art deporte -> Material pedagogico
        ("2.1.2.02.01.003.07", 22),  # Dotacion menajes -> Material pedagogico
        ("2.1.2.02.01.003.08", 7),   # Otros bienes NCP -> Funcionamiento basico
        ("2.1.2.02.02.005.01", 20),  # Mejoras infraestructura
        ("2.1.2.02.02.005.02", 19),  # Construccion/ampliacion
        ("2.1.2.02.02.006.01", 26),  # Viajes estudiantes -> Act pedagogicas
        ("2.1.2.02.02.006.02", 9),   # Acueducto
        ("2.1.2.02.02.006.03", 9),   # Aseo
        ("2.1.2.02.02.006.04", 10),  # Energia
        ("2.1.2.02.02.006.05", 13),  # Gas -> Otros serv publicos
        ("2.1.2.02.02.006.06", 23),  # Transporte escolar
        ("2.1.2.02.02.006.07", 25),  # Alimentacion escolar
        ("2.1.2.02.02.007.01", 14),  # Seguros
        ("2.1.2.02.02.007.02", 86),  # Gastos financieros
        ("2.1.2.02.02.007.03", 8),   # Arrendamiento
        ("2.1.2.02.02.008.01", 12),  # Hosting web -> Internet
        ("2.1.2.02.02.008.02", 11),  # Telefonia
        ("2.1.2.02.02.008.03", 12),  # Internet
        ("2.1.2.02.02.008.04", 16),  # Impresos
        ("2.1.2.02.02.008.05", 15),  # Serv profesionales
        ("2.1.2.02.02.008.06", 15),  # Serv tecnicos
        ("2.1.2.02.02.008.07", 20),  # Mant maquinaria
        ("2.1.2.02.02.009", 26),  # Actividades pedagogicas
    ]

    # Aplicar mapeo a rubros hoja existentes
    rubros_ing = cursor.execute(
        "SELECT codigo FROM rubros_ingresos WHERE es_hoja=1"
    ).fetchall()
    for rubro in rubros_ing:
        cod = rubro[0]
        fuente = _buscar_mapeo_por_prefijo(cod, mapeo_ingresos)
        if fuente is not None:
            cursor.execute(
                "INSERT OR IGNORE INTO mapeo_sifse_ingresos VALUES (?, ?)",
                (cod, fuente)
            )

    rubros_gas = cursor.execute(
        "SELECT codigo FROM rubros_gastos WHERE es_hoja=1"
    ).fetchall()
    for rubro in rubros_gas:
        cod = rubro[0]
        item = _buscar_mapeo_por_prefijo(cod, mapeo_gastos)
        if item is not None:
            cursor.execute(
                "INSERT OR IGNORE INTO mapeo_sifse_gastos VALUES (?, ?)",
                (cod, item)
            )


def _buscar_mapeo_por_prefijo(codigo, mapeo_lista):
    """Busca el mapeo mas especifico (prefijo mas largo) para un codigo de rubro."""
    mejor = None
    mejor_len = 0
    for prefijo, valor in mapeo_lista:
        if codigo == prefijo or codigo.startswith(prefijo + "."):
            if len(prefijo) > mejor_len:
                mejor = valor
                mejor_len = len(prefijo)
    return mejor


# ===================== CUENTAS BANCARIAS =====================
def crear_cuenta_bancaria(banco, tipo_cuenta, numero_cuenta, denominacion=""):
    conn = get_connection()
    conn.execute(
        "INSERT INTO cuentas_bancarias (banco, tipo_cuenta, numero_cuenta, denominacion) "
        "VALUES (?,?,?,?)",
        (banco, tipo_cuenta, numero_cuenta, denominacion)
    )
    conn.commit()
    conn.close()


def listar_cuentas_bancarias(solo_activas=True):
    conn = get_connection()
    if solo_activas:
        rows = conn.execute(
            "SELECT * FROM cuentas_bancarias WHERE estado='ACTIVA' ORDER BY banco, numero_cuenta"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM cuentas_bancarias ORDER BY banco, numero_cuenta"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_cuenta_bancaria(id):
    conn = get_connection()
    row = conn.execute("SELECT * FROM cuentas_bancarias WHERE id=?", (id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def editar_cuenta_bancaria(id, banco, tipo_cuenta, numero_cuenta, denominacion):
    conn = get_connection()
    conn.execute(
        "UPDATE cuentas_bancarias SET banco=?, tipo_cuenta=?, numero_cuenta=?, denominacion=? WHERE id=?",
        (banco, tipo_cuenta, numero_cuenta, denominacion, id)
    )
    conn.commit()
    conn.close()


def desactivar_cuenta_bancaria(id):
    conn = get_connection()
    conn.execute("UPDATE cuentas_bancarias SET estado='INACTIVA' WHERE id=?", (id,))
    conn.commit()
    conn.close()


# ===================== SIFSE - CATALOGOS Y MAPEO =====================
def get_catalogo_fuentes_sifse():
    """Lista todas las fuentes SIFSE disponibles."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT codigo, descripcion FROM catalogo_sifse_fuentes ORDER BY codigo"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_catalogo_items_sifse():
    """Lista todos los items de gasto SIFSE disponibles."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT codigo, descripcion FROM catalogo_sifse_items ORDER BY codigo"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_fuentes_sifse_activas():
    """Solo las fuentes SIFSE que tienen rubros de ingreso mapeados (para dropdown del CDP)."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT DISTINCT f.codigo, f.descripcion "
        "FROM catalogo_sifse_fuentes f "
        "JOIN mapeo_sifse_ingresos m ON f.codigo = m.sifse_fuente "
        "ORDER BY f.codigo"
    ).fetchall()
    conn.close()
    if not rows:
        # Si no hay mapeos, devolver todas las fuentes
        return get_catalogo_fuentes_sifse()
    return [dict(r) for r in rows]


def get_mapeo_sifse_ingreso(codigo_rubro):
    """Retorna el codigo fuente SIFSE para un rubro de ingreso."""
    conn = get_connection()
    r = conn.execute(
        "SELECT sifse_fuente FROM mapeo_sifse_ingresos WHERE codigo_rubro=?",
        (codigo_rubro,)
    ).fetchone()
    conn.close()
    return r["sifse_fuente"] if r else None


def get_mapeo_sifse_gasto(codigo_rubro):
    """Retorna el codigo item SIFSE para un rubro de gasto."""
    conn = get_connection()
    r = conn.execute(
        "SELECT sifse_item FROM mapeo_sifse_gastos WHERE codigo_rubro=?",
        (codigo_rubro,)
    ).fetchone()
    conn.close()
    return r["sifse_item"] if r else None


def set_mapeo_sifse_ingreso(codigo_rubro, sifse_fuente):
    """Asigna o actualiza el mapeo de un rubro de ingreso a una fuente SIFSE."""
    conn = get_connection()
    conn.execute(
        "INSERT OR REPLACE INTO mapeo_sifse_ingresos VALUES (?, ?)",
        (codigo_rubro, sifse_fuente)
    )
    conn.commit()
    conn.close()


def set_mapeo_sifse_gasto(codigo_rubro, sifse_item):
    """Asigna o actualiza el mapeo de un rubro de gasto a un item SIFSE."""
    conn = get_connection()
    conn.execute(
        "INSERT OR REPLACE INTO mapeo_sifse_gastos VALUES (?, ?)",
        (codigo_rubro, sifse_item)
    )
    conn.commit()
    conn.close()


def get_todos_mapeos_ingresos():
    """Retorna todos los rubros de ingreso hoja con su mapeo SIFSE."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT ri.codigo, ri.cuenta, m.sifse_fuente, f.descripcion as desc_sifse "
        "FROM rubros_ingresos ri "
        "LEFT JOIN mapeo_sifse_ingresos m ON ri.codigo = m.codigo_rubro "
        "LEFT JOIN catalogo_sifse_fuentes f ON m.sifse_fuente = f.codigo "
        "WHERE ri.es_hoja=1 ORDER BY ri.codigo"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_todos_mapeos_gastos():
    """Retorna todos los rubros de gasto hoja con su mapeo SIFSE."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT rg.codigo, rg.cuenta, m.sifse_item, i.descripcion as desc_sifse "
        "FROM rubros_gastos rg "
        "LEFT JOIN mapeo_sifse_gastos m ON rg.codigo = m.codigo_rubro "
        "LEFT JOIN catalogo_sifse_items i ON m.sifse_item = i.codigo "
        "WHERE rg.es_hoja=1 ORDER BY rg.codigo"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def restaurar_mapeo_sifse_defecto():
    """Borra mapeos actuales y vuelve a aplicar los mapeos por defecto."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM mapeo_sifse_ingresos")
    c.execute("DELETE FROM mapeo_sifse_gastos")
    # Forzar repoblado
    c.execute("DELETE FROM catalogo_sifse_fuentes")
    c.execute("DELETE FROM catalogo_sifse_items")
    _poblar_catalogos_sifse(c)
    conn.commit()
    conn.close()


# ===================== SIFSE - INFORMES =====================
def informe_sifse_ingresos(trimestre, anio=None):
    """Agrega rubros de ingreso por fuente SIFSE para el trimestre.
    Retorna lista de dicts: fuente, desc_fuente, ppto_inicial, ppto_definitivo, recaudado
    """
    if anio is None:
        anio = get_config("vigencia") or "2026"
    anio = str(anio)

    # Meses del trimestre
    mes_ini = (trimestre - 1) * 3 + 1
    mes_fin = trimestre * 3
    fecha_ini = f"{anio}-01-01"
    if mes_fin >= 12:
        fecha_fin = f"{int(anio)+1}-01-01"
    else:
        fecha_fin = f"{anio}-{mes_fin+1:02d}-01"

    conn = get_connection()
    # Obtener rubros de ingreso hoja con su mapeo
    rows = conn.execute(
        "SELECT ri.codigo, ri.presupuesto_inicial, ri.presupuesto_definitivo, "
        "COALESCE(m.sifse_fuente, 0) as fuente "
        "FROM rubros_ingresos ri "
        "LEFT JOIN mapeo_sifse_ingresos m ON ri.codigo = m.codigo_rubro "
        "WHERE ri.es_hoja=1"
    ).fetchall()

    # Agrupar por fuente SIFSE
    agrupado = {}
    for r in rows:
        fuente = r["fuente"]
        if fuente not in agrupado:
            agrupado[fuente] = {"ppto_inicial": 0, "ppto_definitivo": 0, "recaudado": 0}
        agrupado[fuente]["ppto_inicial"] += r["presupuesto_inicial"]
        agrupado[fuente]["ppto_definitivo"] += r["presupuesto_definitivo"]

        # Recaudado acumulado hasta el fin del trimestre
        rec = conn.execute(
            "SELECT COALESCE(SUM(valor),0) as t FROM recaudo "
            "WHERE codigo_rubro=? AND estado<>'ANULADO' AND fecha >= ? AND fecha < ?",
            (r["codigo"], fecha_ini, fecha_fin)
        ).fetchone()["t"]
        agrupado[fuente]["recaudado"] += rec

    # Obtener descripciones
    resultado = []
    for fuente, datos in sorted(agrupado.items()):
        if fuente == 0:
            desc = "Sin asignar"
        else:
            f_row = conn.execute(
                "SELECT descripcion FROM catalogo_sifse_fuentes WHERE codigo=?", (fuente,)
            ).fetchone()
            desc = f_row["descripcion"] if f_row else f"Fuente {fuente}"
        resultado.append({
            "fuente": fuente,
            "desc_fuente": desc,
            "ppto_inicial": datos["ppto_inicial"],
            "ppto_definitivo": datos["ppto_definitivo"],
            "recaudado": datos["recaudado"],
        })

    conn.close()
    return resultado


def informe_sifse_gastos(trimestre, anio=None):
    """Agrega gastos por cruce fuente x item SIFSE para el trimestre.
    Retorna lista de dicts: fuente, desc_fuente, item, desc_item,
    ppto_inicial, ppto_definitivo, compromisos, obligaciones, pagos
    """
    if anio is None:
        anio = get_config("vigencia") or "2026"
    anio = str(anio)

    mes_ini = (trimestre - 1) * 3 + 1
    mes_fin = trimestre * 3
    fecha_ini = f"{anio}-01-01"
    if mes_fin >= 12:
        fecha_fin = f"{int(anio)+1}-01-01"
    else:
        fecha_fin = f"{anio}-{mes_fin+1:02d}-01"

    conn = get_connection()

    # Obtener rubros de gasto hoja con mapeo
    rubros = conn.execute(
        "SELECT rg.codigo, rg.apropiacion_inicial, rg.apropiacion_definitiva, "
        "COALESCE(m.sifse_item, 0) as item "
        "FROM rubros_gastos rg "
        "LEFT JOIN mapeo_sifse_gastos m ON rg.codigo = m.codigo_rubro "
        "WHERE rg.es_hoja=1"
    ).fetchall()

    # Cruce: para cada rubro hoja, obtener los CDPs con fuente_sifse
    # y agregar compromisos, obligaciones, pagos por (fuente, item)
    agrupado = {}  # key = (fuente, item)

    for rub in rubros:
        codigo = rub["codigo"]
        item = rub["item"]

        # Presupuesto inicial y definitivo se asigna por item (sin fuente)
        # Para el informe SIFSE, la fuente se toma de los CDPs
        # Pero ppto_inicial/definitivo se agrupa solo por item
        # Para simplificar: asignar ppto al cruce (0, item) y gastos al cruce real

        # Compromisos (RPs) con fuente del CDP heredada
        rps = conn.execute(
            "SELECT fuente_sifse, SUM(valor) as total FROM rp "
            "WHERE codigo_rubro=? AND estado<>'ANULADO' AND fecha >= ? AND fecha < ? "
            "GROUP BY fuente_sifse",
            (codigo, fecha_ini, fecha_fin)
        ).fetchall()

        obls = conn.execute(
            "SELECT fuente_sifse, SUM(valor) as total FROM obligacion "
            "WHERE codigo_rubro=? AND estado<>'ANULADA' AND fecha >= ? AND fecha < ? "
            "GROUP BY fuente_sifse",
            (codigo, fecha_ini, fecha_fin)
        ).fetchall()

        pagos = conn.execute(
            "SELECT fuente_sifse, SUM(valor) as total FROM pago "
            "WHERE codigo_rubro=? AND estado<>'ANULADO' AND fecha >= ? AND fecha < ? "
            "GROUP BY fuente_sifse",
            (codigo, fecha_ini, fecha_fin)
        ).fetchall()

        # Recopilar todas las fuentes involucradas
        fuentes_usadas = set()
        rp_por_fuente = {}
        for rp in rps:
            f = rp["fuente_sifse"] or 0
            fuentes_usadas.add(f)
            rp_por_fuente[f] = rp["total"]
        obl_por_fuente = {}
        for obl in obls:
            f = obl["fuente_sifse"] or 0
            fuentes_usadas.add(f)
            obl_por_fuente[f] = obl["total"]
        pago_por_fuente = {}
        for p in pagos:
            f = p["fuente_sifse"] or 0
            fuentes_usadas.add(f)
            pago_por_fuente[f] = p["total"]

        # Si no hay movimientos, registrar ppto con fuente 0
        if not fuentes_usadas:
            fuentes_usadas.add(0)

        for fuente in fuentes_usadas:
            key = (fuente, item)
            if key not in agrupado:
                agrupado[key] = {
                    "ppto_inicial": 0, "ppto_definitivo": 0,
                    "compromisos": 0, "obligaciones": 0, "pagos": 0
                }
            # Solo asignar presupuesto si esta fuente tiene movimientos
            # (distribuir proporcionalmente seria complejo, asignar al primer cruce)
            if fuente in rp_por_fuente or len(fuentes_usadas) == 1:
                agrupado[key]["ppto_inicial"] += rub["apropiacion_inicial"]
                agrupado[key]["ppto_definitivo"] += rub["apropiacion_definitiva"]
            agrupado[key]["compromisos"] += rp_por_fuente.get(fuente, 0)
            agrupado[key]["obligaciones"] += obl_por_fuente.get(fuente, 0)
            agrupado[key]["pagos"] += pago_por_fuente.get(fuente, 0)

    # Construir resultado con descripciones
    resultado = []
    for (fuente, item), datos in sorted(agrupado.items()):
        if fuente == 0:
            desc_fuente = "Sin asignar"
        else:
            f_row = conn.execute(
                "SELECT descripcion FROM catalogo_sifse_fuentes WHERE codigo=?", (fuente,)
            ).fetchone()
            desc_fuente = f_row["descripcion"] if f_row else f"Fuente {fuente}"

        if item == 0:
            desc_item = "Sin asignar"
        else:
            i_row = conn.execute(
                "SELECT descripcion FROM catalogo_sifse_items WHERE codigo=?", (item,)
            ).fetchone()
            desc_item = i_row["descripcion"] if i_row else f"Item {item}"

        resultado.append({
            "fuente": fuente,
            "desc_fuente": desc_fuente,
            "item": item,
            "desc_item": desc_item,
            "ppto_inicial": datos["ppto_inicial"],
            "ppto_definitivo": datos["ppto_definitivo"],
            "compromisos": datos["compromisos"],
            "obligaciones": datos["obligaciones"],
            "pagos": datos["pagos"],
        })

    conn.close()
    return resultado


def exportar_sifse_xls(trimestre, anio=None, destino=None):
    """Genera archivo .xls con formato SIFSE del MEN.
    Hojas: Ingresos_Presupuestales y Gastos_Presupuestales
    """
    try:
        import xlwt
    except ImportError:
        raise ImportError(
            "Se requiere la libreria 'xlwt' para exportar a .xls.\n"
            "Instale con: pip install xlwt"
        )

    if anio is None:
        anio = get_config("vigencia") or "2026"
    anio = str(anio)
    codigo_dane = get_config("codigo_dane") or "186755000015"

    if destino is None:
        destino = os.path.join(os.path.dirname(__file__),
                               f"SIFSE_T{trimestre}_{anio}.xls")

    wb = xlwt.Workbook(encoding="utf-8")

    # --- Hoja 1: Ingresos_Presupuestales ---
    ws_ing = wb.add_sheet("Ingresos_Presupuestales")
    encabezados_ing = ["CODIGO_DANE", "ANIO", "TRIMESTRE", "FUENTE",
                       "PPTO_INICIAL", "PPTO_DEFINITIVO", "RECAUDADO"]
    header_style = xlwt.easyxf('font: bold on; align: horiz center')
    money_style = xlwt.easyxf(num_format_str='#,##0')

    for col, h in enumerate(encabezados_ing):
        ws_ing.write(0, col, h, header_style)

    datos_ing = informe_sifse_ingresos(trimestre, anio)
    fila = 1
    for d in datos_ing:
        if d["fuente"] == 0:
            continue  # Omitir sin asignar
        ws_ing.write(fila, 0, codigo_dane)
        ws_ing.write(fila, 1, int(anio))
        ws_ing.write(fila, 2, trimestre)
        ws_ing.write(fila, 3, d["fuente"])
        ws_ing.write(fila, 4, d["ppto_inicial"], money_style)
        ws_ing.write(fila, 5, d["ppto_definitivo"], money_style)
        ws_ing.write(fila, 6, d["recaudado"], money_style)
        fila += 1

    # --- Hoja 2: Gastos_Presupuestales ---
    ws_gas = wb.add_sheet("Gastos_Presupuestales")
    encabezados_gas = ["CODIGO_DANE", "ANIO", "TRIMESTRE", "FUENTE", "ITEM",
                       "PPTO_INICIAL", "PPTO_DEFINITIVO", "COMPROMISOS",
                       "OBLIGACIONES", "PAGOS"]
    for col, h in enumerate(encabezados_gas):
        ws_gas.write(0, col, h, header_style)

    datos_gas = informe_sifse_gastos(trimestre, anio)
    fila = 1
    for d in datos_gas:
        if d["fuente"] == 0 and d["item"] == 0:
            continue
        ws_gas.write(fila, 0, codigo_dane)
        ws_gas.write(fila, 1, int(anio))
        ws_gas.write(fila, 2, trimestre)
        ws_gas.write(fila, 3, d["fuente"])
        ws_gas.write(fila, 4, d["item"])
        ws_gas.write(fila, 5, d["ppto_inicial"], money_style)
        ws_gas.write(fila, 6, d["ppto_definitivo"], money_style)
        ws_gas.write(fila, 7, d["compromisos"], money_style)
        ws_gas.write(fila, 8, d["obligaciones"], money_style)
        ws_gas.write(fila, 9, d["pagos"], money_style)
        fila += 1

    wb.save(destino)
    return destino


# ===================== CONFIG =====================
def get_config(clave):
    conn = get_connection()
    r = conn.execute("SELECT valor FROM config WHERE clave=?", (clave,)).fetchone()
    conn.close()
    return r["valor"] if r else None


def set_config(clave, valor):
    conn = get_connection()
    conn.execute("INSERT OR REPLACE INTO config VALUES (?, ?)", (clave, str(valor)))
    conn.commit()
    conn.close()


def get_consecutivo(tipo):
    clave = f"consecutivo_{tipo.lower()}"
    conn = get_connection()
    r = conn.execute("SELECT valor FROM config WHERE clave=?", (clave,)).fetchone()
    nuevo = int(r["valor"]) + 1 if r else 1
    conn.execute("UPDATE config SET valor=? WHERE clave=?", (str(nuevo), clave))
    conn.commit()
    conn.close()
    return nuevo


# ===================== RUBROS =====================
def get_rubros_gastos(solo_hojas=False):
    conn = get_connection()
    if solo_hojas:
        rows = conn.execute(
            "SELECT codigo, cuenta, apropiacion_definitiva FROM rubros_gastos WHERE es_hoja=1 ORDER BY codigo"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM rubros_gastos ORDER BY codigo"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_rubro_gasto(codigo):
    conn = get_connection()
    r = conn.execute("SELECT * FROM rubros_gastos WHERE codigo=?", (codigo,)).fetchone()
    conn.close()
    return dict(r) if r else None


def crear_rubro_gasto(codigo, cuenta, apropiacion_definitiva=0, apropiacion_inicial=0):
    """Crea un nuevo rubro de gasto."""
    conn = get_connection()
    existe = conn.execute("SELECT codigo FROM rubros_gastos WHERE codigo=?", (codigo,)).fetchone()
    if existe:
        conn.close()
        raise ValueError(f"El rubro {codigo} ya existe")
    conn.execute(
        "INSERT INTO rubros_gastos (codigo, cuenta, es_hoja, apropiacion_inicial, "
        "adiciones, reducciones, creditos, contracreditos, apropiacion_definitiva) "
        "VALUES (?,?,1,?,0,0,0,0,?)",
        (codigo, cuenta, apropiacion_inicial, apropiacion_definitiva)
    )
    # Si tiene padre, marcarlo como no-hoja
    partes = codigo.rsplit(".", 1)
    if len(partes) > 1:
        padre = partes[0]
        conn.execute("UPDATE rubros_gastos SET es_hoja=0 WHERE codigo=?", (padre,))
    # Recalcular es_hoja para todos
    _recalcular_hojas(conn)
    conn.commit()
    conn.close()


def editar_rubro_gasto(codigo, cuenta=None, apropiacion_definitiva=None, apropiacion_inicial=None):
    """Edita un rubro existente."""
    conn = get_connection()
    rubro = conn.execute("SELECT * FROM rubros_gastos WHERE codigo=?", (codigo,)).fetchone()
    if not rubro:
        conn.close()
        raise ValueError(f"Rubro {codigo} no encontrado")
    if cuenta is not None:
        conn.execute("UPDATE rubros_gastos SET cuenta=? WHERE codigo=?", (cuenta, codigo))
    if apropiacion_inicial is not None:
        conn.execute("UPDATE rubros_gastos SET apropiacion_inicial=? WHERE codigo=?",
                     (apropiacion_inicial, codigo))
    if apropiacion_definitiva is not None:
        conn.execute("UPDATE rubros_gastos SET apropiacion_definitiva=? WHERE codigo=?",
                     (apropiacion_definitiva, codigo))
        # Sincronizar inicial si no hay modificaciones previas
        if apropiacion_inicial is None:
            r = conn.execute("SELECT adiciones, reducciones, creditos, contracreditos "
                             "FROM rubros_gastos WHERE codigo=?", (codigo,)).fetchone()
            if r["adiciones"] == 0 and r["reducciones"] == 0 and r["creditos"] == 0 and r["contracreditos"] == 0:
                conn.execute("UPDATE rubros_gastos SET apropiacion_inicial=? WHERE codigo=?",
                             (apropiacion_definitiva, codigo))
    elif apropiacion_inicial is not None:
        # Si solo se cambio el inicial, recalcular el definitivo
        _recalcular_apropiacion_definitiva_gasto(conn, codigo)
    conn.commit()
    conn.close()


def eliminar_rubro_gasto(codigo):
    """Elimina un rubro si no tiene transacciones."""
    conn = get_connection()
    # Verificar que no tenga CDPs
    tiene_cdp = conn.execute(
        "SELECT COUNT(*) as c FROM cdp WHERE codigo_rubro=?", (codigo,)
    ).fetchone()["c"]
    if tiene_cdp > 0:
        conn.close()
        raise ValueError(f"Rubro {codigo} tiene {tiene_cdp} CDP(s) registrados")
    # Verificar que no tenga hijos
    tiene_hijos = conn.execute(
        "SELECT COUNT(*) as c FROM rubros_gastos WHERE codigo LIKE ?", (codigo + ".%",)
    ).fetchone()["c"]
    if tiene_hijos > 0:
        conn.close()
        raise ValueError(f"Rubro {codigo} tiene {tiene_hijos} sub-rubro(s)")
    conn.execute("DELETE FROM rubros_gastos WHERE codigo=?", (codigo,))
    _recalcular_hojas(conn)
    conn.commit()
    conn.close()


def _recalcular_hojas(conn):
    """Recalcula es_hoja para todos los rubros."""
    codigos = [r["codigo"] for r in conn.execute("SELECT codigo FROM rubros_gastos").fetchall()]
    for codigo in codigos:
        es_hoja = 1
        prefijo = codigo + "."
        for otro in codigos:
            if otro.startswith(prefijo):
                es_hoja = 0
                break
        conn.execute("UPDATE rubros_gastos SET es_hoja=? WHERE codigo=?", (es_hoja, codigo))


def buscar_rubros_gastos(filtro):
    conn = get_connection()
    filtro_like = f"%{filtro}%"
    rows = conn.execute(
        "SELECT codigo, cuenta, apropiacion_definitiva FROM rubros_gastos "
        "WHERE es_hoja=1 AND (codigo LIKE ? OR cuenta LIKE ?) ORDER BY codigo",
        (filtro_like, filtro_like)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ===================== RUBROS INGRESOS =====================
def get_rubros_ingresos(solo_hojas=False):
    conn = get_connection()
    if solo_hojas:
        rows = conn.execute(
            "SELECT codigo, cuenta, presupuesto_definitivo FROM rubros_ingresos WHERE es_hoja=1 ORDER BY codigo"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM rubros_ingresos ORDER BY codigo"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_rubro_ingreso(codigo):
    conn = get_connection()
    r = conn.execute("SELECT * FROM rubros_ingresos WHERE codigo=?", (codigo,)).fetchone()
    conn.close()
    return dict(r) if r else None


def crear_rubro_ingreso(codigo, cuenta, ppto_definitivo=0, ppto_inicial=0):
    conn = get_connection()
    existe = conn.execute("SELECT codigo FROM rubros_ingresos WHERE codigo=?", (codigo,)).fetchone()
    if existe:
        conn.close()
        raise ValueError(f"El rubro {codigo} ya existe")
    conn.execute(
        "INSERT INTO rubros_ingresos (codigo, cuenta, es_hoja, presupuesto_inicial, "
        "adiciones, reducciones, presupuesto_definitivo) "
        "VALUES (?,?,1,?,0,0,?)",
        (codigo, cuenta, ppto_inicial, ppto_definitivo)
    )
    _recalcular_hojas_ingresos(conn)
    conn.commit()
    conn.close()


def editar_rubro_ingreso(codigo, cuenta=None, ppto_definitivo=None, ppto_inicial=None):
    conn = get_connection()
    rubro = conn.execute("SELECT * FROM rubros_ingresos WHERE codigo=?", (codigo,)).fetchone()
    if not rubro:
        conn.close()
        raise ValueError(f"Rubro {codigo} no encontrado")
    if cuenta is not None:
        conn.execute("UPDATE rubros_ingresos SET cuenta=? WHERE codigo=?", (cuenta, codigo))
    if ppto_inicial is not None:
        conn.execute("UPDATE rubros_ingresos SET presupuesto_inicial=? WHERE codigo=?",
                     (ppto_inicial, codigo))
    if ppto_definitivo is not None:
        conn.execute("UPDATE rubros_ingresos SET presupuesto_definitivo=? WHERE codigo=?",
                     (ppto_definitivo, codigo))
        # Sincronizar inicial: si el definitivo se edita directamente y no hay
        # adiciones ni reducciones, el inicial debe ser igual al definitivo
        if ppto_inicial is None:
            r = conn.execute("SELECT adiciones, reducciones FROM rubros_ingresos WHERE codigo=?",
                             (codigo,)).fetchone()
            if r["adiciones"] == 0 and r["reducciones"] == 0:
                conn.execute("UPDATE rubros_ingresos SET presupuesto_inicial=? WHERE codigo=?",
                             (ppto_definitivo, codigo))
    elif ppto_inicial is not None:
        # Si solo se cambio el inicial, recalcular el definitivo
        _recalcular_presupuesto_definitivo_ingreso(conn, codigo)
    conn.commit()
    conn.close()


def eliminar_rubro_ingreso(codigo):
    conn = get_connection()
    tiene_recaudos = conn.execute(
        "SELECT COUNT(*) as c FROM recaudo WHERE codigo_rubro=?", (codigo,)
    ).fetchone()["c"]
    if tiene_recaudos > 0:
        conn.close()
        raise ValueError(f"Rubro {codigo} tiene {tiene_recaudos} recaudo(s) registrados")
    tiene_hijos = conn.execute(
        "SELECT COUNT(*) as c FROM rubros_ingresos WHERE codigo LIKE ?", (codigo + ".%",)
    ).fetchone()["c"]
    if tiene_hijos > 0:
        conn.close()
        raise ValueError(f"Rubro {codigo} tiene {tiene_hijos} sub-rubro(s)")
    conn.execute("DELETE FROM rubros_ingresos WHERE codigo=?", (codigo,))
    _recalcular_hojas_ingresos(conn)
    conn.commit()
    conn.close()


def _recalcular_hojas_ingresos(conn):
    codigos = [r["codigo"] for r in conn.execute("SELECT codigo FROM rubros_ingresos").fetchall()]
    for codigo in codigos:
        es_hoja = 1
        prefijo = codigo + "."
        for otro in codigos:
            if otro.startswith(prefijo):
                es_hoja = 0
                break
        conn.execute("UPDATE rubros_ingresos SET es_hoja=? WHERE codigo=?", (es_hoja, codigo))


def sincronizar_padres_ingresos():
    """Actualiza los valores almacenados de rubros padre sumando sus hojas."""
    conn = get_connection()
    padres = conn.execute(
        "SELECT codigo FROM rubros_ingresos WHERE es_hoja=0 ORDER BY codigo DESC"
    ).fetchall()
    for p in padres:
        codigo = p["codigo"]
        r = conn.execute(
            "SELECT COALESCE(SUM(presupuesto_inicial),0) as pi, "
            "COALESCE(SUM(adiciones),0) as ad, "
            "COALESCE(SUM(reducciones),0) as re, "
            "COALESCE(SUM(presupuesto_definitivo),0) as pd "
            "FROM rubros_ingresos WHERE es_hoja=1 AND codigo LIKE ?",
            (codigo + ".%",)
        ).fetchone()
        conn.execute(
            "UPDATE rubros_ingresos SET presupuesto_inicial=?, adiciones=?, "
            "reducciones=?, presupuesto_definitivo=? WHERE codigo=?",
            (r["pi"], r["ad"], r["re"], r["pd"], codigo)
        )
    conn.commit()
    conn.close()


def sincronizar_padres_gastos():
    """Actualiza los valores almacenados de rubros padre de gastos sumando sus hojas."""
    conn = get_connection()
    padres = conn.execute(
        "SELECT codigo FROM rubros_gastos WHERE es_hoja=0 ORDER BY codigo DESC"
    ).fetchall()
    for p in padres:
        codigo = p["codigo"]
        r = conn.execute(
            "SELECT COALESCE(SUM(apropiacion_inicial),0) as ai, "
            "COALESCE(SUM(adiciones),0) as ad, COALESCE(SUM(reducciones),0) as re, "
            "COALESCE(SUM(creditos),0) as cr, COALESCE(SUM(contracreditos),0) as cc, "
            "COALESCE(SUM(apropiacion_definitiva),0) as apd "
            "FROM rubros_gastos WHERE es_hoja=1 AND codigo LIKE ?",
            (codigo + ".%",)
        ).fetchone()
        conn.execute(
            "UPDATE rubros_gastos SET apropiacion_inicial=?, adiciones=?, reducciones=?, "
            "creditos=?, contracreditos=?, apropiacion_definitiva=? WHERE codigo=?",
            (r["ai"], r["ad"], r["re"], r["cr"], r["cc"], r["apd"], codigo)
        )
    conn.commit()
    conn.close()


def editar_rubro_ingreso_padre(codigo, cuenta=None, ppto_definitivo=None):
    """Edita un rubro padre: propaga el cambio de valor a los rubros hoja proporcionalmente."""
    conn = get_connection()
    rubro = conn.execute("SELECT * FROM rubros_ingresos WHERE codigo=?", (codigo,)).fetchone()
    if not rubro:
        conn.close()
        raise ValueError(f"Rubro {codigo} no encontrado")
    if cuenta is not None:
        conn.execute("UPDATE rubros_ingresos SET cuenta=? WHERE codigo=?", (cuenta, codigo))
    if ppto_definitivo is not None:
        # Obtener hojas hijas
        hojas = conn.execute(
            "SELECT codigo, presupuesto_inicial, presupuesto_definitivo, adiciones, reducciones "
            "FROM rubros_ingresos WHERE es_hoja=1 AND codigo LIKE ?",
            (codigo + ".%",)
        ).fetchall()
        if not hojas:
            conn.close()
            raise ValueError(f"Rubro padre {codigo} no tiene rubros hoja")
        total_actual = sum(h["presupuesto_definitivo"] for h in hojas)
        diferencia = ppto_definitivo - total_actual
        if diferencia != 0:
            if total_actual > 0:
                # Distribuir proporcionalmente entre hojas con valor
                for h in hojas:
                    if h["presupuesto_definitivo"] > 0:
                        proporcion = h["presupuesto_definitivo"] / total_actual
                        nuevo_def = h["presupuesto_definitivo"] + (diferencia * proporcion)
                        nuevo_ini = nuevo_def - h["adiciones"] + h["reducciones"]
                        conn.execute(
                            "UPDATE rubros_ingresos SET presupuesto_definitivo=?, presupuesto_inicial=? "
                            "WHERE codigo=?", (nuevo_def, nuevo_ini, h["codigo"])
                        )
            else:
                # Si ninguna hoja tiene valor, asignar todo a la primera hoja
                h = hojas[0]
                conn.execute(
                    "UPDATE rubros_ingresos SET presupuesto_definitivo=?, presupuesto_inicial=? "
                    "WHERE codigo=?", (ppto_definitivo, ppto_definitivo, h["codigo"])
                )
    conn.commit()
    conn.close()


def editar_rubro_gasto_padre(codigo, cuenta=None, aprop_definitiva=None):
    """Edita un rubro padre de gastos: propaga el cambio a rubros hoja proporcionalmente."""
    conn = get_connection()
    rubro = conn.execute("SELECT * FROM rubros_gastos WHERE codigo=?", (codigo,)).fetchone()
    if not rubro:
        conn.close()
        raise ValueError(f"Rubro {codigo} no encontrado")
    if cuenta is not None:
        conn.execute("UPDATE rubros_gastos SET cuenta=? WHERE codigo=?", (cuenta, codigo))
    if aprop_definitiva is not None:
        hojas = conn.execute(
            "SELECT codigo, apropiacion_inicial, apropiacion_definitiva, "
            "adiciones, reducciones, creditos, contracreditos "
            "FROM rubros_gastos WHERE es_hoja=1 AND codigo LIKE ?",
            (codigo + ".%",)
        ).fetchall()
        if not hojas:
            conn.close()
            raise ValueError(f"Rubro padre {codigo} no tiene rubros hoja")
        total_actual = sum(h["apropiacion_definitiva"] for h in hojas)
        diferencia = aprop_definitiva - total_actual
        if diferencia != 0:
            if total_actual > 0:
                for h in hojas:
                    if h["apropiacion_definitiva"] > 0:
                        proporcion = h["apropiacion_definitiva"] / total_actual
                        nuevo_def = h["apropiacion_definitiva"] + (diferencia * proporcion)
                        nuevo_ini = (nuevo_def - h["adiciones"] + h["reducciones"]
                                     - h["creditos"] + h["contracreditos"])
                        conn.execute(
                            "UPDATE rubros_gastos SET apropiacion_definitiva=?, apropiacion_inicial=? "
                            "WHERE codigo=?", (nuevo_def, nuevo_ini, h["codigo"])
                        )
            else:
                h = hojas[0]
                conn.execute(
                    "UPDATE rubros_gastos SET apropiacion_definitiva=?, apropiacion_inicial=? "
                    "WHERE codigo=?", (aprop_definitiva, aprop_definitiva, h["codigo"])
                )
    conn.commit()
    conn.close()


def buscar_rubros_ingresos(filtro):
    conn = get_connection()
    filtro_like = f"%{filtro}%"
    rows = conn.execute(
        "SELECT codigo, cuenta, presupuesto_definitivo FROM rubros_ingresos "
        "WHERE es_hoja=1 AND (codigo LIKE ? OR cuenta LIKE ?) ORDER BY codigo",
        (filtro_like, filtro_like)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ===================== TERCEROS =====================
def get_terceros():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM terceros ORDER BY nombre").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_tercero(nit):
    conn = get_connection()
    r = conn.execute("SELECT * FROM terceros WHERE nit=?", (nit,)).fetchone()
    conn.close()
    return dict(r) if r else None


def guardar_tercero(nit, dv, nombre, direccion="", telefono="", email="",
                    tipo="Natural", banco="", tipo_cuenta="", no_cuenta=""):
    conn = get_connection()
    conn.execute(
        "INSERT OR REPLACE INTO terceros VALUES (?,?,?,?,?,?,?,?,?,?)",
        (nit, dv, nombre.upper(), direccion, telefono, email, tipo, banco, tipo_cuenta, no_cuenta)
    )
    conn.commit()
    conn.close()


def buscar_terceros(filtro):
    conn = get_connection()
    filtro_like = f"%{filtro}%"
    rows = conn.execute(
        "SELECT * FROM terceros WHERE nit LIKE ? OR nombre LIKE ? ORDER BY nombre",
        (filtro_like, filtro_like)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ===================== CDP =====================
def saldo_disponible_rubro(codigo_rubro):
    conn = get_connection()
    r = conn.execute(
        "SELECT apropiacion_definitiva FROM rubros_gastos WHERE codigo=?",
        (codigo_rubro,)
    ).fetchone()
    aprop = r["apropiacion_definitiva"] if r else 0

    r2 = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM cdp WHERE codigo_rubro=? AND estado<>'ANULADO'",
        (codigo_rubro,)
    ).fetchone()
    total_cdp = r2["total"]
    conn.close()
    return aprop - total_cdp


def saldo_cdp(numero_cdp):
    conn = get_connection()
    r = conn.execute("SELECT valor FROM cdp WHERE numero=?", (numero_cdp,)).fetchone()
    valor_cdp = r["valor"] if r else 0

    r2 = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM rp WHERE cdp_numero=? AND estado<>'ANULADO'",
        (numero_cdp,)
    ).fetchone()
    total_rp = r2["total"]
    conn.close()
    return valor_cdp - total_rp


def registrar_cdp(codigo_rubro, objeto, valor, fuente_sifse=0):
    saldo = saldo_disponible_rubro(codigo_rubro)
    if valor > saldo:
        raise ValueError(f"Valor ({valor:,.0f}) excede saldo disponible ({saldo:,.0f})")
    if valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")

    numero = get_consecutivo("cdp")
    fecha = date.today().strftime("%Y-%m-%d")
    conn = get_connection()
    conn.execute(
        "INSERT INTO cdp (numero, fecha, codigo_rubro, objeto, valor, estado, fuente_sifse) "
        "VALUES (?,?,?,?,?,?,?)",
        (numero, fecha, codigo_rubro, objeto, valor, "ACTIVO", fuente_sifse)
    )
    conn.commit()
    conn.close()
    guardar_concepto(codigo_rubro, objeto)
    return numero, fecha


def get_cdps(estado=None):
    conn = get_connection()
    if estado:
        rows = conn.execute(
            "SELECT c.*, rg.cuenta FROM cdp c JOIN rubros_gastos rg ON c.codigo_rubro=rg.codigo "
            "WHERE c.estado=? ORDER BY c.numero DESC", (estado,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT c.*, rg.cuenta FROM cdp c JOIN rubros_gastos rg ON c.codigo_rubro=rg.codigo "
            "ORDER BY c.numero DESC"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_cdp(numero):
    conn = get_connection()
    r = conn.execute(
        "SELECT c.*, rg.cuenta FROM cdp c JOIN rubros_gastos rg ON c.codigo_rubro=rg.codigo WHERE c.numero=?",
        (numero,)
    ).fetchone()
    conn.close()
    return dict(r) if r else None


def anular_cdp(numero):
    conn = get_connection()
    tiene_rps = conn.execute(
        "SELECT COUNT(*) as cnt FROM rp WHERE cdp_numero=? AND estado<>'ANULADO'", (numero,)
    ).fetchone()["cnt"]
    if tiene_rps > 0:
        conn.close()
        raise ValueError(f"CDP {numero} tiene {tiene_rps} RP(s) activos")
    conn.execute("UPDATE cdp SET estado='ANULADO' WHERE numero=?", (numero,))
    conn.commit()
    conn.close()


def actualizar_estado_cdp(numero):
    s = saldo_cdp(numero)
    conn = get_connection()
    if s <= 0:
        conn.execute("UPDATE cdp SET estado='AGOTADO' WHERE numero=? AND estado='ACTIVO'", (numero,))
    elif s > 0:
        conn.execute("UPDATE cdp SET estado='ACTIVO' WHERE numero=? AND estado='AGOTADO'", (numero,))
    conn.commit()
    conn.close()


def editar_valor_cdp(numero, nuevo_valor):
    conn = get_connection()
    cdp = conn.execute("SELECT * FROM cdp WHERE numero=?", (numero,)).fetchone()
    if not cdp:
        conn.close()
        raise ValueError(f"CDP {numero} no encontrado")
    if cdp["estado"] == "ANULADO":
        conn.close()
        raise ValueError(f"CDP {numero} esta anulado")
    # Valor ya utilizado en RPs
    usado = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM rp WHERE cdp_numero=? AND estado<>'ANULADO'",
        (numero,)
    ).fetchone()["t"]
    if nuevo_valor < usado:
        conn.close()
        raise ValueError(f"El nuevo valor ($ {nuevo_valor:,.0f}) no puede ser menor al valor "
                         f"ya comprometido en RPs ($ {usado:,.0f})")
    if nuevo_valor <= 0:
        conn.close()
        raise ValueError("El valor debe ser mayor a cero")
    # Verificar que no exceda apropiacion
    saldo_rub = saldo_disponible_rubro(cdp["codigo_rubro"])
    disponible = saldo_rub + cdp["valor"]  # devolver el valor actual al saldo
    if nuevo_valor > disponible:
        conn.close()
        raise ValueError(f"El nuevo valor ($ {nuevo_valor:,.0f}) excede el saldo disponible "
                         f"del rubro ($ {disponible:,.0f})")
    conn.execute("UPDATE cdp SET valor=? WHERE numero=?", (nuevo_valor, numero))
    conn.commit()
    conn.close()
    actualizar_estado_cdp(numero)


def editar_cdp(numero, nuevo_valor=None, objeto=None, fuente_sifse=None, item_sifse=None):
    """Edita multiples campos de un CDP: valor, objeto, fuente_sifse, item_sifse."""
    conn = get_connection()
    cdp = conn.execute("SELECT * FROM cdp WHERE numero=?", (numero,)).fetchone()
    if not cdp:
        conn.close()
        raise ValueError(f"CDP {numero} no encontrado")
    if cdp["estado"] == "ANULADO":
        conn.close()
        raise ValueError(f"CDP {numero} esta anulado")

    # Validar valor si se cambia
    if nuevo_valor is not None:
        usado = conn.execute(
            "SELECT COALESCE(SUM(valor), 0) as t FROM rp WHERE cdp_numero=? AND estado<>'ANULADO'",
            (numero,)
        ).fetchone()["t"]
        if nuevo_valor < usado:
            conn.close()
            raise ValueError(f"El nuevo valor ($ {nuevo_valor:,.0f}) no puede ser menor al valor "
                             f"ya comprometido en RPs ($ {usado:,.0f})")
        if nuevo_valor <= 0:
            conn.close()
            raise ValueError("El valor debe ser mayor a cero")
        saldo_rub = saldo_disponible_rubro(cdp["codigo_rubro"])
        disponible = saldo_rub + cdp["valor"]
        if nuevo_valor > disponible:
            conn.close()
            raise ValueError(f"El nuevo valor ($ {nuevo_valor:,.0f}) excede el saldo disponible "
                             f"del rubro ($ {disponible:,.0f})")

    # Construir UPDATE dinamico
    campos = []
    valores = []
    if nuevo_valor is not None:
        campos.append("valor=?")
        valores.append(nuevo_valor)
    if objeto is not None:
        campos.append("objeto=?")
        valores.append(objeto)
    if fuente_sifse is not None:
        campos.append("fuente_sifse=?")
        valores.append(fuente_sifse)
    if item_sifse is not None:
        campos.append("item_sifse=?")
        valores.append(item_sifse)

    if campos:
        valores.append(numero)
        conn.execute(f"UPDATE cdp SET {', '.join(campos)} WHERE numero=?", valores)
        conn.commit()
    conn.close()
    if nuevo_valor is not None:
        actualizar_estado_cdp(numero)


# ===================== RP =====================
def saldo_rp(numero_rp):
    conn = get_connection()
    r = conn.execute("SELECT valor FROM rp WHERE numero=?", (numero_rp,)).fetchone()
    valor_rp = r["valor"] if r else 0
    r2 = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM obligacion WHERE rp_numero=? AND estado<>'ANULADA'",
        (numero_rp,)
    ).fetchone()
    conn.close()
    return valor_rp - r2["total"]


def registrar_rp(cdp_numero, nit_tercero, valor, objeto):
    cdp = get_cdp(cdp_numero)
    if not cdp or cdp["estado"] == "ANULADO":
        raise ValueError(f"CDP {cdp_numero} no existe o esta anulado")
    saldo = saldo_cdp(cdp_numero)
    if valor > saldo:
        raise ValueError(f"Valor ({valor:,.0f}) excede saldo del CDP ({saldo:,.0f})")
    if not get_tercero(nit_tercero):
        raise ValueError(f"Tercero {nit_tercero} no existe")

    numero = get_consecutivo("rp")
    fecha = date.today().strftime("%Y-%m-%d")
    fuente_sifse = cdp.get("fuente_sifse", 0) or 0
    conn = get_connection()
    conn.execute(
        "INSERT INTO rp (numero, fecha, cdp_numero, codigo_rubro, nit_tercero, valor, objeto, estado, fuente_sifse) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (numero, fecha, cdp_numero, cdp["codigo_rubro"], nit_tercero, valor, objeto, "ACTIVO", fuente_sifse)
    )
    conn.commit()
    conn.close()
    actualizar_estado_cdp(cdp_numero)
    return numero, fecha


def get_rp(numero):
    conn = get_connection()
    r = conn.execute(
        "SELECT r.*, rg.cuenta, t.nombre as nombre_tercero FROM rp r "
        "JOIN rubros_gastos rg ON r.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON r.nit_tercero=t.nit WHERE r.numero=?",
        (numero,)
    ).fetchone()
    conn.close()
    return dict(r) if r else None


def get_rps(estado=None):
    conn = get_connection()
    q = ("SELECT r.*, rg.cuenta, t.nombre as nombre_tercero FROM rp r "
         "JOIN rubros_gastos rg ON r.codigo_rubro=rg.codigo "
         "LEFT JOIN terceros t ON r.nit_tercero=t.nit ")
    if estado:
        q += f"WHERE r.estado='{estado}' "
    q += "ORDER BY r.numero DESC"
    rows = conn.execute(q).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def anular_rp(numero):
    conn = get_connection()
    tiene_obl = conn.execute(
        "SELECT COUNT(*) as cnt FROM obligacion WHERE rp_numero=? AND estado<>'ANULADA'", (numero,)
    ).fetchone()["cnt"]
    if tiene_obl > 0:
        conn.close()
        raise ValueError(f"RP {numero} tiene {tiene_obl} obligacion(es) activas")
    rp = get_rp(numero)
    conn.execute("UPDATE rp SET estado='ANULADO' WHERE numero=?", (numero,))
    # Reactivar CDP
    conn.execute("UPDATE cdp SET estado='ACTIVO' WHERE numero=? AND estado='AGOTADO'", (rp["cdp_numero"],))
    conn.commit()
    conn.close()


def actualizar_estado_rp(numero):
    s = saldo_rp(numero)
    conn = get_connection()
    if s <= 0:
        conn.execute("UPDATE rp SET estado='AGOTADO' WHERE numero=? AND estado='ACTIVO'", (numero,))
    elif s > 0:
        conn.execute("UPDATE rp SET estado='ACTIVO' WHERE numero=? AND estado='AGOTADO'", (numero,))
    conn.commit()
    conn.close()


def editar_valor_rp(numero, nuevo_valor):
    rp = get_rp(numero)
    if not rp:
        raise ValueError(f"RP {numero} no encontrado")
    if rp["estado"] == "ANULADO":
        raise ValueError(f"RP {numero} esta anulado")
    # Valor ya usado en obligaciones
    conn = get_connection()
    usado = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM obligacion WHERE rp_numero=? AND estado<>'ANULADA'",
        (numero,)
    ).fetchone()["t"]
    conn.close()
    if nuevo_valor < usado:
        raise ValueError(f"No puede ser menor al valor obligado ($ {usado:,.0f})")
    if nuevo_valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")
    # No puede exceder saldo del CDP
    saldo_c = saldo_cdp(rp["cdp_numero"]) + rp["valor"]  # devolver valor actual
    if nuevo_valor > saldo_c:
        raise ValueError(f"Excede saldo del CDP ($ {saldo_c:,.0f})")
    conn = get_connection()
    conn.execute("UPDATE rp SET valor=? WHERE numero=?", (nuevo_valor, numero))
    conn.commit()
    conn.close()
    actualizar_estado_rp(numero)
    actualizar_estado_cdp(rp["cdp_numero"])


def editar_rp(numero, nuevo_valor=None, objeto=None, fuente_sifse=None, item_sifse=None):
    """Edita multiples campos de un RP: valor, objeto, fuente_sifse, item_sifse."""
    rp = get_rp(numero)
    if not rp:
        raise ValueError(f"RP {numero} no encontrado")
    if rp["estado"] == "ANULADO":
        raise ValueError(f"RP {numero} esta anulado")
    if nuevo_valor is not None:
        conn = get_connection()
        usado = conn.execute(
            "SELECT COALESCE(SUM(valor), 0) as t FROM obligacion WHERE rp_numero=? AND estado<>'ANULADA'",
            (numero,)
        ).fetchone()["t"]
        conn.close()
        if nuevo_valor < usado:
            raise ValueError(f"No puede ser menor al valor obligado ($ {usado:,.0f})")
        if nuevo_valor <= 0:
            raise ValueError("El valor debe ser mayor a cero")
        saldo_c = saldo_cdp(rp["cdp_numero"]) + rp["valor"]
        if nuevo_valor > saldo_c:
            raise ValueError(f"Excede saldo del CDP ($ {saldo_c:,.0f})")
    campos = []
    valores = []
    if nuevo_valor is not None:
        campos.append("valor=?")
        valores.append(nuevo_valor)
    if objeto is not None:
        campos.append("objeto=?")
        valores.append(objeto)
    if fuente_sifse is not None:
        campos.append("fuente_sifse=?")
        valores.append(fuente_sifse)
    if item_sifse is not None:
        campos.append("item_sifse=?")
        valores.append(item_sifse)
    if campos:
        valores.append(numero)
        conn = get_connection()
        conn.execute(f"UPDATE rp SET {', '.join(campos)} WHERE numero=?", valores)
        conn.commit()
        conn.close()
    if nuevo_valor is not None:
        actualizar_estado_rp(numero)
        actualizar_estado_cdp(rp["cdp_numero"])


def editar_valor_obligacion(numero, nuevo_valor):
    obl = get_obligacion(numero)
    if not obl:
        raise ValueError(f"Obligacion {numero} no encontrada")
    if obl["estado"] == "ANULADA":
        raise ValueError(f"Obligacion {numero} esta anulada")
    conn = get_connection()
    pagado = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM pago WHERE obligacion_numero=? AND estado<>'ANULADO'",
        (numero,)
    ).fetchone()["t"]
    conn.close()
    if nuevo_valor < pagado:
        raise ValueError(f"No puede ser menor al valor pagado ($ {pagado:,.0f})")
    if nuevo_valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")
    saldo_r = saldo_rp(obl["rp_numero"]) + obl["valor"]
    if nuevo_valor > saldo_r:
        raise ValueError(f"Excede saldo del RP ($ {saldo_r:,.0f})")
    conn = get_connection()
    conn.execute("UPDATE obligacion SET valor=? WHERE numero=?", (nuevo_valor, numero))
    conn.commit()
    conn.close()
    actualizar_estado_rp(obl["rp_numero"])


def editar_obligacion(numero, nuevo_valor=None, factura=None, fuente_sifse=None, item_sifse=None):
    """Edita multiples campos de una Obligacion: valor, factura, fuente_sifse, item_sifse."""
    obl = get_obligacion(numero)
    if not obl:
        raise ValueError(f"Obligacion {numero} no encontrada")
    if obl["estado"] == "ANULADA":
        raise ValueError(f"Obligacion {numero} esta anulada")
    if nuevo_valor is not None:
        conn = get_connection()
        pagado = conn.execute(
            "SELECT COALESCE(SUM(valor), 0) as t FROM pago WHERE obligacion_numero=? AND estado<>'ANULADO'",
            (numero,)
        ).fetchone()["t"]
        conn.close()
        if nuevo_valor < pagado:
            raise ValueError(f"No puede ser menor al valor pagado ($ {pagado:,.0f})")
        if nuevo_valor <= 0:
            raise ValueError("El valor debe ser mayor a cero")
        saldo_r = saldo_rp(obl["rp_numero"]) + obl["valor"]
        if nuevo_valor > saldo_r:
            raise ValueError(f"Excede saldo del RP ($ {saldo_r:,.0f})")
    campos = []
    valores = []
    if nuevo_valor is not None:
        campos.append("valor=?")
        valores.append(nuevo_valor)
    if factura is not None:
        campos.append("factura=?")
        valores.append(factura)
    if fuente_sifse is not None:
        campos.append("fuente_sifse=?")
        valores.append(fuente_sifse)
    if item_sifse is not None:
        campos.append("item_sifse=?")
        valores.append(item_sifse)
    if campos:
        valores.append(numero)
        conn = get_connection()
        conn.execute(f"UPDATE obligacion SET {', '.join(campos)} WHERE numero=?", valores)
        conn.commit()
        conn.close()
    if nuevo_valor is not None:
        actualizar_estado_rp(obl["rp_numero"])


# ===================== OBLIGACION =====================
def saldo_obligacion(numero_obl):
    conn = get_connection()
    r = conn.execute("SELECT valor FROM obligacion WHERE numero=?", (numero_obl,)).fetchone()
    valor_obl = r["valor"] if r else 0
    r2 = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM pago WHERE obligacion_numero=? AND estado<>'ANULADO'",
        (numero_obl,)
    ).fetchone()
    conn.close()
    return valor_obl - r2["total"]


def registrar_obligacion(rp_numero, valor, factura):
    rp = get_rp(rp_numero)
    if not rp or rp["estado"] == "ANULADO":
        raise ValueError(f"RP {rp_numero} no existe o esta anulado")
    saldo = saldo_rp(rp_numero)
    if valor > saldo:
        raise ValueError(f"Valor ({valor:,.0f}) excede saldo del RP ({saldo:,.0f})")

    numero = get_consecutivo("obligacion")
    fecha = date.today().strftime("%Y-%m-%d")
    fuente_sifse = rp.get("fuente_sifse", 0) or 0
    conn = get_connection()
    conn.execute(
        "INSERT INTO obligacion (numero, fecha, rp_numero, codigo_rubro, nit_tercero, valor, factura, estado, fuente_sifse) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (numero, fecha, rp_numero, rp["codigo_rubro"], rp["nit_tercero"], valor, factura, "ACTIVO", fuente_sifse)
    )
    conn.commit()
    conn.close()
    actualizar_estado_rp(rp_numero)
    guardar_concepto(rp["codigo_rubro"], factura)
    return numero, fecha


def get_obligacion(numero):
    conn = get_connection()
    r = conn.execute(
        "SELECT o.*, rg.cuenta, t.nombre as nombre_tercero FROM obligacion o "
        "JOIN rubros_gastos rg ON o.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON o.nit_tercero=t.nit WHERE o.numero=?",
        (numero,)
    ).fetchone()
    conn.close()
    return dict(r) if r else None


def get_obligaciones(estado=None):
    conn = get_connection()
    q = ("SELECT o.*, rg.cuenta, t.nombre as nombre_tercero FROM obligacion o "
         "JOIN rubros_gastos rg ON o.codigo_rubro=rg.codigo "
         "LEFT JOIN terceros t ON o.nit_tercero=t.nit ")
    if estado:
        q += f"WHERE o.estado='{estado}' "
    q += "ORDER BY o.numero DESC"
    rows = conn.execute(q).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def anular_obligacion(numero):
    conn = get_connection()
    tiene_pagos = conn.execute(
        "SELECT COUNT(*) as cnt FROM pago WHERE obligacion_numero=? AND estado<>'ANULADO'", (numero,)
    ).fetchone()["cnt"]
    if tiene_pagos > 0:
        conn.close()
        raise ValueError(f"Obligacion {numero} tiene {tiene_pagos} pago(s) activos")
    obl = get_obligacion(numero)
    conn.execute("UPDATE obligacion SET estado='ANULADA' WHERE numero=?", (numero,))
    conn.execute("UPDATE rp SET estado='ACTIVO' WHERE numero=? AND estado='AGOTADO'", (obl["rp_numero"],))
    conn.commit()
    conn.close()


# ===================== PAGO =====================
def registrar_pago(obligacion_numero, valor, concepto, medio_pago, no_comprobante, cuenta_bancaria_id=0):
    obl = get_obligacion(obligacion_numero)
    if not obl or obl["estado"] == "ANULADA":
        raise ValueError(f"Obligacion {obligacion_numero} no existe o esta anulada")
    saldo = saldo_obligacion(obligacion_numero)
    if valor > saldo:
        raise ValueError(f"Valor ({valor:,.0f}) excede saldo de la obligacion ({saldo:,.0f})")

    # Validar cupo PAC del mes actual
    mes_actual = date.today().month
    pac_ok, pac_msg = validar_pago_pac(obl["codigo_rubro"], mes_actual, valor)
    if not pac_ok:
        raise ValueError(f"PAC: {pac_msg}")

    numero = get_consecutivo("pago")
    fecha = date.today().strftime("%Y-%m-%d")
    fuente_sifse = obl.get("fuente_sifse", 0) or 0
    conn = get_connection()
    conn.execute(
        "INSERT INTO pago (numero, fecha, obligacion_numero, codigo_rubro, nit_tercero, "
        "valor, concepto, medio_pago, no_comprobante, estado, fuente_sifse, cuenta_bancaria_id) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (numero, fecha, obligacion_numero, obl["codigo_rubro"], obl["nit_tercero"],
         valor, concepto, medio_pago, no_comprobante, "PAGADO", fuente_sifse, cuenta_bancaria_id)
    )
    # Actualizar estado obligacion
    conn.commit()
    conn.close()
    if saldo_obligacion(obligacion_numero) <= 0:
        conn = get_connection()
        conn.execute("UPDATE obligacion SET estado='PAGADA' WHERE numero=? AND estado='ACTIVO'",
                     (obligacion_numero,))
        conn.commit()
        conn.close()
    return numero, fecha


def get_pagos():
    conn = get_connection()
    rows = conn.execute(
        "SELECT p.*, rg.cuenta, t.nombre as nombre_tercero FROM pago p "
        "JOIN rubros_gastos rg ON p.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON p.nit_tercero=t.nit ORDER BY p.numero DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def anular_pago(numero):
    conn = get_connection()
    pago = conn.execute("SELECT * FROM pago WHERE numero=?", (numero,)).fetchone()
    if not pago:
        conn.close()
        raise ValueError(f"Pago {numero} no encontrado")
    conn.execute("UPDATE pago SET estado='ANULADO' WHERE numero=?", (numero,))
    conn.execute("UPDATE obligacion SET estado='ACTIVO' WHERE numero=? AND estado='PAGADA'",
                 (pago["obligacion_numero"],))
    conn.commit()
    conn.close()


def get_pago(numero):
    conn = get_connection()
    r = conn.execute(
        "SELECT p.*, rg.cuenta, t.nombre as nombre_tercero FROM pago p "
        "JOIN rubros_gastos rg ON p.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON p.nit_tercero=t.nit WHERE p.numero=?",
        (numero,)
    ).fetchone()
    conn.close()
    return dict(r) if r else None


def editar_valor_pago(numero, nuevo_valor):
    conn = get_connection()
    pago = conn.execute("SELECT * FROM pago WHERE numero=?", (numero,)).fetchone()
    if not pago:
        conn.close()
        raise ValueError(f"Pago {numero} no encontrado")
    if pago["estado"] == "ANULADO":
        conn.close()
        raise ValueError(f"Pago {numero} esta anulado")
    if nuevo_valor <= 0:
        conn.close()
        raise ValueError("El valor debe ser mayor a cero")
    # No puede exceder saldo de la obligacion
    saldo_o = saldo_obligacion(pago["obligacion_numero"]) + pago["valor"]
    if nuevo_valor > saldo_o:
        conn.close()
        raise ValueError(f"Excede saldo de la obligacion ($ {saldo_o:,.0f})")
    conn.execute("UPDATE pago SET valor=? WHERE numero=?", (nuevo_valor, numero))
    conn.commit()
    conn.close()


def editar_pago(numero, nuevo_valor=None, concepto=None, medio_pago=None,
                no_comprobante=None, cuenta_bancaria_id=None, fuente_sifse=None, item_sifse=None):
    """Edita multiples campos de un Pago."""
    pago = get_pago(numero)
    if not pago:
        raise ValueError(f"Pago {numero} no encontrado")
    if pago["estado"] == "ANULADO":
        raise ValueError(f"Pago {numero} esta anulado")
    if nuevo_valor is not None:
        if nuevo_valor <= 0:
            raise ValueError("El valor debe ser mayor a cero")
        saldo_o = saldo_obligacion(pago["obligacion_numero"]) + pago["valor"]
        if nuevo_valor > saldo_o:
            raise ValueError(f"Excede saldo de la obligacion ($ {saldo_o:,.0f})")
    campos = []
    valores = []
    if nuevo_valor is not None:
        campos.append("valor=?")
        valores.append(nuevo_valor)
    if concepto is not None:
        campos.append("concepto=?")
        valores.append(concepto)
    if medio_pago is not None:
        campos.append("medio_pago=?")
        valores.append(medio_pago)
    if no_comprobante is not None:
        campos.append("no_comprobante=?")
        valores.append(no_comprobante)
    if cuenta_bancaria_id is not None:
        campos.append("cuenta_bancaria_id=?")
        valores.append(cuenta_bancaria_id)
    if fuente_sifse is not None:
        campos.append("fuente_sifse=?")
        valores.append(fuente_sifse)
    if item_sifse is not None:
        campos.append("item_sifse=?")
        valores.append(item_sifse)
    if campos:
        valores.append(numero)
        conn = get_connection()
        conn.execute(f"UPDATE pago SET {', '.join(campos)} WHERE numero=?", valores)
        conn.commit()
        conn.close()


# ===================== CONSULTAS POR RUBRO =====================
def get_cdps_por_rubro(codigo_rubro):
    conn = get_connection()
    rows = conn.execute(
        "SELECT c.*, rg.cuenta FROM cdp c JOIN rubros_gastos rg ON c.codigo_rubro=rg.codigo "
        "WHERE c.codigo_rubro=? AND c.estado<>'ANULADO' ORDER BY c.numero", (codigo_rubro,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_rps_por_rubro(codigo_rubro):
    conn = get_connection()
    rows = conn.execute(
        "SELECT r.*, rg.cuenta, t.nombre as nombre_tercero FROM rp r "
        "JOIN rubros_gastos rg ON r.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON r.nit_tercero=t.nit "
        "WHERE r.codigo_rubro=? AND r.estado<>'ANULADO' ORDER BY r.numero", (codigo_rubro,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_obligaciones_por_rubro(codigo_rubro):
    conn = get_connection()
    rows = conn.execute(
        "SELECT o.*, rg.cuenta, t.nombre as nombre_tercero FROM obligacion o "
        "JOIN rubros_gastos rg ON o.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON o.nit_tercero=t.nit "
        "WHERE o.codigo_rubro=? AND o.estado<>'ANULADA' ORDER BY o.numero", (codigo_rubro,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pagos_por_rubro(codigo_rubro):
    conn = get_connection()
    rows = conn.execute(
        "SELECT p.*, rg.cuenta, t.nombre as nombre_tercero FROM pago p "
        "JOIN rubros_gastos rg ON p.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON p.nit_tercero=t.nit "
        "WHERE p.codigo_rubro=? AND p.estado<>'ANULADO' ORDER BY p.numero", (codigo_rubro,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recaudos_por_rubro(codigo_rubro):
    conn = get_connection()
    rows = conn.execute(
        "SELECT r.*, ri.cuenta FROM recaudo r JOIN rubros_ingresos ri ON r.codigo_rubro=ri.codigo "
        "WHERE r.codigo_rubro=? AND r.estado<>'ANULADO' ORDER BY r.numero", (codigo_rubro,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ===================== RECAUDO =====================
def saldo_por_recaudar(codigo_rubro):
    conn = get_connection()
    r = conn.execute(
        "SELECT presupuesto_definitivo FROM rubros_ingresos WHERE codigo=?",
        (codigo_rubro,)
    ).fetchone()
    ppto = r["presupuesto_definitivo"] if r else 0
    r2 = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM recaudo WHERE codigo_rubro=? AND estado<>'ANULADO'",
        (codigo_rubro,)
    ).fetchone()
    total_recaudo = r2["total"]
    conn.close()
    return ppto - total_recaudo


def registrar_recaudo(codigo_rubro, valor, concepto, no_comprobante, cuenta_bancaria_id=0):
    if valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")
    rubro = get_rubro_ingreso(codigo_rubro)
    if not rubro:
        raise ValueError(f"Rubro de ingreso {codigo_rubro} no existe")

    numero = get_consecutivo("recaudo")
    fecha = date.today().strftime("%Y-%m-%d")
    conn = get_connection()
    conn.execute(
        "INSERT INTO recaudo (numero, fecha, codigo_rubro, valor, concepto, no_comprobante, estado, cuenta_bancaria_id) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (numero, fecha, codigo_rubro, valor, concepto, no_comprobante, "ACTIVO", cuenta_bancaria_id)
    )
    conn.commit()
    conn.close()
    return numero, fecha


def get_recaudos(estado=None):
    conn = get_connection()
    if estado:
        rows = conn.execute(
            "SELECT r.*, ri.cuenta FROM recaudo r JOIN rubros_ingresos ri ON r.codigo_rubro=ri.codigo "
            "WHERE r.estado=? ORDER BY r.numero DESC", (estado,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT r.*, ri.cuenta FROM recaudo r JOIN rubros_ingresos ri ON r.codigo_rubro=ri.codigo "
            "ORDER BY r.numero DESC"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_recaudo(numero):
    conn = get_connection()
    r = conn.execute(
        "SELECT r.*, ri.cuenta FROM recaudo r JOIN rubros_ingresos ri ON r.codigo_rubro=ri.codigo "
        "WHERE r.numero=?", (numero,)
    ).fetchone()
    conn.close()
    return dict(r) if r else None


def anular_recaudo(numero):
    conn = get_connection()
    recaudo = conn.execute("SELECT * FROM recaudo WHERE numero=?", (numero,)).fetchone()
    if not recaudo:
        conn.close()
        raise ValueError(f"Recaudo {numero} no encontrado")
    if recaudo["estado"] == "ANULADO":
        conn.close()
        raise ValueError(f"Recaudo {numero} ya esta anulado")
    conn.execute("UPDATE recaudo SET estado='ANULADO' WHERE numero=?", (numero,))
    conn.commit()
    conn.close()


def editar_valor_recaudo(numero, nuevo_valor):
    conn = get_connection()
    recaudo = conn.execute("SELECT * FROM recaudo WHERE numero=?", (numero,)).fetchone()
    if not recaudo:
        conn.close()
        raise ValueError(f"Recaudo {numero} no encontrado")
    if recaudo["estado"] == "ANULADO":
        conn.close()
        raise ValueError(f"Recaudo {numero} esta anulado")
    if nuevo_valor <= 0:
        conn.close()
        raise ValueError("El valor debe ser mayor a cero")
    conn.execute("UPDATE recaudo SET valor=? WHERE numero=?", (nuevo_valor, numero))
    conn.commit()
    conn.close()


def editar_recaudo(numero, nuevo_valor=None, concepto=None, no_comprobante=None, cuenta_bancaria_id=None):
    """Edita multiples campos de un Recaudo."""
    recaudo = get_recaudo(numero)
    if not recaudo:
        raise ValueError(f"Recaudo {numero} no encontrado")
    if recaudo["estado"] == "ANULADO":
        raise ValueError(f"Recaudo {numero} esta anulado")
    if nuevo_valor is not None:
        if nuevo_valor <= 0:
            raise ValueError("El valor debe ser mayor a cero")
    campos = []
    valores = []
    if nuevo_valor is not None:
        campos.append("valor=?")
        valores.append(nuevo_valor)
    if concepto is not None:
        campos.append("concepto=?")
        valores.append(concepto)
    if no_comprobante is not None:
        campos.append("no_comprobante=?")
        valores.append(no_comprobante)
    if cuenta_bancaria_id is not None:
        campos.append("cuenta_bancaria_id=?")
        valores.append(cuenta_bancaria_id)
    if campos:
        valores.append(numero)
        conn = get_connection()
        conn.execute(f"UPDATE recaudo SET {', '.join(campos)} WHERE numero=?", valores)
        conn.commit()
        conn.close()


# ===================== MODIFICACIONES PRESUPUESTALES =====================
def _recalcular_apropiacion_definitiva_gasto(conn, codigo):
    """Recalcula: inicial + adiciones - reducciones + creditos - contracreditos"""
    r = conn.execute(
        "SELECT apropiacion_inicial, adiciones, reducciones, creditos, contracreditos "
        "FROM rubros_gastos WHERE codigo=?", (codigo,)
    ).fetchone()
    if r:
        nueva = r["apropiacion_inicial"] + r["adiciones"] - r["reducciones"] + r["creditos"] - r["contracreditos"]
        conn.execute("UPDATE rubros_gastos SET apropiacion_definitiva=? WHERE codigo=?", (nueva, codigo))


def _recalcular_presupuesto_definitivo_ingreso(conn, codigo):
    """Recalcula: inicial + adiciones - reducciones"""
    r = conn.execute(
        "SELECT presupuesto_inicial, adiciones, reducciones "
        "FROM rubros_ingresos WHERE codigo=?", (codigo,)
    ).fetchone()
    if r:
        nuevo = r["presupuesto_inicial"] + r["adiciones"] - r["reducciones"]
        conn.execute("UPDATE rubros_ingresos SET presupuesto_definitivo=? WHERE codigo=?", (nuevo, codigo))


def registrar_adicion(codigo_gasto, codigo_ingreso, valor, numero_acto, descripcion=""):
    """Registra una adicion presupuestal que incrementa ambos rubros (gasto e ingreso)."""
    if valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")

    rubro_g = get_rubro_gasto(codigo_gasto)
    if not rubro_g:
        raise ValueError(f"Rubro de gasto {codigo_gasto} no existe")
    if not rubro_g["es_hoja"]:
        raise ValueError(f"Rubro de gasto {codigo_gasto} no es rubro hoja")

    rubro_i = get_rubro_ingreso(codigo_ingreso)
    if not rubro_i:
        raise ValueError(f"Rubro de ingreso {codigo_ingreso} no existe")
    if not rubro_i["es_hoja"]:
        raise ValueError(f"Rubro de ingreso {codigo_ingreso} no es rubro hoja")

    numero = get_consecutivo("modificacion")
    fecha = date.today().strftime("%Y-%m-%d")
    conn = get_connection()
    try:
        # Registrar modificacion
        conn.execute(
            "INSERT INTO modificaciones_presupuestales (id, fecha, tipo, numero_acto, descripcion, valor, estado) "
            "VALUES (?,?,?,?,?,?,?)",
            (numero, fecha, "ADICION", numero_acto, descripcion, valor, "ACTIVO")
        )
        # Detalle gasto
        conn.execute(
            "INSERT INTO detalle_modificacion (id_modificacion, codigo_rubro, tipo_rubro, campo_afectado, valor) "
            "VALUES (?,?,?,?,?)",
            (numero, codigo_gasto, "GASTO", "adiciones", valor)
        )
        # Detalle ingreso
        conn.execute(
            "INSERT INTO detalle_modificacion (id_modificacion, codigo_rubro, tipo_rubro, campo_afectado, valor) "
            "VALUES (?,?,?,?,?)",
            (numero, codigo_ingreso, "INGRESO", "adiciones", valor)
        )
        # Actualizar rubros
        conn.execute("UPDATE rubros_gastos SET adiciones = adiciones + ? WHERE codigo=?", (valor, codigo_gasto))
        _recalcular_apropiacion_definitiva_gasto(conn, codigo_gasto)
        conn.execute("UPDATE rubros_ingresos SET adiciones = adiciones + ? WHERE codigo=?", (valor, codigo_ingreso))
        _recalcular_presupuesto_definitivo_ingreso(conn, codigo_ingreso)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return numero, fecha


def registrar_reduccion(codigo_gasto, codigo_ingreso, valor, numero_acto, descripcion=""):
    """Registra una reduccion presupuestal que disminuye ambos rubros (gasto e ingreso)."""
    if valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")

    rubro_g = get_rubro_gasto(codigo_gasto)
    if not rubro_g:
        raise ValueError(f"Rubro de gasto {codigo_gasto} no existe")
    if not rubro_g["es_hoja"]:
        raise ValueError(f"Rubro de gasto {codigo_gasto} no es rubro hoja")

    rubro_i = get_rubro_ingreso(codigo_ingreso)
    if not rubro_i:
        raise ValueError(f"Rubro de ingreso {codigo_ingreso} no existe")
    if not rubro_i["es_hoja"]:
        raise ValueError(f"Rubro de ingreso {codigo_ingreso} no es rubro hoja")

    # Validar saldo disponible del rubro de gasto
    saldo_g = saldo_disponible_rubro(codigo_gasto)
    if valor > saldo_g:
        raise ValueError(f"Valor ($ {valor:,.0f}) excede saldo disponible del rubro de gasto ($ {saldo_g:,.0f})")

    # Validar que la apropiacion definitiva no quede negativa
    nueva_aprop = (rubro_g["apropiacion_inicial"] + rubro_g["adiciones"]
                   - (rubro_g["reducciones"] + valor)
                   + rubro_g["creditos"] - rubro_g["contracreditos"])
    if nueva_aprop < 0:
        raise ValueError(f"La reduccion dejaria la apropiacion definitiva en negativo ($ {nueva_aprop:,.0f})")

    # Validar ingreso
    nuevo_ppto = (rubro_i["presupuesto_inicial"] + rubro_i["adiciones"]
                  - (rubro_i["reducciones"] + valor))
    if nuevo_ppto < 0:
        raise ValueError(f"La reduccion dejaria el presupuesto de ingreso en negativo ($ {nuevo_ppto:,.0f})")

    numero = get_consecutivo("modificacion")
    fecha = date.today().strftime("%Y-%m-%d")
    conn = get_connection()
    try:
        conn.execute(
            "INSERT INTO modificaciones_presupuestales (id, fecha, tipo, numero_acto, descripcion, valor, estado) "
            "VALUES (?,?,?,?,?,?,?)",
            (numero, fecha, "REDUCCION", numero_acto, descripcion, valor, "ACTIVO")
        )
        conn.execute(
            "INSERT INTO detalle_modificacion (id_modificacion, codigo_rubro, tipo_rubro, campo_afectado, valor) "
            "VALUES (?,?,?,?,?)",
            (numero, codigo_gasto, "GASTO", "reducciones", valor)
        )
        conn.execute(
            "INSERT INTO detalle_modificacion (id_modificacion, codigo_rubro, tipo_rubro, campo_afectado, valor) "
            "VALUES (?,?,?,?,?)",
            (numero, codigo_ingreso, "INGRESO", "reducciones", valor)
        )
        conn.execute("UPDATE rubros_gastos SET reducciones = reducciones + ? WHERE codigo=?", (valor, codigo_gasto))
        _recalcular_apropiacion_definitiva_gasto(conn, codigo_gasto)
        conn.execute("UPDATE rubros_ingresos SET reducciones = reducciones + ? WHERE codigo=?", (valor, codigo_ingreso))
        _recalcular_presupuesto_definitivo_ingreso(conn, codigo_ingreso)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return numero, fecha


def registrar_credito_contracredito(codigo_credito, codigo_contracredito, valor, numero_acto, descripcion=""):
    """Registra un traslado credito/contracredito entre rubros de gasto."""
    if valor <= 0:
        raise ValueError("El valor debe ser mayor a cero")
    if codigo_credito == codigo_contracredito:
        raise ValueError("Los rubros de credito y contracredito deben ser diferentes")

    rubro_cred = get_rubro_gasto(codigo_credito)
    if not rubro_cred:
        raise ValueError(f"Rubro credito {codigo_credito} no existe")
    if not rubro_cred["es_hoja"]:
        raise ValueError(f"Rubro credito {codigo_credito} no es rubro hoja")

    rubro_ccred = get_rubro_gasto(codigo_contracredito)
    if not rubro_ccred:
        raise ValueError(f"Rubro contracredito {codigo_contracredito} no existe")
    if not rubro_ccred["es_hoja"]:
        raise ValueError(f"Rubro contracredito {codigo_contracredito} no es rubro hoja")

    # Validar saldo disponible del rubro que cede (contracredito)
    saldo_ccred = saldo_disponible_rubro(codigo_contracredito)
    if valor > saldo_ccred:
        raise ValueError(f"Valor ($ {valor:,.0f}) excede saldo disponible del rubro contracredito ($ {saldo_ccred:,.0f})")

    # Validar que no quede negativo
    nueva_aprop_ccred = (rubro_ccred["apropiacion_inicial"] + rubro_ccred["adiciones"]
                         - rubro_ccred["reducciones"] + rubro_ccred["creditos"]
                         - (rubro_ccred["contracreditos"] + valor))
    if nueva_aprop_ccred < 0:
        raise ValueError(f"El contracredito dejaria la apropiacion definitiva del rubro en negativo")

    numero = get_consecutivo("modificacion")
    fecha = date.today().strftime("%Y-%m-%d")
    conn = get_connection()
    try:
        conn.execute(
            "INSERT INTO modificaciones_presupuestales (id, fecha, tipo, numero_acto, descripcion, valor, estado) "
            "VALUES (?,?,?,?,?,?,?)",
            (numero, fecha, "CREDITO_CONTRACREDITO", numero_acto, descripcion, valor, "ACTIVO")
        )
        # Credito: el rubro que recibe
        conn.execute(
            "INSERT INTO detalle_modificacion (id_modificacion, codigo_rubro, tipo_rubro, campo_afectado, valor) "
            "VALUES (?,?,?,?,?)",
            (numero, codigo_credito, "GASTO", "creditos", valor)
        )
        # Contracredito: el rubro que cede
        conn.execute(
            "INSERT INTO detalle_modificacion (id_modificacion, codigo_rubro, tipo_rubro, campo_afectado, valor) "
            "VALUES (?,?,?,?,?)",
            (numero, codigo_contracredito, "GASTO", "contracreditos", valor)
        )
        conn.execute("UPDATE rubros_gastos SET creditos = creditos + ? WHERE codigo=?", (valor, codigo_credito))
        _recalcular_apropiacion_definitiva_gasto(conn, codigo_credito)
        conn.execute("UPDATE rubros_gastos SET contracreditos = contracreditos + ? WHERE codigo=?", (valor, codigo_contracredito))
        _recalcular_apropiacion_definitiva_gasto(conn, codigo_contracredito)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return numero, fecha


def listar_modificaciones():
    """Lista todas las modificaciones presupuestales con detalles."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM modificaciones_presupuestales ORDER BY id DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_modificacion(id_mod):
    """Obtiene una modificacion con sus detalles y nombres de rubros."""
    conn = get_connection()
    mod = conn.execute(
        "SELECT * FROM modificaciones_presupuestales WHERE id=?", (id_mod,)
    ).fetchone()
    if not mod:
        conn.close()
        return None

    detalles = conn.execute(
        "SELECT d.*, "
        "COALESCE(rg.cuenta, ri.cuenta) as nombre_rubro "
        "FROM detalle_modificacion d "
        "LEFT JOIN rubros_gastos rg ON d.codigo_rubro = rg.codigo AND d.tipo_rubro = 'GASTO' "
        "LEFT JOIN rubros_ingresos ri ON d.codigo_rubro = ri.codigo AND d.tipo_rubro = 'INGRESO' "
        "WHERE d.id_modificacion=?",
        (id_mod,)
    ).fetchall()
    conn.close()

    resultado = dict(mod)
    resultado["detalles"] = [dict(d) for d in detalles]
    return resultado


# ===================== CONCEPTOS =====================
def guardar_concepto(codigo_rubro, concepto):
    if not concepto.strip():
        return
    conn = get_connection()
    hoy = date.today().strftime("%Y-%m-%d")
    try:
        conn.execute(
            "INSERT INTO conceptos (codigo_rubro, concepto, veces_usado, ultimo_uso) VALUES (?,?,1,?)",
            (codigo_rubro, concepto.strip(), hoy)
        )
    except sqlite3.IntegrityError:
        conn.execute(
            "UPDATE conceptos SET veces_usado=veces_usado+1, ultimo_uso=? "
            "WHERE codigo_rubro=? AND concepto=?",
            (hoy, codigo_rubro, concepto.strip())
        )
    conn.commit()
    conn.close()


def get_conceptos_rubro(codigo_rubro):
    conn = get_connection()
    rows = conn.execute(
        "SELECT concepto, veces_usado FROM conceptos WHERE codigo_rubro=? ORDER BY veces_usado DESC",
        (codigo_rubro,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ===================== TARJETAS =====================
def generar_tarjeta(codigo_rubro):
    rubro = get_rubro_gasto(codigo_rubro)
    if not rubro:
        return None, []

    conn = get_connection()
    movimientos = []

    # CDPs
    for r in conn.execute(
        "SELECT numero, fecha, objeto, valor FROM cdp WHERE codigo_rubro=? AND estado<>'ANULADO' ORDER BY fecha",
        (codigo_rubro,)
    ).fetchall():
        movimientos.append({
            "fecha": r["fecha"], "tipo": "CDP", "numero": r["numero"],
            "nit": "", "tercero": "", "concepto": r["objeto"],
            "v_cdp": r["valor"], "v_rp": 0, "v_obl": 0, "v_pago": 0
        })

    # RPs
    for r in conn.execute(
        "SELECT r.numero, r.fecha, r.nit_tercero, t.nombre, r.objeto, r.valor "
        "FROM rp r LEFT JOIN terceros t ON r.nit_tercero=t.nit "
        "WHERE r.codigo_rubro=? AND r.estado<>'ANULADO' ORDER BY r.fecha",
        (codigo_rubro,)
    ).fetchall():
        movimientos.append({
            "fecha": r["fecha"], "tipo": "RP", "numero": r["numero"],
            "nit": r["nit_tercero"], "tercero": r["nombre"] or "", "concepto": r["objeto"],
            "v_cdp": 0, "v_rp": r["valor"], "v_obl": 0, "v_pago": 0
        })

    # Obligaciones
    for r in conn.execute(
        "SELECT o.numero, o.fecha, o.nit_tercero, t.nombre, o.factura, o.valor "
        "FROM obligacion o LEFT JOIN terceros t ON o.nit_tercero=t.nit "
        "WHERE o.codigo_rubro=? AND o.estado<>'ANULADA' ORDER BY o.fecha",
        (codigo_rubro,)
    ).fetchall():
        movimientos.append({
            "fecha": r["fecha"], "tipo": "OBLIG", "numero": r["numero"],
            "nit": r["nit_tercero"], "tercero": r["nombre"] or "", "concepto": r["factura"],
            "v_cdp": 0, "v_rp": 0, "v_obl": r["valor"], "v_pago": 0
        })

    # Pagos
    for r in conn.execute(
        "SELECT p.numero, p.fecha, p.nit_tercero, t.nombre, p.concepto, p.valor "
        "FROM pago p LEFT JOIN terceros t ON p.nit_tercero=t.nit "
        "WHERE p.codigo_rubro=? AND p.estado<>'ANULADO' ORDER BY p.fecha",
        (codigo_rubro,)
    ).fetchall():
        movimientos.append({
            "fecha": r["fecha"], "tipo": "PAGO", "numero": r["numero"],
            "nit": r["nit_tercero"], "tercero": r["nombre"] or "", "concepto": r["concepto"],
            "v_cdp": 0, "v_rp": 0, "v_obl": 0, "v_pago": r["valor"]
        })

    conn.close()
    movimientos.sort(key=lambda x: x["fecha"])
    return rubro, movimientos


# ===================== RESUMEN DETALLADO POR RUBRO =====================
def resumen_rubro(codigo_rubro, mes_inicio=1, mes_fin=12):
    """Genera resumen detallado de ejecucion presupuestal de un rubro por periodo."""
    rubro = get_rubro_gasto(codigo_rubro)
    if not rubro:
        return None

    conn = get_connection()
    anio = get_config("vigencia") or "2026"

    # Fechas limite del periodo
    fecha_ini_periodo = f"{anio}-{mes_inicio:02d}-01"
    # Usar primer dia del mes siguiente para incluir todo el mes final
    if mes_fin >= 12:
        fecha_fin_periodo = f"{int(anio) + 1}-01-01"
    else:
        fecha_fin_periodo = f"{anio}-{mes_fin + 1:02d}-01"

    # Si es rubro hoja, buscar exacto; si es padre, buscar todos los hijos
    es_hoja = rubro.get("es_hoja", 0)
    if es_hoja:
        filtro_rubro = "codigo_rubro = ?"
        param_rubro = codigo_rubro
    else:
        filtro_rubro = "(codigo_rubro = ? OR codigo_rubro LIKE ?)"
        param_rubro = None  # se maneja abajo

    def sumar(tabla, campo_estado, estado_excluir, antes=True):
        if es_hoja:
            params_base = [codigo_rubro, estado_excluir]
        else:
            params_base = [codigo_rubro, codigo_rubro + ".%", estado_excluir]

        if antes:
            sql = (f"SELECT COALESCE(SUM(valor), 0) as t FROM {tabla} "
                   f"WHERE {filtro_rubro} AND {campo_estado}<>? AND fecha < ?")
            params_base.append(fecha_ini_periodo)
        else:
            sql = (f"SELECT COALESCE(SUM(valor), 0) as t FROM {tabla} "
                   f"WHERE {filtro_rubro} AND {campo_estado}<>? AND fecha >= ? AND fecha < ?")
            params_base.extend([fecha_ini_periodo, fecha_fin_periodo])

        r = conn.execute(sql, params_base).fetchone()
        return r["t"]

    # CDPs (Disponibilidades)
    disp_anteriores = sumar("cdp", "estado", "ANULADO", antes=True)
    disp_periodo = sumar("cdp", "estado", "ANULADO", antes=False)
    total_disp = disp_anteriores + disp_periodo

    # Compromisos (RPs)
    comp_anteriores = sumar("rp", "estado", "ANULADO", antes=True)
    comp_periodo = sumar("rp", "estado", "ANULADO", antes=False)
    total_comp = comp_anteriores + comp_periodo

    # Obligaciones
    obl_anteriores = sumar("obligacion", "estado", "ANULADA", antes=True)
    obl_periodo = sumar("obligacion", "estado", "ANULADA", antes=False)
    total_obl = obl_anteriores + obl_periodo

    # Pagos
    pago_anteriores = sumar("pago", "estado", "ANULADO", antes=True)
    pago_periodo = sumar("pago", "estado", "ANULADO", antes=False)
    total_pago = pago_anteriores + pago_periodo

    # Si es padre, sumar apropiaciones solo de rubros hoja (evitar doble conteo)
    if es_hoja:
        aprop = rubro["apropiacion_definitiva"]
        aprop_ini = rubro["apropiacion_inicial"]
        adiciones = rubro["adiciones"]
        reducciones = rubro["reducciones"]
        creditos = rubro["creditos"]
        contracreditos = rubro["contracreditos"]
    else:
        r_sum = conn.execute(
            "SELECT COALESCE(SUM(apropiacion_definitiva), 0) as ad, "
            "COALESCE(SUM(apropiacion_inicial), 0) as ai, "
            "COALESCE(SUM(adiciones), 0) as adic, "
            "COALESCE(SUM(reducciones), 0) as red, "
            "COALESCE(SUM(creditos), 0) as cred, "
            "COALESCE(SUM(contracreditos), 0) as ccred "
            "FROM rubros_gastos WHERE es_hoja=1 AND codigo LIKE ?",
            (codigo_rubro + ".%",)
        ).fetchone()
        aprop = r_sum["ad"]
        aprop_ini = r_sum["ai"]
        adiciones = r_sum["adic"]
        reducciones = r_sum["red"]
        creditos = r_sum["cred"]
        contracreditos = r_sum["ccred"]

    conn.close()

    return {
        "rubro": rubro,
        "apropiacion_inicial": aprop_ini,
        "adiciones": adiciones,
        "reducciones": reducciones,
        "creditos": creditos,
        "contracreditos": contracreditos,
        "apropiacion_definitiva": aprop,
        # Disponibilidades (CDPs)
        "disp_anteriores": disp_anteriores,
        "disp_periodo": disp_periodo,
        "total_disp": total_disp,
        "saldo_disponible": aprop - total_disp,
        "disp_sin_compromiso": total_disp - total_comp,
        # Compromisos (RPs)
        "comp_anteriores": comp_anteriores,
        "comp_periodo": comp_periodo,
        "total_comp": total_comp,
        "comp_sin_obligacion": total_comp - total_obl,
        # Obligaciones
        "obl_anteriores": obl_anteriores,
        "obl_periodo": obl_periodo,
        "total_obl": total_obl,
        "obl_x_pagar": total_obl - total_pago,
        # Pagos
        "pago_anteriores": pago_anteriores,
        "pago_periodo": pago_periodo,
        "total_pago": total_pago,
        # Apropiacion x afectar
        "aprop_x_afectar": aprop - total_disp,
    }


# ===================== INFORMES =====================
def informe_ejecucion_gastos(mes_consulta=None):
    """Ejecucion presupuestal de gastos - formato catalogo oficial.
    Columnas: Ppto Inicial, Adiciones, Reducciones, Creditos, Contra-Creditos,
    Ppto Definitivo, Comp. Anterior, Comp. Mes, Comp. Acum.,
    Pagos Anterior, Pagos Mes, Pagos Acum., Saldo Aprop, Saldo Comp x Pagar.
    """
    if mes_consulta is None:
        mes_consulta = int(get_config("mes_actual") or 1)

    conn = get_connection()
    anio = get_config("vigencia") or "2026"
    rubros = get_rubros_gastos(solo_hojas=False)
    resultado = []

    fecha_ini_mes = f"{anio}-{mes_consulta:02d}-01"
    if mes_consulta >= 12:
        fecha_fin_mes = f"{int(anio)+1}-01-01"
    else:
        fecha_fin_mes = f"{anio}-{mes_consulta+1:02d}-01"

    for r in rubros:
        codigo = r["codigo"]
        es_hoja = r["es_hoja"]

        if es_hoja:
            filtro = "codigo_rubro = ?"
            params = [codigo]
        else:
            filtro = "(codigo_rubro = ? OR codigo_rubro LIKE ?)"
            params = [codigo, codigo + ".%"]

        # Apropiaciones: si es padre, sumar solo hojas
        if es_hoja:
            ppto_ini = r["apropiacion_inicial"]
            adiciones = r["adiciones"]
            reducciones = r["reducciones"]
            creditos = r["creditos"]
            contracreditos = r["contracreditos"]
            ppto_def = r["apropiacion_definitiva"]
        else:
            rs = conn.execute(
                "SELECT COALESCE(SUM(apropiacion_inicial),0) as ai, "
                "COALESCE(SUM(adiciones),0) as ad, COALESCE(SUM(reducciones),0) as re, "
                "COALESCE(SUM(creditos),0) as cr, COALESCE(SUM(contracreditos),0) as cc, "
                "COALESCE(SUM(apropiacion_definitiva),0) as apd "
                "FROM rubros_gastos WHERE es_hoja=1 AND codigo LIKE ?",
                (codigo + ".%",)
            ).fetchone()
            ppto_ini = rs["ai"]
            adiciones = rs["ad"]
            reducciones = rs["re"]
            creditos = rs["cr"]
            contracreditos = rs["cc"]
            ppto_def = rs["apd"]

        # Compromisos (RPs) anteriores al mes
        comp_ant = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) as t FROM rp WHERE {filtro} AND estado<>'ANULADO' AND fecha < ?",
            params + [fecha_ini_mes]
        ).fetchone()["t"]

        # Compromisos del mes
        comp_mes = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) as t FROM rp WHERE {filtro} AND estado<>'ANULADO' AND fecha >= ? AND fecha < ?",
            params + [fecha_ini_mes, fecha_fin_mes]
        ).fetchone()["t"]

        comp_acum = comp_ant + comp_mes

        # Pagos anteriores al mes
        pago_ant = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) as t FROM pago WHERE {filtro} AND estado<>'ANULADO' AND fecha < ?",
            params + [fecha_ini_mes]
        ).fetchone()["t"]

        # Pagos del mes
        pago_mes = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) as t FROM pago WHERE {filtro} AND estado<>'ANULADO' AND fecha >= ? AND fecha < ?",
            params + [fecha_ini_mes, fecha_fin_mes]
        ).fetchone()["t"]

        pago_acum = pago_ant + pago_mes

        saldo_aprop = ppto_def - comp_acum
        saldo_comp_pagar = comp_acum - pago_acum

        resultado.append({
            "codigo": codigo,
            "cuenta": r["cuenta"],
            "es_hoja": es_hoja,
            "nivel": codigo.count(".") + 1,
            "ppto_inicial": ppto_ini,
            "adiciones": adiciones,
            "reducciones": reducciones,
            "creditos": creditos,
            "contracreditos": contracreditos,
            "ppto_definitivo": ppto_def,
            "comp_anterior": comp_ant,
            "comp_mes": comp_mes,
            "comp_acumulado": comp_acum,
            "pago_anterior": pago_ant,
            "pago_mes": pago_mes,
            "pago_acumulado": pago_acum,
            "saldo_apropiacion": saldo_aprop,
            "saldo_comp_pagar": saldo_comp_pagar,
        })

    conn.close()
    return resultado


def informe_ejecucion_auxiliar(mes_inicio=1, mes_fin=12):
    """Ejecucion auxiliar: detalle de cada movimiento por rubro."""
    conn = get_connection()
    anio = get_config("vigencia") or "2026"
    fecha_ini = f"{anio}-{mes_inicio:02d}-01"
    fecha_fin = f"{int(anio)+1}-01-01" if mes_fin >= 12 else f"{anio}-{mes_fin+1:02d}-01"

    movimientos = []

    # CDPs
    for r in conn.execute(
        "SELECT c.numero, c.fecha, c.codigo_rubro, rg.cuenta, c.objeto, c.valor, c.estado "
        "FROM cdp c JOIN rubros_gastos rg ON c.codigo_rubro=rg.codigo "
        "WHERE c.estado<>'ANULADO' AND c.fecha >= ? AND c.fecha < ? ORDER BY c.fecha",
        (fecha_ini, fecha_fin)
    ).fetchall():
        movimientos.append({"tipo": "CDP", **dict(r)})

    # RPs
    for r in conn.execute(
        "SELECT r.numero, r.fecha, r.codigo_rubro, rg.cuenta, r.objeto, r.valor, r.estado, "
        "r.cdp_numero, t.nombre as tercero "
        "FROM rp r JOIN rubros_gastos rg ON r.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON r.nit_tercero=t.nit "
        "WHERE r.estado<>'ANULADO' AND r.fecha >= ? AND r.fecha < ? ORDER BY r.fecha",
        (fecha_ini, fecha_fin)
    ).fetchall():
        movimientos.append({"tipo": "RP", **dict(r)})

    # Obligaciones
    for r in conn.execute(
        "SELECT o.numero, o.fecha, o.codigo_rubro, rg.cuenta, o.factura as objeto, o.valor, o.estado, "
        "o.rp_numero, t.nombre as tercero "
        "FROM obligacion o JOIN rubros_gastos rg ON o.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON o.nit_tercero=t.nit "
        "WHERE o.estado<>'ANULADA' AND o.fecha >= ? AND o.fecha < ? ORDER BY o.fecha",
        (fecha_ini, fecha_fin)
    ).fetchall():
        movimientos.append({"tipo": "OBLIG", **dict(r)})

    # Pagos
    for r in conn.execute(
        "SELECT p.numero, p.fecha, p.codigo_rubro, rg.cuenta, p.concepto as objeto, p.valor, p.estado, "
        "p.obligacion_numero, t.nombre as tercero "
        "FROM pago p JOIN rubros_gastos rg ON p.codigo_rubro=rg.codigo "
        "LEFT JOIN terceros t ON p.nit_tercero=t.nit "
        "WHERE p.estado<>'ANULADO' AND p.fecha >= ? AND p.fecha < ? ORDER BY p.fecha",
        (fecha_ini, fecha_fin)
    ).fetchall():
        movimientos.append({"tipo": "PAGO", **dict(r)})

    conn.close()
    movimientos.sort(key=lambda x: (x["codigo_rubro"], x["fecha"], x["tipo"]))
    return movimientos


def informe_cadena_presupuestal():
    """Informe encadenado: CDP -> RP -> OBLIGACION -> PAGO."""
    conn = get_connection()
    cadenas = []

    cdps = conn.execute(
        "SELECT c.*, rg.cuenta FROM cdp c "
        "JOIN rubros_gastos rg ON c.codigo_rubro=rg.codigo "
        "WHERE c.estado<>'ANULADO' ORDER BY c.numero"
    ).fetchall()

    for cdp in cdps:
        cdp_dict = dict(cdp)
        cdp_dict["rps"] = []

        rps = conn.execute(
            "SELECT r.*, t.nombre as tercero FROM rp r "
            "LEFT JOIN terceros t ON r.nit_tercero=t.nit "
            "WHERE r.cdp_numero=? AND r.estado<>'ANULADO' ORDER BY r.numero",
            (cdp["numero"],)
        ).fetchall()

        for rp in rps:
            rp_dict = dict(rp)
            rp_dict["obligaciones"] = []

            obls = conn.execute(
                "SELECT o.* FROM obligacion o "
                "WHERE o.rp_numero=? AND o.estado<>'ANULADA' ORDER BY o.numero",
                (rp["numero"],)
            ).fetchall()

            for obl in obls:
                obl_dict = dict(obl)
                obl_dict["pagos"] = []

                pagos = conn.execute(
                    "SELECT p.* FROM pago p "
                    "WHERE p.obligacion_numero=? AND p.estado<>'ANULADO' ORDER BY p.numero",
                    (obl["numero"],)
                ).fetchall()
                obl_dict["pagos"] = [dict(p) for p in pagos]
                rp_dict["obligaciones"].append(obl_dict)

            cdp_dict["rps"].append(rp_dict)
        cadenas.append(cdp_dict)

    conn.close()
    return cadenas


# ===================== INFORMES INGRESOS =====================
def resumen_rubro_ingreso(codigo_rubro, mes_inicio=1, mes_fin=12):
    """Genera resumen de ejecucion de ingresos de un rubro por periodo."""
    rubro = get_rubro_ingreso(codigo_rubro)
    if not rubro:
        return None

    conn = get_connection()
    anio = get_config("vigencia") or "2026"

    fecha_ini_periodo = f"{anio}-{mes_inicio:02d}-01"
    if mes_fin >= 12:
        fecha_fin_periodo = f"{int(anio) + 1}-01-01"
    else:
        fecha_fin_periodo = f"{anio}-{mes_fin + 1:02d}-01"

    es_hoja = rubro.get("es_hoja", 0)
    if es_hoja:
        filtro_rubro = "codigo_rubro = ?"
    else:
        filtro_rubro = "(codigo_rubro = ? OR codigo_rubro LIKE ?)"

    def sumar_recaudos(antes=True):
        if es_hoja:
            params = [codigo_rubro, "ANULADO"]
        else:
            params = [codigo_rubro, codigo_rubro + ".%", "ANULADO"]
        if antes:
            sql = (f"SELECT COALESCE(SUM(valor), 0) as t FROM recaudo "
                   f"WHERE {filtro_rubro} AND estado<>? AND fecha < ?")
            params.append(fecha_ini_periodo)
        else:
            sql = (f"SELECT COALESCE(SUM(valor), 0) as t FROM recaudo "
                   f"WHERE {filtro_rubro} AND estado<>? AND fecha >= ? AND fecha < ?")
            params.extend([fecha_ini_periodo, fecha_fin_periodo])
        return conn.execute(sql, params).fetchone()["t"]

    recaudo_anterior = sumar_recaudos(antes=True)
    recaudo_periodo = sumar_recaudos(antes=False)
    recaudo_acumulado = recaudo_anterior + recaudo_periodo

    if es_hoja:
        ppto_ini = rubro["presupuesto_inicial"]
        adiciones = rubro["adiciones"]
        reducciones = rubro["reducciones"]
        ppto_def = rubro["presupuesto_definitivo"]
    else:
        r_sum = conn.execute(
            "SELECT COALESCE(SUM(presupuesto_definitivo), 0) as pd, "
            "COALESCE(SUM(presupuesto_inicial), 0) as pi, "
            "COALESCE(SUM(adiciones), 0) as ad, "
            "COALESCE(SUM(reducciones), 0) as re "
            "FROM rubros_ingresos WHERE es_hoja=1 AND codigo LIKE ?",
            (codigo_rubro + ".%",)
        ).fetchone()
        ppto_def = r_sum["pd"]
        ppto_ini = r_sum["pi"]
        adiciones = r_sum["ad"]
        reducciones = r_sum["re"]

    conn.close()

    return {
        "rubro": rubro,
        "presupuesto_inicial": ppto_ini,
        "adiciones": adiciones,
        "reducciones": reducciones,
        "presupuesto_definitivo": ppto_def,
        "recaudo_anterior": recaudo_anterior,
        "recaudo_periodo": recaudo_periodo,
        "recaudo_acumulado": recaudo_acumulado,
        "saldo_por_recaudar": ppto_def - recaudo_acumulado,
    }


def informe_ejecucion_ingresos(mes_consulta=None):
    """Ejecucion presupuestal de ingresos - formato catalogo."""
    if mes_consulta is None:
        mes_consulta = int(get_config("mes_actual") or 1)

    conn = get_connection()
    anio = get_config("vigencia") or "2026"
    rubros = get_rubros_ingresos(solo_hojas=False)
    resultado = []

    fecha_ini_mes = f"{anio}-{mes_consulta:02d}-01"
    if mes_consulta >= 12:
        fecha_fin_mes = f"{int(anio)+1}-01-01"
    else:
        fecha_fin_mes = f"{anio}-{mes_consulta+1:02d}-01"

    for r in rubros:
        codigo = r["codigo"]
        es_hoja = r["es_hoja"]

        if es_hoja:
            filtro = "codigo_rubro = ?"
            params = [codigo]
        else:
            filtro = "(codigo_rubro = ? OR codigo_rubro LIKE ?)"
            params = [codigo, codigo + ".%"]

        if es_hoja:
            ppto_ini = r["presupuesto_inicial"]
            adiciones = r["adiciones"]
            reducciones = r["reducciones"]
            ppto_def = r["presupuesto_definitivo"]
        else:
            rs = conn.execute(
                "SELECT COALESCE(SUM(presupuesto_inicial),0) as pi, "
                "COALESCE(SUM(adiciones),0) as ad, COALESCE(SUM(reducciones),0) as re, "
                "COALESCE(SUM(presupuesto_definitivo),0) as pd "
                "FROM rubros_ingresos WHERE es_hoja=1 AND codigo LIKE ?",
                (codigo + ".%",)
            ).fetchone()
            ppto_ini = rs["pi"]
            adiciones = rs["ad"]
            reducciones = rs["re"]
            ppto_def = rs["pd"]

        # Recaudos anteriores al mes
        rec_ant = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) as t FROM recaudo WHERE {filtro} AND estado<>'ANULADO' AND fecha < ?",
            params + [fecha_ini_mes]
        ).fetchone()["t"]

        # Recaudos del mes
        rec_mes = conn.execute(
            f"SELECT COALESCE(SUM(valor),0) as t FROM recaudo WHERE {filtro} AND estado<>'ANULADO' AND fecha >= ? AND fecha < ?",
            params + [fecha_ini_mes, fecha_fin_mes]
        ).fetchone()["t"]

        rec_acum = rec_ant + rec_mes
        saldo_recaudar = ppto_def - rec_acum

        resultado.append({
            "codigo": codigo,
            "cuenta": r["cuenta"],
            "es_hoja": es_hoja,
            "nivel": codigo.count(".") + 1,
            "ppto_inicial": ppto_ini,
            "adiciones": adiciones,
            "reducciones": reducciones,
            "ppto_definitivo": ppto_def,
            "recaudo_anterior": rec_ant,
            "recaudo_mes": rec_mes,
            "recaudo_acumulado": rec_acum,
            "saldo_por_recaudar": saldo_recaudar,
        })

    conn.close()
    return resultado


# ===================== CONSOLIDACION =====================
def consolidar_mes_ingresos():
    """Consolida recaudos del mes actual para rubros de ingresos."""
    mes = int(get_config("mes_actual"))
    anio = int(get_config("vigencia"))
    rubros = get_rubros_ingresos(solo_hojas=True)
    conn = get_connection()

    for rubro in rubros:
        codigo = rubro["codigo"]
        r = conn.execute(
            "SELECT COALESCE(SUM(valor), 0) as total FROM recaudo "
            "WHERE codigo_rubro=? AND estado<>'ANULADO' AND "
            "CAST(strftime('%%m', fecha) AS INTEGER)=?",
            (codigo, mes)
        ).fetchone()
        recaudo = r["total"]

        conn.execute(
            "INSERT OR REPLACE INTO consolidacion_mensual_ingresos "
            "(mes, anio, codigo_rubro, recaudo_mes, fecha_consolidacion) "
            "VALUES (?,?,?,?,?)",
            (mes, anio, codigo, recaudo, date.today().strftime("%Y-%m-%d"))
        )

    conn.commit()
    conn.close()
    return mes, len(rubros)


def consolidar_mes():
    mes = int(get_config("mes_actual"))
    anio = int(get_config("vigencia"))
    rubros = get_rubros_gastos(solo_hojas=True)
    conn = get_connection()

    for rubro in rubros:
        codigo = rubro["codigo"]
        # Compromisos del mes (RPs)
        r = conn.execute(
            "SELECT COALESCE(SUM(valor), 0) as total FROM rp "
            "WHERE codigo_rubro=? AND estado<>'ANULADO' AND "
            "CAST(strftime('%%m', fecha) AS INTEGER)=?",
            (codigo, mes)
        ).fetchone()
        compromisos = r["total"]

        # Pagos del mes
        r = conn.execute(
            "SELECT COALESCE(SUM(valor), 0) as total FROM pago "
            "WHERE codigo_rubro=? AND estado<>'ANULADO' AND "
            "CAST(strftime('%%m', fecha) AS INTEGER)=?",
            (codigo, mes)
        ).fetchone()
        pagos = r["total"]

        conn.execute(
            "INSERT OR REPLACE INTO consolidacion_mensual "
            "(mes, anio, codigo_rubro, compromisos_mes, pagos_mes, fecha_consolidacion) "
            "VALUES (?,?,?,?,?,?)",
            (mes, anio, codigo, compromisos, pagos, date.today().strftime("%Y-%m-%d"))
        )

    conn.commit()
    conn.close()
    return mes, len(rubros)


def cierre_mes():
    mes_actual = int(get_config("mes_actual"))
    consolidar_mes()
    consolidar_mes_ingresos()
    set_config("mes_actual", str(mes_actual + 1))
    return mes_actual


# ===================== RESUMEN =====================
def get_resumen():
    conn = get_connection()
    # Apropiacion definitiva total (de rubros hoja)
    total_aprop = conn.execute(
        "SELECT COALESCE(SUM(apropiacion_definitiva), 0) as t FROM rubros_gastos WHERE es_hoja=1"
    ).fetchone()["t"]
    total_cdp = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM cdp WHERE estado<>'ANULADO'"
    ).fetchone()["t"]
    total_comp = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM rp WHERE estado<>'ANULADO'"
    ).fetchone()["t"]
    total_obl = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM obligacion WHERE estado<>'ANULADA'"
    ).fetchone()["t"]
    total_pago = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM pago WHERE estado<>'ANULADO'"
    ).fetchone()["t"]
    # Ingresos
    total_ppto_ingresos = conn.execute(
        "SELECT COALESCE(SUM(presupuesto_definitivo), 0) as t FROM rubros_ingresos WHERE es_hoja=1"
    ).fetchone()["t"]
    total_recaudado = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as t FROM recaudo WHERE estado<>'ANULADO'"
    ).fetchone()["t"]

    conn.close()

    saldo_por_recaudar = total_ppto_ingresos - total_recaudado
    equilibrio = total_ppto_ingresos - total_aprop

    return {
        "apropiacion": total_aprop,
        "cdp": total_cdp,
        "comprometido": total_comp,
        "obligado": total_obl,
        "pagado": total_pago,
        "saldo_disponible": total_aprop - total_cdp,
        "saldo_por_pagar": total_comp - total_pago,
        "ppto_ingresos": total_ppto_ingresos,
        "recaudado": total_recaudado,
        "saldo_por_recaudar": saldo_por_recaudar,
        "equilibrio": equilibrio,
    }


# ===================== VERIFICAR EQUILIBRIO PRESUPUESTAL =====================
def verificar_equilibrio():
    """Verifica que el total de ingresos sea igual al total de gastos.
    Retorna (total_gastos, total_ingresos, diferencia)."""
    conn = get_connection()
    total_gastos = conn.execute(
        "SELECT COALESCE(SUM(apropiacion_definitiva), 0) as t FROM rubros_gastos WHERE es_hoja=1"
    ).fetchone()["t"]
    total_ingresos = conn.execute(
        "SELECT COALESCE(SUM(presupuesto_definitivo), 0) as t FROM rubros_ingresos WHERE es_hoja=1"
    ).fetchone()["t"]
    conn.close()
    return total_gastos, total_ingresos, total_ingresos - total_gastos


# ===================== IMPORTAR DESDE EXCEL =====================
def importar_catalogo_excel(filepath):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    conn = get_connection()
    count_g, count_i = 0, 0

    # Buscar hojas por nombre (pueden tener espacios al final)
    def buscar_hoja(nombre):
        for sn in wb.sheetnames:
            if sn.strip().upper() == nombre.upper():
                return wb[sn]
        return None

    skip_words = ("total", "totales", "total gastos", "total ingresos",
                  "codigo", "cuenta", "presupuesto", "desequilibrio")

    # Importar GASTOS
    ws = buscar_hoja("GASTOS")
    if ws:
        codigos = []
        for row in ws.iter_rows(min_row=2, max_col=17, values_only=False):
            codigo = str(row[1].value or "").strip()  # Col B
            cuenta = str(row[2].value or "").strip()   # Col C
            if not codigo or codigo.lower() in skip_words:
                continue
            # Verificar que el codigo parece un codigo presupuestal (empieza con digito)
            if not codigo[0].isdigit():
                continue
            aprop_def = float(row[8].value or 0) if row[8].value else 0  # Col I
            codigos.append(codigo)
            conn.execute(
                "INSERT OR REPLACE INTO rubros_gastos "
                "(codigo, cuenta, apropiacion_definitiva, apropiacion_inicial) VALUES (?,?,?,?)",
                (codigo, cuenta, aprop_def, aprop_def)
            )
            count_g += 1

        # Marcar hojas (leaf nodes)
        for codigo in codigos:
            es_hoja = 1
            prefijo = codigo + "."
            for otro in codigos:
                if otro.startswith(prefijo):
                    es_hoja = 0
                    break
            conn.execute("UPDATE rubros_gastos SET es_hoja=? WHERE codigo=?", (es_hoja, codigo))

    # Importar INGRESOS
    ws = buscar_hoja("INGRESOS")
    if ws:
        codigos_i = []
        for row in ws.iter_rows(min_row=2, max_col=10, values_only=False):
            codigo = str(row[1].value or "").strip()
            cuenta = str(row[2].value or "").strip()
            if not codigo or codigo.lower() in skip_words:
                continue
            if not codigo[0].isdigit():
                continue
            # Col G = PRESUPUESTO DEFINITIVO (indice 6)
            ppto_def = float(row[6].value or 0) if row[6].value else 0
            codigos_i.append(codigo)
            conn.execute(
                "INSERT OR REPLACE INTO rubros_ingresos "
                "(codigo, cuenta, presupuesto_definitivo, presupuesto_inicial) VALUES (?,?,?,?)",
                (codigo, cuenta, ppto_def, ppto_def)
            )
            count_i += 1

        # Marcar hojas de ingresos
        for codigo in codigos_i:
            es_hoja = 1
            prefijo = codigo + "."
            for otro in codigos_i:
                if otro.startswith(prefijo):
                    es_hoja = 0
                    break
            conn.execute("UPDATE rubros_ingresos SET es_hoja=? WHERE codigo=?", (es_hoja, codigo))

    conn.commit()
    conn.close()
    wb.close()

    # Verificar equilibrio despues de importar
    total_gastos, total_ingresos, diferencia = verificar_equilibrio()
    if diferencia != 0:
        return count_g, count_i, total_gastos, total_ingresos, diferencia
    return count_g, count_i, total_gastos, total_ingresos, 0


# ===================== IMPORTAR DESDE CSV =====================
import csv

def importar_rubros_gastos_csv(filepath, separador=";"):
    """Importa rubros de gastos desde CSV. Columnas: codigo;cuenta;apropiacion_inicial"""
    errores = []
    cantidad = 0
    conn = get_connection()
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=separador)
            encabezado = next(reader, None)
            if not encabezado or len(encabezado) < 3:
                return 0, ["El archivo no tiene el formato esperado (codigo;cuenta;apropiacion_inicial)"]
            for i, row in enumerate(reader, start=2):
                try:
                    if len(row) < 3 or not row[0].strip():
                        continue
                    codigo = row[0].strip()
                    cuenta = row[1].strip()
                    try:
                        aprop = float(row[2].strip().replace(",", ""))
                    except ValueError:
                        errores.append(f"Fila {i}: valor invalido '{row[2].strip()}'")
                        continue
                    conn.execute(
                        "INSERT OR REPLACE INTO rubros_gastos "
                        "(codigo, cuenta, apropiacion_inicial, apropiacion_definitiva) VALUES (?,?,?,?)",
                        (codigo, cuenta, aprop, aprop)
                    )
                    cantidad += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {e}")
        _recalcular_hojas(conn)
        conn.commit()
    finally:
        conn.close()
    sincronizar_padres_gastos()
    return cantidad, errores


def importar_rubros_ingresos_csv(filepath, separador=";"):
    """Importa rubros de ingresos desde CSV. Columnas: codigo;cuenta;presupuesto_inicial"""
    errores = []
    cantidad = 0
    conn = get_connection()
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=separador)
            encabezado = next(reader, None)
            if not encabezado or len(encabezado) < 3:
                return 0, ["El archivo no tiene el formato esperado (codigo;cuenta;presupuesto_inicial)"]
            for i, row in enumerate(reader, start=2):
                try:
                    if len(row) < 3 or not row[0].strip():
                        continue
                    codigo = row[0].strip()
                    cuenta = row[1].strip()
                    try:
                        ppto = float(row[2].strip().replace(",", ""))
                    except ValueError:
                        errores.append(f"Fila {i}: valor invalido '{row[2].strip()}'")
                        continue
                    conn.execute(
                        "INSERT OR REPLACE INTO rubros_ingresos "
                        "(codigo, cuenta, presupuesto_inicial, presupuesto_definitivo) VALUES (?,?,?,?)",
                        (codigo, cuenta, ppto, ppto)
                    )
                    cantidad += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {e}")
        _recalcular_hojas_ingresos(conn)
        conn.commit()
    finally:
        conn.close()
    sincronizar_padres_ingresos()
    return cantidad, errores


def importar_terceros_csv(filepath, separador=";"):
    """Importa terceros desde CSV. Columnas: nit;dv;nombre;direccion;telefono;email;tipo;banco;tipo_cuenta;no_cuenta"""
    errores = []
    cantidad = 0
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=separador)
            encabezado = next(reader, None)
            if not encabezado or len(encabezado) < 3:
                return 0, ["El archivo no tiene el formato esperado (nit;dv;nombre;...)"]
            for i, row in enumerate(reader, start=2):
                try:
                    if len(row) < 3 or not row[0].strip():
                        continue
                    nit = row[0].strip()
                    dv = row[1].strip() if len(row) > 1 else ""
                    nombre = row[2].strip() if len(row) > 2 else ""
                    direccion = row[3].strip() if len(row) > 3 else ""
                    telefono = row[4].strip() if len(row) > 4 else ""
                    email = row[5].strip() if len(row) > 5 else ""
                    tipo = row[6].strip() if len(row) > 6 and row[6].strip() else "Natural"
                    banco = row[7].strip() if len(row) > 7 else ""
                    tipo_cuenta = row[8].strip() if len(row) > 8 else ""
                    no_cuenta = row[9].strip() if len(row) > 9 else ""
                    if not nombre:
                        errores.append(f"Fila {i}: nombre vacio para NIT {nit}")
                        continue
                    guardar_tercero(nit, dv, nombre, direccion, telefono, email,
                                    tipo, banco, tipo_cuenta, no_cuenta)
                    cantidad += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {e}")
    except Exception as e:
        return 0, [f"Error al leer archivo: {e}"]
    return cantidad, errores


def importar_conceptos_csv(filepath, separador=";"):
    """Importa conceptos desde CSV. Columnas: codigo_rubro;concepto"""
    errores = []
    cantidad = 0
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=separador)
            encabezado = next(reader, None)
            if not encabezado or len(encabezado) < 2:
                return 0, ["El archivo no tiene el formato esperado (codigo_rubro;concepto)"]
            for i, row in enumerate(reader, start=2):
                try:
                    if len(row) < 2 or not row[0].strip():
                        continue
                    codigo_rubro = row[0].strip()
                    concepto = row[1].strip()
                    if not concepto:
                        errores.append(f"Fila {i}: concepto vacio")
                        continue
                    guardar_concepto(codigo_rubro, concepto)
                    cantidad += 1
                except Exception as e:
                    errores.append(f"Fila {i}: {e}")
    except Exception as e:
        return 0, [f"Error al leer archivo: {e}"]
    return cantidad, errores


def exportar_plantillas_csv(directorio):
    """Genera 4 archivos CSV de plantilla con encabezados y una fila de ejemplo."""
    plantillas = {
        "rubros_gastos.csv": (
            ["codigo", "cuenta", "apropiacion_inicial"],
            ["2.1.1.01.01", "Sueldos Personal", "1000000"]
        ),
        "rubros_ingresos.csv": (
            ["codigo", "cuenta", "presupuesto_inicial"],
            ["1.1.1.01.01", "Transferencias SGP", "1000000"]
        ),
        "terceros.csv": (
            ["nit", "dv", "nombre", "direccion", "telefono", "email", "tipo", "banco", "tipo_cuenta", "no_cuenta"],
            ["900123456", "7", "EMPRESA EJEMPLO SAS", "Calle 1 # 2-3", "3001234567", "correo@ejemplo.com", "Juridica", "Bancolombia", "Ahorros", "12345678901"]
        ),
        "conceptos.csv": (
            ["codigo_rubro", "concepto"],
            ["2.1.1.01.01", "Pago nomina mes de enero"]
        ),
    }
    archivos_creados = []
    for nombre, (encabezados, ejemplo) in plantillas.items():
        ruta = os.path.join(directorio, nombre)
        with open(ruta, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(encabezados)
            writer.writerow(ejemplo)
        archivos_creados.append(nombre)
    return archivos_creados


# ===================== BACKUP / RESTAURACION =====================
def crear_backup(destino):
    """Copia la BD al directorio destino con timestamp en el nombre."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    nombre = f"backup_{timestamp}.db"
    ruta_backup = os.path.join(destino, nombre)
    shutil.copy2(DB_PATH, ruta_backup)
    set_config("ultimo_backup", datetime.now().strftime("%Y-%m-%d %H:%M"))
    return ruta_backup


def restaurar_backup(origen):
    """Restaura un archivo .db como la BD principal. Retorna True/False."""
    if not os.path.exists(origen):
        return False
    try:
        # Verificar que es un SQLite valido
        conn_test = sqlite3.connect(origen)
        conn_test.execute("SELECT name FROM sqlite_master WHERE type='table'")
        conn_test.close()
    except Exception:
        return False
    shutil.copy2(origen, DB_PATH)
    return True


def get_info_db():
    """Retorna info de la BD: tamano, fecha modificacion, conteo por tabla."""
    info = {}
    if os.path.exists(DB_PATH):
        stat = os.stat(DB_PATH)
        info["tamano"] = stat.st_size
        info["fecha_modificacion"] = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")
    else:
        info["tamano"] = 0
        info["fecha_modificacion"] = "N/A"
    conn = get_connection()
    for tabla in ("cdp", "rp", "obligacion", "pago", "recaudo"):
        try:
            r = conn.execute(f"SELECT COUNT(*) as c FROM {tabla} WHERE estado != 'ANULADO'").fetchone()
            info[tabla] = r["c"]
        except Exception:
            info[tabla] = 0
    conn.close()
    return info


# ===================== PAC =====================
def inicializar_pac(codigo_rubro):
    """Crea 12 registros con valor 0 para un rubro."""
    conn = get_connection()
    for mes in range(1, 13):
        conn.execute(
            "INSERT OR IGNORE INTO pac (codigo_rubro, mes, valor_programado) VALUES (?, ?, 0)",
            (codigo_rubro, mes)
        )
    conn.commit()
    conn.close()


def get_pac(codigo_rubro):
    """Retorna lista de 12 registros {mes, valor_programado}."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT mes, valor_programado FROM pac WHERE codigo_rubro=? ORDER BY mes",
        (codigo_rubro,)
    ).fetchall()
    conn.close()
    if not rows:
        return []
    return [{"mes": r["mes"], "valor_programado": r["valor_programado"]} for r in rows]


def set_pac(codigo_rubro, mes, valor):
    """Actualiza valor programado de un mes."""
    conn = get_connection()
    conn.execute(
        "INSERT OR REPLACE INTO pac (codigo_rubro, mes, valor_programado) VALUES (?, ?, ?)",
        (codigo_rubro, mes, valor)
    )
    conn.commit()
    conn.close()


def set_pac_completo(codigo_rubro, valores_mensuales):
    """Recibe lista de 12 valores, actualiza todos."""
    conn = get_connection()
    for mes, valor in enumerate(valores_mensuales, 1):
        conn.execute(
            "INSERT OR REPLACE INTO pac (codigo_rubro, mes, valor_programado) VALUES (?, ?, ?)",
            (codigo_rubro, mes, valor)
        )
    conn.commit()
    conn.close()


def get_pac_disponible(codigo_rubro, mes):
    """Retorna: valor_programado_mes - pagos_realizados_mes."""
    conn = get_connection()
    r = conn.execute(
        "SELECT valor_programado FROM pac WHERE codigo_rubro=? AND mes=?",
        (codigo_rubro, mes)
    ).fetchone()
    if not r:
        conn.close()
        return None
    programado = r["valor_programado"]
    # Pagos activos del mes para este rubro
    mes_str = f"{mes:02d}"
    r2 = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM pago "
        "WHERE codigo_rubro=? AND estado='PAGADO' AND substr(fecha, 6, 2)=?",
        (codigo_rubro, mes_str)
    ).fetchone()
    conn.close()
    return programado - r2["total"]


def get_pagos_mes_rubro(codigo_rubro, mes):
    """Retorna total de pagos activos de un rubro en un mes."""
    conn = get_connection()
    mes_str = f"{mes:02d}"
    r = conn.execute(
        "SELECT COALESCE(SUM(valor), 0) as total FROM pago "
        "WHERE codigo_rubro=? AND estado='PAGADO' AND substr(fecha, 6, 2)=?",
        (codigo_rubro, mes_str)
    ).fetchone()
    conn.close()
    return r["total"]


def get_resumen_pac():
    """Retorna todos los rubros hoja con su PAC y ejecucion."""
    rubros = get_rubros_gastos(solo_hojas=True)
    resultado = []
    for rubro in rubros:
        codigo = rubro["codigo"]
        pac = get_pac(codigo)
        total_pac = sum(p["valor_programado"] for p in pac) if pac else 0
        resultado.append({
            "codigo": codigo,
            "cuenta": rubro["cuenta"],
            "apropiacion_definitiva": rubro["apropiacion_definitiva"],
            "total_pac": total_pac,
            "pac_configurado": len(pac) > 0,
            "pac": pac,
        })
    return resultado


def distribuir_pac_uniforme(codigo_rubro):
    """Distribuye la apropiacion definitiva en 12 partes iguales."""
    rubro = get_rubro_gasto(codigo_rubro)
    if not rubro:
        raise ValueError(f"Rubro {codigo_rubro} no encontrado")
    total = rubro["apropiacion_definitiva"]
    valor_mes = round(total / 12, 2)
    # Ajustar ultimo mes para que sume exacto
    valores = [valor_mes] * 11
    valores.append(round(total - valor_mes * 11, 2))
    set_pac_completo(codigo_rubro, valores)
    return valores


def validar_pago_pac(codigo_rubro, mes, valor):
    """Verifica si hay cupo PAC disponible. Retorna (ok, mensaje)."""
    pac = get_pac(codigo_rubro)
    if not pac:
        return True, "PAC no configurado para este rubro"
    disponible = get_pac_disponible(codigo_rubro, mes)
    if disponible is None:
        return True, "PAC no configurado para este mes"
    if valor > disponible:
        return False, (f"Valor ($ {valor:,.0f}) excede el cupo PAC del mes "
                       f"(disponible: $ {disponible:,.0f})")
    return True, "OK"
