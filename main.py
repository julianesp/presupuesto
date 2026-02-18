"""
main.py - Sistema Presupuestal IE 2026
Interfaz grafica con tkinter + SQLite
"""
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import os
import sys
import importlib
import webbrowser
import tempfile
from datetime import date, datetime

import database as db

# Colores del tema
COLOR_BG = "#F0F4F8"
COLOR_PRIMARY = "#003366"
COLOR_SUCCESS = "#006633"
COLOR_WARNING = "#993300"
COLOR_DANGER = "#C00000"
COLOR_PURPLE = "#660066"
COLOR_ACCENT = "#006699"
COLOR_WHITE = "#FFFFFF"
COLOR_LIGHT = "#E8EDF2"

MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sistema Presupuestal IE 2026")
        self.geometry("1100x700")
        self.configure(bg=COLOR_BG)
        self.state("zoomed")

        db.init_db()
        self._verificar_catalogo()
        self._crear_menu_superior()
        self._crear_contenido()

    def _verificar_catalogo(self):
        rubros = db.get_rubros_gastos(solo_hojas=True)
        if not rubros:
            resp = messagebox.askyesno(
                "Importar Catalogo",
                "No se encontraron rubros en la base de datos.\n\n"
                "Desea importar el catalogo desde el archivo Excel existente?"
            )
            if resp:
                self._importar_catalogo()

    def _importar_catalogo(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo de catalogo presupuestal",
            initialdir=os.path.dirname(os.path.dirname(__file__)),
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")]
        )
        if filepath:
            try:
                g, i, total_gas, total_ing, diferencia = db.importar_catalogo_excel(filepath)
                msg = (f"Rubros de gastos importados: {g}\n"
                       f"Rubros de ingresos importados: {i}\n\n"
                       f"Total Gastos (Aprop. Definitiva): $ {total_gas:,.0f}\n"
                       f"Total Ingresos (Ppto. Definitivo): $ {total_ing:,.0f}\n")
                if diferencia != 0:
                    msg += (f"\nDIFERENCIA: $ {diferencia:,.0f}\n\n"
                            f"ADVERTENCIA: No hay equilibrio presupuestal.\n"
                            f"Los ingresos y gastos deben ser iguales.")
                    messagebox.showwarning("Importacion - SIN EQUILIBRIO", msg)
                else:
                    msg += "\nEQUILIBRIO PRESUPUESTAL: OK"
                    messagebox.showinfo("Importacion Exitosa", msg)
            except Exception as e:
                messagebox.showerror("Error", f"Error al importar:\n{e}")

    def _importar_csv(self, tipo):
        formatos = {
            "gastos": ("Rubros de Gastos",
                       "codigo;cuenta;apropiacion_inicial",
                       "Ejemplo: 2.1.1.01.01;Sueldos Personal;1000000"),
            "ingresos": ("Rubros de Ingresos",
                         "codigo;cuenta;presupuesto_inicial",
                         "Ejemplo: 1.1.1.01.01;Transferencias SGP;1000000"),
            "terceros": ("Terceros",
                         "nit;dv;nombre;direccion;telefono;email;tipo;banco;tipo_cuenta;no_cuenta",
                         "Campos desde 'direccion' son opcionales.\n"
                         "Ejemplo: 900123456;7;EMPRESA EJEMPLO SAS;Calle 1;300123;correo@ej.com;Juridica;Bancolombia;Ahorros;123456"),
            "conceptos": ("Conceptos",
                          "codigo_rubro;concepto",
                          "Ejemplo: 2.1.1.01.01;Pago nomina mes de enero"),
        }
        titulo, columnas, ejemplo = formatos[tipo]
        messagebox.showinfo(
            f"Formato CSV - {titulo}",
            f"El archivo CSV debe tener separador punto y coma (;)\n"
            f"y codificacion UTF-8.\n\n"
            f"Columnas:\n{columnas}\n\n{ejemplo}\n\n"
            f"La primera fila (encabezados) sera ignorada."
        )
        filepath = filedialog.askopenfilename(
            title=f"Seleccionar archivo CSV de {titulo}",
            filetypes=[("CSV", "*.csv"), ("Texto", "*.txt"), ("Todos", "*.*")]
        )
        if not filepath:
            return
        try:
            if tipo == "gastos":
                cantidad, errores = db.importar_rubros_gastos_csv(filepath)
            elif tipo == "ingresos":
                cantidad, errores = db.importar_rubros_ingresos_csv(filepath)
            elif tipo == "terceros":
                cantidad, errores = db.importar_terceros_csv(filepath)
            elif tipo == "conceptos":
                cantidad, errores = db.importar_conceptos_csv(filepath)
            msg = f"Registros importados: {cantidad}"
            if errores:
                msg += f"\n\nErrores ({len(errores)}):\n" + "\n".join(errores[:20])
                if len(errores) > 20:
                    msg += f"\n... y {len(errores) - 20} errores mas."
            if tipo in ("gastos", "ingresos"):
                total_g, total_i, dif = db.verificar_equilibrio()
                msg += (f"\n\nVerificacion de equilibrio:\n"
                        f"Total Gastos: $ {total_g:,.0f}\n"
                        f"Total Ingresos: $ {total_i:,.0f}\n"
                        f"Diferencia: $ {dif:,.0f}")
                if dif != 0:
                    messagebox.showwarning("Importacion CSV - SIN EQUILIBRIO", msg)
                else:
                    messagebox.showinfo("Importacion CSV Exitosa", msg)
            else:
                if errores:
                    messagebox.showwarning("Importacion CSV con errores", msg)
                else:
                    messagebox.showinfo("Importacion CSV Exitosa", msg)
        except Exception as e:
            messagebox.showerror("Error", f"Error al importar CSV:\n{e}")

    def _descargar_plantillas_csv(self):
        directorio = filedialog.askdirectory(title="Seleccionar carpeta para guardar plantillas CSV")
        if not directorio:
            return
        try:
            archivos = db.exportar_plantillas_csv(directorio)
            messagebox.showinfo(
                "Plantillas CSV",
                f"Se crearon {len(archivos)} plantillas en:\n{directorio}\n\n"
                + "\n".join(f"  - {a}" for a in archivos)
            )
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear plantillas:\n{e}")

    def _crear_menu_superior(self):
        menubar = tk.Menu(self)

        # 1. Menu Plan Presupuestal
        m_plan = tk.Menu(menubar, tearoff=0)
        m_plan.add_command(label="Gestion Rubros de Gastos", command=self._gestion_rubros)
        m_plan.add_command(label="Gestion Rubros de Ingresos", command=self._gestion_rubros_ingresos)
        m_plan.add_separator()
        m_plan.add_command(label="Importar Catalogo Excel", command=self._importar_catalogo)
        # Submenu Importar Archivo Plano (CSV)
        m_csv = tk.Menu(m_plan, tearoff=0)
        m_csv.add_command(label="Rubros de Gastos (CSV)", command=lambda: self._importar_csv("gastos"))
        m_csv.add_command(label="Rubros de Ingresos (CSV)", command=lambda: self._importar_csv("ingresos"))
        m_csv.add_command(label="Terceros (CSV)", command=lambda: self._importar_csv("terceros"))
        m_csv.add_command(label="Conceptos (CSV)", command=lambda: self._importar_csv("conceptos"))
        m_csv.add_separator()
        m_csv.add_command(label="Descargar Plantillas CSV", command=self._descargar_plantillas_csv)
        m_plan.add_cascade(label="Importar Archivo Plano", menu=m_csv)
        m_plan.add_command(label="Resumen Presupuestal", command=self._ver_resumen_detallado)
        m_plan.add_separator()
        m_plan.add_command(label="Plan Anualizado de Caja (PAC)", command=self._gestionar_pac)
        menubar.add_cascade(label="Plan Presupuestal", menu=m_plan)

        # 2. Menu Gestion de Terceros
        m_terceros = tk.Menu(menubar, tearoff=0)
        m_terceros.add_command(label="Registrar Tercero", command=self._registrar_tercero)
        m_terceros.add_command(label="Ver Terceros", command=lambda: self._ver_listado("TERCEROS"))
        menubar.add_cascade(label="Gestion de Terceros", menu=m_terceros)

        # 3. Menu Registro de Movimiento (orden de ejecucion presupuestal)
        m_registro = tk.Menu(menubar, tearoff=0)
        m_registro.add_command(label="1. Disponibilidades (CDP)", command=self._registrar_cdp)
        m_registro.add_command(label="2. Registros Presupuestales (RP)", command=self._registrar_rp)
        m_registro.add_command(label="3. Obligaciones", command=self._registrar_obligacion)
        m_registro.add_command(label="4. Pagos", command=self._registrar_pago)
        m_registro.add_separator()
        m_registro.add_command(label="5. Ingresos / Recaudos", command=self._registrar_recaudo)
        m_registro.add_separator()
        m_registro.add_command(label="6. Adicion Presupuestal", command=self._registrar_adicion)
        m_registro.add_command(label="7. Reduccion Presupuestal", command=self._registrar_reduccion)
        m_registro.add_command(label="8. Credito / Contracredito", command=self._registrar_credito_contracredito)
        menubar.add_cascade(label="Registro de Movimiento", menu=m_registro)

        # 4. Menu Consultas / Listados
        m_consultas = tk.Menu(menubar, tearoff=0)
        m_consultas.add_command(label="Ver CDPs", command=lambda: self._ver_listado("CDP"))
        m_consultas.add_command(label="Ver RPs", command=lambda: self._ver_listado("RP"))
        m_consultas.add_command(label="Ver Obligaciones", command=lambda: self._ver_listado("OBLIGACION"))
        m_consultas.add_command(label="Ver Pagos", command=lambda: self._ver_listado("PAGO"))
        m_consultas.add_separator()
        m_consultas.add_command(label="Ver Recaudos", command=self._seleccionar_recaudo)
        m_consultas.add_separator()
        m_consultas.add_command(label="Tarjeta Presupuestal", command=self._ver_tarjeta)
        m_consultas.add_separator()
        m_consultas.add_command(label="Ver Modificaciones Presupuestales", command=self._ver_modificaciones)
        menubar.add_cascade(label="Consultas", menu=m_consultas)

        # 5. Menu Generacion de Informes
        m_informes = tk.Menu(menubar, tearoff=0)
        m_informes.add_command(label="Informes y Ejecuciones", command=self._menu_informes)
        m_informes.add_separator()
        m_informes.add_command(label="Informe SIFSE (Trimestral)", command=self._generar_informe_sifse)
        m_informes.add_command(label="Informe SIA Contraloria (Acumulado Anual)", command=self._generar_informe_sia)
        menubar.add_cascade(label="Generacion de Informes", menu=m_informes)

        # 6. Menu Procesos Presupuestales
        m_procesos = tk.Menu(menubar, tearoff=0)
        m_procesos.add_command(label="Consolidar Mes", command=self._consolidar_mes)
        m_procesos.add_command(label="Cierre de Mes", command=self._cierre_mes)
        menubar.add_cascade(label="Procesos Presupuestales", menu=m_procesos)

        # 7. Menu Configuracion
        m_config = tk.Menu(menubar, tearoff=0)
        m_config.add_command(label="Configuracion General", command=self._ventana_config)
        m_config.add_command(label="Cuentas Bancarias", command=self._gestionar_cuentas_bancarias)
        m_config.add_command(label="Mapeo SIFSE", command=self._configurar_mapeo_sifse)
        m_config.add_separator()
        m_config.add_command(label="Crear Copia de Seguridad", command=self._crear_backup)
        m_config.add_command(label="Restaurar Copia de Seguridad", command=self._restaurar_backup)
        m_config.add_separator()
        m_config.add_command(label="Salir", command=self.quit)
        menubar.add_cascade(label="Configuracion", menu=m_config)

        self.config(menu=menubar)

    def _crear_contenido(self):
        # Frame principal con scroll
        self._scroll_canvas = tk.Canvas(self, bg=COLOR_BG, highlightthickness=0)
        self._scrollbar = tk.Scrollbar(self, orient="vertical", command=self._scroll_canvas.yview)
        self.main_frame = tk.Frame(self._scroll_canvas, bg=COLOR_BG)

        self.main_frame.bind("<Configure>",
            lambda e: self._scroll_canvas.configure(scrollregion=self._scroll_canvas.bbox("all")))
        self._scroll_canvas.create_window((0, 0), window=self.main_frame, anchor="nw", tags="main_win")
        self._scroll_canvas.configure(yscrollcommand=self._scrollbar.set)

        self._scrollbar.pack(side="right", fill="y")
        self._scroll_canvas.pack(side="left", fill="both", expand=True, padx=20, pady=10)

        # Ajustar ancho del frame interno al canvas
        def _on_canvas_configure(event):
            self._scroll_canvas.itemconfig("main_win", width=event.width)
        self._scroll_canvas.bind("<Configure>", _on_canvas_configure)

        # Scroll con rueda del mouse
        def _on_mousewheel(event):
            self._scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self._scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        self._mostrar_dashboard()

    def _limpiar_main(self):
        for w in self.main_frame.winfo_children():
            w.destroy()
        # Resetear scroll al inicio
        self._scroll_canvas.yview_moveto(0)

    def _mostrar_dashboard(self):
        db.sincronizar_padres_gastos()
        db.sincronizar_padres_ingresos()
        self._limpiar_main()
        f = self.main_frame

        # Titulo con datos de la institucion
        nombre_ie = db.get_config("institucion") or "INSTITUCION EDUCATIVA"
        nit_ie = db.get_config("nit_institucion") or ""
        rector = db.get_config("rector") or ""
        tesorero = db.get_config("tesorero") or ""
        vigencia = db.get_config("vigencia") or "2026"

        tk.Label(f, text="SISTEMA PRESUPUESTAL", font=("Segoe UI", 22, "bold"),
                 bg=COLOR_BG, fg=COLOR_PRIMARY).pack(pady=(10, 0))
        tk.Label(f, text=nombre_ie, font=("Segoe UI", 14, "bold"),
                 bg=COLOR_BG, fg=COLOR_ACCENT).pack(pady=(0, 0))
        if nit_ie and not nit_ie.startswith("("):
            tk.Label(f, text=f"NIT: {nit_ie}", font=("Segoe UI", 10),
                     bg=COLOR_BG, fg="#555").pack()

        frame_info_ie = tk.Frame(f, bg=COLOR_BG)
        frame_info_ie.pack(pady=(2, 0))
        tk.Label(frame_info_ie, text=f"Vigencia: {vigencia}", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_BG, fg=COLOR_PRIMARY).pack(side="left", padx=10)
        mes = int(db.get_config("mes_actual") or 1)
        if 1 <= mes <= 12:
            tk.Label(frame_info_ie, text=f"Mes activo: {MESES[mes]}", font=("Segoe UI", 10),
                     bg=COLOR_BG, fg="#555").pack(side="left", padx=10)

        if (rector and not rector.startswith("(")) or (tesorero and not tesorero.startswith("(")):
            frame_cargos = tk.Frame(f, bg=COLOR_BG)
            frame_cargos.pack(pady=(0, 2))
            if rector and not rector.startswith("("):
                tk.Label(frame_cargos, text=f"Rector: {rector}", font=("Segoe UI", 9),
                         bg=COLOR_BG, fg="#444").pack(side="left", padx=10)
            if tesorero and not tesorero.startswith("("):
                tk.Label(frame_cargos, text=f"Tesorero: {tesorero}", font=("Segoe UI", 9),
                         bg=COLOR_BG, fg="#444").pack(side="left", padx=10)

        # === 1. PLAN PRESUPUESTAL ===
        lbl_plan = tk.Label(f, text="1. PLAN PRESUPUESTAL", font=("Segoe UI", 9, "bold"),
                            bg=COLOR_BG, fg="#2E7D32")
        lbl_plan.pack(anchor="w", padx=15, pady=(8, 1))

        frame_plan = tk.Frame(f, bg=COLOR_BG)
        frame_plan.pack(pady=1)

        botones_plan = [
            ("RUBROS DE\nGASTOS", "#2E7D32", self._gestion_rubros),
            ("RUBROS DE\nINGRESOS", "#006633", self._gestion_rubros_ingresos),
            ("RESUMEN\nPRESUPUESTAL", "#336699", self._ver_resumen_detallado),
            ("BASE DE\nTERCEROS", COLOR_ACCENT, self._registrar_tercero),
            ("PAC", "#8B4513", self._gestionar_pac),
        ]
        for i, (texto, color, cmd) in enumerate(botones_plan):
            btn = tk.Button(frame_plan, text=texto, bg=color, fg=COLOR_WHITE,
                            font=("Segoe UI", 9, "bold"), width=15, height=2,
                            cursor="hand2", relief="raised", bd=2, command=cmd)
            btn.grid(row=0, column=i, padx=4, pady=2)

        # === 2. REGISTRO DE MOVIMIENTO (orden de ejecucion) ===
        lbl_reg = tk.Label(f, text="2. REGISTRO DE MOVIMIENTO", font=("Segoe UI", 9, "bold"),
                           bg=COLOR_BG, fg=COLOR_PRIMARY)
        lbl_reg.pack(anchor="w", padx=15, pady=(8, 1))

        frame_btns = tk.Frame(f, bg=COLOR_BG)
        frame_btns.pack(pady=1)

        botones_reg = [
            ("1. CDP", COLOR_SUCCESS, self._registrar_cdp),
            ("2. RP", COLOR_PRIMARY, self._registrar_rp),
            ("3. OBLIGACION", COLOR_WARNING, self._registrar_obligacion),
            ("4. PAGO", COLOR_PURPLE, self._registrar_pago),
            ("5. RECAUDO", "#008B45", self._registrar_recaudo),
        ]
        for i, (texto, color, cmd) in enumerate(botones_reg):
            btn = tk.Button(frame_btns, text=texto, bg=color, fg=COLOR_WHITE,
                            font=("Segoe UI", 10, "bold"), width=15, height=2,
                            cursor="hand2", relief="raised", bd=2, command=cmd)
            btn.grid(row=0, column=i, padx=4, pady=2)

        # === MODIFICACIONES PRESUPUESTALES ===
        lbl_mod = tk.Label(f, text="MODIFICACIONES PRESUPUESTALES", font=("Segoe UI", 9, "bold"),
                           bg=COLOR_BG, fg=COLOR_DANGER)
        lbl_mod.pack(anchor="w", padx=15, pady=(8, 1))

        frame_mod = tk.Frame(f, bg=COLOR_BG)
        frame_mod.pack(pady=1)

        botones_mod = [
            ("6. ADICION", "#B8860B", self._registrar_adicion),
            ("7. REDUCCION", "#8B4513", self._registrar_reduccion),
            ("8. CRED/CONTRACRED", "#800080", self._registrar_credito_contracredito),
            ("MODIFICACIONES", "#555555", self._ver_modificaciones),
        ]
        for i, (texto, color, cmd) in enumerate(botones_mod):
            btn = tk.Button(frame_mod, text=texto, bg=color, fg=COLOR_WHITE,
                            font=("Segoe UI", 9, "bold"), width=15, height=2,
                            cursor="hand2", relief="raised", bd=2, command=cmd)
            btn.grid(row=0, column=i, padx=4, pady=2)

        # === 3. INFORMES Y PROCESOS ===
        lbl_inf = tk.Label(f, text="3. INFORMES Y PROCESOS", font=("Segoe UI", 9, "bold"),
                           bg=COLOR_BG, fg="#4B0082")
        lbl_inf.pack(anchor="w", padx=15, pady=(8, 1))

        frame_btns2 = tk.Frame(f, bg=COLOR_BG)
        frame_btns2.pack(pady=1)

        botones_inf = [
            ("INFORMES", "#4B0082", self._menu_informes),
            ("TARJETAS", COLOR_ACCENT, self._ver_tarjeta),
            ("LISTADOS", COLOR_ACCENT, self._menu_listados),
            ("RECAUDOS", "#2E8B57", self._seleccionar_recaudo),
            ("CONSOLIDAR", COLOR_DANGER, self._consolidar_mes),
            ("CIERRE MES", "#8B0000", self._cierre_mes),
            ("ACTUALIZAR", "#555555", self._reiniciar_app),
        ]
        for i, (texto, color, cmd) in enumerate(botones_inf):
            btn = tk.Button(frame_btns2, text=texto, bg=color, fg=COLOR_WHITE,
                            font=("Segoe UI", 9, "bold"), width=13, height=2,
                            cursor="hand2", relief="raised", bd=2, command=cmd)
            btn.grid(row=0, column=i, padx=3, pady=2)

        # === RESUMEN - Gastos e Ingresos lado a lado ===
        resumen = db.get_resumen()

        frame_resumenes = tk.Frame(f, bg=COLOR_BG)
        frame_resumenes.pack(pady=(10, 5), fill="x")
        frame_resumenes.columnconfigure(0, weight=1)
        frame_resumenes.columnconfigure(1, weight=1)

        # --- Columna izquierda: GASTOS ---
        frame_resumen = tk.LabelFrame(frame_resumenes, text=" RESUMEN GASTOS ",
                                       font=("Segoe UI", 10, "bold"),
                                       bg=COLOR_WHITE, fg=COLOR_PRIMARY, padx=10, pady=5)
        frame_resumen.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        datos = [
            ("Aprop. Definitiva:", resumen["apropiacion"], COLOR_PRIMARY),
            ("Total CDPs:", resumen["cdp"], COLOR_SUCCESS),
            ("Comprometido (RPs):", resumen["comprometido"], COLOR_PRIMARY),
            ("Total Obligado:", resumen["obligado"], COLOR_WARNING),
            ("Total Pagado:", resumen["pagado"], COLOR_PURPLE),
            ("Saldo Disponible:", resumen["saldo_disponible"], COLOR_SUCCESS),
            ("Saldo por Pagar:", resumen["saldo_por_pagar"], COLOR_ACCENT),
        ]
        for i, (label, valor, default_color) in enumerate(datos):
            tk.Label(frame_resumen, text=label, font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE, anchor="e").grid(row=i, column=0, sticky="e", padx=(0, 5))
            color = COLOR_DANGER if valor < 0 else default_color
            tk.Label(frame_resumen, text=f"$ {valor:,.0f}", font=("Segoe UI", 11, "bold"),
                     bg=COLOR_WHITE, fg=color).grid(row=i, column=1, sticky="w")

        # --- Columna derecha: INGRESOS ---
        frame_ingresos = tk.LabelFrame(frame_resumenes, text=" RESUMEN INGRESOS ",
                                        font=("Segoe UI", 10, "bold"),
                                        bg=COLOR_WHITE, fg="#006633", padx=10, pady=5)
        frame_ingresos.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        datos_ing = [
            ("Ppto. Definitivo:", resumen["ppto_ingresos"], "#006633"),
            ("Total Recaudado:", resumen["recaudado"], "#008B45"),
            ("Saldo por Recaudar:", resumen["saldo_por_recaudar"], COLOR_ACCENT),
        ]
        for i, (label, valor, default_color) in enumerate(datos_ing):
            tk.Label(frame_ingresos, text=label, font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE, anchor="e").grid(row=i, column=0, sticky="e", padx=(0, 5))
            color = COLOR_DANGER if valor < 0 else default_color
            tk.Label(frame_ingresos, text=f"$ {valor:,.0f}", font=("Segoe UI", 11, "bold"),
                     bg=COLOR_WHITE, fg=color).grid(row=i, column=1, sticky="w")

        # --- Equilibrio presupuestal ---
        equilibrio = resumen["equilibrio"]
        if equilibrio == 0:
            estado_eq = "EQUILIBRADO"
            color_eq = "#006633"
            bg_eq = "#E8F5E9"
        elif equilibrio > 0:
            estado_eq = "SUPERAVIT - SIN EQUILIBRIO"
            color_eq = COLOR_DANGER
            bg_eq = "#FFEEEE"
        else:
            estado_eq = "DEFICIT - SIN EQUILIBRIO"
            color_eq = COLOR_DANGER
            bg_eq = "#FFEEEE"

        frame_eq = tk.Frame(f, bg=bg_eq, padx=10, pady=5)
        frame_eq.pack(fill="x", pady=(3, 5))
        tk.Label(frame_eq, text=f"EQUILIBRIO (Ingresos - Gastos):  $ {equilibrio:,.0f}  -  {estado_eq}",
                 font=("Segoe UI", 11, "bold"), bg=bg_eq, fg=color_eq).pack()
        if equilibrio != 0:
            tk.Label(frame_eq,
                     text="Los ingresos y gastos DEBEN ser iguales para el equilibrio presupuestal",
                     font=("Segoe UI", 9, "bold"), bg=bg_eq, fg=COLOR_DANGER).pack()

        # --- Indicador de ultimo backup ---
        ultimo_backup = db.get_config("ultimo_backup")
        if ultimo_backup:
            txt_backup = f"Ultima copia de seguridad: {ultimo_backup}"
            color_bk = "#006633"
        else:
            txt_backup = "Sin copias de seguridad - Se recomienda crear una"
            color_bk = COLOR_DANGER
        frame_bk = tk.Frame(f, bg=COLOR_LIGHT, padx=8, pady=3)
        frame_bk.pack(fill="x", pady=(0, 5))
        tk.Label(frame_bk, text=txt_backup, font=("Segoe UI", 9),
                 bg=COLOR_LIGHT, fg=color_bk).pack(side="left")
        tk.Button(frame_bk, text="Crear Backup", font=("Segoe UI", 8),
                  bg=COLOR_ACCENT, fg=COLOR_WHITE, cursor="hand2",
                  command=self._crear_backup).pack(side="right")

    # ===================== DIALOGO BUSCAR RUBRO =====================
    def _buscar_rubro(self):
        dialog = tk.Toplevel(self)
        dialog.title("Buscar Rubro Presupuestal")
        dialog.geometry("560x480")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"codigo": None}

        tk.Label(dialog, text="BUSCAR RUBRO DE GASTO", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        # Busqueda
        frame_busq = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_busq.pack(fill="x", padx=15)
        tk.Label(frame_busq, text="Escriba codigo o nombre:", font=("Segoe UI", 9),
                 bg=COLOR_WHITE, fg="#666").pack(anchor="w")

        frame_input = tk.Frame(frame_busq, bg=COLOR_WHITE)
        frame_input.pack(fill="x")
        entry_buscar = tk.Entry(frame_input, font=("Segoe UI", 12))
        entry_buscar.pack(side="left", fill="x", expand=True, ipady=3)

        # Contador
        lbl_count = tk.Label(dialog, text="", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#999")
        lbl_count.pack(anchor="w", padx=15)

        # Lista
        frame_lista = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("codigo", "cuenta", "apropiacion")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings", height=12)
        tree.heading("codigo", text="Codigo")
        tree.heading("cuenta", text="Cuenta")
        tree.heading("apropiacion", text="Aprop. Definitiva")
        tree.column("codigo", width=180)
        tree.column("cuenta", width=250)
        tree.column("apropiacion", width=100, anchor="e")

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Info seleccion
        lbl_info = tk.Label(dialog, text="", font=("Segoe UI", 9, "bold"),
                            bg=COLOR_WHITE, fg=COLOR_SUCCESS)
        lbl_info.pack(anchor="w", padx=15)

        def filtrar(*args):
            filtro = entry_buscar.get().strip()
            tree.delete(*tree.get_children())
            if filtro:
                rubros = db.buscar_rubros_gastos(filtro)
            else:
                rubros = db.get_rubros_gastos(solo_hojas=True)
            for r in rubros:
                tree.insert("", "end", values=(
                    r["codigo"], r["cuenta"], f"{r['apropiacion_definitiva']:,.0f}"
                ))
            total = len(db.get_rubros_gastos(solo_hojas=True))
            lbl_count.config(text=f"Rubros encontrados: {len(rubros)} de {total}")

        def on_select(event):
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0])["values"]
                codigo = str(vals[0])
                saldo = db.saldo_disponible_rubro(codigo)
                color = COLOR_SUCCESS if saldo > 0 else COLOR_DANGER
                lbl_info.config(text=f"{codigo} - Saldo disponible: $ {saldo:,.0f}", fg=color)

        def aceptar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un rubro", parent=dialog)
                return
            resultado["codigo"] = str(tree.item(sel[0])["values"][0])
            dialog.destroy()

        def on_doble_click(event):
            aceptar()

        entry_buscar.bind("<KeyRelease>", filtrar)
        tree.bind("<<TreeviewSelect>>", on_select)
        tree.bind("<Double-1>", on_doble_click)

        # Botones
        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)

        filtrar()
        entry_buscar.focus_set()
        self.wait_window(dialog)
        return resultado["codigo"]

    # ===================== DIALOGO CONCEPTO =====================
    def _seleccionar_concepto(self, codigo_rubro, titulo_rubro, saldo):
        dialog = tk.Toplevel(self)
        dialog.title("Objeto del Gasto / Concepto")
        dialog.geometry("520x420")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"concepto": None}

        tk.Label(dialog, text=f"Rubro: {codigo_rubro} - {titulo_rubro}",
                 font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY,
                 wraplength=480).pack(anchor="w", padx=15, pady=(10, 0))
        tk.Label(dialog, text=f"Saldo disponible: $ {saldo:,.0f}",
                 font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_SUCCESS).pack(
            anchor="w", padx=15)

        # Conceptos anteriores
        frame_ant = tk.LabelFrame(dialog, text=" Conceptos usados anteriormente: ",
                                   font=("Segoe UI", 9, "bold"), bg=COLOR_WHITE, padx=10, pady=5)
        frame_ant.pack(fill="both", expand=True, padx=15, pady=5)

        conceptos = db.get_conceptos_rubro(codigo_rubro)
        listbox = tk.Listbox(frame_ant, font=("Segoe UI", 10), height=8)
        listbox.pack(fill="both", expand=True)

        if conceptos:
            for c in conceptos:
                listbox.insert("end", c["concepto"])
            frame_ant.config(text=f" Conceptos anteriores ({len(conceptos)} guardados): ")
        else:
            frame_ant.config(text=" No hay conceptos guardados para este rubro ")

        # Nuevo concepto
        frame_nuevo = tk.LabelFrame(dialog, text=" O escriba un concepto nuevo: ",
                                     font=("Segoe UI", 9, "bold"), bg=COLOR_WHITE, padx=10, pady=5)
        frame_nuevo.pack(fill="x", padx=15, pady=5)
        txt_nuevo = tk.Text(frame_nuevo, font=("Segoe UI", 11), height=2, wrap="word")
        txt_nuevo.pack(fill="x")

        def on_list_select(event):
            txt_nuevo.delete("1.0", "end")

        def on_text_change(event):
            if txt_nuevo.get("1.0", "end").strip():
                listbox.selection_clear(0, "end")

        listbox.bind("<<ListboxSelect>>", on_list_select)
        txt_nuevo.bind("<KeyRelease>", on_text_change)

        def aceptar():
            texto = txt_nuevo.get("1.0", "end").strip()
            if texto:
                resultado["concepto"] = texto
            elif listbox.curselection():
                resultado["concepto"] = listbox.get(listbox.curselection()[0])
            else:
                messagebox.showwarning("Concepto", "Seleccione o escriba un concepto", parent=dialog)
                return
            dialog.destroy()

        def on_doble_click(event):
            aceptar()

        listbox.bind("<Double-1>", on_doble_click)

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)

        self.wait_window(dialog)
        return resultado["concepto"]

    # ===================== REGISTRAR CDP =====================
    def _registrar_cdp(self):
        codigo = self._buscar_rubro()
        if not codigo:
            return

        rubro = db.get_rubro_gasto(codigo)
        saldo = db.saldo_disponible_rubro(codigo)
        if saldo <= 0:
            messagebox.showerror("Sin Saldo", f"El rubro {codigo} no tiene saldo disponible.\n"
                                 f"Saldo: $ {saldo:,.0f}")
            return

        objeto = self._seleccionar_concepto(codigo, rubro["cuenta"], saldo)
        if not objeto:
            return

        # Selector de fuente SIFSE
        fuente_sifse = self._seleccionar_fuente_sifse()
        if fuente_sifse is None:
            return

        valor_str = simpledialog.askstring(
            "Registrar CDP - Valor",
            f"Rubro: {codigo} - {rubro['cuenta']}\n"
            f"Objeto: {objeto}\n"
            f"Saldo disponible: $ {saldo:,.0f}\n\n"
            f"Ingrese el VALOR del CDP:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor numerico invalido")
            return

        if valor <= 0:
            messagebox.showerror("Error", "El valor debe ser mayor a cero")
            return
        if valor > saldo:
            messagebox.showerror("Error", f"Valor ($ {valor:,.0f}) excede saldo ($ {saldo:,.0f})")
            return

        # Obtener descripcion de la fuente para la confirmacion
        fuentes = db.get_fuentes_sifse_activas()
        desc_fuente = "Sin asignar"
        for f in fuentes:
            if f["codigo"] == fuente_sifse:
                desc_fuente = f"{f['codigo']} - {f['descripcion']}"
                break

        if not messagebox.askyesno("Confirmar CDP",
                                    f"CONFIRMAR REGISTRO DE CDP\n\n"
                                    f"Rubro: {codigo} - {rubro['cuenta']}\n"
                                    f"Objeto: {objeto}\n"
                                    f"Fuente SIFSE: {desc_fuente}\n"
                                    f"Valor: $ {valor:,.0f}"):
            return

        try:
            num, fecha = db.registrar_cdp(codigo, objeto, valor, fuente_sifse)
            messagebox.showinfo("CDP Registrado",
                                f"CDP REGISTRADO EXITOSAMENTE\n\n"
                                f"No. CDP: {num}\n"
                                f"Fecha: {fecha}\n"
                                f"Rubro: {codigo}\n"
                                f"Fuente SIFSE: {desc_fuente}\n"
                                f"Valor: $ {valor:,.0f}\n"
                                f"Nuevo saldo: $ {saldo - valor:,.0f}")
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _seleccionar_fuente_sifse(self):
        """Muestra dialogo para seleccionar fuente SIFSE. Retorna codigo o None si cancela."""
        fuentes = db.get_fuentes_sifse_activas()
        if not fuentes:
            messagebox.showwarning("SIFSE", "No hay fuentes SIFSE configuradas.\n"
                                   "Configure el mapeo en Configuracion > Mapeo SIFSE.")
            return 0

        dialog = tk.Toplevel(self)
        dialog.title("Seleccionar Fuente SIFSE")
        dialog.geometry("450x180")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"valor": None}

        tk.Label(dialog, text="FUENTE DE FINANCIACION (SIFSE)",
                 font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))
        tk.Label(dialog, text="Seleccione la fuente de financiacion para este gasto:",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(pady=(0, 5))

        opciones = [f"{f['codigo']} - {f['descripcion']}" for f in fuentes]
        fuente_var = tk.StringVar(value=opciones[0] if opciones else "")
        cmb = ttk.Combobox(dialog, textvariable=fuente_var, values=opciones,
                           state="readonly", width=50, font=("Segoe UI", 10))
        cmb.pack(padx=20, pady=5)

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(pady=10)

        def aceptar():
            sel = fuente_var.get()
            if sel:
                resultado["valor"] = int(sel.split(" - ")[0])
            dialog.destroy()

        def cancelar():
            dialog.destroy()

        tk.Button(frame_btns, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=cancelar).pack(side="left", padx=5)

        self.wait_window(dialog)
        return resultado["valor"]

    # ===================== SELECCIONAR CDP =====================
    def _dialogo_editar_cdp(self, numero, parent=None, callback_refresh=None):
        """Dialogo completo para editar un CDP: valor, objeto y fuente SIFSE."""
        cdp = db.get_cdp(numero)
        if not cdp:
            messagebox.showerror("Error", f"CDP {numero} no encontrado")
            return
        if cdp["estado"] == "ANULADO":
            messagebox.showinfo("Info", f"CDP {numero} esta anulado, no se puede editar")
            return

        dlg = tk.Toplevel(parent or self)
        dlg.title(f"Editar CDP No. {numero}")
        dlg.geometry("520x400")
        dlg.transient(parent or self)
        dlg.grab_set()
        dlg.configure(bg=COLOR_WHITE)

        tk.Label(dlg, text=f"EDITAR CDP No. {numero}", font=("Segoe UI", 13, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)

        tk.Label(dlg, text=f"Rubro: {cdp['codigo_rubro']} - {cdp['cuenta']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w")
        tk.Label(dlg, text=f"Fecha: {cdp['fecha']}  |  Estado: {cdp['estado']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w", pady=(0, 8))

        frame_campos = tk.Frame(dlg, bg=COLOR_WHITE)
        frame_campos.pack(padx=15, pady=5, fill="x")

        # Valor
        tk.Label(frame_campos, text="Valor:", bg=COLOR_WHITE,
                 font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=6)
        entry_valor = tk.Entry(frame_campos, font=("Segoe UI", 10), width=20)
        entry_valor.grid(row=0, column=1, padx=5, pady=6, sticky="w")
        entry_valor.insert(0, f"{cdp['valor']:,.0f}")

        # Objeto del gasto
        tk.Label(frame_campos, text="Objeto:", bg=COLOR_WHITE,
                 font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="ne", padx=5, pady=6)
        text_objeto = tk.Text(frame_campos, font=("Segoe UI", 10), width=40, height=3, wrap="word")
        text_objeto.grid(row=1, column=1, padx=5, pady=6, sticky="w")
        text_objeto.insert("1.0", cdp["objeto"])

        # Fuente SIFSE (ingresos)
        tk.Label(frame_campos, text="Fuente SIFSE\n(Ingresos):", bg=COLOR_WHITE,
                 font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="e", padx=5, pady=6)
        fuentes = db.get_fuentes_sifse_activas()
        opciones_fuente = ["0 - Sin asignar"] + [f"{f['codigo']} - {f['descripcion']}" for f in fuentes]
        fuente_var = tk.StringVar()
        combo_fuente = ttk.Combobox(frame_campos, textvariable=fuente_var,
                     values=opciones_fuente, state="readonly", width=38)
        combo_fuente.grid(row=2, column=1, padx=5, pady=6, sticky="w")
        fuente_actual = cdp.get("fuente_sifse", 0) or 0
        for i, op in enumerate(opciones_fuente):
            if op.startswith(f"{fuente_actual} -"):
                combo_fuente.current(i)
                break
        else:
            combo_fuente.current(0)

        # Item SIFSE (gastos)
        tk.Label(frame_campos, text="Item SIFSE\n(Gastos):", bg=COLOR_WHITE,
                 font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="e", padx=5, pady=6)
        items = db.get_catalogo_items_sifse()
        opciones_item = ["0 - Sin asignar"] + [f"{it['codigo']} - {it['descripcion']}" for it in items]
        item_var = tk.StringVar()
        combo_item = ttk.Combobox(frame_campos, textvariable=item_var,
                     values=opciones_item, state="readonly", width=38)
        combo_item.grid(row=3, column=1, padx=5, pady=6, sticky="w")
        # Pre-seleccionar: primero del CDP, si no del mapeo del rubro
        item_actual = cdp.get("item_sifse", 0) or 0
        if not item_actual:
            item_actual = db.get_mapeo_sifse_gasto(cdp["codigo_rubro"]) or 0
        for i, op in enumerate(opciones_item):
            if op.startswith(f"{item_actual} -"):
                combo_item.current(i)
                break
        else:
            combo_item.current(0)

        def guardar():
            # Parsear valor
            val_str = entry_valor.get().replace(",", "").replace(".", "").strip()
            try:
                nuevo_valor = float(val_str)
            except ValueError:
                messagebox.showerror("Error", "Valor numerico invalido", parent=dlg)
                return

            nuevo_objeto = text_objeto.get("1.0", "end").strip()
            if not nuevo_objeto:
                messagebox.showwarning("Datos", "El objeto del gasto no puede estar vacio", parent=dlg)
                return

            # Parsear fuente e item
            sel_fuente = fuente_var.get()
            nueva_fuente = int(sel_fuente.split(" - ")[0]) if sel_fuente else 0
            sel_item = item_var.get()
            nuevo_item = int(sel_item.split(" - ")[0]) if sel_item else 0

            try:
                db.editar_cdp(numero, nuevo_valor=nuevo_valor, objeto=nuevo_objeto,
                              fuente_sifse=nueva_fuente, item_sifse=nuevo_item)
                messagebox.showinfo("Editado",
                    f"CDP {numero} actualizado exitosamente", parent=dlg)
                dlg.destroy()
                if callback_refresh:
                    callback_refresh()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        tk.Button(dlg, text="Guardar Cambios", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 11, "bold"), width=18, cursor="hand2",
                  command=guardar).pack(pady=15)

    def _dialogo_editar_rp(self, numero, parent=None, callback_refresh=None):
        """Dialogo completo para editar un RP: valor, objeto, fuente y item SIFSE."""
        rp = db.get_rp(numero)
        if not rp:
            messagebox.showerror("Error", f"RP {numero} no encontrado")
            return
        if rp["estado"] == "ANULADO":
            messagebox.showinfo("Info", f"RP {numero} esta anulado, no se puede editar")
            return

        dlg = tk.Toplevel(parent or self)
        dlg.title(f"Editar RP No. {numero}")
        dlg.geometry("520x420")
        dlg.transient(parent or self)
        dlg.grab_set()
        dlg.configure(bg=COLOR_WHITE)

        tk.Label(dlg, text=f"EDITAR RP No. {numero}", font=("Segoe UI", 13, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)
        tk.Label(dlg, text=f"Rubro: {rp['codigo_rubro']} - {rp['cuenta']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w")
        tk.Label(dlg, text=f"Tercero: {rp.get('nombre_tercero','')}  |  CDP: {rp['cdp_numero']}  |  {rp['estado']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w", pady=(0, 8))

        fc = tk.Frame(dlg, bg=COLOR_WHITE)
        fc.pack(padx=15, pady=5, fill="x")

        tk.Label(fc, text="Valor:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        entry_valor = tk.Entry(fc, font=("Segoe UI", 10), width=20)
        entry_valor.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        entry_valor.insert(0, f"{rp['valor']:,.0f}")

        tk.Label(fc, text="Objeto:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="ne", padx=5, pady=5)
        text_objeto = tk.Text(fc, font=("Segoe UI", 10), width=40, height=3, wrap="word")
        text_objeto.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        text_objeto.insert("1.0", rp.get("objeto", ""))

        tk.Label(fc, text="Fuente SIFSE\n(Ingresos):", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        fuentes = db.get_fuentes_sifse_activas()
        opciones_fuente = ["0 - Sin asignar"] + [f"{f['codigo']} - {f['descripcion']}" for f in fuentes]
        fuente_var = tk.StringVar()
        combo_fuente = ttk.Combobox(fc, textvariable=fuente_var, values=opciones_fuente, state="readonly", width=38)
        combo_fuente.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        fuente_actual = rp.get("fuente_sifse", 0) or 0
        for i, op in enumerate(opciones_fuente):
            if op.startswith(f"{fuente_actual} -"):
                combo_fuente.current(i)
                break
        else:
            combo_fuente.current(0)

        tk.Label(fc, text="Item SIFSE\n(Gastos):", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        items = db.get_catalogo_items_sifse()
        opciones_item = ["0 - Sin asignar"] + [f"{it['codigo']} - {it['descripcion']}" for it in items]
        item_var = tk.StringVar()
        combo_item = ttk.Combobox(fc, textvariable=item_var, values=opciones_item, state="readonly", width=38)
        combo_item.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        item_actual = rp.get("item_sifse", 0) or 0
        if not item_actual:
            item_actual = db.get_mapeo_sifse_gasto(rp["codigo_rubro"]) or 0
        for i, op in enumerate(opciones_item):
            if op.startswith(f"{item_actual} -"):
                combo_item.current(i)
                break
        else:
            combo_item.current(0)

        def guardar():
            val_str = entry_valor.get().replace(",", "").replace(".", "").strip()
            try:
                nuevo_valor = float(val_str)
            except ValueError:
                messagebox.showerror("Error", "Valor numerico invalido", parent=dlg)
                return
            nuevo_objeto = text_objeto.get("1.0", "end").strip()
            if not nuevo_objeto:
                messagebox.showwarning("Datos", "El objeto no puede estar vacio", parent=dlg)
                return
            nueva_fuente = int(fuente_var.get().split(" - ")[0]) if fuente_var.get() else 0
            nuevo_item = int(item_var.get().split(" - ")[0]) if item_var.get() else 0
            try:
                db.editar_rp(numero, nuevo_valor=nuevo_valor, objeto=nuevo_objeto,
                             fuente_sifse=nueva_fuente, item_sifse=nuevo_item)
                messagebox.showinfo("Editado", f"RP {numero} actualizado exitosamente", parent=dlg)
                dlg.destroy()
                if callback_refresh:
                    callback_refresh()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        tk.Button(dlg, text="Guardar Cambios", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 11, "bold"), width=18, cursor="hand2", command=guardar).pack(pady=15)

    def _dialogo_editar_obligacion(self, numero, parent=None, callback_refresh=None):
        """Dialogo completo para editar una Obligacion: valor, factura, fuente y item SIFSE."""
        obl = db.get_obligacion(numero)
        if not obl:
            messagebox.showerror("Error", f"Obligacion {numero} no encontrada")
            return
        if obl["estado"] == "ANULADA":
            messagebox.showinfo("Info", f"Obligacion {numero} esta anulada, no se puede editar")
            return

        dlg = tk.Toplevel(parent or self)
        dlg.title(f"Editar Obligacion No. {numero}")
        dlg.geometry("520x400")
        dlg.transient(parent or self)
        dlg.grab_set()
        dlg.configure(bg=COLOR_WHITE)

        tk.Label(dlg, text=f"EDITAR OBLIGACION No. {numero}", font=("Segoe UI", 13, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)
        tk.Label(dlg, text=f"Rubro: {obl['codigo_rubro']} - {obl['cuenta']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w")
        tk.Label(dlg, text=f"Tercero: {obl.get('nombre_tercero','')}  |  RP: {obl['rp_numero']}  |  {obl['estado']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w", pady=(0, 8))

        fc = tk.Frame(dlg, bg=COLOR_WHITE)
        fc.pack(padx=15, pady=5, fill="x")

        tk.Label(fc, text="Valor:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        entry_valor = tk.Entry(fc, font=("Segoe UI", 10), width=20)
        entry_valor.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        entry_valor.insert(0, f"{obl['valor']:,.0f}")

        tk.Label(fc, text="Factura:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        entry_factura = tk.Entry(fc, font=("Segoe UI", 10), width=32)
        entry_factura.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        entry_factura.insert(0, obl.get("factura", ""))

        tk.Label(fc, text="Fuente SIFSE\n(Ingresos):", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        fuentes = db.get_fuentes_sifse_activas()
        opciones_fuente = ["0 - Sin asignar"] + [f"{f['codigo']} - {f['descripcion']}" for f in fuentes]
        fuente_var = tk.StringVar()
        combo_fuente = ttk.Combobox(fc, textvariable=fuente_var, values=opciones_fuente, state="readonly", width=38)
        combo_fuente.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        fuente_actual = obl.get("fuente_sifse", 0) or 0
        for i, op in enumerate(opciones_fuente):
            if op.startswith(f"{fuente_actual} -"):
                combo_fuente.current(i)
                break
        else:
            combo_fuente.current(0)

        tk.Label(fc, text="Item SIFSE\n(Gastos):", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        items = db.get_catalogo_items_sifse()
        opciones_item = ["0 - Sin asignar"] + [f"{it['codigo']} - {it['descripcion']}" for it in items]
        item_var = tk.StringVar()
        combo_item = ttk.Combobox(fc, textvariable=item_var, values=opciones_item, state="readonly", width=38)
        combo_item.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        item_actual = obl.get("item_sifse", 0) or 0
        if not item_actual:
            item_actual = db.get_mapeo_sifse_gasto(obl["codigo_rubro"]) or 0
        for i, op in enumerate(opciones_item):
            if op.startswith(f"{item_actual} -"):
                combo_item.current(i)
                break
        else:
            combo_item.current(0)

        def guardar():
            val_str = entry_valor.get().replace(",", "").replace(".", "").strip()
            try:
                nuevo_valor = float(val_str)
            except ValueError:
                messagebox.showerror("Error", "Valor numerico invalido", parent=dlg)
                return
            nueva_factura = entry_factura.get().strip()
            nueva_fuente = int(fuente_var.get().split(" - ")[0]) if fuente_var.get() else 0
            nuevo_item = int(item_var.get().split(" - ")[0]) if item_var.get() else 0
            try:
                db.editar_obligacion(numero, nuevo_valor=nuevo_valor, factura=nueva_factura,
                                     fuente_sifse=nueva_fuente, item_sifse=nuevo_item)
                messagebox.showinfo("Editado", f"Obligacion {numero} actualizada exitosamente", parent=dlg)
                dlg.destroy()
                if callback_refresh:
                    callback_refresh()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        tk.Button(dlg, text="Guardar Cambios", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 11, "bold"), width=18, cursor="hand2", command=guardar).pack(pady=15)

    def _dialogo_editar_pago(self, numero, parent=None, callback_refresh=None):
        """Dialogo completo para editar un Pago: valor, concepto, medio, comprobante, cuenta, SIFSE."""
        pago = db.get_pago(numero)
        if not pago:
            messagebox.showerror("Error", f"Pago {numero} no encontrado")
            return
        if pago["estado"] == "ANULADO":
            messagebox.showinfo("Info", f"Pago {numero} esta anulado, no se puede editar")
            return

        dlg = tk.Toplevel(parent or self)
        dlg.title(f"Editar Pago No. {numero}")
        dlg.geometry("530x480")
        dlg.transient(parent or self)
        dlg.grab_set()
        dlg.configure(bg=COLOR_WHITE)

        tk.Label(dlg, text=f"EDITAR PAGO No. {numero}", font=("Segoe UI", 13, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)
        tk.Label(dlg, text=f"Rubro: {pago['codigo_rubro']} - {pago['cuenta']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w")
        tk.Label(dlg, text=f"Tercero: {pago.get('nombre_tercero','')}  |  Oblig: {pago['obligacion_numero']}  |  {pago['estado']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w", pady=(0, 8))

        fc = tk.Frame(dlg, bg=COLOR_WHITE)
        fc.pack(padx=15, pady=5, fill="x")

        row = 0
        tk.Label(fc, text="Valor:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        entry_valor = tk.Entry(fc, font=("Segoe UI", 10), width=20)
        entry_valor.grid(row=row, column=1, padx=5, pady=4, sticky="w")
        entry_valor.insert(0, f"{pago['valor']:,.0f}")

        row += 1
        tk.Label(fc, text="Concepto:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        entry_concepto = tk.Entry(fc, font=("Segoe UI", 10), width=38)
        entry_concepto.grid(row=row, column=1, padx=5, pady=4, sticky="w")
        entry_concepto.insert(0, pago.get("concepto", ""))

        row += 1
        tk.Label(fc, text="Medio Pago:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        medio_var = tk.StringVar(value=pago.get("medio_pago", "Transferencia"))
        ttk.Combobox(fc, textvariable=medio_var, values=["Transferencia", "Cheque"], state="readonly", width=36).grid(row=row, column=1, padx=5, pady=4, sticky="w")

        row += 1
        tk.Label(fc, text="No. Comprobante:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        entry_comprob = tk.Entry(fc, font=("Segoe UI", 10), width=20)
        entry_comprob.grid(row=row, column=1, padx=5, pady=4, sticky="w")
        entry_comprob.insert(0, pago.get("no_comprobante", ""))

        row += 1
        tk.Label(fc, text="Cuenta Bancaria:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        cuentas = db.listar_cuentas_bancarias()
        cuenta_opciones = {"0 - Sin asignar": 0}
        cuenta_opciones.update({f"{c['banco']} - {c['tipo_cuenta']} - {c['numero_cuenta']}": c["id"] for c in cuentas})
        cuenta_var = tk.StringVar()
        combo_cuenta = ttk.Combobox(fc, textvariable=cuenta_var, values=list(cuenta_opciones.keys()), state="readonly", width=36)
        combo_cuenta.grid(row=row, column=1, padx=5, pady=4, sticky="w")
        cuenta_actual = pago.get("cuenta_bancaria_id", 0) or 0
        sel_idx = 0
        for i, (k, v) in enumerate(cuenta_opciones.items()):
            if v == cuenta_actual:
                sel_idx = i
                break
        combo_cuenta.current(sel_idx)

        row += 1
        tk.Label(fc, text="Fuente SIFSE\n(Ingresos):", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        fuentes = db.get_fuentes_sifse_activas()
        opciones_fuente = ["0 - Sin asignar"] + [f"{f['codigo']} - {f['descripcion']}" for f in fuentes]
        fuente_var = tk.StringVar()
        combo_fuente = ttk.Combobox(fc, textvariable=fuente_var, values=opciones_fuente, state="readonly", width=36)
        combo_fuente.grid(row=row, column=1, padx=5, pady=4, sticky="w")
        fuente_actual = pago.get("fuente_sifse", 0) or 0
        for i, op in enumerate(opciones_fuente):
            if op.startswith(f"{fuente_actual} -"):
                combo_fuente.current(i)
                break
        else:
            combo_fuente.current(0)

        row += 1
        tk.Label(fc, text="Item SIFSE\n(Gastos):", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="e", padx=5, pady=4)
        items_sifse = db.get_catalogo_items_sifse()
        opciones_item = ["0 - Sin asignar"] + [f"{it['codigo']} - {it['descripcion']}" for it in items_sifse]
        item_var = tk.StringVar()
        combo_item = ttk.Combobox(fc, textvariable=item_var, values=opciones_item, state="readonly", width=36)
        combo_item.grid(row=row, column=1, padx=5, pady=4, sticky="w")
        item_actual = pago.get("item_sifse", 0) or 0
        if not item_actual:
            item_actual = db.get_mapeo_sifse_gasto(pago["codigo_rubro"]) or 0
        for i, op in enumerate(opciones_item):
            if op.startswith(f"{item_actual} -"):
                combo_item.current(i)
                break
        else:
            combo_item.current(0)

        def guardar():
            val_str = entry_valor.get().replace(",", "").replace(".", "").strip()
            try:
                nuevo_valor = float(val_str)
            except ValueError:
                messagebox.showerror("Error", "Valor numerico invalido", parent=dlg)
                return
            nueva_fuente = int(fuente_var.get().split(" - ")[0]) if fuente_var.get() else 0
            nuevo_item = int(item_var.get().split(" - ")[0]) if item_var.get() else 0
            nueva_cuenta = cuenta_opciones.get(cuenta_var.get(), 0)
            try:
                db.editar_pago(numero, nuevo_valor=nuevo_valor,
                               concepto=entry_concepto.get().strip(),
                               medio_pago=medio_var.get(),
                               no_comprobante=entry_comprob.get().strip(),
                               cuenta_bancaria_id=nueva_cuenta,
                               fuente_sifse=nueva_fuente, item_sifse=nuevo_item)
                messagebox.showinfo("Editado", f"Pago {numero} actualizado exitosamente", parent=dlg)
                dlg.destroy()
                if callback_refresh:
                    callback_refresh()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        tk.Button(dlg, text="Guardar Cambios", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 11, "bold"), width=18, cursor="hand2", command=guardar).pack(pady=12)

    def _dialogo_editar_recaudo(self, numero, parent=None, callback_refresh=None):
        """Dialogo completo para editar un Recaudo: valor, concepto, comprobante, cuenta bancaria."""
        rec = db.get_recaudo(numero)
        if not rec:
            messagebox.showerror("Error", f"Recaudo {numero} no encontrado")
            return
        if rec["estado"] == "ANULADO":
            messagebox.showinfo("Info", f"Recaudo {numero} esta anulado, no se puede editar")
            return

        dlg = tk.Toplevel(parent or self)
        dlg.title(f"Editar Recaudo No. {numero}")
        dlg.geometry("500x320")
        dlg.transient(parent or self)
        dlg.grab_set()
        dlg.configure(bg=COLOR_WHITE)

        tk.Label(dlg, text=f"EDITAR RECAUDO No. {numero}", font=("Segoe UI", 13, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)
        tk.Label(dlg, text=f"Rubro: {rec['codigo_rubro']} - {rec['cuenta']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w")
        tk.Label(dlg, text=f"Fecha: {rec['fecha']}  |  {rec['estado']}",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(padx=15, anchor="w", pady=(0, 8))

        fc = tk.Frame(dlg, bg=COLOR_WHITE)
        fc.pack(padx=15, pady=5, fill="x")

        tk.Label(fc, text="Valor:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        entry_valor = tk.Entry(fc, font=("Segoe UI", 10), width=20)
        entry_valor.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        entry_valor.insert(0, f"{rec['valor']:,.0f}")

        tk.Label(fc, text="Concepto:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        entry_concepto = tk.Entry(fc, font=("Segoe UI", 10), width=38)
        entry_concepto.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        entry_concepto.insert(0, rec.get("concepto", ""))

        tk.Label(fc, text="No. Comprobante:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        entry_comprob = tk.Entry(fc, font=("Segoe UI", 10), width=20)
        entry_comprob.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        entry_comprob.insert(0, rec.get("no_comprobante", ""))

        tk.Label(fc, text="Cuenta Bancaria:", bg=COLOR_WHITE, font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="e", padx=5, pady=5)
        cuentas = db.listar_cuentas_bancarias()
        cuenta_opciones = {"0 - Sin asignar": 0}
        cuenta_opciones.update({f"{c['banco']} - {c['tipo_cuenta']} - {c['numero_cuenta']}": c["id"] for c in cuentas})
        cuenta_var = tk.StringVar()
        combo_cuenta = ttk.Combobox(fc, textvariable=cuenta_var, values=list(cuenta_opciones.keys()), state="readonly", width=36)
        combo_cuenta.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        cuenta_actual = rec.get("cuenta_bancaria_id", 0) or 0
        sel_idx = 0
        for i, (k, v) in enumerate(cuenta_opciones.items()):
            if v == cuenta_actual:
                sel_idx = i
                break
        combo_cuenta.current(sel_idx)

        def guardar():
            val_str = entry_valor.get().replace(",", "").replace(".", "").strip()
            try:
                nuevo_valor = float(val_str)
            except ValueError:
                messagebox.showerror("Error", "Valor numerico invalido", parent=dlg)
                return
            nueva_cuenta = cuenta_opciones.get(cuenta_var.get(), 0)
            try:
                db.editar_recaudo(numero, nuevo_valor=nuevo_valor,
                                  concepto=entry_concepto.get().strip(),
                                  no_comprobante=entry_comprob.get().strip(),
                                  cuenta_bancaria_id=nueva_cuenta)
                messagebox.showinfo("Editado", f"Recaudo {numero} actualizado exitosamente", parent=dlg)
                dlg.destroy()
                if callback_refresh:
                    callback_refresh()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        tk.Button(dlg, text="Guardar Cambios", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 11, "bold"), width=18, cursor="hand2", command=guardar).pack(pady=15)

    def _seleccionar_cdp(self):
        dialog = tk.Toplevel(self)
        dialog.title("Seleccionar CDP")
        dialog.geometry("700x450")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"numero": None}

        tk.Label(dialog, text="SELECCIONAR CDP DISPONIBLE", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        # Lista de CDPs
        frame_lista = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("No", "Fecha", "Rubro", "Objeto", "Fuente", "Valor", "Saldo CDP", "Estado")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings", height=12)
        tree.heading("No", text="No. CDP")
        tree.heading("Fecha", text="Fecha")
        tree.heading("Rubro", text="Rubro")
        tree.heading("Objeto", text="Objeto del Gasto")
        tree.heading("Fuente", text="Fuente SIFSE")
        tree.heading("Valor", text="Valor")
        tree.heading("Saldo CDP", text="Saldo Disp.")
        tree.heading("Estado", text="Estado")
        tree.column("No", width=55, anchor="center")
        tree.column("Fecha", width=80, anchor="center")
        tree.column("Rubro", width=120)
        tree.column("Objeto", width=150)
        tree.column("Fuente", width=70, anchor="center")
        tree.column("Valor", width=90, anchor="e")
        tree.column("Saldo CDP", width=90, anchor="e")
        tree.column("Estado", width=65, anchor="center")

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def cargar_cdps():
            tree.delete(*tree.get_children())
            cdps = db.get_cdps()
            count = 0
            for d in cdps:
                if d["estado"] == "ANULADO":
                    continue
                saldo = db.saldo_cdp(d["numero"])
                tag = "agotado" if saldo <= 0 else "activo"
                fuente = d.get("fuente_sifse", 0) or 0
                tree.insert("", "end", values=(
                    d["numero"], d["fecha"], d["codigo_rubro"],
                    d["objeto"][:35], fuente if fuente else "-",
                    f"{d['valor']:,.0f}", f"{saldo:,.0f}", d["estado"]
                ), tags=(tag,))
                count += 1
            tree.tag_configure("agotado", foreground="#999999")
            tree.tag_configure("activo", foreground="#000000")
            lbl_info.config(text=f"CDPs disponibles: {count}")
            lbl_detalle.config(text="Seleccione un CDP de la lista", fg=COLOR_SUCCESS)

        lbl_info = tk.Label(dialog, text="", font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#666")
        lbl_info.pack(anchor="w", padx=15)

        lbl_detalle = tk.Label(dialog, text="Seleccione un CDP de la lista",
                               font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_SUCCESS)
        lbl_detalle.pack(anchor="w", padx=15, pady=3)

        def on_select(event):
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0])["values"]
                saldo_val = str(vals[5]).replace(",", "")
                color = COLOR_SUCCESS if float(saldo_val) > 0 else COLOR_DANGER
                lbl_detalle.config(
                    text=f"CDP {vals[0]} | {vals[2]} | Saldo: $ {vals[5]} | {vals[6]}",
                    fg=color)

        def anular_cdp():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un CDP", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            if not messagebox.askyesno("Anular CDP",
                    f"Desea ANULAR el CDP No. {num}?\nEl valor se liberara al rubro.", parent=dialog):
                return
            try:
                db.anular_cdp(num)
                messagebox.showinfo("Anulado", f"CDP {num} anulado", parent=dialog)
                cargar_cdps()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dialog)

        def editar_cdp_completo():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un CDP", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            self._dialogo_editar_cdp(num, dialog, cargar_cdps)

        def aceptar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un CDP", parent=dialog)
                return
            vals = tree.item(sel[0])["values"]
            saldo_val = float(str(vals[5]).replace(",", ""))
            if saldo_val <= 0:
                messagebox.showerror("Sin saldo", "Este CDP no tiene saldo disponible", parent=dialog)
                return
            resultado["numero"] = int(vals[0])
            dialog.destroy()

        tree.bind("<<TreeviewSelect>>", on_select)
        tree.bind("<Double-1>", lambda e: aceptar())

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Anular CDP", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=12, command=anular_cdp).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Editar CDP", bg=COLOR_WARNING, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=12, command=editar_cdp_completo).pack(side="left", padx=5)

        cargar_cdps()
        self.wait_window(dialog)
        return resultado["numero"]

    # ===================== SELECCIONAR RP =====================
    def _seleccionar_rp(self):
        dialog = tk.Toplevel(self)
        dialog.title("Seleccionar RP")
        dialog.geometry("750x450")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"numero": None}

        tk.Label(dialog, text="SELECCIONAR REGISTRO PRESUPUESTAL", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        frame_lista = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("No", "Fecha", "CDP", "Rubro", "Tercero", "Valor", "Saldo RP", "Estado")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings", height=12)
        tree.heading("No", text="No. RP")
        tree.heading("Fecha", text="Fecha")
        tree.heading("CDP", text="CDP")
        tree.heading("Rubro", text="Rubro")
        tree.heading("Tercero", text="Tercero")
        tree.heading("Valor", text="Valor")
        tree.heading("Saldo RP", text="Saldo Disp.")
        tree.heading("Estado", text="Estado")
        tree.column("No", width=50, anchor="center")
        tree.column("Fecha", width=80, anchor="center")
        tree.column("CDP", width=45, anchor="center")
        tree.column("Rubro", width=120)
        tree.column("Tercero", width=160)
        tree.column("Valor", width=85, anchor="e")
        tree.column("Saldo RP", width=85, anchor="e")
        tree.column("Estado", width=65, anchor="center")

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def cargar_rps():
            tree.delete(*tree.get_children())
            rps = db.get_rps()
            count = 0
            for d in rps:
                if d["estado"] == "ANULADO":
                    continue
                saldo = db.saldo_rp(d["numero"])
                tag = "agotado" if saldo <= 0 else "activo"
                tree.insert("", "end", values=(
                    d["numero"], d["fecha"], d["cdp_numero"], d["codigo_rubro"],
                    (d["nombre_tercero"] or "")[:30], f"{d['valor']:,.0f}",
                    f"{saldo:,.0f}", d["estado"]
                ), tags=(tag,))
                count += 1
            tree.tag_configure("agotado", foreground="#999999")
            lbl_detalle.config(text=f"RPs disponibles: {count}", fg=COLOR_SUCCESS)

        lbl_detalle = tk.Label(dialog, text="",
                               font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_SUCCESS)
        lbl_detalle.pack(anchor="w", padx=15, pady=3)

        def on_select(event):
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0])["values"]
                saldo_val = str(vals[6]).replace(",", "")
                color = COLOR_SUCCESS if float(saldo_val) > 0 else COLOR_DANGER
                lbl_detalle.config(
                    text=f"RP {vals[0]} | Tercero: {vals[4]} | Saldo: $ {vals[6]}",
                    fg=color)

        def anular_rp():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un RP", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            if not messagebox.askyesno("Anular RP",
                    f"Desea ANULAR el RP No. {num}?\nEl valor se devolvera al CDP.", parent=dialog):
                return
            try:
                db.anular_rp(num)
                messagebox.showinfo("Anulado", f"RP {num} anulado", parent=dialog)
                cargar_rps()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dialog)

        def editar_rp_completo():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un RP", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            self._dialogo_editar_rp(num, dialog, cargar_rps)

        def aceptar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un RP", parent=dialog)
                return
            vals = tree.item(sel[0])["values"]
            saldo_val = float(str(vals[6]).replace(",", ""))
            if saldo_val <= 0:
                messagebox.showerror("Sin saldo", "Este RP no tiene saldo disponible", parent=dialog)
                return
            resultado["numero"] = int(vals[0])
            dialog.destroy()

        tree.bind("<<TreeviewSelect>>", on_select)
        tree.bind("<Double-1>", lambda e: aceptar())

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Anular RP", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=12, command=anular_rp).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Editar RP", bg=COLOR_WARNING, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=12, command=editar_rp_completo).pack(side="left", padx=5)

        cargar_rps()
        self.wait_window(dialog)
        return resultado["numero"]

    # ===================== SELECCIONAR OBLIGACION =====================
    def _seleccionar_obligacion(self):
        dialog = tk.Toplevel(self)
        dialog.title("Seleccionar Obligacion")
        dialog.geometry("750x450")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"numero": None}

        tk.Label(dialog, text="SELECCIONAR OBLIGACION", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        frame_lista = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("No", "Fecha", "RP", "Rubro", "Tercero", "Valor", "Saldo", "Estado")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings", height=12)
        tree.heading("No", text="No. Oblig")
        tree.heading("Fecha", text="Fecha")
        tree.heading("RP", text="RP")
        tree.heading("Rubro", text="Rubro")
        tree.heading("Tercero", text="Tercero")
        tree.heading("Valor", text="Valor")
        tree.heading("Saldo", text="Saldo Disp.")
        tree.heading("Estado", text="Estado")
        tree.column("No", width=55, anchor="center")
        tree.column("Fecha", width=80, anchor="center")
        tree.column("RP", width=45, anchor="center")
        tree.column("Rubro", width=120)
        tree.column("Tercero", width=160)
        tree.column("Valor", width=85, anchor="e")
        tree.column("Saldo", width=85, anchor="e")
        tree.column("Estado", width=65, anchor="center")

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def cargar_obls():
            tree.delete(*tree.get_children())
            obls = db.get_obligaciones()
            count = 0
            for d in obls:
                if d["estado"] == "ANULADA":
                    continue
                saldo = db.saldo_obligacion(d["numero"])
                tag = "pagada" if saldo <= 0 else "activo"
                tree.insert("", "end", values=(
                    d["numero"], d["fecha"], d["rp_numero"], d["codigo_rubro"],
                    (d["nombre_tercero"] or "")[:30], f"{d['valor']:,.0f}",
                    f"{saldo:,.0f}", d["estado"]
                ), tags=(tag,))
                count += 1
            tree.tag_configure("pagada", foreground="#999999")
            lbl_detalle.config(text=f"Obligaciones disponibles: {count}", fg=COLOR_SUCCESS)

        lbl_detalle = tk.Label(dialog, text="",
                               font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_SUCCESS)
        lbl_detalle.pack(anchor="w", padx=15, pady=3)

        def on_select(event):
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0])["values"]
                saldo_val = str(vals[6]).replace(",", "")
                color = COLOR_SUCCESS if float(saldo_val) > 0 else COLOR_DANGER
                lbl_detalle.config(
                    text=f"Oblig. {vals[0]} | Tercero: {vals[4]} | Saldo: $ {vals[6]}",
                    fg=color)

        def anular_obl():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione una obligacion", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            if not messagebox.askyesno("Anular Obligacion",
                    f"Desea ANULAR la Obligacion No. {num}?\nEl valor se devolvera al RP.", parent=dialog):
                return
            try:
                db.anular_obligacion(num)
                messagebox.showinfo("Anulada", f"Obligacion {num} anulada", parent=dialog)
                cargar_obls()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dialog)

        def editar_obl_completo():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione una obligacion", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            self._dialogo_editar_obligacion(num, dialog, cargar_obls)

        def aceptar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione una obligacion", parent=dialog)
                return
            vals = tree.item(sel[0])["values"]
            saldo_val = float(str(vals[6]).replace(",", ""))
            if saldo_val <= 0:
                messagebox.showerror("Sin saldo", "Esta obligacion ya esta pagada", parent=dialog)
                return
            resultado["numero"] = int(vals[0])
            dialog.destroy()

        tree.bind("<<TreeviewSelect>>", on_select)
        tree.bind("<Double-1>", lambda e: aceptar())

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Anular Oblig.", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=12, command=anular_obl).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Editar Oblig.", bg=COLOR_WARNING, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=12, command=editar_obl_completo).pack(side="left", padx=5)

        cargar_obls()
        self.wait_window(dialog)
        return resultado["numero"]

    # ===================== REGISTRAR RP =====================
    def _registrar_rp(self):
        cdp_num = self._seleccionar_cdp()
        if not cdp_num:
            return

        cdp = db.get_cdp(cdp_num)
        saldo = db.saldo_cdp(cdp_num)

        nit = simpledialog.askstring(
            "Registrar RP - Tercero",
            f"CDP: {cdp_num} - {cdp['cuenta']}\n"
            f"Objeto: {cdp['objeto']}\n"
            f"Saldo CDP: $ {saldo:,.0f}\n\n"
            f"Ingrese NIT/CC del tercero:",
            parent=self
        )
        if not nit:
            return

        tercero = db.get_tercero(nit.strip())
        if not tercero:
            if messagebox.askyesno("Tercero", f"Tercero {nit} no existe. Desea registrarlo?"):
                nombre = simpledialog.askstring("Tercero", "Nombre o Razon Social:", parent=self)
                if nombre:
                    db.guardar_tercero(nit.strip(), "", nombre)
                else:
                    return
            else:
                return
            tercero = db.get_tercero(nit.strip())

        valor_str = simpledialog.askstring(
            "Registrar RP - Valor",
            f"CDP: {cdp_num} - {cdp['cuenta']}\n"
            f"Tercero: {tercero['nombre']}\n"
            f"Saldo CDP: $ {saldo:,.0f}\n\nIngrese el VALOR del RP:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor invalido")
            return

        if valor <= 0 or valor > saldo:
            messagebox.showerror("Error", "Valor invalido o excede saldo")
            return

        objeto = simpledialog.askstring("Registrar RP", "Objeto/Contrato:", parent=self)
        if not objeto:
            return

        if not messagebox.askyesno("Confirmar RP",
                                    f"CDP: {cdp_num}\nTercero: {tercero['nombre']}\n"
                                    f"Valor: $ {valor:,.0f}\nObjeto: {objeto}"):
            return

        try:
            num, fecha = db.registrar_rp(cdp_num, nit.strip(), valor, objeto)
            messagebox.showinfo("RP Registrado",
                                f"RP REGISTRADO\n\nNo. RP: {num}\nValor: $ {valor:,.0f}")
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ===================== REGISTRAR OBLIGACION =====================
    def _registrar_obligacion(self):
        rp_num = self._seleccionar_rp()
        if not rp_num:
            return

        rp = db.get_rp(rp_num)
        saldo = db.saldo_rp(rp_num)

        valor_str = simpledialog.askstring(
            "Obligacion - Valor",
            f"RP: {rp_num} - {rp['nombre_tercero']}\n"
            f"Rubro: {rp['codigo_rubro']} - {rp['cuenta']}\n"
            f"Saldo RP: $ {saldo:,.0f}\n\nIngrese el VALOR:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor invalido")
            return

        if valor <= 0 or valor > saldo:
            messagebox.showerror("Error", "Valor invalido o excede saldo")
            return

        factura = simpledialog.askstring("Obligacion", "Factura/Concepto:", parent=self)
        if not factura:
            return

        if not messagebox.askyesno("Confirmar",
                                    f"RP: {rp_num}\nTercero: {rp['nombre_tercero']}\n"
                                    f"Valor: $ {valor:,.0f}\nFactura: {factura}"):
            return

        try:
            num, fecha = db.registrar_obligacion(rp_num, valor, factura)
            messagebox.showinfo("Obligacion", f"OBLIGACION REGISTRADA\n\nNo.: {num}\nValor: $ {valor:,.0f}")
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ===================== REGISTRAR PAGO =====================
    def _registrar_pago(self):
        obl_num = self._seleccionar_obligacion()
        if not obl_num:
            return

        obl = db.get_obligacion(obl_num)
        saldo = db.saldo_obligacion(obl_num)

        valor_str = simpledialog.askstring(
            "Pago - Valor",
            f"Obligacion: {obl_num} - {obl['nombre_tercero']}\n"
            f"Rubro: {obl['codigo_rubro']} - {obl['cuenta']}\n"
            f"Saldo: $ {saldo:,.0f}\n\nIngrese el VALOR del pago:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor invalido")
            return

        if valor <= 0 or valor > saldo:
            messagebox.showerror("Error", "Valor invalido o excede saldo")
            return

        concepto = simpledialog.askstring("Pago", "Concepto del pago:", parent=self)
        if not concepto:
            return

        # Dialogo para seleccionar cuenta bancaria y medio de pago
        dlg_pago = tk.Toplevel(self)
        dlg_pago.title("Medio de Pago y Cuenta Bancaria")
        dlg_pago.geometry("450x280")
        dlg_pago.transient(self)
        dlg_pago.grab_set()
        dlg_pago.configure(bg=COLOR_WHITE)

        tk.Label(dlg_pago, text="DATOS DEL PAGO", font=("Segoe UI", 12, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=8)

        frame_campos = tk.Frame(dlg_pago, bg=COLOR_WHITE)
        frame_campos.pack(padx=15, pady=5, fill="x")

        tk.Label(frame_campos, text="Medio de Pago:", bg=COLOR_WHITE,
                 font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        medio_var = tk.StringVar(value="Transferencia")
        ttk.Combobox(frame_campos, textvariable=medio_var,
                     values=["Transferencia", "Cheque"], state="readonly",
                     width=30).grid(row=0, column=1, padx=5, pady=5)

        tk.Label(frame_campos, text="Cuenta Bancaria:", bg=COLOR_WHITE,
                 font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        cuentas = db.listar_cuentas_bancarias()
        cuenta_opciones = {f"{c['banco']} - {c['tipo_cuenta']} - {c['numero_cuenta']}": c["id"] for c in cuentas}
        cuenta_var = tk.StringVar()
        combo_cuenta = ttk.Combobox(frame_campos, textvariable=cuenta_var,
                     values=list(cuenta_opciones.keys()), state="readonly",
                     width=30)
        combo_cuenta.grid(row=1, column=1, padx=5, pady=5)
        if cuenta_opciones:
            combo_cuenta.current(0)

        tk.Label(frame_campos, text="No. Comprobante:", bg=COLOR_WHITE,
                 font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=5, pady=5)
        entry_comprob = tk.Entry(frame_campos, font=("Segoe UI", 10), width=32)
        entry_comprob.grid(row=2, column=1, padx=5, pady=5)

        resultado = {"ok": False}

        def confirmar():
            if not entry_comprob.get().strip():
                messagebox.showwarning("Datos", "Ingrese No. de Comprobante", parent=dlg_pago)
                return
            if not cuenta_var.get() and cuentas:
                messagebox.showwarning("Datos", "Seleccione una cuenta bancaria", parent=dlg_pago)
                return
            resultado["ok"] = True
            dlg_pago.destroy()

        tk.Button(dlg_pago, text="Registrar Pago", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 11, "bold"), width=18, cursor="hand2",
                  command=confirmar).pack(pady=15)

        if not cuentas:
            tk.Label(dlg_pago, text="No hay cuentas bancarias registradas.\nRegistrelas en Configuracion > Cuentas Bancarias.",
                     font=("Segoe UI", 9), bg=COLOR_WHITE, fg=COLOR_DANGER).pack()

        self.wait_window(dlg_pago)
        if not resultado["ok"]:
            return

        medio_pago = medio_var.get()
        comprobante = entry_comprob.get().strip()
        cuenta_id = cuenta_opciones.get(cuenta_var.get(), 0)

        if not messagebox.askyesno("Confirmar Pago",
                                    f"Obligacion: {obl_num}\nTercero: {obl['nombre_tercero']}\n"
                                    f"Valor: $ {valor:,.0f}\nMedio: {medio_pago}\n"
                                    f"Cuenta: {cuenta_var.get() or 'Sin asignar'}"):
            return

        try:
            num, fecha = db.registrar_pago(obl_num, valor, concepto, medio_pago, comprobante, cuenta_id)
            messagebox.showinfo("Pago", f"PAGO REGISTRADO\n\nNo.: {num}\nValor: $ {valor:,.0f}")
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ===================== TERCEROS =====================
    def _registrar_tercero(self):
        dialog = tk.Toplevel(self)
        dialog.title("Registrar Tercero")
        dialog.geometry("450x380")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        campos = [
            ("NIT/CC:", "nit"), ("DV:", "dv"), ("Nombre/Razon Social:", "nombre"),
            ("Direccion:", "dir"), ("Telefono:", "tel"), ("Email:", "email"),
        ]
        entries = {}
        for i, (label, key) in enumerate(campos):
            tk.Label(dialog, text=label, bg=COLOR_WHITE, font=("Segoe UI", 10),
                     anchor="e").grid(row=i, column=0, padx=10, pady=3, sticky="e")
            e = tk.Entry(dialog, font=("Segoe UI", 10), width=35)
            e.grid(row=i, column=1, padx=10, pady=3)
            entries[key] = e

        tk.Label(dialog, text="Tipo:", bg=COLOR_WHITE, font=("Segoe UI", 10),
                 anchor="e").grid(row=6, column=0, padx=10, pady=3, sticky="e")
        tipo_var = tk.StringVar(value="Natural")
        ttk.Combobox(dialog, textvariable=tipo_var, values=["Natural", "Juridica"],
                     state="readonly", width=33).grid(row=6, column=1, padx=10, pady=3)

        def guardar():
            nit = entries["nit"].get().strip()
            nombre = entries["nombre"].get().strip()
            if not nit or not nombre:
                messagebox.showwarning("Datos", "NIT y Nombre son obligatorios", parent=dialog)
                return
            db.guardar_tercero(nit, entries["dv"].get().strip(), nombre,
                               entries["dir"].get().strip(), entries["tel"].get().strip(),
                               entries["email"].get().strip(), tipo_var.get())
            messagebox.showinfo("Tercero", f"Tercero {nombre} registrado", parent=dialog)
            dialog.destroy()

        tk.Button(dialog, text="Guardar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=15, command=guardar).grid(
            row=8, column=1, pady=15, sticky="e", padx=10)

    # ===================== MENU LISTADOS =====================
    def _menu_listados(self):
        dialog = tk.Toplevel(self)
        dialog.title("Ver Listados")
        dialog.geometry("350x300")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="SELECCIONE UN LISTADO", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=15)

        opciones = [
            ("CDPs", "CDP", COLOR_SUCCESS),
            ("Registros Presupuestales (RP)", "RP", COLOR_PRIMARY),
            ("Obligaciones", "OBLIGACION", COLOR_WARNING),
            ("Pagos", "PAGO", COLOR_PURPLE),
            ("Recaudos", "RECAUDO", "#008B45"),
            ("Terceros", "TERCEROS", COLOR_ACCENT),
        ]
        for texto, tipo, color in opciones:
            tk.Button(dialog, text=texto, bg=color, fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=35, height=1,
                      cursor="hand2", command=lambda t=tipo, d=dialog: (d.destroy(), self._ver_listado(t))
                      ).pack(pady=3, padx=20)

    # ===================== VER LISTADOS =====================
    def _ver_listado(self, tipo):
        self._limpiar_main()
        f = self.main_frame

        tk.Button(f, text="< Volver al Menu", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9), cursor="hand2").pack(anchor="w")

        tk.Label(f, text=f"LISTADO DE {tipo}S", font=("Segoe UI", 16, "bold"),
                 bg=COLOR_BG, fg=COLOR_PRIMARY).pack(pady=5)

        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=5)

        if tipo == "CDP":
            cols = ("No", "Fecha", "Rubro", "Cuenta", "Objeto", "Valor", "Estado")
            data = db.get_cdps()
            widths = [50, 90, 160, 200, 250, 100, 80]
            rows = [(d["numero"], d["fecha"], d["codigo_rubro"], d["cuenta"],
                     d["objeto"], f"{d['valor']:,.0f}", d["estado"]) for d in data]
        elif tipo == "RP":
            cols = ("No", "Fecha", "CDP", "Rubro", "Tercero", "Valor", "Estado")
            data = db.get_rps()
            widths = [50, 90, 50, 160, 200, 100, 80]
            rows = [(d["numero"], d["fecha"], d["cdp_numero"], d["codigo_rubro"],
                     d["nombre_tercero"], f"{d['valor']:,.0f}", d["estado"]) for d in data]
        elif tipo == "OBLIGACION":
            cols = ("No", "Fecha", "RP", "Rubro", "Tercero", "Valor", "Factura", "Estado")
            data = db.get_obligaciones()
            widths = [50, 90, 50, 140, 180, 100, 100, 80]
            rows = [(d["numero"], d["fecha"], d["rp_numero"], d["codigo_rubro"],
                     d["nombre_tercero"], f"{d['valor']:,.0f}", d["factura"], d["estado"]) for d in data]
        elif tipo == "PAGO":
            cols = ("No", "Fecha", "Oblig", "Tercero", "Valor", "Medio", "Comprob", "Estado")
            data = db.get_pagos()
            widths = [50, 90, 50, 200, 100, 100, 100, 80]
            rows = [(d["numero"], d["fecha"], d["obligacion_numero"],
                     d["nombre_tercero"], f"{d['valor']:,.0f}", d["medio_pago"],
                     d["no_comprobante"], d["estado"]) for d in data]
        elif tipo == "RECAUDO":
            cols = ("No", "Fecha", "Rubro", "Cuenta", "Concepto", "Comprobante", "Valor", "Estado")
            data = db.get_recaudos()
            widths = [50, 90, 140, 180, 180, 100, 100, 80]
            rows = [(d["numero"], d["fecha"], d["codigo_rubro"], d["cuenta"],
                     d["concepto"], d["no_comprobante"], f"{d['valor']:,.0f}", d["estado"]) for d in data]
        elif tipo == "TERCEROS":
            cols = ("NIT", "Nombre", "Tipo", "Telefono", "Email")
            data = db.get_terceros()
            widths = [100, 300, 80, 120, 200]
            rows = [(d["nit"], d["nombre"], d["tipo"], d["telefono"], d["email"]) for d in data]

        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=20)
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center" if w < 100 else "w")

        for row in rows:
            tree.insert("", "end", values=row)

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Barra inferior con info y botones de accion
        frame_bottom = tk.Frame(f, bg=COLOR_BG)
        frame_bottom.pack(fill="x", pady=5)
        tk.Label(frame_bottom, text=f"Total registros: {len(rows)}", font=("Segoe UI", 9),
                 bg=COLOR_BG, fg="#666").pack(side="left")

        if tipo in ("CDP", "RP", "OBLIGACION", "PAGO", "RECAUDO"):
            def anular_seleccionado():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Seleccione", "Seleccione un registro para anular")
                    return
                vals = tree.item(sel[0])["values"]
                numero = int(vals[0])
                estado = str(vals[-1])
                if estado in ("ANULADO", "ANULADA"):
                    messagebox.showinfo("Info", f"{tipo} {numero} ya esta anulado")
                    return
                if not messagebox.askyesno("Anular",
                        f"Desea ANULAR el {tipo} No. {numero}?\n\nEsta accion no se puede deshacer."):
                    return
                try:
                    if tipo == "CDP":
                        db.anular_cdp(numero)
                    elif tipo == "RP":
                        db.anular_rp(numero)
                    elif tipo == "OBLIGACION":
                        db.anular_obligacion(numero)
                    elif tipo == "PAGO":
                        db.anular_pago(numero)
                    elif tipo == "RECAUDO":
                        db.anular_recaudo(numero)
                    messagebox.showinfo("Anulado", f"{tipo} {numero} anulado exitosamente")
                    self._ver_listado(tipo)
                except Exception as e:
                    messagebox.showerror("Error", str(e))

            tk.Button(frame_bottom, text="Anular Seleccionado", bg=COLOR_DANGER, fg=COLOR_WHITE,
                      font=("Segoe UI", 9, "bold"), command=anular_seleccionado).pack(side="right", padx=5)

            def editar_seleccionado():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Seleccione", "Seleccione un registro para editar")
                    return
                vals = tree.item(sel[0])["values"]
                numero = int(vals[0])
                estado = str(vals[-1])
                if estado in ("ANULADO", "ANULADA"):
                    messagebox.showinfo("Info", f"{tipo} {numero} esta anulado, no se puede editar")
                    return
                refresh = lambda: self._ver_listado(tipo)
                if tipo == "CDP":
                    self._dialogo_editar_cdp(numero, callback_refresh=refresh)
                elif tipo == "RP":
                    self._dialogo_editar_rp(numero, callback_refresh=refresh)
                elif tipo == "OBLIGACION":
                    self._dialogo_editar_obligacion(numero, callback_refresh=refresh)
                elif tipo == "PAGO":
                    self._dialogo_editar_pago(numero, callback_refresh=refresh)
                elif tipo == "RECAUDO":
                    self._dialogo_editar_recaudo(numero, callback_refresh=refresh)

            nombres_btn = {"CDP": "Editar CDP", "RP": "Editar RP", "OBLIGACION": "Editar Oblig.",
                           "PAGO": "Editar Pago", "RECAUDO": "Editar Recaudo"}
            tk.Button(frame_bottom, text=nombres_btn.get(tipo, "Editar"), bg=COLOR_WARNING, fg=COLOR_WHITE,
                      font=("Segoe UI", 9, "bold"), command=editar_seleccionado).pack(side="right", padx=5)

            def imprimir_seleccionado():
                sel = tree.selection()
                if not sel:
                    messagebox.showwarning("Seleccione", "Seleccione un registro para imprimir")
                    return
                vals = tree.item(sel[0])["values"]
                numero = int(vals[0])
                self._imprimir_comprobante(tipo, numero)

            if tipo != "RECAUDO":
                tk.Button(frame_bottom, text="Imprimir", bg=COLOR_PRIMARY, fg=COLOR_WHITE,
                          font=("Segoe UI", 9, "bold"), command=imprimir_seleccionado).pack(side="right", padx=5)

    # ===================== CONSULTA DE DOCUMENTOS POR RUBRO =====================
    def _ver_documentos_rubro(self, tipo_doc, codigo_rubro):
        if not codigo_rubro:
            return

        dialog = tk.Toplevel(self)
        dialog.title(f"{tipo_doc}s del Rubro {codigo_rubro}")
        dialog.geometry("850x450")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text=f"{tipo_doc}S - Rubro: {codigo_rubro}",
                 font=("Segoe UI", 13, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        frame_tree = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_tree.pack(fill="both", expand=True, padx=10, pady=5)

        if tipo_doc == "CDP":
            cols = ("No", "Fecha", "Objeto", "Valor", "Estado")
            widths = [60, 90, 350, 120, 80]
            data = db.get_cdps_por_rubro(codigo_rubro)
            rows = [(d["numero"], d["fecha"], d["objeto"],
                     f"$ {d['valor']:,.0f}", d["estado"]) for d in data]
        elif tipo_doc == "RP":
            cols = ("No", "Fecha", "CDP", "Tercero", "Valor", "Estado")
            widths = [60, 90, 60, 280, 120, 80]
            data = db.get_rps_por_rubro(codigo_rubro)
            rows = [(d["numero"], d["fecha"], d["cdp_numero"],
                     d["nombre_tercero"], f"$ {d['valor']:,.0f}", d["estado"]) for d in data]
        elif tipo_doc == "OBLIGACION":
            cols = ("No", "Fecha", "RP", "Tercero", "Factura", "Valor", "Estado")
            widths = [60, 90, 60, 220, 100, 120, 80]
            data = db.get_obligaciones_por_rubro(codigo_rubro)
            rows = [(d["numero"], d["fecha"], d["rp_numero"],
                     d["nombre_tercero"], d["factura"], f"$ {d['valor']:,.0f}", d["estado"]) for d in data]
        elif tipo_doc == "PAGO":
            cols = ("No", "Fecha", "Oblig", "Tercero", "Valor", "Medio", "Estado")
            widths = [60, 90, 60, 220, 120, 100, 80]
            data = db.get_pagos_por_rubro(codigo_rubro)
            rows = [(d["numero"], d["fecha"], d["obligacion_numero"],
                     d["nombre_tercero"], f"$ {d['valor']:,.0f}", d["medio_pago"], d["estado"]) for d in data]
        elif tipo_doc == "RECAUDO":
            cols = ("No", "Fecha", "Concepto", "Comprobante", "Valor", "Estado")
            widths = [60, 90, 250, 120, 120, 80]
            data = db.get_recaudos_por_rubro(codigo_rubro)
            rows = [(d["numero"], d["fecha"], d["concepto"],
                     d["no_comprobante"], f"$ {d['valor']:,.0f}", d["estado"]) for d in data]
        else:
            dialog.destroy()
            return

        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=15)
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center" if w <= 90 else "w")

        for row in rows:
            tree.insert("", "end", values=row)

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Barra inferior
        frame_bot = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_bot.pack(fill="x", padx=10, pady=8)

        total = sum(d["valor"] for d in data)
        tk.Label(frame_bot, text=f"Total: $ {total:,.0f}  ({len(data)} registros)",
                 font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(side="left")

        def imprimir_sel():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un registro para imprimir", parent=dialog)
                return
            numero = int(tree.item(sel[0])["values"][0])
            self._imprimir_comprobante(tipo_doc, numero)

        if tipo_doc != "RECAUDO":
            tk.Button(frame_bot, text="Imprimir Seleccionado", bg=COLOR_PRIMARY, fg=COLOR_WHITE,
                      font=("Segoe UI", 9, "bold"), cursor="hand2",
                      command=imprimir_sel).pack(side="right", padx=5)

        # Doble clic para imprimir
        def on_doble_clic(event):
            if tipo_doc != "RECAUDO":
                sel = tree.selection()
                if sel:
                    numero = int(tree.item(sel[0])["values"][0])
                    self._imprimir_comprobante(tipo_doc, numero)

        tree.bind("<Double-1>", on_doble_clic)

        tk.Button(frame_bot, text="Cerrar", bg="#666", fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), command=dialog.destroy).pack(side="right", padx=5)

    # ===================== IMPRESION DE COMPROBANTES =====================
    def _imprimir_comprobante(self, tipo, numero):
        institucion = db.get_config("institucion") or "INSTITUCION EDUCATIVA"
        nit_ie = db.get_config("nit_institucion") or ""
        vigencia = db.get_config("vigencia") or "2026"
        rector = db.get_config("rector") or ""
        tesorero = db.get_config("tesorero") or ""

        if tipo == "CDP":
            datos = db.get_cdp(numero)
            if not datos:
                messagebox.showerror("Error", f"CDP {numero} no encontrado")
                return
            titulo = "CERTIFICADO DE DISPONIBILIDAD PRESUPUESTAL"
            contenido = f"""
            <tr><th>No. CDP</th><td>{datos['numero']}</td></tr>
            <tr><th>Fecha</th><td>{datos['fecha']}</td></tr>
            <tr><th>Codigo Rubro</th><td>{datos['codigo_rubro']}</td></tr>
            <tr><th>Cuenta</th><td>{datos['cuenta']}</td></tr>
            <tr><th>Objeto del Gasto</th><td>{datos['objeto']}</td></tr>
            <tr><th>Valor</th><td class="valor">$ {datos['valor']:,.0f}</td></tr>
            <tr><th>Estado</th><td>{datos['estado']}</td></tr>
            """
            firmas = f"""
            <div class="firma">
                <div class="linea-firma"></div>
                <div class="nombre-firma">{tesorero}</div>
                <div class="cargo-firma">Tesorero(a) / Pagador(a)</div>
            </div>
            """

        elif tipo == "RP":
            datos = db.get_rp(numero)
            if not datos:
                messagebox.showerror("Error", f"RP {numero} no encontrado")
                return
            titulo = "REGISTRO PRESUPUESTAL DE COMPROMISO"
            contenido = f"""
            <tr><th>No. RP</th><td>{datos['numero']}</td></tr>
            <tr><th>Fecha</th><td>{datos['fecha']}</td></tr>
            <tr><th>No. CDP Asociado</th><td>{datos['cdp_numero']}</td></tr>
            <tr><th>Codigo Rubro</th><td>{datos['codigo_rubro']}</td></tr>
            <tr><th>Cuenta</th><td>{datos['cuenta']}</td></tr>
            <tr><th>NIT Tercero</th><td>{datos['nit_tercero']}</td></tr>
            <tr><th>Nombre Tercero</th><td>{datos['nombre_tercero']}</td></tr>
            <tr><th>Objeto / Contrato</th><td>{datos['objeto']}</td></tr>
            <tr><th>Valor</th><td class="valor">$ {datos['valor']:,.0f}</td></tr>
            <tr><th>Estado</th><td>{datos['estado']}</td></tr>
            """
            firmas = f"""
            <div class="firma">
                <div class="linea-firma"></div>
                <div class="nombre-firma">{tesorero}</div>
                <div class="cargo-firma">Tesorero(a) / Pagador(a)</div>
            </div>
            """

        elif tipo == "OBLIGACION":
            datos = db.get_obligacion(numero)
            if not datos:
                messagebox.showerror("Error", f"Obligacion {numero} no encontrada")
                return
            titulo = "OBLIGACION PRESUPUESTAL"
            contenido = f"""
            <tr><th>No. Obligacion</th><td>{datos['numero']}</td></tr>
            <tr><th>Fecha</th><td>{datos['fecha']}</td></tr>
            <tr><th>No. RP Asociado</th><td>{datos['rp_numero']}</td></tr>
            <tr><th>Codigo Rubro</th><td>{datos['codigo_rubro']}</td></tr>
            <tr><th>Cuenta</th><td>{datos['cuenta']}</td></tr>
            <tr><th>NIT Tercero</th><td>{datos['nit_tercero']}</td></tr>
            <tr><th>Nombre Tercero</th><td>{datos['nombre_tercero']}</td></tr>
            <tr><th>Factura / Concepto</th><td>{datos['factura']}</td></tr>
            <tr><th>Valor</th><td class="valor">$ {datos['valor']:,.0f}</td></tr>
            <tr><th>Estado</th><td>{datos['estado']}</td></tr>
            """
            firmas = f"""
            <div class="firma">
                <div class="linea-firma"></div>
                <div class="nombre-firma">{tesorero}</div>
                <div class="cargo-firma">Tesorero(a) / Pagador(a)</div>
            </div>
            """

        elif tipo == "PAGO":
            datos = db.get_pago(numero)
            if not datos:
                messagebox.showerror("Error", f"Pago {numero} no encontrado")
                return
            titulo = "ORDEN DE PAGO"
            contenido = f"""
            <tr><th>No. Pago</th><td>{datos['numero']}</td></tr>
            <tr><th>Fecha</th><td>{datos['fecha']}</td></tr>
            <tr><th>No. Obligacion</th><td>{datos['obligacion_numero']}</td></tr>
            <tr><th>Codigo Rubro</th><td>{datos['codigo_rubro']}</td></tr>
            <tr><th>Cuenta</th><td>{datos['cuenta']}</td></tr>
            <tr><th>NIT Tercero</th><td>{datos['nit_tercero']}</td></tr>
            <tr><th>Nombre Tercero</th><td>{datos['nombre_tercero']}</td></tr>
            <tr><th>Concepto</th><td>{datos['concepto']}</td></tr>
            <tr><th>Valor</th><td class="valor">$ {datos['valor']:,.0f}</td></tr>
            <tr><th>Medio de Pago</th><td>{datos['medio_pago']}</td></tr>
            <tr><th>No. Comprobante Egreso</th><td>{datos['no_comprobante']}</td></tr>
            <tr><th>Estado</th><td>{datos['estado']}</td></tr>
            """
            firmas = f"""
            <div class="firma">
                <div class="linea-firma"></div>
                <div class="nombre-firma">{rector}</div>
                <div class="cargo-firma">Rector(a) / Ordenador(a) del Gasto</div>
            </div>
            <div class="firma">
                <div class="linea-firma"></div>
                <div class="nombre-firma">{tesorero}</div>
                <div class="cargo-firma">Tesorero(a) / Pagador(a)</div>
            </div>
            """
        else:
            return

        html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{titulo} No. {numero}</title>
<style>
    @media print {{
        body {{ margin: 1.5cm; }}
        .no-print {{ display: none; }}
    }}
    body {{
        font-family: Arial, Helvetica, sans-serif;
        font-size: 12px;
        color: #333;
        max-width: 800px;
        margin: 20px auto;
        padding: 20px;
    }}
    .encabezado {{
        text-align: center;
        border-bottom: 2px solid #003366;
        padding-bottom: 10px;
        margin-bottom: 20px;
    }}
    .encabezado h2 {{
        margin: 0;
        color: #003366;
        font-size: 16px;
    }}
    .encabezado h3 {{
        margin: 5px 0;
        color: #003366;
        font-size: 18px;
    }}
    .encabezado p {{
        margin: 2px 0;
        font-size: 11px;
        color: #555;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        margin: 15px 0;
    }}
    th, td {{
        border: 1px solid #ccc;
        padding: 8px 12px;
        text-align: left;
    }}
    th {{
        background-color: #f0f4f8;
        color: #003366;
        width: 200px;
        font-weight: bold;
    }}
    .valor {{
        font-weight: bold;
        font-size: 14px;
        color: #006633;
    }}
    .firmas {{
        display: flex;
        justify-content: space-around;
        margin-top: 60px;
        page-break-inside: avoid;
    }}
    .firma {{
        text-align: center;
        min-width: 250px;
    }}
    .linea-firma {{
        border-top: 1px solid #333;
        width: 250px;
        margin: 0 auto 5px;
    }}
    .nombre-firma {{
        font-weight: bold;
        font-size: 12px;
    }}
    .cargo-firma {{
        font-size: 11px;
        color: #555;
    }}
    .btn-imprimir {{
        display: block;
        margin: 20px auto;
        padding: 10px 30px;
        background: #003366;
        color: white;
        border: none;
        font-size: 14px;
        cursor: pointer;
        border-radius: 5px;
    }}
    .btn-imprimir:hover {{
        background: #004488;
    }}
</style>
</head>
<body>
    <div class="encabezado">
        <h2>{institucion}</h2>
        <p>NIT: {nit_ie}</p>
        <p>Vigencia Fiscal: {vigencia}</p>
        <h3>{titulo}</h3>
    </div>
    <table>
        {contenido}
    </table>
    <div class="firmas">
        {firmas}
    </div>
    <button class="btn-imprimir no-print" onclick="window.print()">Imprimir (Ctrl+P)</button>
</body>
</html>"""

        try:
            tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False,
                                              encoding='utf-8', prefix=f'{tipo}_{numero}_')
            tmp.write(html)
            tmp.close()
            webbrowser.open('file:///' + tmp.name.replace('\\', '/'))
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar comprobante:\n{e}")

    # ===================== RESUMEN PRESUPUESTAL DETALLADO =====================
    def _ver_resumen_detallado(self):
        db.sincronizar_padres_gastos()
        db.sincronizar_padres_ingresos()
        self._limpiar_main()
        f = self.main_frame

        # Barra superior
        frame_top = tk.Frame(f, bg=COLOR_PRIMARY, padx=10, pady=8)
        frame_top.pack(fill="x")

        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9), cursor="hand2").pack(side="left")
        tk.Label(frame_top, text="RESUMEN PRESUPUESTAL POR RUBRO",
                 font=("Segoe UI", 14, "bold"), bg=COLOR_PRIMARY, fg=COLOR_WHITE).pack(side="left", padx=20)

        # Filtros: Mes Inicial / Mes Final
        frame_filtros = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=8)
        frame_filtros.pack(fill="x")

        tk.Label(frame_filtros, text="Mes Inicial:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=(10, 5))
        mes_ini_var = tk.StringVar(value="Enero")
        cmb_ini = ttk.Combobox(frame_filtros, textvariable=mes_ini_var,
                               values=MESES[1:], state="readonly", width=12)
        cmb_ini.pack(side="left")

        tk.Label(frame_filtros, text="Mes Final:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=(20, 5))
        mes_actual = int(db.get_config("mes_actual") or 1)
        mes_fin_default = MESES[mes_actual] if 1 <= mes_actual <= 12 else "Diciembre"
        mes_fin_var = tk.StringVar(value=mes_fin_default)
        cmb_fin = ttk.Combobox(frame_filtros, textvariable=mes_fin_var,
                               values=MESES[1:], state="readonly", width=12)
        cmb_fin.pack(side="left")

        tk.Label(frame_filtros, text=f"Vigencia: {db.get_config('vigencia')}",
                 font=("Segoe UI", 10, "bold"), bg=COLOR_LIGHT, fg=COLOR_PRIMARY).pack(side="right", padx=10)

        # Contenido principal: izquierda (arbol) + derecha (detalle)
        frame_body = tk.Frame(f, bg=COLOR_BG)
        frame_body.pack(fill="both", expand=True, pady=5)

        # === PANEL IZQUIERDO: Arbol de rubros ===
        frame_izq = tk.Frame(frame_body, bg=COLOR_WHITE, bd=1, relief="solid")
        frame_izq.pack(side="left", fill="both", expand=False, padx=(5, 3))

        # Selector de tipo: Gastos o Ingresos
        frame_tipo = tk.Frame(frame_izq, bg=COLOR_WHITE)
        frame_tipo.pack(fill="x", padx=8, pady=5)

        tipo_var = tk.StringVar(value="GASTOS")
        tk.Radiobutton(frame_tipo, text="Gastos", variable=tipo_var, value="GASTOS",
                        font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY,
                        selectcolor=COLOR_WHITE).pack(side="left", padx=(0, 10))
        tk.Radiobutton(frame_tipo, text="Ingresos", variable=tipo_var, value="INGRESOS",
                        font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg="#006633",
                        selectcolor=COLOR_WHITE).pack(side="left")

        frame_arbol = tk.Frame(frame_izq, bg=COLOR_WHITE)
        frame_arbol.pack(fill="both", expand=True, padx=5, pady=3)

        tree_rubros = ttk.Treeview(frame_arbol, show="tree", height=20)
        tree_rubros.column("#0", width=320)
        vsb_rubros = ttk.Scrollbar(frame_arbol, orient="vertical", command=tree_rubros.yview)
        tree_rubros.configure(yscrollcommand=vsb_rubros.set)
        tree_rubros.pack(side="left", fill="both", expand=True)
        vsb_rubros.pack(side="right", fill="y")

        def cargar_arbol_rubros():
            tree_rubros.delete(*tree_rubros.get_children())
            tipo = tipo_var.get()
            if tipo == "GASTOS":
                rubros = db.get_rubros_gastos(solo_hojas=False)
            else:
                rubros = db.get_rubros_ingresos(solo_hojas=False)
            nodos = {}
            for r in rubros:
                codigo = r["codigo"]
                texto = f"{codigo}  {r['cuenta']}"
                partes = codigo.rsplit(".", 1)
                padre_codigo = partes[0] if len(partes) > 1 else ""
                padre_id = nodos.get(padre_codigo, "")
                nodo_id = tree_rubros.insert(padre_id, "end", text=texto,
                                              values=(codigo,), open=False)
                nodos[codigo] = nodo_id

        # Cargar inicialmente gastos
        cargar_arbol_rubros()

        # === PANEL DERECHO: Detalle del rubro ===
        frame_der = tk.Frame(frame_body, bg=COLOR_WHITE, bd=1, relief="solid")
        frame_der.pack(side="left", fill="both", expand=True, padx=(3, 5))

        # Info del rubro seleccionado
        frame_info = tk.Frame(frame_der, bg="#E8EDF2", padx=10, pady=8)
        frame_info.pack(fill="x")

        lbl_codigo = tk.Label(frame_info, text="Codigo:", font=("Segoe UI", 10, "bold"),
                              bg="#E8EDF2", fg=COLOR_PRIMARY)
        lbl_codigo.grid(row=0, column=0, sticky="e", padx=5)
        lbl_codigo_val = tk.Label(frame_info, text="", font=("Segoe UI", 10),
                                   bg="#E8EDF2")
        lbl_codigo_val.grid(row=0, column=1, sticky="w")

        lbl_nombre = tk.Label(frame_info, text="Nombre:", font=("Segoe UI", 10, "bold"),
                              bg="#E8EDF2", fg=COLOR_PRIMARY)
        lbl_nombre.grid(row=1, column=0, sticky="e", padx=5)
        lbl_nombre_val = tk.Label(frame_info, text="", font=("Segoe UI", 10),
                                   bg="#E8EDF2", wraplength=400, justify="left")
        lbl_nombre_val.grid(row=1, column=1, sticky="w")

        lbl_aprop = tk.Label(frame_info, text="Presupuesto:", font=("Segoe UI", 10, "bold"),
                             bg="#E8EDF2", fg=COLOR_PRIMARY)
        lbl_aprop.grid(row=2, column=0, sticky="e", padx=5)
        lbl_aprop_val = tk.Label(frame_info, text="", font=("Segoe UI", 12, "bold"),
                                  bg="#E8EDF2", fg=COLOR_SUCCESS)
        lbl_aprop_val.grid(row=2, column=1, sticky="w")

        # Frame de detalle con secciones
        frame_detalle = tk.Frame(frame_der, bg=COLOR_WHITE)
        frame_detalle.pack(fill="both", expand=True, padx=10, pady=5)

        # Labels placeholder para datos
        detalle_labels = {}
        rubro_actual = {"codigo": None, "tipo": None}

        # Claves que abren documentos al hacer clic
        CLAVES_DOCS = {
            "disp_ant": "CDP", "disp_per": "CDP", "total_disp": "CDP",
            "disp_sin_comp": "CDP",
            "comp_ant": "RP", "comp_per": "RP", "total_comp": "RP",
            "comp_sin_obl": "RP",
            "obl_ant": "OBLIGACION", "obl_per": "OBLIGACION", "total_obl": "OBLIGACION",
            "obl_x_pagar": "OBLIGACION",
            "pago_ant": "PAGO", "pago_per": "PAGO", "total_pago": "PAGO",
            "rec_ant": "RECAUDO", "rec_per": "RECAUDO", "rec_acum": "RECAUDO",
        }

        def crear_seccion(parent, titulo, filas, row_start, col_start, color_titulo):
            lbl_tit = tk.Label(parent, text=f"  {titulo}  ", font=("Segoe UI", 10, "bold"),
                               bg=color_titulo, fg=COLOR_WHITE)
            lbl_tit.grid(row=row_start, column=col_start, columnspan=2, sticky="ew", pady=(8, 2))
            for i, (nombre, clave) in enumerate(filas):
                bold = "bold" if nombre.startswith("Total") or nombre.startswith("Saldo") else "normal"
                tk.Label(parent, text=nombre, font=("Segoe UI", 9, bold),
                         bg=COLOR_WHITE, anchor="e").grid(
                    row=row_start + 1 + i, column=col_start, sticky="e", padx=(5, 5), pady=1)
                es_link = clave in CLAVES_DOCS
                lbl = tk.Label(parent, text="0.00", font=("Segoe UI", 10, bold),
                               bg=COLOR_WHITE, fg=COLOR_PRIMARY, anchor="e", width=18,
                               cursor="hand2" if es_link else "")
                lbl.grid(row=row_start + 1 + i, column=col_start + 1, sticky="e", padx=5, pady=1)
                if es_link:
                    tipo_doc = CLAVES_DOCS[clave]
                    if bold == "bold":
                        lbl.bind("<Enter>", lambda e, l=lbl: l.config(font=("Segoe UI", 10, "underline bold")))
                        lbl.bind("<Leave>", lambda e, l=lbl: l.config(font=("Segoe UI", 10, "bold")))
                    else:
                        lbl.bind("<Enter>", lambda e, l=lbl: l.config(font=("Segoe UI", 10, "underline")))
                        lbl.bind("<Leave>", lambda e, l=lbl: l.config(font=("Segoe UI", 10, "normal")))
                    lbl.bind("<Button-1>", lambda e, td=tipo_doc: self._ver_documentos_rubro(td, rubro_actual["codigo"]) if rubro_actual["codigo"] else None)
                detalle_labels[clave] = lbl

        def construir_secciones_gastos():
            for w in frame_detalle.winfo_children():
                w.destroy()
            detalle_labels.clear()
            crear_seccion(frame_detalle, "Apropiacion", [
                ("Apropiacion Inicial", "aprop_ini"),
                ("Adiciones", "adiciones"),
                ("Reducciones", "reducciones"),
                ("Creditos", "creditos"),
                ("Contra-Creditos", "contracreditos"),
                ("Apropiacion Definitiva", "aprop_def"),
            ], 0, 0, "#336699")
            crear_seccion(frame_detalle, "Disponibilidades (CDPs)", [
                ("Disponibilidades anteriores", "disp_ant"),
                ("Disponibilidades periodo", "disp_per"),
                ("Total Disponibilidades", "total_disp"),
                ("Saldo Disponible", "saldo_disp"),
                ("Disp. sin compromiso", "disp_sin_comp"),
                ("Apropiacion x afectar", "aprop_x_afectar"),
            ], 8, 0, "#336699")
            crear_seccion(frame_detalle, "Compromisos (RPs)", [
                ("Compromisos anteriores", "comp_ant"),
                ("Compromisos periodo", "comp_per"),
                ("Total Compromisos", "total_comp"),
                ("Compromisos sin obligacion", "comp_sin_obl"),
            ], 0, 2, "#663399")
            crear_seccion(frame_detalle, "Obligaciones", [
                ("Obligaciones anteriores", "obl_ant"),
                ("Obligaciones periodo", "obl_per"),
                ("Total Obligaciones", "total_obl"),
                ("Obligaciones x pagar", "obl_x_pagar"),
            ], 6, 2, "#663399")
            crear_seccion(frame_detalle, "Pagos", [
                ("Pagos anteriores", "pago_ant"),
                ("Pagos periodo", "pago_per"),
                ("Total Pagos", "total_pago"),
            ], 12, 2, "#663399")

        def construir_secciones_ingresos():
            for w in frame_detalle.winfo_children():
                w.destroy()
            detalle_labels.clear()
            crear_seccion(frame_detalle, "Presupuesto", [
                ("Presupuesto Inicial", "ppto_ini"),
                ("Adiciones", "adiciones"),
                ("Reducciones", "reducciones"),
                ("Presupuesto Definitivo", "ppto_def"),
            ], 0, 0, "#006633")
            crear_seccion(frame_detalle, "Recaudos", [
                ("Recaudo anterior", "rec_ant"),
                ("Recaudo periodo", "rec_per"),
                ("Recaudo acumulado", "rec_acum"),
                ("Saldo por recaudar", "saldo_rec"),
            ], 0, 2, "#2E8B57")

        # Inicializar con gastos
        construir_secciones_gastos()

        def actualizar_detalle(codigo):
            mes_i = MESES.index(mes_ini_var.get()) if mes_ini_var.get() in MESES else 1
            mes_f = MESES.index(mes_fin_var.get()) if mes_fin_var.get() in MESES else 12
            tipo = tipo_var.get()
            rubro_actual["codigo"] = codigo
            rubro_actual["tipo"] = tipo

            if tipo == "GASTOS":
                res = db.resumen_rubro(codigo, mes_i, mes_f)
                if not res:
                    return
                rubro = res["rubro"]
                lbl_codigo_val.config(text=rubro["codigo"])
                lbl_nombre_val.config(text=rubro["cuenta"])
                lbl_aprop.config(text="Apropiacion:")
                lbl_aprop_val.config(text=f"$ {res['apropiacion_definitiva']:,.0f}")

                mapeo = {
                    "aprop_ini": res["apropiacion_inicial"],
                    "adiciones": res["adiciones"],
                    "reducciones": res["reducciones"],
                    "creditos": res["creditos"],
                    "contracreditos": res["contracreditos"],
                    "aprop_def": res["apropiacion_definitiva"],
                    "disp_ant": res["disp_anteriores"],
                    "disp_per": res["disp_periodo"],
                    "total_disp": res["total_disp"],
                    "saldo_disp": res["saldo_disponible"],
                    "disp_sin_comp": res["disp_sin_compromiso"],
                    "aprop_x_afectar": res["aprop_x_afectar"],
                    "comp_ant": res["comp_anteriores"],
                    "comp_per": res["comp_periodo"],
                    "total_comp": res["total_comp"],
                    "comp_sin_obl": res["comp_sin_obligacion"],
                    "obl_ant": res["obl_anteriores"],
                    "obl_per": res["obl_periodo"],
                    "total_obl": res["total_obl"],
                    "obl_x_pagar": res["obl_x_pagar"],
                    "pago_ant": res["pago_anteriores"],
                    "pago_per": res["pago_periodo"],
                    "total_pago": res["total_pago"],
                }
            else:
                res = db.resumen_rubro_ingreso(codigo, mes_i, mes_f)
                if not res:
                    return
                rubro = res["rubro"]
                lbl_codigo_val.config(text=rubro["codigo"])
                lbl_nombre_val.config(text=rubro["cuenta"])
                lbl_aprop.config(text="Presupuesto:")
                lbl_aprop_val.config(text=f"$ {res['presupuesto_definitivo']:,.0f}")

                mapeo = {
                    "ppto_ini": res["presupuesto_inicial"],
                    "adiciones": res["adiciones"],
                    "reducciones": res["reducciones"],
                    "ppto_def": res["presupuesto_definitivo"],
                    "rec_ant": res["recaudo_anterior"],
                    "rec_per": res["recaudo_periodo"],
                    "rec_acum": res["recaudo_acumulado"],
                    "saldo_rec": res["saldo_por_recaudar"],
                }

            def fmt(val):
                color = COLOR_DANGER if val < 0 else COLOR_PRIMARY
                return f"$ {val:,.0f}", color

            for clave, valor in mapeo.items():
                if clave in detalle_labels:
                    texto, color = fmt(valor)
                    if clave in ("saldo_disp", "saldo_rec"):
                        color = COLOR_SUCCESS if valor > 0 else COLOR_DANGER
                        detalle_labels[clave].config(text=texto, fg=color,
                                                      font=("Segoe UI", 12, "bold"))
                    elif clave in ("aprop_def", "ppto_def", "total_disp", "total_comp",
                                   "total_obl", "total_pago", "rec_acum"):
                        detalle_labels[clave].config(text=texto, fg=color,
                                                      font=("Segoe UI", 10, "bold"))
                    else:
                        detalle_labels[clave].config(text=texto, fg=color)

        def on_cambio_tipo(*args):
            tipo = tipo_var.get()
            if tipo == "GASTOS":
                construir_secciones_gastos()
            else:
                construir_secciones_ingresos()
            cargar_arbol_rubros()
            # Limpiar info
            lbl_codigo_val.config(text="")
            lbl_nombre_val.config(text="")
            lbl_aprop_val.config(text="")

        tipo_var.trace_add("write", on_cambio_tipo)

        def on_select_rubro(event):
            sel = tree_rubros.selection()
            if sel:
                item = tree_rubros.item(sel[0])
                if item["values"]:
                    codigo = str(item["values"][0])
                    actualizar_detalle(codigo)

        def on_cambio_mes(*args):
            sel = tree_rubros.selection()
            if sel:
                item = tree_rubros.item(sel[0])
                if item["values"]:
                    actualizar_detalle(str(item["values"][0]))

        tree_rubros.bind("<<TreeviewSelect>>", on_select_rubro)
        cmb_ini.bind("<<ComboboxSelected>>", on_cambio_mes)
        cmb_fin.bind("<<ComboboxSelected>>", on_cambio_mes)

    # ===================== TARJETA PRESUPUESTAL =====================
    def _ver_tarjeta(self):
        codigo = self._buscar_rubro()
        if not codigo:
            return

        rubro, movimientos = db.generar_tarjeta(codigo)
        if not rubro:
            messagebox.showerror("Error", "Rubro no encontrado")
            return

        self._limpiar_main()
        f = self.main_frame

        tk.Button(f, text="< Volver al Menu", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9), cursor="hand2").pack(anchor="w")

        tk.Label(f, text="TARJETA PRESUPUESTAL DE GASTOS - VIGENCIA 2026",
                 font=("Segoe UI", 14, "bold"), bg=COLOR_BG, fg=COLOR_PRIMARY).pack(pady=(5, 0))
        tk.Label(f, text=f"Rubro: {codigo} - {rubro['cuenta']}",
                 font=("Segoe UI", 11, "bold"), bg=COLOR_BG).pack()
        tk.Label(f, text=f"Apropiacion Definitiva: $ {rubro['apropiacion_definitiva']:,.0f}",
                 font=("Segoe UI", 11), bg=COLOR_BG, fg=COLOR_SUCCESS).pack()

        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=5)

        cols = ("Fecha", "Tipo", "No.", "NIT", "Tercero", "Concepto",
                "V.CDP", "V.RP", "V.Oblig", "V.Pago")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=18)
        widths = [85, 50, 50, 90, 160, 200, 90, 90, 90, 90]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="e" if col.startswith("V.") else "w")

        totales = {"v_cdp": 0, "v_rp": 0, "v_obl": 0, "v_pago": 0}
        for m in movimientos:
            tree.insert("", "end", values=(
                m["fecha"], m["tipo"], m["numero"], m["nit"], m["tercero"], m["concepto"],
                f"{m['v_cdp']:,.0f}" if m["v_cdp"] else "",
                f"{m['v_rp']:,.0f}" if m["v_rp"] else "",
                f"{m['v_obl']:,.0f}" if m["v_obl"] else "",
                f"{m['v_pago']:,.0f}" if m["v_pago"] else "",
            ))
            for k in totales:
                totales[k] += m[k]

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Doble clic para ver/imprimir documento
        tipo_map = {"CDP": "CDP", "RP": "RP", "OBL": "OBLIGACION", "PAG": "PAGO"}
        def on_doble_clic_tarjeta(event):
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0])["values"]
            tipo_mov = str(vals[1])  # CDP, RP, OBL, PAG
            numero = int(vals[2])
            tipo_doc = tipo_map.get(tipo_mov)
            if tipo_doc:
                self._imprimir_comprobante(tipo_doc, numero)

        tree.bind("<Double-1>", on_doble_clic_tarjeta)

        # Totales
        frame_tot = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_tot.pack(fill="x")
        labels_tot = [
            (f"CDPs: $ {totales['v_cdp']:,.0f}", COLOR_SUCCESS),
            (f"RPs: $ {totales['v_rp']:,.0f}", COLOR_PRIMARY),
            (f"Oblig: $ {totales['v_obl']:,.0f}", COLOR_WARNING),
            (f"Pagos: $ {totales['v_pago']:,.0f}", COLOR_PURPLE),
            (f"Saldo Disp: $ {rubro['apropiacion_definitiva'] - totales['v_cdp']:,.0f}", COLOR_DANGER),
        ]
        for i, (txt, color) in enumerate(labels_tot):
            tk.Label(frame_tot, text=txt, font=("Segoe UI", 10, "bold"),
                     bg=COLOR_LIGHT, fg=color).pack(side="left", padx=15)

        tk.Label(frame_tot, text="(Doble clic en un registro para imprimir)",
                 font=("Segoe UI", 8, "italic"), bg=COLOR_LIGHT, fg="#888").pack(side="right", padx=10)

    # ===================== CONSOLIDACION =====================
    def _consolidar_mes(self):
        mes = int(db.get_config("mes_actual"))
        if not messagebox.askyesno("Consolidar",
                                    f"Consolidar mes de {MESES[mes]}?\n\n"
                                    f"Se actualizaran compromisos, pagos y recaudos en la base de datos."):
            return
        mes_r, rubros = db.consolidar_mes()
        mes_ri, rubros_i = db.consolidar_mes_ingresos()
        messagebox.showinfo("Consolidacion",
                            f"Consolidacion completada\n\nMes: {MESES[mes_r]}\n"
                            f"Rubros gastos: {rubros}\n"
                            f"Rubros ingresos: {rubros_i}")

    def _cierre_mes(self):
        mes = int(db.get_config("mes_actual"))
        if mes > 12:
            messagebox.showinfo("Cierre", "Ya se cerraron todos los meses de la vigencia")
            return
        if not messagebox.askyesno("Cierre de Mes",
                                    f"CIERRE DE MES - {MESES[mes]}\n\n"
                                    f"Se consolidaran datos y se avanzara al siguiente mes.\n\n"
                                    f"Desea continuar?"):
            return
        mes_cerrado = db.cierre_mes()
        nuevo = mes_cerrado + 1
        msg_nuevo = MESES[nuevo] if nuevo <= 12 else "FIN DE VIGENCIA"
        messagebox.showinfo("Cierre",
                            f"Cierre completado\n\nMes cerrado: {MESES[mes_cerrado]}\n"
                            f"Nuevo mes: {msg_nuevo}")
        self._mostrar_dashboard()

    # ===================== CONFIGURACION =====================
    def _ventana_config(self):
        dialog = tk.Toplevel(self)
        dialog.title("Configuracion")
        dialog.geometry("500x350")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        campos = [
            ("Vigencia:", "vigencia"),
            ("Institucion Educativa:", "institucion"),
            ("NIT Institucion:", "nit_institucion"),
            ("Codigo DANE:", "codigo_dane"),
            ("Rector/Director:", "rector"),
            ("Tesorero/Pagador:", "tesorero"),
            ("Mes Actual:", "mes_actual"),
        ]
        entries = {}
        for i, (label, clave) in enumerate(campos):
            tk.Label(dialog, text=label, bg=COLOR_WHITE, font=("Segoe UI", 10),
                     anchor="e").grid(row=i, column=0, padx=10, pady=5, sticky="e")
            e = tk.Entry(dialog, font=("Segoe UI", 10), width=40)
            e.insert(0, db.get_config(clave) or "")
            e.grid(row=i, column=1, padx=10, pady=5)
            entries[clave] = e

        def guardar():
            for clave, entry in entries.items():
                db.set_config(clave, entry.get().strip())
            messagebox.showinfo("Config", "Configuracion guardada", parent=dialog)
            dialog.destroy()
            self._mostrar_dashboard()

        tk.Button(dialog, text="Guardar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=15, command=guardar).grid(
            row=len(campos) + 1, column=1, pady=15, sticky="e", padx=10)


    # ===================== CUENTAS BANCARIAS =====================
    def _gestionar_cuentas_bancarias(self):
        dialog = tk.Toplevel(self)
        dialog.title("Cuentas Bancarias de la Institucion")
        dialog.geometry("750x450")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="CUENTAS BANCARIAS DE LA INSTITUCION",
                 font=("Segoe UI", 14, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)

        frame_tree = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_tree.pack(fill="both", expand=True, padx=10)

        cols = ("ID", "Banco", "Tipo Cuenta", "No. Cuenta", "Denominacion", "Estado")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=12)
        tree.heading("ID", text="ID")
        tree.column("ID", width=40, anchor="center")
        tree.heading("Banco", text="Banco")
        tree.column("Banco", width=150)
        tree.heading("Tipo Cuenta", text="Tipo Cuenta")
        tree.column("Tipo Cuenta", width=90, anchor="center")
        tree.heading("No. Cuenta", text="No. Cuenta")
        tree.column("No. Cuenta", width=130)
        tree.heading("Denominacion", text="Denominacion")
        tree.column("Denominacion", width=200)
        tree.heading("Estado", text="Estado")
        tree.column("Estado", width=80, anchor="center")

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def cargar():
            tree.delete(*tree.get_children())
            cuentas = db.listar_cuentas_bancarias(solo_activas=False)
            for c in cuentas:
                tag = "inactiva" if c["estado"] == "INACTIVA" else ""
                tree.insert("", "end", values=(
                    c["id"], c["banco"], c["tipo_cuenta"],
                    c["numero_cuenta"], c["denominacion"], c["estado"]
                ), tags=(tag,))
            tree.tag_configure("inactiva", foreground="#999")

        def formulario_cuenta(datos=None):
            form = tk.Toplevel(dialog)
            form.title("Editar Cuenta" if datos else "Nueva Cuenta Bancaria")
            form.geometry("400x250")
            form.transient(dialog)
            form.grab_set()
            form.configure(bg=COLOR_WHITE)

            campos = [("Banco:", "banco"), ("Tipo Cuenta:", "tipo_cuenta"),
                      ("No. Cuenta:", "numero_cuenta"), ("Denominacion:", "denominacion")]
            entries = {}
            for i, (label, key) in enumerate(campos):
                tk.Label(form, text=label, bg=COLOR_WHITE, font=("Segoe UI", 10),
                         anchor="e").grid(row=i, column=0, padx=10, pady=5, sticky="e")
                if key == "tipo_cuenta":
                    var = tk.StringVar(value=datos["tipo_cuenta"] if datos else "Ahorros")
                    ttk.Combobox(form, textvariable=var, values=["Ahorros", "Corriente"],
                                 state="readonly", width=30).grid(row=i, column=1, padx=10, pady=5)
                    entries[key] = var
                else:
                    e = tk.Entry(form, font=("Segoe UI", 10), width=32)
                    e.grid(row=i, column=1, padx=10, pady=5)
                    if datos:
                        e.insert(0, datos.get(key, ""))
                    entries[key] = e

            def guardar():
                banco = entries["banco"].get().strip() if isinstance(entries["banco"], tk.Entry) else entries["banco"].get()
                tipo = entries["tipo_cuenta"].get()
                num = entries["numero_cuenta"].get().strip() if isinstance(entries["numero_cuenta"], tk.Entry) else entries["numero_cuenta"].get()
                denom = entries["denominacion"].get().strip() if isinstance(entries["denominacion"], tk.Entry) else entries["denominacion"].get()
                if not banco or not num:
                    messagebox.showwarning("Datos", "Banco y No. Cuenta son obligatorios", parent=form)
                    return
                if datos:
                    db.editar_cuenta_bancaria(datos["id"], banco, tipo, num, denom)
                    messagebox.showinfo("Cuenta", "Cuenta actualizada", parent=form)
                else:
                    db.crear_cuenta_bancaria(banco, tipo, num, denom)
                    messagebox.showinfo("Cuenta", "Cuenta creada", parent=form)
                form.destroy()
                cargar()

            tk.Button(form, text="Guardar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=15, command=guardar).grid(
                row=len(campos), column=1, pady=15, sticky="e", padx=10)

        def agregar():
            formulario_cuenta()

        def editar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione una cuenta para editar", parent=dialog)
                return
            vals = tree.item(sel[0])["values"]
            cuenta = db.get_cuenta_bancaria(int(vals[0]))
            if cuenta:
                formulario_cuenta(cuenta)

        def desactivar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione una cuenta para desactivar", parent=dialog)
                return
            vals = tree.item(sel[0])["values"]
            if vals[5] == "INACTIVA":
                messagebox.showinfo("Info", "La cuenta ya esta inactiva", parent=dialog)
                return
            if messagebox.askyesno("Confirmar", f"Desactivar cuenta {vals[3]} de {vals[1]}?", parent=dialog):
                db.desactivar_cuenta_bancaria(int(vals[0]))
                cargar()

        frame_btn = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btn.pack(pady=10)
        tk.Button(frame_btn, text="Agregar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, cursor="hand2",
                  command=agregar).pack(side="left", padx=5)
        tk.Button(frame_btn, text="Editar", bg=COLOR_PRIMARY, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, cursor="hand2",
                  command=editar).pack(side="left", padx=5)
        tk.Button(frame_btn, text="Desactivar", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, cursor="hand2",
                  command=desactivar).pack(side="left", padx=5)

        cargar()

    # ===================== GESTION DE RUBROS =====================
    def _gestion_rubros(self):
        db.sincronizar_padres_gastos()
        self._limpiar_main()
        f = self.main_frame

        frame_top = tk.Frame(f, bg="#2E7D32", padx=10, pady=8)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text="GESTION DE RUBROS PRESUPUESTALES",
                 font=("Segoe UI", 13, "bold"), bg="#2E7D32", fg=COLOR_WHITE).pack(side="left", padx=15)

        # Barra de busqueda
        frame_busq = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_busq.pack(fill="x")
        tk.Label(frame_busq, text="Buscar:", font=("Segoe UI", 10),
                 bg=COLOR_LIGHT).pack(side="left", padx=5)
        entry_busq = tk.Entry(frame_busq, font=("Segoe UI", 10), width=30)
        entry_busq.pack(side="left", padx=5)

        # Tabla de rubros
        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=3, padx=5)

        cols = ("Codigo", "Cuenta", "Hoja", "Aprop. Inicial", "Aprop. Definitiva")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=20)
        tree.heading("Codigo", text="Codigo")
        tree.heading("Cuenta", text="Cuenta")
        tree.heading("Hoja", text="Hoja")
        tree.heading("Aprop. Inicial", text="Aprop. Inicial")
        tree.heading("Aprop. Definitiva", text="Aprop. Definitiva")
        tree.column("Codigo", width=200)
        tree.column("Cuenta", width=350)
        tree.column("Hoja", width=50, anchor="center")
        tree.column("Aprop. Inicial", width=120, anchor="e")
        tree.column("Aprop. Definitiva", width=130, anchor="e")

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tree.tag_configure("padre", background="#E8EDF5", font=("Segoe UI", 9, "bold"))
        tree.tag_configure("hoja", background=COLOR_WHITE)

        lbl_info = tk.Label(f, text="", font=("Segoe UI", 9), bg=COLOR_BG, fg="#666")
        lbl_info.pack(anchor="w", padx=10)

        def cargar_rubros(filtro=""):
            tree.delete(*tree.get_children())
            rubros = db.get_rubros_gastos(solo_hojas=False)
            count = 0
            for r in rubros:
                if filtro and filtro.upper() not in r["codigo"].upper() and filtro.upper() not in r["cuenta"].upper():
                    continue
                tag = "hoja" if r["es_hoja"] else "padre"
                tree.insert("", "end", values=(
                    r["codigo"], r["cuenta"],
                    "Si" if r["es_hoja"] else "",
                    f"{r['apropiacion_inicial']:,.0f}" if r["apropiacion_inicial"] else "",
                    f"{r['apropiacion_definitiva']:,.0f}" if r["apropiacion_definitiva"] else "",
                ), tags=(tag,))
                count += 1
            lbl_info.config(text=f"Total rubros: {count}")

        def on_buscar(*args):
            cargar_rubros(entry_busq.get().strip())

        entry_busq.bind("<KeyRelease>", on_buscar)

        # Botones de accion
        frame_btns = tk.Frame(f, bg=COLOR_BG, pady=5)
        frame_btns.pack(fill="x", padx=10)

        def crear_rubro():
            dialog = tk.Toplevel(self)
            dialog.title("Crear Nuevo Rubro")
            dialog.geometry("500x280")
            dialog.transient(self)
            dialog.grab_set()
            dialog.configure(bg=COLOR_WHITE)

            tk.Label(dialog, text="CREAR NUEVO RUBRO", font=("Segoe UI", 13, "bold"),
                     bg=COLOR_WHITE, fg="#2E7D32").pack(pady=10)

            frame_campos = tk.Frame(dialog, bg=COLOR_WHITE, padx=20)
            frame_campos.pack(fill="x")

            tk.Label(frame_campos, text="Codigo:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=0, column=0, sticky="e", padx=5, pady=5)
            e_codigo = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_codigo.grid(row=0, column=1, pady=5)

            # Precargar codigo si hay seleccion
            sel = tree.selection()
            if sel:
                codigo_padre = str(tree.item(sel[0])["values"][0])
                e_codigo.insert(0, codigo_padre + ".")

            tk.Label(frame_campos, text="Cuenta/Nombre:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=1, column=0, sticky="e", padx=5, pady=5)
            e_cuenta = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_cuenta.grid(row=1, column=1, pady=5)

            tk.Label(frame_campos, text="Aprop. Inicial:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=2, column=0, sticky="e", padx=5, pady=5)
            e_aprop_ini = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_aprop_ini.insert(0, "0")
            e_aprop_ini.grid(row=2, column=1, pady=5)

            tk.Label(frame_campos, text="Aprop. Definitiva:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=3, column=0, sticky="e", padx=5, pady=5)
            e_aprop_def = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_aprop_def.insert(0, "0")
            e_aprop_def.grid(row=3, column=1, pady=5)

            def guardar():
                codigo = e_codigo.get().strip()
                cuenta = e_cuenta.get().strip()
                if not codigo or not cuenta:
                    messagebox.showwarning("Datos", "Codigo y Cuenta son obligatorios", parent=dialog)
                    return
                try:
                    aprop_ini = float(e_aprop_ini.get().replace(",", "").replace(".", "").strip() or "0")
                    aprop_def = float(e_aprop_def.get().replace(",", "").replace(".", "").strip() or "0")
                    # Verificar equilibrio antes de crear
                    if aprop_def > 0:
                        total_gas, total_ing, _ = db.verificar_equilibrio()
                        nuevo_total_gas = total_gas + aprop_def
                        nueva_diferencia = total_ing - nuevo_total_gas
                        if nueva_diferencia != 0:
                            msg = (f"Con este nuevo rubro:\n"
                                   f"Total Gastos: $ {nuevo_total_gas:,.0f}\n"
                                   f"Total Ingresos: $ {total_ing:,.0f}\n"
                                   f"Diferencia: $ {nueva_diferencia:,.0f}\n\n"
                                   f"No habra equilibrio presupuestal.\n"
                                   f"Desea continuar?")
                            if not messagebox.askyesno("Advertencia Equilibrio", msg, parent=dialog):
                                return
                    db.crear_rubro_gasto(codigo, cuenta, aprop_def, aprop_ini)
                    messagebox.showinfo("Creado", f"Rubro {codigo} creado exitosamente", parent=dialog)
                    dialog.destroy()
                    cargar_rubros(entry_busq.get().strip())
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)

            tk.Button(dialog, text="Crear Rubro", bg="#2E7D32", fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=15, command=guardar).pack(pady=15)

        def editar_rubro():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un rubro para editar")
                return
            vals = tree.item(sel[0])["values"]
            codigo = str(vals[0])
            rubro = db.get_rubro_gasto(codigo)

            dialog = tk.Toplevel(self)
            dialog.title(f"Editar Rubro {codigo}")
            dialog.geometry("500x280")
            dialog.transient(self)
            dialog.grab_set()
            dialog.configure(bg=COLOR_WHITE)

            es_hoja = rubro.get("es_hoja", 0)
            tk.Label(dialog, text=f"EDITAR RUBRO: {codigo}", font=("Segoe UI", 13, "bold"),
                     bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)

            if not es_hoja:
                tk.Label(dialog, text="(Rubro padre: el cambio de valor se distribuye\n"
                         "proporcionalmente entre los rubros hoja.)",
                         font=("Segoe UI", 9, "italic"), bg=COLOR_WHITE, fg="#999").pack()

            frame_campos = tk.Frame(dialog, bg=COLOR_WHITE, padx=20)
            frame_campos.pack(fill="x")

            tk.Label(frame_campos, text="Codigo:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=0, column=0, sticky="e", padx=5, pady=5)
            tk.Label(frame_campos, text=codigo, font=("Segoe UI", 11),
                     bg=COLOR_WHITE).grid(row=0, column=1, sticky="w", pady=5)

            tk.Label(frame_campos, text="Cuenta/Nombre:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=1, column=0, sticky="e", padx=5, pady=5)
            e_cuenta = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_cuenta.insert(0, rubro["cuenta"])
            e_cuenta.grid(row=1, column=1, pady=5)

            tk.Label(frame_campos, text="Aprop. Inicial:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=2, column=0, sticky="e", padx=5, pady=5)
            e_aprop_ini = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_aprop_ini.insert(0, f"{rubro['apropiacion_inicial']:,.0f}")
            e_aprop_ini.grid(row=2, column=1, pady=5)
            if not es_hoja:
                e_aprop_ini.config(state="disabled")

            tk.Label(frame_campos, text="Aprop. Definitiva:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=3, column=0, sticky="e", padx=5, pady=5)
            e_aprop_def = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_aprop_def.insert(0, f"{rubro['apropiacion_definitiva']:,.0f}")
            e_aprop_def.grid(row=3, column=1, pady=5)

            def guardar():
                cuenta = e_cuenta.get().strip()
                if not cuenta:
                    messagebox.showwarning("Datos", "Cuenta es obligatoria", parent=dialog)
                    return
                try:
                    aprop_def = float(e_aprop_def.get().replace(",", "").replace(".", "").strip() or "0")
                    # Verificar equilibrio antes de guardar
                    total_gas, total_ing, _ = db.verificar_equilibrio()
                    if not es_hoja:
                        conn_tmp = db.get_connection()
                        valor_anterior = conn_tmp.execute(
                            "SELECT COALESCE(SUM(apropiacion_definitiva),0) as t "
                            "FROM rubros_gastos WHERE es_hoja=1 AND codigo LIKE ?",
                            (codigo + ".%",)
                        ).fetchone()["t"]
                        conn_tmp.close()
                    else:
                        valor_anterior = rubro["apropiacion_definitiva"]
                    nuevo_total_gas = total_gas - valor_anterior + aprop_def
                    nueva_diferencia = total_ing - nuevo_total_gas
                    if nueva_diferencia != 0:
                        msg = (f"Con este cambio:\n"
                               f"Total Gastos: $ {nuevo_total_gas:,.0f}\n"
                               f"Total Ingresos: $ {total_ing:,.0f}\n"
                               f"Diferencia: $ {nueva_diferencia:,.0f}\n\n"
                               f"No habra equilibrio presupuestal.\n"
                               f"Desea continuar?")
                        if not messagebox.askyesno("Advertencia Equilibrio", msg, parent=dialog):
                            return
                    if es_hoja:
                        aprop_ini = float(e_aprop_ini.get().replace(",", "").replace(".", "").strip() or "0")
                        db.editar_rubro_gasto(codigo, cuenta, aprop_def, aprop_ini)
                    else:
                        db.editar_rubro_gasto_padre(codigo, cuenta, aprop_def)
                    messagebox.showinfo("Editado", f"Rubro {codigo} actualizado", parent=dialog)
                    dialog.destroy()
                    cargar_rubros(entry_busq.get().strip())
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)

            tk.Button(dialog, text="Guardar Cambios", bg=COLOR_PRIMARY, fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=15, command=guardar).pack(pady=15)

        def eliminar_rubro():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un rubro para eliminar")
                return
            codigo = str(tree.item(sel[0])["values"][0])
            if not messagebox.askyesno("Eliminar",
                    f"Desea ELIMINAR el rubro {codigo}?\n\nEsta accion no se puede deshacer."):
                return
            try:
                db.eliminar_rubro_gasto(codigo)
                messagebox.showinfo("Eliminado", f"Rubro {codigo} eliminado")
                cargar_rubros(entry_busq.get().strip())
            except Exception as e:
                messagebox.showerror("Error", str(e))

        tk.Button(frame_btns, text="Crear Rubro", bg="#2E7D32", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=crear_rubro).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Editar Rubro", bg=COLOR_PRIMARY, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=editar_rubro).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Eliminar Rubro", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=eliminar_rubro).pack(side="left", padx=5)

        cargar_rubros()

    # ===================== MENU DE INFORMES =====================
    def _menu_informes(self):
        dialog = tk.Toplevel(self)
        dialog.title("Generacion de Informes")
        dialog.geometry("420x500")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="GENERACION DE INFORMES", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=10)

        # Filtros de mes
        frame_mes = tk.Frame(dialog, bg=COLOR_LIGHT, padx=10, pady=8)
        frame_mes.pack(fill="x", padx=15)

        tk.Label(frame_mes, text="Mes Inicial:", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=3)
        mes_ini_var = tk.StringVar(value="Enero")
        ttk.Combobox(frame_mes, textvariable=mes_ini_var,
                     values=MESES[1:], state="readonly", width=10).pack(side="left")
        tk.Label(frame_mes, text="Mes Final:", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=(15, 3))
        mes_actual = int(db.get_config("mes_actual") or 1)
        mes_fin_var = tk.StringVar(value=MESES[mes_actual] if 1 <= mes_actual <= 12 else "Diciembre")
        ttk.Combobox(frame_mes, textvariable=mes_fin_var,
                     values=MESES[1:], state="readonly", width=10).pack(side="left")

        def get_meses():
            mi = MESES.index(mes_ini_var.get()) if mes_ini_var.get() in MESES else 1
            mf = MESES.index(mes_fin_var.get()) if mes_fin_var.get() in MESES else 12
            return mi, mf

        # Arbol de informes
        frame_tree = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_tree.pack(fill="both", expand=True, padx=15, pady=10)

        tree = ttk.Treeview(frame_tree, show="tree", height=14)
        tree.column("#0", width=370)

        n_ejec = tree.insert("", "end", text="Ejecuciones Presupuestales", open=True)
        tree.insert(n_ejec, "end", text="Ejecucion de Gastos (Formato Catalogo)", values=("gastos",))
        tree.insert(n_ejec, "end", text="Auxiliar - Detalle Movimientos", values=("auxiliar",))
        tree.insert(n_ejec, "end", text="Cadena Presupuestal (CDP>RP>Oblig>Pago)", values=("cadena",))

        n_ing = tree.insert("", "end", text="Ejecuciones de Ingresos", open=True)
        tree.insert(n_ing, "end", text="Ejecucion de Ingresos (Formato Catalogo)", values=("ingresos",))
        tree.insert(n_ing, "end", text="Equilibrio Presupuestal (Ingresos vs Gastos)", values=("equilibrio",))

        n_ccpet = tree.insert("", "end", text="Ejecuciones CCPET", open=True)
        tree.insert(n_ccpet, "end", text="Consulta por Rubro y Periodo", values=("ccpet",))

        tree.pack(fill="both", expand=True)

        def generar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un tipo de informe", parent=dialog)
                return
            vals = tree.item(sel[0])["values"]
            if not vals:
                return
            tipo = str(vals[0])
            mi, mf = get_meses()
            dialog.destroy()

            if tipo == "gastos":
                self._informe_ejecucion_gastos(mi, mf)
            elif tipo == "auxiliar":
                self._informe_auxiliar(mi, mf)
            elif tipo == "cadena":
                self._informe_cadena()
            elif tipo == "ingresos":
                self._informe_ejecucion_ingresos(mi, mf)
            elif tipo == "equilibrio":
                self._informe_equilibrio()
            elif tipo == "ccpet":
                self._ver_resumen_detallado()

        tree.bind("<Double-1>", lambda e: generar())

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Generar Informe", bg="#4B0082", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=18, command=generar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)

    # ===================== INFORME EJECUCION DE GASTOS =====================
    def _informe_ejecucion_gastos(self, mes_ini, mes_fin):
        self._limpiar_main()
        f = self.main_frame

        # Barra superior
        frame_top = tk.Frame(f, bg=COLOR_PRIMARY, padx=10, pady=6)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text=f"EJECUCION PRESUPUESTAL DE GASTOS - VIGENCIA {db.get_config('vigencia')}",
                 font=("Segoe UI", 12, "bold"), bg=COLOR_PRIMARY, fg=COLOR_WHITE).pack(side="left", padx=10)

        # Selector de mes
        frame_mes = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_mes.pack(fill="x")
        tk.Label(frame_mes, text="Mes de consulta:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=5)
        mes_var = tk.StringVar(value=MESES[mes_fin] if 1 <= mes_fin <= 12 else "Enero")
        cmb_mes = ttk.Combobox(frame_mes, textvariable=mes_var,
                               values=MESES[1:], state="readonly", width=12)
        cmb_mes.pack(side="left", padx=5)

        institucion = db.get_config("institucion") or ""
        tk.Label(frame_mes, text=institucion, font=("Segoe UI", 9),
                 bg=COLOR_LIGHT, fg="#555").pack(side="right", padx=10)

        # Tabla con scroll horizontal y vertical
        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=3, padx=3)

        cols = ("Codigo", "Cuenta",
                "Ppto Inicial", "Adiciones", "Reducciones", "Creditos", "Contra-Cred", "Ppto Definitivo",
                "Comp Anterior", "Comp Mes", "Comp Acumulado",
                "Pago Anterior", "Pago Mes", "Pago Acumulado",
                "Saldo Aprop", "Saldo Comp x Pagar")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=22)

        widths = [150, 200, 95, 80, 80, 80, 80, 100, 95, 90, 100, 95, 90, 100, 100, 110]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            anchor = "e" if col not in ("Codigo", "Cuenta") else "w"
            tree.column(col, width=w, anchor=anchor, minwidth=60)

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame_tree, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        # Estilos por nivel
        tree.tag_configure("n1", background="#C5D9F1", font=("Segoe UI", 10, "bold"))
        tree.tag_configure("n2", background="#D9E2F3", font=("Segoe UI", 9, "bold"))
        tree.tag_configure("n3", background="#E8EDF5")
        tree.tag_configure("hoja", background=COLOR_WHITE)

        def cargar_datos():
            tree.delete(*tree.get_children())
            mes = MESES.index(mes_var.get()) if mes_var.get() in MESES else 1
            data = db.informe_ejecucion_gastos(mes)

            for d in data:
                nivel = d["nivel"]
                if nivel <= 2:
                    tag = "n1"
                elif nivel <= 3:
                    tag = "n2"
                elif d["es_hoja"]:
                    tag = "hoja"
                else:
                    tag = "n3"

                def fmt(v):
                    return f"{v:,.0f}" if v else ""

                tree.insert("", "end", values=(
                    d["codigo"], d["cuenta"],
                    fmt(d["ppto_inicial"]), fmt(d["adiciones"]),
                    fmt(d["reducciones"]), fmt(d["creditos"]),
                    fmt(d["contracreditos"]), fmt(d["ppto_definitivo"]),
                    fmt(d["comp_anterior"]), fmt(d["comp_mes"]),
                    fmt(d["comp_acumulado"]),
                    fmt(d["pago_anterior"]), fmt(d["pago_mes"]),
                    fmt(d["pago_acumulado"]),
                    fmt(d["saldo_apropiacion"]), fmt(d["saldo_comp_pagar"]),
                ), tags=(tag,))

        def exportar_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

            mes = MESES.index(mes_var.get()) if mes_var.get() in MESES else 1
            vigencia = db.get_config("vigencia") or "2026"
            inst = db.get_config("institucion") or ""

            path = filedialog.asksaveasfilename(
                title="Exportar Ejecucion de Gastos",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"Ejecucion_Gastos_{MESES[mes]}_{vigencia}.xlsx"
            )
            if not path:
                return

            wb_exp = Workbook()
            ws = wb_exp.active
            ws.title = "GASTOS"

            # Estilos
            borde = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            fill_header = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            fill_sub = PatternFill(start_color="336699", end_color="336699", fill_type="solid")
            font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            font_titulo = Font(name="Calibri", size=14, bold=True, color="003366")
            font_normal = Font(name="Calibri", size=9)
            font_bold = Font(name="Calibri", size=9, bold=True)
            fill_n1 = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
            fill_n2 = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            fill_n3 = PatternFill(start_color="E8EDF5", end_color="E8EDF5", fill_type="solid")
            fmt_num = '#,##0'

            # Fila 1: Titulo
            ws.merge_cells("A1:P1")
            ws["A1"] = f"EJECUCION PRESUPUESTAL DE GASTOS - VIGENCIA {vigencia}"
            ws["A1"].font = font_titulo
            ws["A1"].alignment = Alignment(horizontal="center")

            # Fila 2: Institucion y mes
            ws.merge_cells("A2:H2")
            ws["A2"] = inst
            ws["A2"].font = Font(name="Calibri", size=11, bold=True)
            ws.merge_cells("I2:P2")
            ws["I2"] = f"Mes de consulta: {MESES[mes]}"
            ws["I2"].font = Font(name="Calibri", size=11, bold=True)
            ws["I2"].alignment = Alignment(horizontal="right")

            # Fila 3: Encabezados principales
            headers_r3 = ["", "CODIGO", "CUENTA", "PRESUPUESTO INICIAL", "ADICIONES",
                          "REDUCCIONES", "CREDITOS", "CONTRA CREDITOS",
                          "PRESUPUESTO DEFINITIVO", "COMPROMISOS", "", "",
                          "PAGOS", "", "", "SALDO FINAL", ""]
            for j, h in enumerate(headers_r3, 1):
                cell = ws.cell(row=3, column=j, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = borde

            # Merge de encabezados agrupados
            ws.merge_cells("J3:L3")
            ws.merge_cells("M3:O3")
            ws.merge_cells("P3:Q3")

            # Fila 4: Sub-encabezados
            headers_r4 = ["", "", "", "", "", "", "", "", "",
                          "SALDO ANTERIOR", "COMPROMISOS DEL MES", "COMPROMISOS ACUMULADOS",
                          "SALDO ANTERIOR", "PAGOS DEL MES", "PAGOS ACUMULADOS",
                          "SALDO DE APROPIACION", "SALDO COMPROMETIDO POR PAGAR"]
            for j, h in enumerate(headers_r4, 1):
                cell = ws.cell(row=4, column=j, value=h)
                cell.font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
                cell.fill = fill_sub
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = borde

            # Anchos de columna
            col_widths = [2, 22, 40, 16, 12, 12, 12, 14, 18, 16, 16, 18, 16, 14, 16, 18, 22]
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + j) if j <= 26 else ""].width = w

            # Datos
            data = db.informe_ejecucion_gastos(mes)
            for i, d in enumerate(data):
                row = i + 5
                nivel = d["nivel"]

                ws.cell(row=row, column=2, value=d["codigo"])
                ws.cell(row=row, column=3, value=d["cuenta"])
                ws.cell(row=row, column=4, value=d["ppto_inicial"] or None)
                ws.cell(row=row, column=5, value=d["adiciones"] or None)
                ws.cell(row=row, column=6, value=d["reducciones"] or None)
                ws.cell(row=row, column=7, value=d["creditos"] or None)
                ws.cell(row=row, column=8, value=d["contracreditos"] or None)

                # Col I: PRESUPUESTO DEFINITIVO = D + E - F + G - H (formula)
                ws.cell(row=row, column=9).value = f"=D{row}+E{row}-F{row}+G{row}-H{row}"

                # Compromisos
                ws.cell(row=row, column=10, value=d["comp_anterior"] or None)
                ws.cell(row=row, column=11, value=d["comp_mes"] or None)

                # Col L: COMPROMISOS ACUMULADOS = J + K (formula)
                ws.cell(row=row, column=12).value = f"=J{row}+K{row}"

                # Pagos
                ws.cell(row=row, column=13, value=d["pago_anterior"] or None)
                ws.cell(row=row, column=14, value=d["pago_mes"] or None)

                # Col O: PAGOS ACUMULADOS = M + N (formula)
                ws.cell(row=row, column=15).value = f"=M{row}+N{row}"

                # Col P: SALDO DE APROPIACION = I - L (formula)
                ws.cell(row=row, column=16).value = f"=I{row}-L{row}"

                # Col Q: SALDO COMPROMETIDO POR PAGAR = L - O (formula)
                ws.cell(row=row, column=17).value = f"=L{row}-O{row}"

                # Formato por nivel
                if nivel <= 2:
                    fill = fill_n1
                    font = font_bold
                elif nivel <= 3:
                    fill = fill_n2
                    font = font_bold
                elif not d["es_hoja"]:
                    fill = fill_n3
                    font = font_normal
                else:
                    fill = PatternFill()
                    font = font_normal

                for j in range(1, 18):
                    cell = ws.cell(row=row, column=j)
                    cell.border = borde
                    cell.font = font
                    if fill.start_color and fill.start_color.rgb != "00000000":
                        cell.fill = fill
                    if j >= 4:
                        cell.number_format = fmt_num
                        cell.alignment = Alignment(horizontal="right")

            # Fila de TOTAL GASTOS
            total_row = len(data) + 5
            ws.cell(row=total_row, column=2, value="TOTAL GASTOS")
            ws.cell(row=total_row, column=2).font = Font(name="Calibri", size=11, bold=True)
            for j in range(4, 18):
                col_letter = chr(64 + j)
                ws.cell(row=total_row, column=j).value = f"=SUBTOTAL(9,{col_letter}5:{col_letter}{total_row-1})"
                ws.cell(row=total_row, column=j).font = Font(name="Calibri", size=10, bold=True)
                ws.cell(row=total_row, column=j).number_format = fmt_num
                ws.cell(row=total_row, column=j).border = borde
                ws.cell(row=total_row, column=j).fill = PatternFill(
                    start_color="003366", end_color="003366", fill_type="solid")
                ws.cell(row=total_row, column=j).font = Font(
                    name="Calibri", size=10, bold=True, color="FFFFFF")

            # Inmovilizar paneles
            ws.freeze_panes = "D5"

            # Configurar impresion
            ws.sheet_properties.pageSetUpPr = None
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "3:4"

            wb_exp.save(path)
            messagebox.showinfo("Exportado",
                f"Ejecucion exportada a Excel con formulas:\n{path}")

        # Barra inferior
        frame_bot = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_bot.pack(fill="x")
        tk.Button(frame_bot, text="Exportar a Excel (con formulas)", bg="#006633", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), cursor="hand2", command=exportar_excel).pack(side="right", padx=5)

        cmb_mes.bind("<<ComboboxSelected>>", lambda e: cargar_datos())
        cargar_datos()

    # ===================== INFORME AUXILIAR =====================
    def _informe_auxiliar(self, mes_ini, mes_fin):
        self._limpiar_main()
        f = self.main_frame

        frame_top = tk.Frame(f, bg=COLOR_WARNING, padx=10, pady=8)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text=f"EJECUCION AUXILIAR - {MESES[mes_ini]} a {MESES[mes_fin]} {db.get_config('vigencia')}",
                 font=("Segoe UI", 13, "bold"), bg=COLOR_WARNING, fg=COLOR_WHITE).pack(side="left", padx=15)

        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=5, padx=5)

        cols = ("Tipo", "No.", "Fecha", "Rubro", "Cuenta", "Tercero", "Concepto", "Valor", "Estado")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=22)
        widths = [50, 45, 80, 140, 180, 160, 180, 100, 70]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            anchor = "e" if col == "Valor" else "w"
            tree.column(col, width=w, anchor=anchor)

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tree.tag_configure("CDP", background="#E8F5E9")
        tree.tag_configure("RP", background="#E3F2FD")
        tree.tag_configure("OBLIG", background="#FFF3E0")
        tree.tag_configure("PAGO", background="#F3E5F5")

        data = db.informe_ejecucion_auxiliar(mes_ini, mes_fin)
        total = 0
        for d in data:
            tercero = d.get("tercero", "") or ""
            tree.insert("", "end", values=(
                d["tipo"], d["numero"], d["fecha"], d["codigo_rubro"],
                d["cuenta"][:35], tercero[:30], (d.get("objeto") or "")[:35],
                f"{d['valor']:,.0f}", d["estado"]
            ), tags=(d["tipo"],))

        # Totales por tipo
        frame_tot = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_tot.pack(fill="x")
        tipos = {"CDP": 0, "RP": 0, "OBLIG": 0, "PAGO": 0}
        for d in data:
            tipos[d["tipo"]] += d["valor"]
        colores = {"CDP": COLOR_SUCCESS, "RP": COLOR_PRIMARY, "OBLIG": COLOR_WARNING, "PAGO": COLOR_PURPLE}
        for tipo, val in tipos.items():
            tk.Label(frame_tot, text=f"{tipo}: $ {val:,.0f}", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_LIGHT, fg=colores[tipo]).pack(side="left", padx=15)
        tk.Label(frame_tot, text=f"Total movimientos: {len(data)}", font=("Segoe UI", 9),
                 bg=COLOR_LIGHT, fg="#666").pack(side="right")

    # ===================== INFORME CADENA PRESUPUESTAL =====================
    def _informe_cadena(self):
        self._limpiar_main()
        f = self.main_frame

        frame_top = tk.Frame(f, bg=COLOR_ACCENT, padx=10, pady=8)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text=f"CADENA PRESUPUESTAL - CDP > RP > OBLIGACION > PAGO",
                 font=("Segoe UI", 13, "bold"), bg=COLOR_ACCENT, fg=COLOR_WHITE).pack(side="left", padx=15)

        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=5, padx=5)

        tree = ttk.Treeview(frame_tree, columns=("Detalle", "Valor", "Estado"),
                            show="tree headings", height=22)
        tree.heading("#0", text="Documento")
        tree.heading("Detalle", text="Detalle")
        tree.heading("Valor", text="Valor")
        tree.heading("Estado", text="Estado")
        tree.column("#0", width=250)
        tree.column("Detalle", width=400)
        tree.column("Valor", width=120, anchor="e")
        tree.column("Estado", width=80, anchor="center")

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tree.tag_configure("cdp", background="#E8F5E9", font=("Segoe UI", 10, "bold"))
        tree.tag_configure("rp", background="#E3F2FD")
        tree.tag_configure("obl", background="#FFF3E0")
        tree.tag_configure("pago", background="#F3E5F5")

        cadenas = db.informe_cadena_presupuestal()
        for cdp in cadenas:
            cdp_id = tree.insert("", "end",
                text=f"CDP {cdp['numero']} - {cdp['fecha']}",
                values=(f"{cdp['codigo_rubro']} | {cdp['objeto'][:50]}",
                        f"$ {cdp['valor']:,.0f}", cdp["estado"]),
                open=True, tags=("cdp",))

            for rp in cdp.get("rps", []):
                rp_id = tree.insert(cdp_id, "end",
                    text=f"  RP {rp['numero']} - {rp['fecha']}",
                    values=(f"{rp.get('tercero', '')} | {rp['objeto'][:40]}",
                            f"$ {rp['valor']:,.0f}", rp["estado"]),
                    open=True, tags=("rp",))

                for obl in rp.get("obligaciones", []):
                    obl_id = tree.insert(rp_id, "end",
                        text=f"    Oblig. {obl['numero']} - {obl['fecha']}",
                        values=(obl.get("factura", ""),
                                f"$ {obl['valor']:,.0f}", obl["estado"]),
                        open=True, tags=("obl",))

                    for pago in obl.get("pagos", []):
                        tree.insert(obl_id, "end",
                            text=f"      Pago {pago['numero']} - {pago['fecha']}",
                            values=(pago.get("concepto", ""),
                                    f"$ {pago['valor']:,.0f}", pago["estado"]),
                            tags=("pago",))

    # ===================== BUSCAR RUBRO INGRESO =====================
    def _buscar_rubro_ingreso(self):
        dialog = tk.Toplevel(self)
        dialog.title("Buscar Rubro de Ingreso")
        dialog.geometry("560x480")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        resultado = {"codigo": None}

        tk.Label(dialog, text="BUSCAR RUBRO DE INGRESO", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg="#006633").pack(pady=(10, 5))

        frame_busq = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_busq.pack(fill="x", padx=15)
        tk.Label(frame_busq, text="Escriba codigo o nombre:", font=("Segoe UI", 9),
                 bg=COLOR_WHITE, fg="#666").pack(anchor="w")

        frame_input = tk.Frame(frame_busq, bg=COLOR_WHITE)
        frame_input.pack(fill="x")
        entry_buscar = tk.Entry(frame_input, font=("Segoe UI", 12))
        entry_buscar.pack(side="left", fill="x", expand=True, ipady=3)

        lbl_count = tk.Label(dialog, text="", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#999")
        lbl_count.pack(anchor="w", padx=15)

        frame_lista = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("codigo", "cuenta", "ppto_def")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings", height=12)
        tree.heading("codigo", text="Codigo")
        tree.heading("cuenta", text="Cuenta")
        tree.heading("ppto_def", text="Ppto. Definitivo")
        tree.column("codigo", width=180)
        tree.column("cuenta", width=250)
        tree.column("ppto_def", width=100, anchor="e")

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        lbl_info = tk.Label(dialog, text="", font=("Segoe UI", 9, "bold"),
                            bg=COLOR_WHITE, fg="#008B45")
        lbl_info.pack(anchor="w", padx=15)

        def filtrar(*args):
            filtro = entry_buscar.get().strip()
            tree.delete(*tree.get_children())
            if filtro:
                rubros = db.buscar_rubros_ingresos(filtro)
            else:
                rubros = db.get_rubros_ingresos(solo_hojas=True)
            for r in rubros:
                tree.insert("", "end", values=(
                    r["codigo"], r["cuenta"], f"{r['presupuesto_definitivo']:,.0f}"
                ))
            total = len(db.get_rubros_ingresos(solo_hojas=True))
            lbl_count.config(text=f"Rubros encontrados: {len(rubros)} de {total}")

        def on_select(event):
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0])["values"]
                codigo = str(vals[0])
                saldo = db.saldo_por_recaudar(codigo)
                color = "#008B45" if saldo > 0 else COLOR_DANGER
                lbl_info.config(text=f"{codigo} - Saldo por recaudar: $ {saldo:,.0f}", fg=color)

        def aceptar():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un rubro", parent=dialog)
                return
            resultado["codigo"] = str(tree.item(sel[0])["values"][0])
            dialog.destroy()

        entry_buscar.bind("<KeyRelease>", filtrar)
        tree.bind("<<TreeviewSelect>>", on_select)
        tree.bind("<Double-1>", lambda e: aceptar())

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Aceptar", bg="#008B45", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=12, command=aceptar).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)

        filtrar()
        entry_buscar.focus_set()
        self.wait_window(dialog)
        return resultado["codigo"]

    # ===================== REGISTRAR RECAUDO =====================
    def _registrar_recaudo(self):
        codigo = self._buscar_rubro_ingreso()
        if not codigo:
            return

        rubro = db.get_rubro_ingreso(codigo)
        saldo = db.saldo_por_recaudar(codigo)

        concepto = simpledialog.askstring(
            "Recaudo - Concepto",
            f"Rubro: {codigo} - {rubro['cuenta']}\n"
            f"Ppto. Definitivo: $ {rubro['presupuesto_definitivo']:,.0f}\n"
            f"Saldo por recaudar: $ {saldo:,.0f}\n\n"
            f"Ingrese el CONCEPTO o fuente del recaudo:",
            parent=self
        )
        if not concepto:
            return

        comprobante = simpledialog.askstring(
            "Recaudo - Comprobante",
            f"Rubro: {codigo}\nConcepto: {concepto}\n\n"
            f"Ingrese No. de Comprobante de Ingreso:",
            parent=self
        )
        if comprobante is None:
            return

        valor_str = simpledialog.askstring(
            "Recaudo - Valor",
            f"Rubro: {codigo} - {rubro['cuenta']}\n"
            f"Concepto: {concepto}\n"
            f"Saldo por recaudar: $ {saldo:,.0f}\n\n"
            f"Ingrese el VALOR del recaudo:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor numerico invalido")
            return

        if valor <= 0:
            messagebox.showerror("Error", "El valor debe ser mayor a cero")
            return

        # Advertencia si excede saldo (no bloquea)
        advertencia = ""
        if valor > saldo:
            advertencia = (f"\nADVERTENCIA: El valor ($ {valor:,.0f}) EXCEDE el saldo "
                          f"por recaudar ($ {saldo:,.0f}).\n"
                          f"El ingreso superaria el presupuesto estimado.\n")

        # Seleccionar cuenta bancaria para el recaudo
        cuentas = db.listar_cuentas_bancarias()
        cuenta_id = 0
        if cuentas:
            dlg_cuenta = tk.Toplevel(self)
            dlg_cuenta.title("Cuenta Bancaria del Recaudo")
            dlg_cuenta.geometry("420x150")
            dlg_cuenta.transient(self)
            dlg_cuenta.grab_set()
            dlg_cuenta.configure(bg=COLOR_WHITE)

            tk.Label(dlg_cuenta, text="Seleccione la cuenta bancaria donde se recibio el ingreso:",
                     font=("Segoe UI", 10), bg=COLOR_WHITE).pack(pady=10)
            cuenta_opciones = {f"{c['banco']} - {c['tipo_cuenta']} - {c['numero_cuenta']}": c["id"] for c in cuentas}
            cuenta_var = tk.StringVar()
            combo = ttk.Combobox(dlg_cuenta, textvariable=cuenta_var,
                         values=list(cuenta_opciones.keys()), state="readonly", width=40)
            combo.pack(padx=15, pady=5)
            if cuenta_opciones:
                combo.current(0)
            resultado_cuenta = {"ok": False}

            def ok_cuenta():
                resultado_cuenta["ok"] = True
                dlg_cuenta.destroy()

            tk.Button(dlg_cuenta, text="Aceptar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=12, command=ok_cuenta).pack(pady=10)
            self.wait_window(dlg_cuenta)
            if not resultado_cuenta["ok"]:
                return
            cuenta_id = cuenta_opciones.get(cuenta_var.get(), 0)

        if not messagebox.askyesno("Confirmar Recaudo",
                                    f"CONFIRMAR REGISTRO DE RECAUDO\n\n"
                                    f"Rubro: {codigo} - {rubro['cuenta']}\n"
                                    f"Concepto: {concepto}\n"
                                    f"Comprobante: {comprobante}\n"
                                    f"Valor: $ {valor:,.0f}"
                                    f"{advertencia}"):
            return

        try:
            num, fecha = db.registrar_recaudo(codigo, valor, concepto, comprobante, cuenta_id)
            messagebox.showinfo("Recaudo Registrado",
                                f"RECAUDO REGISTRADO EXITOSAMENTE\n\n"
                                f"No. Recaudo: {num}\n"
                                f"Fecha: {fecha}\n"
                                f"Rubro: {codigo}\n"
                                f"Valor: $ {valor:,.0f}\n"
                                f"Nuevo saldo por recaudar: $ {saldo - valor:,.0f}")
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ===================== SELECCIONAR / GESTIONAR RECAUDOS =====================
    def _seleccionar_recaudo(self):
        dialog = tk.Toplevel(self)
        dialog.title("Gestion de Recaudos")
        dialog.geometry("800x500")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="GESTION DE RECAUDOS", font=("Segoe UI", 14, "bold"),
                 bg=COLOR_WHITE, fg="#006633").pack(pady=(10, 5))

        frame_lista = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("No", "Fecha", "Rubro", "Cuenta", "Concepto", "Comprobante", "Valor", "Estado")
        tree = ttk.Treeview(frame_lista, columns=cols, show="headings", height=14)
        tree.heading("No", text="No.")
        tree.heading("Fecha", text="Fecha")
        tree.heading("Rubro", text="Rubro")
        tree.heading("Cuenta", text="Cuenta")
        tree.heading("Concepto", text="Concepto")
        tree.heading("Comprobante", text="Comprobante")
        tree.heading("Valor", text="Valor")
        tree.heading("Estado", text="Estado")
        tree.column("No", width=50, anchor="center")
        tree.column("Fecha", width=85, anchor="center")
        tree.column("Rubro", width=120)
        tree.column("Cuenta", width=150)
        tree.column("Concepto", width=150)
        tree.column("Comprobante", width=80, anchor="center")
        tree.column("Valor", width=100, anchor="e")
        tree.column("Estado", width=70, anchor="center")

        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        lbl_detalle = tk.Label(dialog, text="",
                               font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg="#008B45")
        lbl_detalle.pack(anchor="w", padx=15, pady=3)

        def cargar_recaudos():
            tree.delete(*tree.get_children())
            recaudos = db.get_recaudos()
            total_activo = 0
            for d in recaudos:
                tag = "anulado" if d["estado"] == "ANULADO" else "activo"
                tree.insert("", "end", values=(
                    d["numero"], d["fecha"], d["codigo_rubro"],
                    d["cuenta"][:30], d["concepto"][:30],
                    d["no_comprobante"], f"{d['valor']:,.0f}", d["estado"]
                ), tags=(tag,))
                if d["estado"] != "ANULADO":
                    total_activo += d["valor"]
            tree.tag_configure("anulado", foreground="#999999")
            tree.tag_configure("activo", foreground="#000000")
            lbl_detalle.config(text=f"Total recaudos: {len(recaudos)} | Total activo: $ {total_activo:,.0f}")

        def on_select(event):
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0])["values"]
                lbl_detalle.config(
                    text=f"Recaudo {vals[0]} | {vals[2]} - {vals[3]} | $ {vals[6]} | {vals[7]}")

        def anular():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un recaudo", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            if not messagebox.askyesno("Anular Recaudo",
                    f"Desea ANULAR el Recaudo No. {num}?", parent=dialog):
                return
            try:
                db.anular_recaudo(num)
                messagebox.showinfo("Anulado", f"Recaudo {num} anulado", parent=dialog)
                cargar_recaudos()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dialog)

        def editar_recaudo_completo():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un recaudo", parent=dialog)
                return
            num = int(tree.item(sel[0])["values"][0])
            self._dialogo_editar_recaudo(num, dialog, cargar_recaudos)

        tree.bind("<<TreeviewSelect>>", on_select)

        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Cerrar", font=("Segoe UI", 10), width=12,
                  command=dialog.destroy).pack(side="right", padx=5)
        tk.Button(frame_btns, text="Anular Recaudo", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=14, command=anular).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Editar Recaudo", bg=COLOR_WARNING, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), width=14, command=editar_recaudo_completo).pack(side="left", padx=5)

        cargar_recaudos()

    # ===================== GESTION RUBROS INGRESOS =====================
    def _gestion_rubros_ingresos(self):
        db.sincronizar_padres_ingresos()
        self._limpiar_main()
        f = self.main_frame

        frame_top = tk.Frame(f, bg="#006633", padx=10, pady=8)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text="GESTION DE RUBROS DE INGRESOS",
                 font=("Segoe UI", 13, "bold"), bg="#006633", fg=COLOR_WHITE).pack(side="left", padx=15)

        frame_busq = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_busq.pack(fill="x")
        tk.Label(frame_busq, text="Buscar:", font=("Segoe UI", 10),
                 bg=COLOR_LIGHT).pack(side="left", padx=5)
        entry_busq = tk.Entry(frame_busq, font=("Segoe UI", 10), width=30)
        entry_busq.pack(side="left", padx=5)

        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=3, padx=5)

        cols = ("Codigo", "Cuenta", "Hoja", "Ppto. Inicial", "Ppto. Definitivo")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=20)
        tree.heading("Codigo", text="Codigo")
        tree.heading("Cuenta", text="Cuenta")
        tree.heading("Hoja", text="Hoja")
        tree.heading("Ppto. Inicial", text="Ppto. Inicial")
        tree.heading("Ppto. Definitivo", text="Ppto. Definitivo")
        tree.column("Codigo", width=200)
        tree.column("Cuenta", width=350)
        tree.column("Hoja", width=50, anchor="center")
        tree.column("Ppto. Inicial", width=120, anchor="e")
        tree.column("Ppto. Definitivo", width=130, anchor="e")

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tree.tag_configure("padre", background="#E8F5E9", font=("Segoe UI", 9, "bold"))
        tree.tag_configure("hoja", background=COLOR_WHITE)

        lbl_info = tk.Label(f, text="", font=("Segoe UI", 9), bg=COLOR_BG, fg="#666")
        lbl_info.pack(anchor="w", padx=10)

        def cargar_rubros(filtro=""):
            tree.delete(*tree.get_children())
            rubros = db.get_rubros_ingresos(solo_hojas=False)
            count = 0
            for r in rubros:
                if filtro and filtro.upper() not in r["codigo"].upper() and filtro.upper() not in r["cuenta"].upper():
                    continue
                tag = "hoja" if r["es_hoja"] else "padre"
                tree.insert("", "end", values=(
                    r["codigo"], r["cuenta"],
                    "Si" if r["es_hoja"] else "",
                    f"{r['presupuesto_inicial']:,.0f}" if r["presupuesto_inicial"] else "",
                    f"{r['presupuesto_definitivo']:,.0f}" if r["presupuesto_definitivo"] else "",
                ), tags=(tag,))
                count += 1
            lbl_info.config(text=f"Total rubros de ingresos: {count}")

        def on_buscar(*args):
            cargar_rubros(entry_busq.get().strip())

        entry_busq.bind("<KeyRelease>", on_buscar)

        frame_btns = tk.Frame(f, bg=COLOR_BG, pady=5)
        frame_btns.pack(fill="x", padx=10)

        def crear_rubro():
            dialog = tk.Toplevel(self)
            dialog.title("Crear Rubro de Ingreso")
            dialog.geometry("500x280")
            dialog.transient(self)
            dialog.grab_set()
            dialog.configure(bg=COLOR_WHITE)

            tk.Label(dialog, text="CREAR RUBRO DE INGRESO", font=("Segoe UI", 13, "bold"),
                     bg=COLOR_WHITE, fg="#006633").pack(pady=10)

            frame_campos = tk.Frame(dialog, bg=COLOR_WHITE, padx=20)
            frame_campos.pack(fill="x")

            tk.Label(frame_campos, text="Codigo:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=0, column=0, sticky="e", padx=5, pady=5)
            e_codigo = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_codigo.grid(row=0, column=1, pady=5)

            sel = tree.selection()
            if sel:
                codigo_padre = str(tree.item(sel[0])["values"][0])
                e_codigo.insert(0, codigo_padre + ".")

            tk.Label(frame_campos, text="Cuenta/Nombre:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=1, column=0, sticky="e", padx=5, pady=5)
            e_cuenta = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_cuenta.grid(row=1, column=1, pady=5)

            tk.Label(frame_campos, text="Ppto. Inicial:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=2, column=0, sticky="e", padx=5, pady=5)
            e_ppto_ini = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_ppto_ini.insert(0, "0")
            e_ppto_ini.grid(row=2, column=1, pady=5)

            tk.Label(frame_campos, text="Ppto. Definitivo:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=3, column=0, sticky="e", padx=5, pady=5)
            e_ppto_def = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_ppto_def.insert(0, "0")
            e_ppto_def.grid(row=3, column=1, pady=5)

            def guardar():
                codigo = e_codigo.get().strip()
                cuenta = e_cuenta.get().strip()
                if not codigo or not cuenta:
                    messagebox.showwarning("Datos", "Codigo y Cuenta son obligatorios", parent=dialog)
                    return
                try:
                    ppto_ini = float(e_ppto_ini.get().replace(",", "").replace(".", "").strip() or "0")
                    ppto_def = float(e_ppto_def.get().replace(",", "").replace(".", "").strip() or "0")
                    # Verificar equilibrio antes de crear
                    if ppto_def > 0:
                        total_gas, total_ing, _ = db.verificar_equilibrio()
                        nuevo_total_ing = total_ing + ppto_def
                        nueva_diferencia = nuevo_total_ing - total_gas
                        if nueva_diferencia != 0:
                            msg = (f"Con este nuevo rubro:\n"
                                   f"Total Gastos: $ {total_gas:,.0f}\n"
                                   f"Total Ingresos: $ {nuevo_total_ing:,.0f}\n"
                                   f"Diferencia: $ {nueva_diferencia:,.0f}\n\n"
                                   f"No habra equilibrio presupuestal.\n"
                                   f"Desea continuar?")
                            if not messagebox.askyesno("Advertencia Equilibrio", msg, parent=dialog):
                                return
                    db.crear_rubro_ingreso(codigo, cuenta, ppto_def, ppto_ini)
                    messagebox.showinfo("Creado", f"Rubro {codigo} creado exitosamente", parent=dialog)
                    dialog.destroy()
                    cargar_rubros(entry_busq.get().strip())
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)

            tk.Button(dialog, text="Crear Rubro", bg="#006633", fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=15, command=guardar).pack(pady=15)

        def editar_rubro():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un rubro para editar")
                return
            vals = tree.item(sel[0])["values"]
            codigo = str(vals[0])
            rubro = db.get_rubro_ingreso(codigo)

            dialog = tk.Toplevel(self)
            dialog.title(f"Editar Rubro Ingreso {codigo}")
            dialog.geometry("500x280")
            dialog.transient(self)
            dialog.grab_set()
            dialog.configure(bg=COLOR_WHITE)

            es_hoja = rubro.get("es_hoja", 0)
            tk.Label(dialog, text=f"EDITAR RUBRO: {codigo}", font=("Segoe UI", 13, "bold"),
                     bg=COLOR_WHITE, fg="#006633").pack(pady=10)

            if not es_hoja:
                tk.Label(dialog, text="(Rubro padre: el cambio de valor se distribuye\n"
                         "proporcionalmente entre los rubros hoja.)",
                         font=("Segoe UI", 9, "italic"), bg=COLOR_WHITE, fg="#999").pack()

            frame_campos = tk.Frame(dialog, bg=COLOR_WHITE, padx=20)
            frame_campos.pack(fill="x")

            tk.Label(frame_campos, text="Codigo:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=0, column=0, sticky="e", padx=5, pady=5)
            tk.Label(frame_campos, text=codigo, font=("Segoe UI", 11),
                     bg=COLOR_WHITE).grid(row=0, column=1, sticky="w", pady=5)

            tk.Label(frame_campos, text="Cuenta/Nombre:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=1, column=0, sticky="e", padx=5, pady=5)
            e_cuenta = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_cuenta.insert(0, rubro["cuenta"])
            e_cuenta.grid(row=1, column=1, pady=5)

            tk.Label(frame_campos, text="Ppto. Inicial:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=2, column=0, sticky="e", padx=5, pady=5)
            e_ppto_ini = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_ppto_ini.insert(0, f"{rubro['presupuesto_inicial']:,.0f}")
            e_ppto_ini.grid(row=2, column=1, pady=5)
            if not es_hoja:
                e_ppto_ini.config(state="disabled")

            tk.Label(frame_campos, text="Ppto. Definitivo:", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).grid(row=3, column=0, sticky="e", padx=5, pady=5)
            e_ppto_def = tk.Entry(frame_campos, font=("Segoe UI", 11), width=30)
            e_ppto_def.insert(0, f"{rubro['presupuesto_definitivo']:,.0f}")
            e_ppto_def.grid(row=3, column=1, pady=5)

            def guardar():
                cuenta = e_cuenta.get().strip()
                if not cuenta:
                    messagebox.showwarning("Datos", "Cuenta es obligatoria", parent=dialog)
                    return
                try:
                    ppto_def = float(e_ppto_def.get().replace(",", "").replace(".", "").strip() or "0")
                    # Verificar equilibrio antes de guardar
                    total_gas, total_ing, _ = db.verificar_equilibrio()
                    # Para padre, calcular total actual sumando hojas
                    if not es_hoja:
                        conn_tmp = db.get_connection()
                        valor_anterior = conn_tmp.execute(
                            "SELECT COALESCE(SUM(presupuesto_definitivo),0) as t "
                            "FROM rubros_ingresos WHERE es_hoja=1 AND codigo LIKE ?",
                            (codigo + ".%",)
                        ).fetchone()["t"]
                        conn_tmp.close()
                    else:
                        valor_anterior = rubro["presupuesto_definitivo"]
                    nuevo_total_ing = total_ing - valor_anterior + ppto_def
                    nueva_diferencia = nuevo_total_ing - total_gas
                    if nueva_diferencia != 0:
                        msg = (f"Con este cambio:\n"
                               f"Total Gastos: $ {total_gas:,.0f}\n"
                               f"Total Ingresos: $ {nuevo_total_ing:,.0f}\n"
                               f"Diferencia: $ {nueva_diferencia:,.0f}\n\n"
                               f"No habra equilibrio presupuestal.\n"
                               f"Desea continuar?")
                        if not messagebox.askyesno("Advertencia Equilibrio", msg, parent=dialog):
                            return
                    if es_hoja:
                        ppto_ini = float(e_ppto_ini.get().replace(",", "").replace(".", "").strip() or "0")
                        db.editar_rubro_ingreso(codigo, cuenta, ppto_def, ppto_ini)
                    else:
                        db.editar_rubro_ingreso_padre(codigo, cuenta, ppto_def)
                    messagebox.showinfo("Editado", f"Rubro {codigo} actualizado", parent=dialog)
                    dialog.destroy()
                    cargar_rubros(entry_busq.get().strip())
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)

            tk.Button(dialog, text="Guardar Cambios", bg="#006633", fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), width=15, command=guardar).pack(pady=15)

        def eliminar_rubro():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Seleccione", "Seleccione un rubro para eliminar")
                return
            codigo = str(tree.item(sel[0])["values"][0])
            if not messagebox.askyesno("Eliminar",
                    f"Desea ELIMINAR el rubro {codigo}?\n\nEsta accion no se puede deshacer."):
                return
            try:
                db.eliminar_rubro_ingreso(codigo)
                messagebox.showinfo("Eliminado", f"Rubro {codigo} eliminado")
                cargar_rubros(entry_busq.get().strip())
            except Exception as e:
                messagebox.showerror("Error", str(e))

        tk.Button(frame_btns, text="Crear Rubro", bg="#006633", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=crear_rubro).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Editar Rubro", bg="#008B45", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=editar_rubro).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Eliminar Rubro", bg=COLOR_DANGER, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=eliminar_rubro).pack(side="left", padx=5)

        cargar_rubros()

    # ===================== INFORME EJECUCION DE INGRESOS =====================
    def _informe_ejecucion_ingresos(self, mes_ini, mes_fin):
        self._limpiar_main()
        f = self.main_frame

        frame_top = tk.Frame(f, bg="#006633", padx=10, pady=6)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text=f"EJECUCION PRESUPUESTAL DE INGRESOS - VIGENCIA {db.get_config('vigencia')}",
                 font=("Segoe UI", 12, "bold"), bg="#006633", fg=COLOR_WHITE).pack(side="left", padx=10)

        frame_mes = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_mes.pack(fill="x")
        tk.Label(frame_mes, text="Mes de consulta:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=5)
        mes_var = tk.StringVar(value=MESES[mes_fin] if 1 <= mes_fin <= 12 else "Enero")
        cmb_mes = ttk.Combobox(frame_mes, textvariable=mes_var,
                               values=MESES[1:], state="readonly", width=12)
        cmb_mes.pack(side="left", padx=5)

        institucion = db.get_config("institucion") or ""
        tk.Label(frame_mes, text=institucion, font=("Segoe UI", 9),
                 bg=COLOR_LIGHT, fg="#555").pack(side="right", padx=10)

        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, pady=3, padx=3)

        cols = ("Codigo", "Cuenta",
                "Ppto Inicial", "Adiciones", "Reducciones", "Ppto Definitivo",
                "Recaudo Anterior", "Recaudo Mes", "Recaudo Acumulado",
                "Saldo x Recaudar")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=22)

        widths = [150, 220, 100, 90, 90, 110, 110, 100, 110, 110]
        for col, w in zip(cols, widths):
            tree.heading(col, text=col)
            anchor = "e" if col not in ("Codigo", "Cuenta") else "w"
            tree.column(col, width=w, anchor=anchor, minwidth=60)

        vsb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame_tree, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right", fill="y")
        tree.pack(side="left", fill="both", expand=True)

        tree.tag_configure("n1", background="#C8E6C9", font=("Segoe UI", 10, "bold"))
        tree.tag_configure("n2", background="#E8F5E9", font=("Segoe UI", 9, "bold"))
        tree.tag_configure("n3", background="#F1F8E9")
        tree.tag_configure("hoja", background=COLOR_WHITE)

        def cargar_datos():
            tree.delete(*tree.get_children())
            mes = MESES.index(mes_var.get()) if mes_var.get() in MESES else 1
            data = db.informe_ejecucion_ingresos(mes)

            for d in data:
                nivel = d["nivel"]
                if nivel <= 2:
                    tag = "n1"
                elif nivel <= 3:
                    tag = "n2"
                elif d["es_hoja"]:
                    tag = "hoja"
                else:
                    tag = "n3"

                def fmt(v):
                    return f"{v:,.0f}" if v else ""

                tree.insert("", "end", values=(
                    d["codigo"], d["cuenta"],
                    fmt(d["ppto_inicial"]), fmt(d["adiciones"]),
                    fmt(d["reducciones"]), fmt(d["ppto_definitivo"]),
                    fmt(d["recaudo_anterior"]), fmt(d["recaudo_mes"]),
                    fmt(d["recaudo_acumulado"]), fmt(d["saldo_por_recaudar"]),
                ), tags=(tag,))

        def exportar_excel():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            mes = MESES.index(mes_var.get()) if mes_var.get() in MESES else 1
            vigencia = db.get_config("vigencia") or "2026"
            inst = db.get_config("institucion") or ""

            path = filedialog.asksaveasfilename(
                title="Exportar Ejecucion de Ingresos",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile=f"Ejecucion_Ingresos_{MESES[mes]}_{vigencia}.xlsx"
            )
            if not path:
                return

            wb_exp = Workbook()
            ws = wb_exp.active
            ws.title = "INGRESOS"

            borde = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
            fill_header = PatternFill(start_color="006633", end_color="006633", fill_type="solid")
            fill_sub = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            font_titulo = Font(name="Calibri", size=14, bold=True, color="006633")
            font_normal = Font(name="Calibri", size=9)
            font_bold = Font(name="Calibri", size=9, bold=True)
            fill_n1 = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            fill_n2 = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            fill_n3 = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
            fmt_num = '#,##0'

            # Fila 1: Titulo
            ws.merge_cells("A1:J1")
            ws["A1"] = f"EJECUCION PRESUPUESTAL DE INGRESOS - VIGENCIA {vigencia}"
            ws["A1"].font = font_titulo
            ws["A1"].alignment = Alignment(horizontal="center")

            # Fila 2
            ws.merge_cells("A2:E2")
            ws["A2"] = inst
            ws["A2"].font = Font(name="Calibri", size=11, bold=True)
            ws.merge_cells("F2:J2")
            ws["F2"] = f"Mes de consulta: {MESES[mes]}"
            ws["F2"].font = Font(name="Calibri", size=11, bold=True)
            ws["F2"].alignment = Alignment(horizontal="right")

            # Fila 3: Encabezados
            headers = ["", "CODIGO", "CUENTA", "PRESUPUESTO INICIAL", "ADICIONES",
                       "REDUCCIONES", "PRESUPUESTO DEFINITIVO",
                       "RECAUDO ANTERIOR", "RECAUDO DEL MES", "RECAUDO ACUMULADO",
                       "SALDO POR RECAUDAR"]
            for j, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=j, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = borde

            col_widths = [2, 22, 40, 16, 14, 14, 18, 16, 16, 18, 18]
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + j) if j <= 26 else ""].width = w

            data = db.informe_ejecucion_ingresos(mes)
            for i, d in enumerate(data):
                row = i + 4
                nivel = d["nivel"]

                ws.cell(row=row, column=2, value=d["codigo"])
                ws.cell(row=row, column=3, value=d["cuenta"])
                ws.cell(row=row, column=4, value=d["ppto_inicial"] or None)
                ws.cell(row=row, column=5, value=d["adiciones"] or None)
                ws.cell(row=row, column=6, value=d["reducciones"] or None)

                # Col G: PRESUPUESTO DEFINITIVO = D + E - F (formula)
                ws.cell(row=row, column=7).value = f"=D{row}+E{row}-F{row}"

                ws.cell(row=row, column=8, value=d["recaudo_anterior"] or None)
                ws.cell(row=row, column=9, value=d["recaudo_mes"] or None)

                # Col J: RECAUDO ACUMULADO = H + I (formula)
                ws.cell(row=row, column=10).value = f"=H{row}+I{row}"

                # Col K: SALDO POR RECAUDAR = G - J (formula)
                ws.cell(row=row, column=11).value = f"=G{row}-J{row}"

                if nivel <= 2:
                    fill = fill_n1
                    font = font_bold
                elif nivel <= 3:
                    fill = fill_n2
                    font = font_bold
                elif not d["es_hoja"]:
                    fill = fill_n3
                    font = font_normal
                else:
                    fill = PatternFill()
                    font = font_normal

                for j in range(1, 12):
                    cell = ws.cell(row=row, column=j)
                    cell.border = borde
                    cell.font = font
                    if fill.start_color and fill.start_color.rgb != "00000000":
                        cell.fill = fill
                    if j >= 4:
                        cell.number_format = fmt_num
                        cell.alignment = Alignment(horizontal="right")

            # Fila de TOTAL INGRESOS
            total_row = len(data) + 4
            ws.cell(row=total_row, column=2, value="TOTAL INGRESOS")
            ws.cell(row=total_row, column=2).font = Font(name="Calibri", size=11, bold=True)
            for j in range(4, 12):
                col_letter = chr(64 + j)
                ws.cell(row=total_row, column=j).value = f"=SUBTOTAL(9,{col_letter}4:{col_letter}{total_row-1})"
                ws.cell(row=total_row, column=j).font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                ws.cell(row=total_row, column=j).number_format = fmt_num
                ws.cell(row=total_row, column=j).border = borde
                ws.cell(row=total_row, column=j).fill = PatternFill(
                    start_color="006633", end_color="006633", fill_type="solid")

            ws.freeze_panes = "D4"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "3:3"

            wb_exp.save(path)
            messagebox.showinfo("Exportado",
                f"Ejecucion de ingresos exportada a Excel:\n{path}")

        frame_bot = tk.Frame(f, bg=COLOR_LIGHT, padx=10, pady=5)
        frame_bot.pack(fill="x")
        tk.Button(frame_bot, text="Exportar a Excel (con formulas)", bg="#006633", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), cursor="hand2", command=exportar_excel).pack(side="right", padx=5)

        cmb_mes.bind("<<ComboboxSelected>>", lambda e: cargar_datos())
        cargar_datos()

    # ===================== INFORME EQUILIBRIO PRESUPUESTAL =====================
    def _informe_equilibrio(self):
        self._limpiar_main()
        f = self.main_frame

        frame_top = tk.Frame(f, bg="#4B0082", padx=10, pady=8)
        frame_top.pack(fill="x")
        tk.Button(frame_top, text="< Volver", command=self._mostrar_dashboard,
                  font=("Segoe UI", 9)).pack(side="left")
        tk.Label(frame_top, text=f"EQUILIBRIO PRESUPUESTAL - VIGENCIA {db.get_config('vigencia')}",
                 font=("Segoe UI", 13, "bold"), bg="#4B0082", fg=COLOR_WHITE).pack(side="left", padx=15)

        resumen = db.get_resumen()

        # Frame principal
        frame_body = tk.Frame(f, bg=COLOR_BG)
        frame_body.pack(fill="both", expand=True, padx=20, pady=15)

        # Seccion INGRESOS
        frame_ing = tk.LabelFrame(frame_body, text=" INGRESOS ",
                                   font=("Segoe UI", 13, "bold"),
                                   bg=COLOR_WHITE, fg="#006633", padx=25, pady=15)
        frame_ing.pack(fill="x", pady=5)

        datos_ing = [
            ("Presupuesto Definitivo Ingresos:", resumen["ppto_ingresos"], "#006633"),
            ("Total Recaudado:", resumen["recaudado"], "#008B45"),
            ("Saldo por Recaudar:", resumen["saldo_por_recaudar"], COLOR_ACCENT),
        ]
        for i, (label, valor, color) in enumerate(datos_ing):
            tk.Label(frame_ing, text=label, font=("Segoe UI", 12, "bold"),
                     bg=COLOR_WHITE, anchor="e").grid(row=i, column=0, sticky="e", padx=(0, 15))
            fg = COLOR_DANGER if valor < 0 else color
            tk.Label(frame_ing, text=f"$ {valor:,.0f}", font=("Segoe UI", 14, "bold"),
                     bg=COLOR_WHITE, fg=fg).grid(row=i, column=1, sticky="w")

        # Seccion GASTOS
        frame_gas = tk.LabelFrame(frame_body, text=" GASTOS ",
                                   font=("Segoe UI", 13, "bold"),
                                   bg=COLOR_WHITE, fg=COLOR_PRIMARY, padx=25, pady=15)
        frame_gas.pack(fill="x", pady=5)

        datos_gas = [
            ("Apropiacion Definitiva Gastos:", resumen["apropiacion"], COLOR_PRIMARY),
            ("Total Comprometido:", resumen["comprometido"], COLOR_PRIMARY),
            ("Total Pagado:", resumen["pagado"], COLOR_PURPLE),
            ("Saldo Disponible:", resumen["saldo_disponible"], COLOR_SUCCESS),
        ]
        for i, (label, valor, color) in enumerate(datos_gas):
            tk.Label(frame_gas, text=label, font=("Segoe UI", 12, "bold"),
                     bg=COLOR_WHITE, anchor="e").grid(row=i, column=0, sticky="e", padx=(0, 15))
            fg = COLOR_DANGER if valor < 0 else color
            tk.Label(frame_gas, text=f"$ {valor:,.0f}", font=("Segoe UI", 14, "bold"),
                     bg=COLOR_WHITE, fg=fg).grid(row=i, column=1, sticky="w")

        # Seccion EQUILIBRIO
        equilibrio = resumen["equilibrio"]
        if equilibrio == 0:
            estado = "EQUILIBRADO"
            color_eq = "#006633"
        elif equilibrio > 0:
            estado = "SUPERAVIT"
            color_eq = "#008B45"
        else:
            estado = "DEFICIT"
            color_eq = COLOR_DANGER

        frame_eq = tk.LabelFrame(frame_body, text=" RESULTADO ",
                                  font=("Segoe UI", 13, "bold"),
                                  bg=COLOR_WHITE, fg=color_eq, padx=25, pady=20)
        frame_eq.pack(fill="x", pady=10)

        tk.Label(frame_eq, text="Ingresos - Gastos:", font=("Segoe UI", 12, "bold"),
                 bg=COLOR_WHITE, anchor="e").grid(row=0, column=0, sticky="e", padx=(0, 15))
        tk.Label(frame_eq, text=f"$ {equilibrio:,.0f}", font=("Segoe UI", 16, "bold"),
                 bg=COLOR_WHITE, fg=color_eq).grid(row=0, column=1, sticky="w")

        tk.Label(frame_eq, text="Estado:", font=("Segoe UI", 12, "bold"),
                 bg=COLOR_WHITE, anchor="e").grid(row=1, column=0, sticky="e", padx=(0, 15))
        tk.Label(frame_eq, text=estado, font=("Segoe UI", 18, "bold"),
                 bg=COLOR_WHITE, fg=color_eq).grid(row=1, column=1, sticky="w")

        # Ejecucion comparativa
        pct_recaudo = (resumen["recaudado"] / resumen["ppto_ingresos"] * 100) if resumen["ppto_ingresos"] else 0
        pct_gasto = (resumen["pagado"] / resumen["apropiacion"] * 100) if resumen["apropiacion"] else 0

        tk.Label(frame_eq, text=f"Ejecucion Ingresos: {pct_recaudo:.1f}%",
                 font=("Segoe UI", 11), bg=COLOR_WHITE, fg="#006633").grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))
        tk.Label(frame_eq, text=f"Ejecucion Gastos: {pct_gasto:.1f}%",
                 font=("Segoe UI", 11), bg=COLOR_WHITE, fg=COLOR_PRIMARY).grid(row=3, column=0, columnspan=2, sticky="w")

    # ===================== MODIFICACIONES PRESUPUESTALES =====================
    def _registrar_adicion(self):
        """Registra una adicion presupuestal (aumenta gasto e ingreso)."""
        messagebox.showinfo(
            "Adicion Presupuestal",
            "ADICION PRESUPUESTAL\n\n"
            "Una adicion incrementa tanto el presupuesto de gastos como el de ingresos.\n"
            "Paso 1: Seleccione el rubro de GASTO a adicionar.\n"
            "Paso 2: Seleccione el rubro de INGRESO correspondiente.\n"
            "Paso 3: Ingrese el valor y datos del acto administrativo."
        )

        # Paso 1: Rubro de gasto
        codigo_gasto = self._buscar_rubro()
        if not codigo_gasto:
            return
        rubro_g = db.get_rubro_gasto(codigo_gasto)

        # Paso 2: Rubro de ingreso
        codigo_ingreso = self._buscar_rubro_ingreso()
        if not codigo_ingreso:
            return
        rubro_i = db.get_rubro_ingreso(codigo_ingreso)

        # Paso 3: Valor
        valor_str = simpledialog.askstring(
            "Adicion - Valor",
            f"Rubro Gasto: {codigo_gasto} - {rubro_g['cuenta']}\n"
            f"Rubro Ingreso: {codigo_ingreso} - {rubro_i['cuenta']}\n\n"
            f"Ingrese el VALOR de la adicion:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor numerico invalido")
            return
        if valor <= 0:
            messagebox.showerror("Error", "El valor debe ser mayor a cero")
            return

        # Paso 4: Numero de acto
        numero_acto = simpledialog.askstring(
            "Adicion - Acto Administrativo",
            "Ingrese el numero del acto administrativo\n(Acuerdo, Resolucion, etc.):",
            parent=self
        )
        if not numero_acto:
            numero_acto = ""

        # Paso 5: Descripcion
        descripcion = simpledialog.askstring(
            "Adicion - Descripcion",
            "Descripcion de la adicion (opcional):",
            parent=self
        )
        if not descripcion:
            descripcion = ""

        # Paso 6: Confirmacion
        nueva_aprop_g = rubro_g["apropiacion_definitiva"] + valor
        nuevo_ppto_i = rubro_i["presupuesto_definitivo"] + valor
        if not messagebox.askyesno(
            "Confirmar Adicion",
            f"CONFIRMAR ADICION PRESUPUESTAL\n\n"
            f"Rubro Gasto: {codigo_gasto} - {rubro_g['cuenta']}\n"
            f"  Aprop. Definitiva actual: $ {rubro_g['apropiacion_definitiva']:,.0f}\n"
            f"  Nueva Aprop. Definitiva:  $ {nueva_aprop_g:,.0f}\n\n"
            f"Rubro Ingreso: {codigo_ingreso} - {rubro_i['cuenta']}\n"
            f"  Ppto. Definitivo actual: $ {rubro_i['presupuesto_definitivo']:,.0f}\n"
            f"  Nuevo Ppto. Definitivo:  $ {nuevo_ppto_i:,.0f}\n\n"
            f"Valor: $ {valor:,.0f}\n"
            f"Acto: {numero_acto}\n"
            f"Descripcion: {descripcion}"
        ):
            return

        try:
            num, fecha = db.registrar_adicion(codigo_gasto, codigo_ingreso, valor, numero_acto, descripcion)
            messagebox.showinfo(
                "Adicion Registrada",
                f"ADICION REGISTRADA EXITOSAMENTE\n\n"
                f"No. Modificacion: {num}\n"
                f"Fecha: {fecha}\n"
                f"Valor: $ {valor:,.0f}\n"
                f"Rubro Gasto: {codigo_gasto}\n"
                f"Rubro Ingreso: {codigo_ingreso}"
            )
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _registrar_reduccion(self):
        """Registra una reduccion presupuestal (disminuye gasto e ingreso)."""
        messagebox.showinfo(
            "Reduccion Presupuestal",
            "REDUCCION PRESUPUESTAL\n\n"
            "Una reduccion disminuye tanto el presupuesto de gastos como el de ingresos.\n"
            "El valor no puede exceder el saldo disponible del rubro.\n\n"
            "Paso 1: Seleccione el rubro de GASTO a reducir.\n"
            "Paso 2: Seleccione el rubro de INGRESO correspondiente.\n"
            "Paso 3: Ingrese el valor y datos del acto administrativo."
        )

        # Paso 1: Rubro de gasto
        codigo_gasto = self._buscar_rubro()
        if not codigo_gasto:
            return
        rubro_g = db.get_rubro_gasto(codigo_gasto)
        saldo_g = db.saldo_disponible_rubro(codigo_gasto)

        if saldo_g <= 0:
            messagebox.showerror("Sin Saldo",
                                 f"El rubro {codigo_gasto} no tiene saldo disponible.\n"
                                 f"Saldo: $ {saldo_g:,.0f}")
            return

        # Paso 2: Rubro de ingreso
        codigo_ingreso = self._buscar_rubro_ingreso()
        if not codigo_ingreso:
            return
        rubro_i = db.get_rubro_ingreso(codigo_ingreso)

        # Paso 3: Valor
        valor_str = simpledialog.askstring(
            "Reduccion - Valor",
            f"Rubro Gasto: {codigo_gasto} - {rubro_g['cuenta']}\n"
            f"  Saldo disponible: $ {saldo_g:,.0f}\n\n"
            f"Rubro Ingreso: {codigo_ingreso} - {rubro_i['cuenta']}\n\n"
            f"Ingrese el VALOR de la reduccion:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor numerico invalido")
            return
        if valor <= 0:
            messagebox.showerror("Error", "El valor debe ser mayor a cero")
            return
        if valor > saldo_g:
            messagebox.showerror("Error",
                                 f"Valor ($ {valor:,.0f}) excede saldo disponible ($ {saldo_g:,.0f})")
            return

        # Paso 4: Numero de acto
        numero_acto = simpledialog.askstring(
            "Reduccion - Acto Administrativo",
            "Ingrese el numero del acto administrativo\n(Acuerdo, Resolucion, etc.):",
            parent=self
        )
        if not numero_acto:
            numero_acto = ""

        # Paso 5: Descripcion
        descripcion = simpledialog.askstring(
            "Reduccion - Descripcion",
            "Descripcion de la reduccion (opcional):",
            parent=self
        )
        if not descripcion:
            descripcion = ""

        # Paso 6: Confirmacion
        nueva_aprop_g = rubro_g["apropiacion_definitiva"] - valor
        nuevo_ppto_i = rubro_i["presupuesto_definitivo"] - valor
        if not messagebox.askyesno(
            "Confirmar Reduccion",
            f"CONFIRMAR REDUCCION PRESUPUESTAL\n\n"
            f"Rubro Gasto: {codigo_gasto} - {rubro_g['cuenta']}\n"
            f"  Aprop. Definitiva actual: $ {rubro_g['apropiacion_definitiva']:,.0f}\n"
            f"  Nueva Aprop. Definitiva:  $ {nueva_aprop_g:,.0f}\n\n"
            f"Rubro Ingreso: {codigo_ingreso} - {rubro_i['cuenta']}\n"
            f"  Ppto. Definitivo actual: $ {rubro_i['presupuesto_definitivo']:,.0f}\n"
            f"  Nuevo Ppto. Definitivo:  $ {nuevo_ppto_i:,.0f}\n\n"
            f"Valor: $ {valor:,.0f}\n"
            f"Acto: {numero_acto}\n"
            f"Descripcion: {descripcion}"
        ):
            return

        try:
            num, fecha = db.registrar_reduccion(codigo_gasto, codigo_ingreso, valor, numero_acto, descripcion)
            messagebox.showinfo(
                "Reduccion Registrada",
                f"REDUCCION REGISTRADA EXITOSAMENTE\n\n"
                f"No. Modificacion: {num}\n"
                f"Fecha: {fecha}\n"
                f"Valor: $ {valor:,.0f}\n"
                f"Rubro Gasto: {codigo_gasto}\n"
                f"Rubro Ingreso: {codigo_ingreso}"
            )
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _registrar_credito_contracredito(self):
        """Registra un traslado credito/contracredito entre rubros de gasto."""
        messagebox.showinfo(
            "Credito / Contracredito",
            "CREDITO / CONTRACREDITO\n\n"
            "Un credito/contracredito traslada recursos entre rubros de GASTOS.\n"
            "El rubro CREDITO recibe los recursos (aumenta).\n"
            "El rubro CONTRACREDITO cede los recursos (disminuye).\n"
            "No afecta ingresos ni el total de gastos.\n\n"
            "Paso 1: Seleccione el rubro que RECIBE (credito).\n"
            "Paso 2: Seleccione el rubro que CEDE (contracredito).\n"
            "Paso 3: Ingrese el valor y datos del acto."
        )

        # Paso 1: Rubro credito (recibe)
        messagebox.showinfo("Paso 1", "Seleccione el rubro que RECIBE los recursos (CREDITO)")
        codigo_credito = self._buscar_rubro()
        if not codigo_credito:
            return
        rubro_cred = db.get_rubro_gasto(codigo_credito)

        # Paso 2: Rubro contracredito (cede)
        messagebox.showinfo("Paso 2", "Seleccione el rubro que CEDE los recursos (CONTRACREDITO)")
        codigo_contracredito = self._buscar_rubro()
        if not codigo_contracredito:
            return
        if codigo_credito == codigo_contracredito:
            messagebox.showerror("Error", "Los rubros de credito y contracredito deben ser diferentes")
            return
        rubro_ccred = db.get_rubro_gasto(codigo_contracredito)
        saldo_ccred = db.saldo_disponible_rubro(codigo_contracredito)

        if saldo_ccred <= 0:
            messagebox.showerror("Sin Saldo",
                                 f"El rubro {codigo_contracredito} no tiene saldo disponible.\n"
                                 f"Saldo: $ {saldo_ccred:,.0f}")
            return

        # Paso 3: Valor
        valor_str = simpledialog.askstring(
            "Credito/Contracredito - Valor",
            f"Rubro CREDITO (recibe): {codigo_credito} - {rubro_cred['cuenta']}\n"
            f"  Aprop. actual: $ {rubro_cred['apropiacion_definitiva']:,.0f}\n\n"
            f"Rubro CONTRACREDITO (cede): {codigo_contracredito} - {rubro_ccred['cuenta']}\n"
            f"  Aprop. actual: $ {rubro_ccred['apropiacion_definitiva']:,.0f}\n"
            f"  Saldo disponible: $ {saldo_ccred:,.0f}\n\n"
            f"Ingrese el VALOR del traslado:",
            parent=self
        )
        if not valor_str:
            return
        try:
            valor = float(valor_str.replace(",", "").replace(".", "").strip())
        except ValueError:
            messagebox.showerror("Error", "Valor numerico invalido")
            return
        if valor <= 0:
            messagebox.showerror("Error", "El valor debe ser mayor a cero")
            return
        if valor > saldo_ccred:
            messagebox.showerror("Error",
                                 f"Valor ($ {valor:,.0f}) excede saldo disponible del contracredito ($ {saldo_ccred:,.0f})")
            return

        # Paso 4: Numero de acto
        numero_acto = simpledialog.askstring(
            "Credito/Contracredito - Acto Administrativo",
            "Ingrese el numero del acto administrativo\n(Acuerdo, Resolucion, etc.):",
            parent=self
        )
        if not numero_acto:
            numero_acto = ""

        # Paso 5: Descripcion
        descripcion = simpledialog.askstring(
            "Credito/Contracredito - Descripcion",
            "Descripcion del traslado (opcional):",
            parent=self
        )
        if not descripcion:
            descripcion = ""

        # Paso 6: Confirmacion
        nueva_aprop_cred = rubro_cred["apropiacion_definitiva"] + valor
        nueva_aprop_ccred = rubro_ccred["apropiacion_definitiva"] - valor
        if not messagebox.askyesno(
            "Confirmar Credito/Contracredito",
            f"CONFIRMAR CREDITO / CONTRACREDITO\n\n"
            f"Rubro CREDITO (recibe): {codigo_credito} - {rubro_cred['cuenta']}\n"
            f"  Aprop. actual: $ {rubro_cred['apropiacion_definitiva']:,.0f}\n"
            f"  Nueva Aprop.:  $ {nueva_aprop_cred:,.0f}\n\n"
            f"Rubro CONTRACREDITO (cede): {codigo_contracredito} - {rubro_ccred['cuenta']}\n"
            f"  Aprop. actual: $ {rubro_ccred['apropiacion_definitiva']:,.0f}\n"
            f"  Nueva Aprop.:  $ {nueva_aprop_ccred:,.0f}\n\n"
            f"Valor del traslado: $ {valor:,.0f}\n"
            f"Acto: {numero_acto}\n"
            f"Descripcion: {descripcion}"
        ):
            return

        try:
            num, fecha = db.registrar_credito_contracredito(
                codigo_credito, codigo_contracredito, valor, numero_acto, descripcion
            )
            messagebox.showinfo(
                "Credito/Contracredito Registrado",
                f"CREDITO/CONTRACREDITO REGISTRADO EXITOSAMENTE\n\n"
                f"No. Modificacion: {num}\n"
                f"Fecha: {fecha}\n"
                f"Valor: $ {valor:,.0f}\n"
                f"Rubro Credito: {codigo_credito}\n"
                f"Rubro Contracredito: {codigo_contracredito}"
            )
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _ver_modificaciones(self):
        """Vista de historial de modificaciones presupuestales."""
        self._limpiar_main()
        f = self.main_frame

        tk.Label(f, text="MODIFICACIONES PRESUPUESTALES", font=("Segoe UI", 18, "bold"),
                 bg=COLOR_BG, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        # Boton volver
        tk.Button(f, text="Volver al Dashboard", bg=COLOR_ACCENT, fg=COLOR_WHITE,
                  font=("Segoe UI", 9, "bold"), cursor="hand2",
                  command=self._mostrar_dashboard).pack(anchor="w", padx=15, pady=5)

        modificaciones = db.listar_modificaciones()
        if not modificaciones:
            tk.Label(f, text="No hay modificaciones presupuestales registradas.",
                     font=("Segoe UI", 12), bg=COLOR_BG, fg="#666").pack(pady=30)
            return

        # Treeview
        frame_tree = tk.Frame(f, bg=COLOR_BG)
        frame_tree.pack(fill="both", expand=True, padx=15, pady=5)

        cols = ("no", "fecha", "tipo", "acto", "valor", "descripcion", "estado")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=20)
        tree.heading("no", text="No.")
        tree.heading("fecha", text="Fecha")
        tree.heading("tipo", text="Tipo")
        tree.heading("acto", text="Acto Adm.")
        tree.heading("valor", text="Valor")
        tree.heading("descripcion", text="Descripcion")
        tree.heading("estado", text="Estado")
        tree.column("no", width=50, anchor="center")
        tree.column("fecha", width=90, anchor="center")
        tree.column("tipo", width=160, anchor="center")
        tree.column("acto", width=120)
        tree.column("valor", width=120, anchor="e")
        tree.column("descripcion", width=250)
        tree.column("estado", width=80, anchor="center")

        scrollbar = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for m in modificaciones:
            tipo_display = m["tipo"].replace("_", " / ")
            tree.insert("", "end", iid=str(m["id"]), values=(
                m["id"], m["fecha"], tipo_display, m["numero_acto"],
                f"$ {m['valor']:,.0f}", m["descripcion"], m["estado"]
            ))

        def ver_detalle(event):
            sel = tree.selection()
            if not sel:
                return
            id_mod = int(sel[0])
            mod = db.get_modificacion(id_mod)
            if not mod:
                return

            tipo_display = mod["tipo"].replace("_", " / ")
            detalle_msg = (
                f"MODIFICACION No. {mod['id']}\n\n"
                f"Fecha: {mod['fecha']}\n"
                f"Tipo: {tipo_display}\n"
                f"Acto Administrativo: {mod['numero_acto']}\n"
                f"Valor: $ {mod['valor']:,.0f}\n"
                f"Descripcion: {mod['descripcion']}\n"
                f"Estado: {mod['estado']}\n\n"
                f"RUBROS AFECTADOS:\n"
            )
            for d in mod.get("detalles", []):
                nombre = d.get("nombre_rubro", "")
                detalle_msg += (
                    f"  - {d['tipo_rubro']}: {d['codigo_rubro']} - {nombre}\n"
                    f"    Campo: {d['campo_afectado']} | Valor: $ {d['valor']:,.0f}\n"
                )

            messagebox.showinfo("Detalle de Modificacion", detalle_msg)

        tree.bind("<Double-1>", ver_detalle)

        tk.Label(f, text="Doble-click en una fila para ver detalle con rubros afectados",
                 font=("Segoe UI", 9, "italic"), bg=COLOR_BG, fg="#666").pack(pady=5)

    # ===================== COPIAS DE SEGURIDAD =====================
    def _crear_backup(self):
        destino = filedialog.askdirectory(
            title="Seleccionar carpeta para guardar la copia de seguridad",
            initialdir=os.path.dirname(os.path.dirname(__file__))
        )
        if not destino:
            return
        try:
            ruta = db.crear_backup(destino)
            tamano = os.path.getsize(ruta)
            tamano_mb = tamano / (1024 * 1024)
            messagebox.showinfo(
                "Copia de Seguridad",
                f"Copia creada exitosamente.\n\n"
                f"Archivo: {ruta}\n"
                f"Tamano: {tamano_mb:.2f} MB"
            )
            self._mostrar_dashboard()
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear copia de seguridad:\n{e}")

    def _restaurar_backup(self):
        ok = messagebox.askokcancel(
            "Restaurar Copia de Seguridad",
            "ADVERTENCIA: Esta accion reemplazara TODOS los datos actuales\n"
            "con los datos del archivo de respaldo seleccionado.\n\n"
            "Desea continuar?"
        )
        if not ok:
            return
        origen = filedialog.askopenfilename(
            title="Seleccionar archivo de respaldo",
            filetypes=[("Base de datos", "*.db"), ("Todos", "*.*")],
            initialdir=os.path.dirname(os.path.dirname(__file__))
        )
        if not origen:
            return
        tamano = os.path.getsize(origen)
        tamano_mb = tamano / (1024 * 1024)
        fecha_mod = datetime.fromtimestamp(os.path.getmtime(origen)).strftime("%Y-%m-%d %H:%M")
        confirmar = messagebox.askyesno(
            "Confirmar Restauracion",
            f"Archivo seleccionado:\n{origen}\n\n"
            f"Tamano: {tamano_mb:.2f} MB\n"
            f"Fecha de modificacion: {fecha_mod}\n\n"
            f"Desea restaurar este archivo?\n"
            f"Los datos actuales seran reemplazados."
        )
        if not confirmar:
            return
        exito = db.restaurar_backup(origen)
        if exito:
            messagebox.showinfo("Restauracion Exitosa",
                                "La base de datos fue restaurada correctamente.\n"
                                "El sistema se reiniciara.")
            self._reiniciar_app()
        else:
            messagebox.showerror("Error", "No se pudo restaurar el archivo seleccionado.\n"
                                 "Verifique que sea un archivo de base de datos valido.")

    # ===================== PAC =====================
    def _gestionar_pac(self):
        self._limpiar_main()
        f = self.main_frame

        tk.Label(f, text="PLAN ANUALIZADO DE CAJA (PAC)", font=("Segoe UI", 16, "bold"),
                 bg=COLOR_BG, fg=COLOR_PRIMARY).pack(pady=(10, 5))
        tk.Label(f, text="Distribucion mensual de la apropiacion para control de pagos",
                 font=("Segoe UI", 10), bg=COLOR_BG, fg="#666").pack(pady=(0, 10))

        # Frame principal con dos paneles
        frame_principal = tk.Frame(f, bg=COLOR_BG)
        frame_principal.pack(fill="both", expand=True, padx=10)

        # === Panel izquierdo: lista de rubros ===
        frame_izq = tk.LabelFrame(frame_principal, text=" Rubros de Gasto ",
                                   font=("Segoe UI", 10, "bold"),
                                   bg=COLOR_WHITE, fg=COLOR_PRIMARY, padx=5, pady=5)
        frame_izq.pack(side="left", fill="y", padx=(0, 5))

        # Buscador
        tk.Label(frame_izq, text="Buscar:", font=("Segoe UI", 9),
                 bg=COLOR_WHITE).pack(anchor="w")
        entry_buscar = tk.Entry(frame_izq, font=("Segoe UI", 10), width=30)
        entry_buscar.pack(fill="x", pady=(0, 5))

        # Lista de rubros
        frame_lista = tk.Frame(frame_izq, bg=COLOR_WHITE)
        frame_lista.pack(fill="both", expand=True)

        cols_rubros = ("codigo", "cuenta", "apropiacion")
        tree_rubros = ttk.Treeview(frame_lista, columns=cols_rubros, show="headings", height=18)
        tree_rubros.heading("codigo", text="Codigo")
        tree_rubros.heading("cuenta", text="Cuenta")
        tree_rubros.heading("apropiacion", text="Aprop. Def.")
        tree_rubros.column("codigo", width=140)
        tree_rubros.column("cuenta", width=180)
        tree_rubros.column("apropiacion", width=100, anchor="e")

        sb_rubros = ttk.Scrollbar(frame_lista, orient="vertical", command=tree_rubros.yview)
        tree_rubros.configure(yscrollcommand=sb_rubros.set)
        tree_rubros.pack(side="left", fill="both", expand=True)
        sb_rubros.pack(side="right", fill="y")

        # === Panel derecho: grilla PAC ===
        frame_der = tk.LabelFrame(frame_principal, text=" Distribucion Mensual PAC ",
                                   font=("Segoe UI", 10, "bold"),
                                   bg=COLOR_WHITE, fg=COLOR_PRIMARY, padx=10, pady=5)
        frame_der.pack(side="left", fill="both", expand=True)

        lbl_rubro_sel = tk.Label(frame_der, text="Seleccione un rubro de la lista",
                                  font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#999")
        lbl_rubro_sel.pack(anchor="w", pady=(0, 5))

        lbl_aprop = tk.Label(frame_der, text="", font=("Segoe UI", 10, "bold"),
                              bg=COLOR_WHITE, fg=COLOR_PRIMARY)
        lbl_aprop.pack(anchor="w")

        # Grilla de meses
        frame_grilla = tk.Frame(frame_der, bg=COLOR_WHITE)
        frame_grilla.pack(fill="x", pady=5)

        # Encabezados
        tk.Label(frame_grilla, text="Mes", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT, width=12, anchor="w").grid(row=0, column=0, padx=1, pady=1)
        tk.Label(frame_grilla, text="Valor Programado", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT, width=16, anchor="e").grid(row=0, column=1, padx=1, pady=1)
        tk.Label(frame_grilla, text="Pagos Ejecutados", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT, width=16, anchor="e").grid(row=0, column=2, padx=1, pady=1)
        tk.Label(frame_grilla, text="Saldo Disponible", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT, width=16, anchor="e").grid(row=0, column=3, padx=1, pady=1)

        entries_pac = []
        lbls_pagos = []
        lbls_saldo = []
        for i in range(12):
            mes_num = i + 1
            tk.Label(frame_grilla, text=MESES[mes_num], font=("Segoe UI", 9),
                     bg=COLOR_WHITE, width=12, anchor="w").grid(row=i+1, column=0, padx=1, pady=1)
            e = tk.Entry(frame_grilla, font=("Segoe UI", 9), width=16, justify="right")
            e.grid(row=i+1, column=1, padx=1, pady=1)
            e.insert(0, "0")
            entries_pac.append(e)
            lp = tk.Label(frame_grilla, text="$ 0", font=("Segoe UI", 9),
                          bg=COLOR_WHITE, width=16, anchor="e")
            lp.grid(row=i+1, column=2, padx=1, pady=1)
            lbls_pagos.append(lp)
            ls = tk.Label(frame_grilla, text="$ 0", font=("Segoe UI", 9),
                          bg=COLOR_WHITE, width=16, anchor="e")
            ls.grid(row=i+1, column=3, padx=1, pady=1)
            lbls_saldo.append(ls)

        # Fila total
        tk.Label(frame_grilla, text="TOTAL", font=("Segoe UI", 9, "bold"),
                 bg=COLOR_LIGHT, width=12, anchor="w").grid(row=13, column=0, padx=1, pady=1)
        lbl_total_pac = tk.Label(frame_grilla, text="$ 0", font=("Segoe UI", 9, "bold"),
                                  bg=COLOR_LIGHT, width=16, anchor="e")
        lbl_total_pac.grid(row=13, column=1, padx=1, pady=1)
        lbl_total_pagos = tk.Label(frame_grilla, text="$ 0", font=("Segoe UI", 9, "bold"),
                                    bg=COLOR_LIGHT, width=16, anchor="e")
        lbl_total_pagos.grid(row=13, column=2, padx=1, pady=1)
        lbl_total_saldo = tk.Label(frame_grilla, text="$ 0", font=("Segoe UI", 9, "bold"),
                                    bg=COLOR_LIGHT, width=16, anchor="e")
        lbl_total_saldo.grid(row=13, column=3, padx=1, pady=1)

        # Advertencia
        lbl_advertencia = tk.Label(frame_der, text="", font=("Segoe UI", 9, "bold"),
                                    bg=COLOR_WHITE, fg=COLOR_DANGER)
        lbl_advertencia.pack(anchor="w")

        # Estado del rubro seleccionado
        estado = {"codigo": None}

        def cargar_rubros(filtro=""):
            tree_rubros.delete(*tree_rubros.get_children())
            if filtro:
                rubros = db.buscar_rubros_gastos(filtro)
            else:
                rubros = db.get_rubros_gastos(solo_hojas=True)
            for r in rubros:
                tree_rubros.insert("", "end", values=(
                    r["codigo"], r["cuenta"], f"{r['apropiacion_definitiva']:,.0f}"
                ))

        def actualizar_totales(*args):
            total_p = 0
            total_pg = 0
            codigo = estado["codigo"]
            for i in range(12):
                try:
                    val = float(entries_pac[i].get().replace(",", "").replace("$", "").strip())
                except ValueError:
                    val = 0
                total_p += val
                pagos = db.get_pagos_mes_rubro(codigo, i + 1) if codigo else 0
                saldo = val - pagos
                total_pg += pagos
                lbls_pagos[i].config(text=f"$ {pagos:,.0f}")
                if saldo < 0:
                    lbls_saldo[i].config(text=f"$ {saldo:,.0f}", fg=COLOR_DANGER)
                else:
                    lbls_saldo[i].config(text=f"$ {saldo:,.0f}", fg=COLOR_SUCCESS)
            lbl_total_pac.config(text=f"$ {total_p:,.0f}")
            lbl_total_pagos.config(text=f"$ {total_pg:,.0f}")
            lbl_total_saldo.config(text=f"$ {total_p - total_pg:,.0f}")
            # Advertencia si no cuadra con apropiacion
            if codigo:
                rubro = db.get_rubro_gasto(codigo)
                if rubro:
                    dif = total_p - rubro["apropiacion_definitiva"]
                    if abs(dif) > 0.01:
                        lbl_advertencia.config(
                            text=f"ADVERTENCIA: La suma del PAC ($ {total_p:,.0f}) difiere "
                                 f"de la apropiacion definitiva ($ {rubro['apropiacion_definitiva']:,.0f}) "
                                 f"en $ {dif:,.0f}")
                    else:
                        lbl_advertencia.config(text="PAC cuadra con la apropiacion definitiva",
                                               fg=COLOR_SUCCESS)

        def seleccionar_rubro(event):
            sel = tree_rubros.selection()
            if not sel:
                return
            vals = tree_rubros.item(sel[0])["values"]
            codigo = str(vals[0])
            estado["codigo"] = codigo
            rubro = db.get_rubro_gasto(codigo)
            lbl_rubro_sel.config(text=f"{codigo} - {rubro['cuenta']}", fg=COLOR_PRIMARY)
            lbl_aprop.config(text=f"Apropiacion Definitiva: $ {rubro['apropiacion_definitiva']:,.0f}")
            # Cargar PAC existente
            pac = db.get_pac(codigo)
            for i in range(12):
                entries_pac[i].delete(0, "end")
                if pac and i < len(pac):
                    entries_pac[i].insert(0, f"{pac[i]['valor_programado']:,.0f}")
                else:
                    entries_pac[i].insert(0, "0")
            actualizar_totales()

        def distribuir_uniforme():
            codigo = estado["codigo"]
            if not codigo:
                messagebox.showwarning("PAC", "Seleccione un rubro primero")
                return
            valores = db.distribuir_pac_uniforme(codigo)
            for i in range(12):
                entries_pac[i].delete(0, "end")
                entries_pac[i].insert(0, f"{valores[i]:,.0f}")
            actualizar_totales()
            messagebox.showinfo("PAC", "Apropiacion distribuida uniformemente en 12 meses")

        def guardar_pac():
            codigo = estado["codigo"]
            if not codigo:
                messagebox.showwarning("PAC", "Seleccione un rubro primero")
                return
            valores = []
            for i in range(12):
                try:
                    val = float(entries_pac[i].get().replace(",", "").replace("$", "").strip())
                except ValueError:
                    val = 0
                valores.append(val)
            db.set_pac_completo(codigo, valores)
            actualizar_totales()
            messagebox.showinfo("PAC", f"PAC guardado para rubro {codigo}")

        # Bindings
        entry_buscar.bind("<KeyRelease>", lambda e: cargar_rubros(entry_buscar.get().strip()))
        tree_rubros.bind("<<TreeviewSelect>>", seleccionar_rubro)
        for e in entries_pac:
            e.bind("<FocusOut>", actualizar_totales)

        # Botones
        frame_btns = tk.Frame(frame_der, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", pady=10)
        tk.Button(frame_btns, text="Distribuir Uniforme", bg="#8B4513", fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), cursor="hand2",
                  command=distribuir_uniforme).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Guardar PAC", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), cursor="hand2",
                  command=guardar_pac).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Volver al Inicio", bg="#555555", fg=COLOR_WHITE,
                  font=("Segoe UI", 10), cursor="hand2",
                  command=self._mostrar_dashboard).pack(side="right", padx=5)

        # Cargar rubros
        cargar_rubros()

    # ===================== INFORME SIFSE =====================
    def _generar_informe_sifse(self):
        """Pantalla para generar el informe SIFSE trimestral."""
        dialog = tk.Toplevel(self)
        dialog.title("Informe SIFSE - Trimestral")
        dialog.geometry("950x650")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="INFORME SIFSE - SISTEMA DE INFORMACION FSE",
                 font=("Segoe UI", 14, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))

        # Filtros
        frame_filtro = tk.Frame(dialog, bg=COLOR_LIGHT, padx=10, pady=8)
        frame_filtro.pack(fill="x", padx=15)

        tk.Label(frame_filtro, text="Trimestre:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=5)
        trim_var = tk.StringVar(value="1")
        ttk.Combobox(frame_filtro, textvariable=trim_var,
                     values=["1", "2", "3", "4"], state="readonly", width=5).pack(side="left")

        tk.Label(frame_filtro, text="Ano:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=(15, 5))
        anio_var = tk.StringVar(value=db.get_config("vigencia") or "2026")
        ttk.Combobox(frame_filtro, textvariable=anio_var,
                     values=["2025", "2026", "2027"], state="readonly", width=8).pack(side="left")

        dane = db.get_config("codigo_dane") or "186755000015"
        tk.Label(frame_filtro, text=f"DANE: {dane}", font=("Segoe UI", 9),
                 bg=COLOR_LIGHT, fg="#555").pack(side="right", padx=10)

        # Notebook con pestanas de preview
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True, padx=15, pady=5)

        # Pestana Ingresos
        frame_ing = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_ing, text="Ingresos Presupuestales")

        cols_ing = ("Fuente", "Descripcion", "Ppto Inicial", "Ppto Definitivo", "Recaudado")
        tree_ing = ttk.Treeview(frame_ing, columns=cols_ing, show="headings", height=10)
        for c in cols_ing:
            tree_ing.heading(c, text=c)
        tree_ing.column("Fuente", width=60, anchor="center")
        tree_ing.column("Descripcion", width=300)
        tree_ing.column("Ppto Inicial", width=120, anchor="e")
        tree_ing.column("Ppto Definitivo", width=120, anchor="e")
        tree_ing.column("Recaudado", width=120, anchor="e")
        vsb_ing = ttk.Scrollbar(frame_ing, orient="vertical", command=tree_ing.yview)
        tree_ing.configure(yscrollcommand=vsb_ing.set)
        tree_ing.pack(side="left", fill="both", expand=True)
        vsb_ing.pack(side="right", fill="y")

        # Pestana Gastos
        frame_gas = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_gas, text="Gastos Presupuestales")

        cols_gas = ("Fuente", "Item", "Desc Item", "Ppto Inicial", "Ppto Definitivo",
                    "Compromisos", "Obligaciones", "Pagos")
        tree_gas = ttk.Treeview(frame_gas, columns=cols_gas, show="headings", height=10)
        for c in cols_gas:
            tree_gas.heading(c, text=c)
        tree_gas.column("Fuente", width=55, anchor="center")
        tree_gas.column("Item", width=45, anchor="center")
        tree_gas.column("Desc Item", width=200)
        tree_gas.column("Ppto Inicial", width=100, anchor="e")
        tree_gas.column("Ppto Definitivo", width=100, anchor="e")
        tree_gas.column("Compromisos", width=100, anchor="e")
        tree_gas.column("Obligaciones", width=100, anchor="e")
        tree_gas.column("Pagos", width=100, anchor="e")
        vsb_gas = ttk.Scrollbar(frame_gas, orient="vertical", command=tree_gas.yview)
        tree_gas.configure(yscrollcommand=vsb_gas.set)
        hsb_gas = ttk.Scrollbar(frame_gas, orient="horizontal", command=tree_gas.xview)
        tree_gas.configure(xscrollcommand=hsb_gas.set)
        hsb_gas.pack(side="bottom", fill="x")
        tree_gas.pack(side="left", fill="both", expand=True)
        vsb_gas.pack(side="right", fill="y")

        lbl_advertencia = tk.Label(dialog, text="", font=("Segoe UI", 9), bg=COLOR_WHITE, fg=COLOR_WARNING)
        lbl_advertencia.pack(anchor="w", padx=15)

        def cargar_preview():
            trimestre = int(trim_var.get())
            anio = anio_var.get()

            # Ingresos
            tree_ing.delete(*tree_ing.get_children())
            datos_ing = db.informe_sifse_ingresos(trimestre, anio)
            sin_asignar_ing = False
            for d in datos_ing:
                if d["fuente"] == 0:
                    sin_asignar_ing = True
                tree_ing.insert("", "end", values=(
                    d["fuente"] if d["fuente"] else "N/A",
                    d["desc_fuente"],
                    f"{d['ppto_inicial']:,.0f}",
                    f"{d['ppto_definitivo']:,.0f}",
                    f"{d['recaudado']:,.0f}",
                ))

            # Gastos
            tree_gas.delete(*tree_gas.get_children())
            datos_gas = db.informe_sifse_gastos(trimestre, anio)
            sin_asignar_gas = False
            for d in datos_gas:
                if d["fuente"] == 0 or d["item"] == 0:
                    sin_asignar_gas = True
                tree_gas.insert("", "end", values=(
                    d["fuente"] if d["fuente"] else "N/A",
                    d["item"] if d["item"] else "N/A",
                    d["desc_item"],
                    f"{d['ppto_inicial']:,.0f}",
                    f"{d['ppto_definitivo']:,.0f}",
                    f"{d['compromisos']:,.0f}",
                    f"{d['obligaciones']:,.0f}",
                    f"{d['pagos']:,.0f}",
                ))

            advertencias = []
            if sin_asignar_ing:
                advertencias.append("Hay rubros de ingreso sin mapeo SIFSE")
            if sin_asignar_gas:
                advertencias.append("Hay rubros/CDPs de gasto sin asignacion SIFSE")
            if advertencias:
                lbl_advertencia.config(text="ADVERTENCIA: " + " | ".join(advertencias))
            else:
                lbl_advertencia.config(text="")

        def exportar():
            trimestre = int(trim_var.get())
            anio = anio_var.get()
            destino = filedialog.asksaveasfilename(
                title="Guardar archivo SIFSE",
                defaultextension=".xls",
                initialfile=f"SIFSE_T{trimestre}_{anio}.xls",
                filetypes=[("Excel 97-2003", "*.xls")],
                parent=dialog
            )
            if not destino:
                return
            try:
                archivo = db.exportar_sifse_xls(trimestre, anio, destino)
                messagebox.showinfo("SIFSE Exportado",
                                    f"Archivo generado exitosamente:\n{archivo}",
                                    parent=dialog)
            except ImportError as e:
                messagebox.showerror("Error", str(e), parent=dialog)
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar:\n{e}", parent=dialog)

        # Botones
        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Vista Previa", bg=COLOR_ACCENT, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=cargar_preview).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Exportar .xls", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=exportar).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Cerrar", font=("Segoe UI", 10), width=10,
                  command=dialog.destroy).pack(side="right", padx=5)

        # Cargar preview inicial
        cargar_preview()

    # ===================== INFORME SIA CONTRALORIA =====================
    def _generar_informe_sia(self):
        """Pantalla para generar los informes SIA Contraloria (acumulado anual)."""
        dialog = tk.Toplevel(self)
        dialog.title("Informe SIA Contraloria - Acumulado Anual")
        dialog.geometry("1050x700")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="INFORME SIA - CONTRALORIA",
                 font=("Segoe UI", 14, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 2))
        tk.Label(dialog, text="ACUMULADO ANUAL - Consolidado de Enero hasta el mes seleccionado",
                 font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg=COLOR_WARNING).pack(pady=(0, 2))
        tk.Label(dialog, text="Formatos F03, F7B, F08A, F09 y F13A para rendicion de cuentas",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack(pady=(0, 5))

        # Filtros
        frame_filtro = tk.Frame(dialog, bg=COLOR_LIGHT, padx=10, pady=8)
        frame_filtro.pack(fill="x", padx=15)

        tk.Label(frame_filtro, text="Rendido hasta:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=5)
        mes_actual = int(db.get_config("mes_actual") or 1)
        mes_var = tk.StringVar(value=str(mes_actual))
        meses_lista = [str(i) for i in range(1, 13)]
        ttk.Combobox(frame_filtro, textvariable=mes_var,
                     values=meses_lista, state="readonly", width=5).pack(side="left")

        tk.Label(frame_filtro, text="Ano:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_LIGHT).pack(side="left", padx=(15, 5))
        anio_var = tk.StringVar(value=db.get_config("vigencia") or "2026")
        ttk.Combobox(frame_filtro, textvariable=anio_var,
                     values=["2025", "2026", "2027"], state="readonly", width=8).pack(side="left")

        nombres_mes = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                       5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
                       9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
        tk.Label(frame_filtro, text=f"Institucion: {db.get_config('institucion') or ''}",
                 font=("Segoe UI", 9), bg=COLOR_LIGHT, fg="#555").pack(side="right", padx=10)

        # Formato a generar
        frame_formato = tk.Frame(dialog, bg=COLOR_WHITE, padx=10, pady=5)
        frame_formato.pack(fill="x", padx=15)

        tk.Label(frame_formato, text="Formato:", font=("Segoe UI", 10, "bold"),
                 bg=COLOR_WHITE).pack(side="left", padx=5)
        formato_var = tk.StringVar(value="TODOS")
        formatos_opciones = ["TODOS", "F03 - Movimiento de Bancos", "F7B - Formato de Pagos",
                             "F08A_GASTOS - Modificaciones Gastos",
                             "F08A_INGRESOS - Modificaciones Ingresos",
                             "F09 - PAC", "F13A - Contratacion"]
        ttk.Combobox(frame_formato, textvariable=formato_var,
                     values=formatos_opciones, state="readonly", width=40).pack(side="left", padx=5)

        # Notebook con pestanas de preview
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True, padx=15, pady=5)

        # --- Pestana F03: Movimiento de Bancos ---
        frame_f03 = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_f03, text="F03 - Bancos")
        cols_f03 = ("Banco", "No. Cuenta", "Denominacion", "Fuente", "Saldo Inicial",
                    "Ingresos", "Egresos", "Notas Deb", "Notas Cred", "Saldo Libros", "Saldo Extractos")
        tree_f03 = ttk.Treeview(frame_f03, columns=cols_f03, show="headings", height=6)
        for c in cols_f03:
            tree_f03.heading(c, text=c)
        tree_f03.column("Banco", width=100)
        tree_f03.column("No. Cuenta", width=100)
        tree_f03.column("Denominacion", width=150)
        tree_f03.column("Fuente", width=130)
        for col in ("Saldo Inicial", "Ingresos", "Egresos", "Notas Deb", "Notas Cred", "Saldo Libros", "Saldo Extractos"):
            tree_f03.column(col, width=95, anchor="e")
        vsb_f03 = ttk.Scrollbar(frame_f03, orient="vertical", command=tree_f03.yview)
        tree_f03.configure(yscrollcommand=vsb_f03.set)
        tree_f03.pack(side="left", fill="both", expand=True)
        vsb_f03.pack(side="right", fill="y")

        # --- Pestana F7B: Pagos ---
        frame_f7b = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_f7b, text="F7B - Pagos")
        cols_f7b = ("Fecha", "Codigo", "No. RP", "Clase Pago", "Beneficiario",
                    "NIT", "Detalle", "Valor", "Neto Pagado")
        tree_f7b = ttk.Treeview(frame_f7b, columns=cols_f7b, show="headings", height=6)
        for c in cols_f7b:
            tree_f7b.heading(c, text=c)
        tree_f7b.column("Fecha", width=85, anchor="center")
        tree_f7b.column("Codigo", width=100)
        tree_f7b.column("No. RP", width=55, anchor="center")
        tree_f7b.column("Clase Pago", width=90)
        tree_f7b.column("Beneficiario", width=180)
        tree_f7b.column("NIT", width=90, anchor="center")
        tree_f7b.column("Detalle", width=180)
        tree_f7b.column("Valor", width=100, anchor="e")
        tree_f7b.column("Neto Pagado", width=100, anchor="e")
        vsb_f7b = ttk.Scrollbar(frame_f7b, orient="vertical", command=tree_f7b.yview)
        tree_f7b.configure(yscrollcommand=vsb_f7b.set)
        hsb_f7b = ttk.Scrollbar(frame_f7b, orient="horizontal", command=tree_f7b.xview)
        tree_f7b.configure(xscrollcommand=hsb_f7b.set)
        hsb_f7b.pack(side="bottom", fill="x")
        tree_f7b.pack(side="left", fill="both", expand=True)
        vsb_f7b.pack(side="right", fill="y")

        # --- Pestana F08A GASTOS: Modificaciones al Presupuesto de Gastos ---
        frame_f08a_gas = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_f08a_gas, text="F08A - Modif. Gastos")
        cols_f08a = ("Codigo Rubro", "Nombre Rubro", "Acto Administrativo", "Fecha", "Adicion", "Reduccion")
        tree_f08a_gas = ttk.Treeview(frame_f08a_gas, columns=cols_f08a, show="headings", height=6)
        for c in cols_f08a:
            tree_f08a_gas.heading(c, text=c)
        tree_f08a_gas.column("Codigo Rubro", width=150)
        tree_f08a_gas.column("Nombre Rubro", width=180)
        tree_f08a_gas.column("Acto Administrativo", width=150)
        tree_f08a_gas.column("Fecha", width=90, anchor="center")
        tree_f08a_gas.column("Adicion", width=120, anchor="e")
        tree_f08a_gas.column("Reduccion", width=120, anchor="e")
        vsb_f08a_gas = ttk.Scrollbar(frame_f08a_gas, orient="vertical", command=tree_f08a_gas.yview)
        tree_f08a_gas.configure(yscrollcommand=vsb_f08a_gas.set)
        tree_f08a_gas.pack(side="left", fill="both", expand=True)
        vsb_f08a_gas.pack(side="right", fill="y")

        # --- Pestana F08A INGRESOS: Modificaciones al Presupuesto de Ingresos ---
        frame_f08a_ing = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_f08a_ing, text="F08A - Modif. Ingresos")
        tree_f08a_ing = ttk.Treeview(frame_f08a_ing, columns=cols_f08a, show="headings", height=6)
        for c in cols_f08a:
            tree_f08a_ing.heading(c, text=c)
        tree_f08a_ing.column("Codigo Rubro", width=150)
        tree_f08a_ing.column("Nombre Rubro", width=180)
        tree_f08a_ing.column("Acto Administrativo", width=150)
        tree_f08a_ing.column("Fecha", width=90, anchor="center")
        tree_f08a_ing.column("Adicion", width=120, anchor="e")
        tree_f08a_ing.column("Reduccion", width=120, anchor="e")
        vsb_f08a_ing = ttk.Scrollbar(frame_f08a_ing, orient="vertical", command=tree_f08a_ing.yview)
        tree_f08a_ing.configure(yscrollcommand=vsb_f08a_ing.set)
        tree_f08a_ing.pack(side="left", fill="both", expand=True)
        vsb_f08a_ing.pack(side="right", fill="y")

        # --- Pestana F09: PAC ---
        frame_f09 = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_f09, text="F09 - PAC")
        cols_f09 = ("Codigo", "Nombre Rubro", "PAC Periodo", "Anticipos",
                    "Adiciones", "Reducciones", "Aplazamientos", "PAC Situado", "Pago")
        tree_f09 = ttk.Treeview(frame_f09, columns=cols_f09, show="headings", height=6)
        for c in cols_f09:
            tree_f09.heading(c, text=c)
        tree_f09.column("Codigo", width=130)
        tree_f09.column("Nombre Rubro", width=200)
        for col in ("PAC Periodo", "Anticipos", "Adiciones", "Reducciones", "Aplazamientos", "PAC Situado", "Pago"):
            tree_f09.column(col, width=95, anchor="e")
        vsb_f09 = ttk.Scrollbar(frame_f09, orient="vertical", command=tree_f09.yview)
        tree_f09.configure(yscrollcommand=vsb_f09.set)
        hsb_f09 = ttk.Scrollbar(frame_f09, orient="horizontal", command=tree_f09.xview)
        tree_f09.configure(xscrollcommand=hsb_f09.set)
        hsb_f09.pack(side="bottom", fill="x")
        tree_f09.pack(side="left", fill="both", expand=True)
        vsb_f09.pack(side="right", fill="y")

        # --- Pestana F13A: Contratacion ---
        frame_f13a = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_f13a, text="F13A - Contratacion")
        cols_f13a = ("No. Contrato", "Objeto", "Valor", "Contratista",
                     "NIT", "No. CDP", "Fecha CDP", "No. RP", "Fecha RP", "Pagos")
        tree_f13a = ttk.Treeview(frame_f13a, columns=cols_f13a, show="headings", height=6)
        for c in cols_f13a:
            tree_f13a.heading(c, text=c)
        tree_f13a.column("No. Contrato", width=85, anchor="center")
        tree_f13a.column("Objeto", width=200)
        tree_f13a.column("Valor", width=100, anchor="e")
        tree_f13a.column("Contratista", width=160)
        tree_f13a.column("NIT", width=90, anchor="center")
        tree_f13a.column("No. CDP", width=60, anchor="center")
        tree_f13a.column("Fecha CDP", width=85, anchor="center")
        tree_f13a.column("No. RP", width=55, anchor="center")
        tree_f13a.column("Fecha RP", width=85, anchor="center")
        tree_f13a.column("Pagos", width=100, anchor="e")
        vsb_f13a = ttk.Scrollbar(frame_f13a, orient="vertical", command=tree_f13a.yview)
        tree_f13a.configure(yscrollcommand=vsb_f13a.set)
        hsb_f13a = ttk.Scrollbar(frame_f13a, orient="horizontal", command=tree_f13a.xview)
        tree_f13a.configure(xscrollcommand=hsb_f13a.set)
        hsb_f13a.pack(side="bottom", fill="x")
        tree_f13a.pack(side="left", fill="both", expand=True)
        vsb_f13a.pack(side="right", fill="y")

        # Label de estado
        lbl_estado = tk.Label(dialog, text="", font=("Segoe UI", 9), bg=COLOR_WHITE, fg=COLOR_SUCCESS)
        lbl_estado.pack(anchor="w", padx=15)

        def cargar_preview():
            mes = int(mes_var.get())
            anio = anio_var.get()
            conn = db.get_connection()

            # ACUMULADO ANUAL: siempre desde enero 1 hasta fin del mes seleccionado
            fecha_ini = f"{anio}-01-01"
            if mes >= 12:
                fecha_fin = f"{int(anio) + 1}-01-01"
            else:
                fecha_fin = f"{anio}-{mes + 1:02d}-01"

            nombre_mes = nombres_mes.get(mes, str(mes))
            institucion = db.get_config("institucion") or "INSTITUCION EDUCATIVA"

            # --- F03: Bancos (acumulado enero a mes) ---
            tree_f03.delete(*tree_f03.get_children())
            cuentas_banco = conn.execute(
                "SELECT * FROM cuentas_bancarias WHERE estado='ACTIVA' ORDER BY banco, numero_cuenta"
            ).fetchall()
            if cuentas_banco:
                for cb in cuentas_banco:
                    egresos = conn.execute(
                        "SELECT COALESCE(SUM(valor),0) as t FROM pago "
                        "WHERE estado<>'ANULADO' AND fecha>=? AND fecha<? AND cuenta_bancaria_id=?",
                        (fecha_ini, fecha_fin, cb["id"])
                    ).fetchone()["t"]
                    ingresos = conn.execute(
                        "SELECT COALESCE(SUM(valor),0) as t FROM recaudo "
                        "WHERE estado<>'ANULADO' AND fecha>=? AND fecha<? AND cuenta_bancaria_id=?",
                        (fecha_ini, fecha_fin, cb["id"])
                    ).fetchone()["t"]
                    tree_f03.insert("", "end", values=(
                        cb["banco"], cb["numero_cuenta"],
                        cb["denominacion"] or institucion, "FSE",
                        "$0", f"${ingresos:,.0f}", f"${egresos:,.0f}",
                        "$0", "$0", f"${ingresos - egresos:,.0f}", f"${ingresos - egresos:,.0f}"
                    ))
            else:
                tree_f03.insert("", "end", values=(
                    "(Sin cuentas bancarias)", "", institucion, "FSE",
                    "$0", "$0", "$0", "$0", "$0", "$0", "$0"
                ))

            # --- F7B: Pagos (acumulado enero a mes) ---
            tree_f7b.delete(*tree_f7b.get_children())
            pagos = conn.execute(
                "SELECT p.fecha, p.codigo_rubro, p.valor, p.concepto, p.medio_pago, p.no_comprobante, "
                "r.numero as rp_numero, t.nombre as beneficiario, t.nit "
                "FROM pago p "
                "JOIN obligacion o ON p.obligacion_numero = o.numero "
                "JOIN rp r ON o.rp_numero = r.numero "
                "JOIN terceros t ON p.nit_tercero = t.nit "
                "WHERE p.estado <> 'ANULADO' AND p.fecha >= ? AND p.fecha < ? "
                "ORDER BY p.fecha, p.numero", (fecha_ini, fecha_fin)
            ).fetchall()
            total_pagos_f7b = 0
            for p in pagos:
                total_pagos_f7b += p["valor"]
                tree_f7b.insert("", "end", values=(
                    p["fecha"], p["codigo_rubro"], p["rp_numero"], p["medio_pago"],
                    p["beneficiario"], p["nit"], p["concepto"],
                    f"${p['valor']:,.0f}", f"${p['valor']:,.0f}"
                ))
            if not pagos:
                tree_f7b.insert("", "end", values=("(Sin pagos Ene-" + nombre_mes + ")", "", "", "", "", "", "", "", ""))

            # --- F08A: Modificaciones separadas GASTOS e INGRESOS (acumulado enero a mes) ---
            tree_f08a_gas.delete(*tree_f08a_gas.get_children())
            tree_f08a_ing.delete(*tree_f08a_ing.get_children())
            mods = conn.execute(
                "SELECT m.fecha, m.numero_acto, m.id, d.codigo_rubro, d.campo_afectado, "
                "d.valor, d.tipo_rubro, "
                "COALESCE(rg.cuenta, ri.cuenta) as nombre_rubro "
                "FROM modificaciones_presupuestales m "
                "JOIN detalle_modificacion d ON m.id = d.id_modificacion "
                "LEFT JOIN rubros_gastos rg ON d.codigo_rubro = rg.codigo AND d.tipo_rubro = 'GASTO' "
                "LEFT JOIN rubros_ingresos ri ON d.codigo_rubro = ri.codigo AND d.tipo_rubro = 'INGRESO' "
                "WHERE m.estado = 'ACTIVO' AND m.fecha >= ? AND m.fecha < ? "
                "ORDER BY d.tipo_rubro, m.fecha, d.codigo_rubro", (fecha_ini, fecha_fin)
            ).fetchall()
            n_mods_gas = 0
            n_mods_ing = 0
            for m in mods:
                campo = m["campo_afectado"].lower()
                adicion = m["valor"] if "adicion" in campo else 0
                reduccion = m["valor"] if "reduccion" in campo else 0
                fila = (
                    m["codigo_rubro"], m["nombre_rubro"] or "",
                    m["numero_acto"] or f"ACTO-{m['id']}",
                    m["fecha"], f"${adicion:,.0f}", f"${reduccion:,.0f}"
                )
                if m["tipo_rubro"] == "GASTO":
                    tree_f08a_gas.insert("", "end", values=fila)
                    n_mods_gas += 1
                else:
                    tree_f08a_ing.insert("", "end", values=fila)
                    n_mods_ing += 1
            if n_mods_gas == 0:
                tree_f08a_gas.insert("", "end", values=("(Sin modif. gastos Ene-" + nombre_mes + ")", "", "", "", "", ""))
            if n_mods_ing == 0:
                tree_f08a_ing.insert("", "end", values=("(Sin modif. ingresos Ene-" + nombre_mes + ")", "", "", "", "", ""))

            # --- F09: PAC (acumulado enero a mes) ---
            tree_f09.delete(*tree_f09.get_children())
            rubros = conn.execute(
                "SELECT codigo, cuenta, adiciones, reducciones FROM rubros_gastos WHERE es_hoja=1 ORDER BY codigo"
            ).fetchall()
            for r in rubros:
                # PAC acumulado: suma de meses 1 a mes rendido
                pac_acum = conn.execute(
                    "SELECT COALESCE(SUM(valor_programado),0) as vp FROM pac "
                    "WHERE codigo_rubro=? AND mes >= 1 AND mes <= ?",
                    (r["codigo"], mes)
                ).fetchone()["vp"]
                # Pagos acumulados: enero a mes rendido
                pago_acum = conn.execute(
                    "SELECT COALESCE(SUM(valor),0) as t FROM pago "
                    "WHERE codigo_rubro=? AND estado<>'ANULADO' AND fecha>=? AND fecha<?",
                    (r["codigo"], fecha_ini, fecha_fin)
                ).fetchone()["t"]
                tree_f09.insert("", "end", values=(
                    r["codigo"], r["cuenta"], f"${pac_acum:,.0f}", "$0",
                    f"${r['adiciones']:,.0f}", f"${r['reducciones']:,.0f}",
                    "$0", f"${pac_acum:,.0f}", f"${pago_acum:,.0f}"
                ))

            # --- F13A: Contratacion (acumulado enero a mes) ---
            tree_f13a.delete(*tree_f13a.get_children())
            rps = conn.execute(
                "SELECT r.numero as rp_numero, r.fecha as rp_fecha, r.codigo_rubro, "
                "r.valor as rp_valor, r.objeto, "
                "c.numero as cdp_numero, c.fecha as cdp_fecha, "
                "t.nombre as contratista, t.nit "
                "FROM rp r JOIN cdp c ON r.cdp_numero = c.numero "
                "JOIN terceros t ON r.nit_tercero = t.nit "
                "WHERE r.estado <> 'ANULADO' AND r.fecha >= ? AND r.fecha < ? "
                "ORDER BY r.numero", (fecha_ini, fecha_fin)
            ).fetchall()
            for rp in rps:
                pagos_rp = conn.execute(
                    "SELECT COALESCE(SUM(p.valor),0) as t FROM pago p "
                    "JOIN obligacion o ON p.obligacion_numero = o.numero "
                    "WHERE o.rp_numero = ? AND p.estado <> 'ANULADO' "
                    "AND p.fecha >= ? AND p.fecha < ?",
                    (rp["rp_numero"], fecha_ini, fecha_fin)
                ).fetchone()["t"]
                tree_f13a.insert("", "end", values=(
                    f"RP-{rp['rp_numero']}", rp["objeto"], f"${rp['rp_valor']:,.0f}",
                    rp["contratista"], rp["nit"],
                    rp["cdp_numero"], rp["cdp_fecha"],
                    rp["rp_numero"], rp["rp_fecha"], f"${pagos_rp:,.0f}"
                ))
            if not rps:
                tree_f13a.insert("", "end", values=("(Sin contratos Ene-" + nombre_mes + ")", "", "", "", "", "", "", "", "", ""))

            conn.close()

            n_pagos = len(pagos)
            n_rps = len(rps)
            lbl_estado.config(
                text=f"ACUMULADO Enero a {nombre_mes} {anio}  |  "
                     f"Pagos: {n_pagos}  |  Modif. Gastos: {n_mods_gas}  |  "
                     f"Modif. Ingresos: {n_mods_ing}  |  Contratos: {n_rps}  |  "
                     f"Rubros PAC: {len(rubros)}",
                fg=COLOR_SUCCESS
            )

        def exportar_csv():
            mes = int(mes_var.get())
            anio = anio_var.get()
            formato_sel = formato_var.get()

            destino = filedialog.askdirectory(
                title="Seleccionar carpeta para guardar archivos SIA",
                parent=dialog
            )
            if not destino:
                return

            # Importar el generador
            import importlib.util
            sia_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     "..", "SIA CONTRALORIA", "informes", "generar_informes_sia.py")
            # Usar ruta absoluta
            sia_path = os.path.normpath(os.path.join(
                os.path.dirname(db.DB_PATH), "SIA CONTRALORIA", "informes", "generar_informes_sia.py"
            ))

            try:
                spec = importlib.util.spec_from_file_location("generar_informes_sia", sia_path)
                sia_mod = importlib.util.module_from_spec(spec)
                # Sobrescribir directorio de salida
                spec.loader.exec_module(sia_mod)
                sia_mod.SALIDA_DIR = destino

                fmt_clave = None
                if formato_sel != "TODOS":
                    fmt_clave = formato_sel.split(" - ")[0].strip()

                resultados = sia_mod.generar_todos(mes=mes, formato=fmt_clave)

                archivos_ok = [r for r in resultados if r[4] == "OK"]
                lista_archivos = "\n".join([f"  - {os.path.basename(r[2])} ({r[3]} registros)" for r in archivos_ok])
                messagebox.showinfo("SIA Exportado",
                                    f"Se generaron {len(archivos_ok)} archivo(s) CSV:\n\n"
                                    f"{lista_archivos}\n\n"
                                    f"Carpeta: {destino}",
                                    parent=dialog)
                lbl_estado.config(text=f"Exportacion completada: {len(archivos_ok)} archivos en {destino}", fg=COLOR_SUCCESS)
            except Exception as e:
                messagebox.showerror("Error", f"Error al generar archivos SIA:\n{e}", parent=dialog)

        # Botones
        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)
        tk.Button(frame_btns, text="Vista Previa", bg=COLOR_ACCENT, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=cargar_preview).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Exportar CSV", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), width=14, command=exportar_csv).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Cerrar", font=("Segoe UI", 10), width=10,
                  command=dialog.destroy).pack(side="right", padx=5)

        # Cargar preview inicial
        cargar_preview()

    # ===================== CONFIGURAR MAPEO SIFSE =====================
    def _configurar_mapeo_sifse(self):
        """Pantalla de administracion del mapeo CCPET -> SIFSE."""
        dialog = tk.Toplevel(self)
        dialog.title("Configuracion Mapeo SIFSE")
        dialog.geometry("900x550")
        dialog.transient(self)
        dialog.grab_set()
        dialog.configure(bg=COLOR_WHITE)

        tk.Label(dialog, text="MAPEO CCPET -> SIFSE",
                 font=("Segoe UI", 14, "bold"), bg=COLOR_WHITE, fg=COLOR_PRIMARY).pack(pady=(10, 5))
        tk.Label(dialog, text="Doble-click en un rubro para editar su mapeo SIFSE",
                 font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555").pack()

        notebook = ttk.Notebook(dialog)
        notebook.pack(fill="both", expand=True, padx=15, pady=5)

        # --- Pestana Ingresos ---
        frame_ing = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_ing, text="Ingresos -> Fuente SIFSE")

        cols_ing = ("Codigo CCPET", "Cuenta", "Fuente SIFSE", "Descripcion SIFSE")
        tree_ing = ttk.Treeview(frame_ing, columns=cols_ing, show="headings", height=15)
        tree_ing.heading("Codigo CCPET", text="Codigo CCPET")
        tree_ing.heading("Cuenta", text="Cuenta")
        tree_ing.heading("Fuente SIFSE", text="Fuente SIFSE")
        tree_ing.heading("Descripcion SIFSE", text="Descripcion SIFSE")
        tree_ing.column("Codigo CCPET", width=180)
        tree_ing.column("Cuenta", width=250)
        tree_ing.column("Fuente SIFSE", width=90, anchor="center")
        tree_ing.column("Descripcion SIFSE", width=250)
        vsb_i = ttk.Scrollbar(frame_ing, orient="vertical", command=tree_ing.yview)
        tree_ing.configure(yscrollcommand=vsb_i.set)
        tree_ing.pack(side="left", fill="both", expand=True)
        vsb_i.pack(side="right", fill="y")

        tree_ing.tag_configure("sin_mapeo", background="#FFE0E0")
        tree_ing.tag_configure("mapeado", background=COLOR_WHITE)

        # --- Pestana Gastos ---
        frame_gas = tk.Frame(notebook, bg=COLOR_WHITE)
        notebook.add(frame_gas, text="Gastos -> Item SIFSE")

        cols_gas = ("Codigo CCPET", "Cuenta", "Item SIFSE", "Descripcion SIFSE")
        tree_gas = ttk.Treeview(frame_gas, columns=cols_gas, show="headings", height=15)
        tree_gas.heading("Codigo CCPET", text="Codigo CCPET")
        tree_gas.heading("Cuenta", text="Cuenta")
        tree_gas.heading("Item SIFSE", text="Item SIFSE")
        tree_gas.heading("Descripcion SIFSE", text="Descripcion SIFSE")
        tree_gas.column("Codigo CCPET", width=180)
        tree_gas.column("Cuenta", width=250)
        tree_gas.column("Item SIFSE", width=90, anchor="center")
        tree_gas.column("Descripcion SIFSE", width=250)
        vsb_g = ttk.Scrollbar(frame_gas, orient="vertical", command=tree_gas.yview)
        tree_gas.configure(yscrollcommand=vsb_g.set)
        tree_gas.pack(side="left", fill="both", expand=True)
        vsb_g.pack(side="right", fill="y")

        tree_gas.tag_configure("sin_mapeo", background="#FFE0E0")
        tree_gas.tag_configure("mapeado", background=COLOR_WHITE)

        def cargar_ingresos():
            tree_ing.delete(*tree_ing.get_children())
            mapeos = db.get_todos_mapeos_ingresos()
            for m in mapeos:
                fuente = m["sifse_fuente"] if m["sifse_fuente"] else ""
                desc = m["desc_sifse"] or ""
                tag = "mapeado" if fuente else "sin_mapeo"
                tree_ing.insert("", "end", values=(
                    m["codigo"], m["cuenta"], fuente, desc
                ), tags=(tag,))

        def cargar_gastos():
            tree_gas.delete(*tree_gas.get_children())
            mapeos = db.get_todos_mapeos_gastos()
            for m in mapeos:
                item = m["sifse_item"] if m["sifse_item"] else ""
                desc = m["desc_sifse"] or ""
                tag = "mapeado" if item else "sin_mapeo"
                tree_gas.insert("", "end", values=(
                    m["codigo"], m["cuenta"], item, desc
                ), tags=(tag,))

        def editar_mapeo_ingreso(event):
            sel = tree_ing.selection()
            if not sel:
                return
            vals = tree_ing.item(sel[0])["values"]
            codigo = str(vals[0])

            fuentes = db.get_catalogo_fuentes_sifse()
            opciones = ["(Sin asignar)"] + [f"{f['codigo']} - {f['descripcion']}" for f in fuentes]

            ed_dialog = tk.Toplevel(dialog)
            ed_dialog.title(f"Mapeo Ingreso: {codigo}")
            ed_dialog.geometry("450x150")
            ed_dialog.transient(dialog)
            ed_dialog.grab_set()
            ed_dialog.configure(bg=COLOR_WHITE)

            tk.Label(ed_dialog, text=f"Rubro: {codigo}", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).pack(pady=(10, 5))
            tk.Label(ed_dialog, text="Fuente SIFSE:", font=("Segoe UI", 9),
                     bg=COLOR_WHITE).pack()
            sel_var = tk.StringVar(value=opciones[0])
            # Preseleccionar valor actual
            actual = vals[2]
            if actual:
                for op in opciones:
                    if op.startswith(f"{actual} -"):
                        sel_var.set(op)
                        break
            cmb = ttk.Combobox(ed_dialog, textvariable=sel_var, values=opciones,
                               state="readonly", width=50)
            cmb.pack(padx=20, pady=5)

            def guardar_mapeo():
                sel_txt = sel_var.get()
                if sel_txt == "(Sin asignar)":
                    # Borrar mapeo
                    conn = db.get_connection()
                    conn.execute("DELETE FROM mapeo_sifse_ingresos WHERE codigo_rubro=?", (codigo,))
                    conn.commit()
                    conn.close()
                else:
                    fuente_cod = int(sel_txt.split(" - ")[0])
                    db.set_mapeo_sifse_ingreso(codigo, fuente_cod)
                ed_dialog.destroy()
                cargar_ingresos()

            tk.Button(ed_dialog, text="Guardar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), command=guardar_mapeo).pack(pady=5)

        def editar_mapeo_gasto(event):
            sel = tree_gas.selection()
            if not sel:
                return
            vals = tree_gas.item(sel[0])["values"]
            codigo = str(vals[0])

            items = db.get_catalogo_items_sifse()
            opciones = ["(Sin asignar)"] + [f"{i['codigo']} - {i['descripcion']}" for i in items]

            ed_dialog = tk.Toplevel(dialog)
            ed_dialog.title(f"Mapeo Gasto: {codigo}")
            ed_dialog.geometry("450x150")
            ed_dialog.transient(dialog)
            ed_dialog.grab_set()
            ed_dialog.configure(bg=COLOR_WHITE)

            tk.Label(ed_dialog, text=f"Rubro: {codigo}", font=("Segoe UI", 10, "bold"),
                     bg=COLOR_WHITE).pack(pady=(10, 5))
            tk.Label(ed_dialog, text="Item SIFSE:", font=("Segoe UI", 9),
                     bg=COLOR_WHITE).pack()
            sel_var = tk.StringVar(value=opciones[0])
            actual = vals[2]
            if actual:
                for op in opciones:
                    if op.startswith(f"{actual} -"):
                        sel_var.set(op)
                        break
            cmb = ttk.Combobox(ed_dialog, textvariable=sel_var, values=opciones,
                               state="readonly", width=50)
            cmb.pack(padx=20, pady=5)

            def guardar_mapeo():
                sel_txt = sel_var.get()
                if sel_txt == "(Sin asignar)":
                    conn = db.get_connection()
                    conn.execute("DELETE FROM mapeo_sifse_gastos WHERE codigo_rubro=?", (codigo,))
                    conn.commit()
                    conn.close()
                else:
                    item_cod = int(sel_txt.split(" - ")[0])
                    db.set_mapeo_sifse_gasto(codigo, item_cod)
                ed_dialog.destroy()
                cargar_gastos()

            tk.Button(ed_dialog, text="Guardar", bg=COLOR_SUCCESS, fg=COLOR_WHITE,
                      font=("Segoe UI", 10, "bold"), command=guardar_mapeo).pack(pady=5)

        tree_ing.bind("<Double-1>", editar_mapeo_ingreso)
        tree_gas.bind("<Double-1>", editar_mapeo_gasto)

        # Botones
        frame_btns = tk.Frame(dialog, bg=COLOR_WHITE)
        frame_btns.pack(fill="x", padx=15, pady=10)

        def restaurar_defecto():
            if messagebox.askyesno("Restaurar", "Esto borrara los mapeos actuales y aplicara "
                                   "los mapeos por defecto.\n\nDesea continuar?", parent=dialog):
                db.restaurar_mapeo_sifse_defecto()
                cargar_ingresos()
                cargar_gastos()
                messagebox.showinfo("Mapeo", "Mapeo por defecto restaurado", parent=dialog)

        tk.Button(frame_btns, text="Restaurar Mapeo por Defecto", bg=COLOR_WARNING, fg=COLOR_WHITE,
                  font=("Segoe UI", 10, "bold"), command=restaurar_defecto).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Cerrar", font=("Segoe UI", 10), width=10,
                  command=dialog.destroy).pack(side="right", padx=5)

        # Cargar datos
        cargar_ingresos()
        cargar_gastos()

    def _reiniciar_app(self):
        importlib.reload(db)
        # Recargar main.py para aplicar cambios en la interfaz
        import main as mod_main
        importlib.reload(mod_main)
        # Reconstruir metodos de la clase desde el modulo recargado
        for name in dir(mod_main.App):
            if callable(getattr(mod_main.App, name)) and not name.startswith("__"):
                try:
                    setattr(self.__class__, name, getattr(mod_main.App, name))
                except Exception:
                    pass
        self._mostrar_dashboard()
        messagebox.showinfo("Actualizado", "Sistema actualizado correctamente")


if __name__ == "__main__":
    app = App()
    app.mainloop()
